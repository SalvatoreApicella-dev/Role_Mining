"""Regression guard: role modeling export must produce a styled XLSX workbook."""

import os
import subprocess
import sys
import tempfile
import textwrap
import unittest
from pathlib import Path


class RoleModelingXlsxExportTest(unittest.TestCase):
    def test_export_endpoint_generates_xlsx_with_removed_rows_in_red(self):
        backend_dir = Path(__file__).resolve().parent
        script = textwrap.dedent(
            """
            from io import BytesIO

            from fastapi.testclient import TestClient
            from openpyxl import load_workbook

            from app.server import app

            client = TestClient(app)
            login = client.post(
                "/api/auth/login",
                json={"domain": "example.internal", "username": "admin", "password": "admin123"},
            )
            assert login.status_code == 200, login.text
            token = login.json()["access_token"]
            headers = {"Authorization": f"Bearer {token}"}

            response = client.post(
                "/api/role-modeling/export/xlsx",
                headers=headers,
                json={
                    "filename": "role_modeling_test.xlsx",
                    "sheet_name": "New Model",
                    "rows": [
                        {"business_role": "Finance AP", "role": "SAP_AP", "status": "active", "highlight": "", "note": ""},
                        {"business_role": "Legacy", "role": "LEGACY_X", "status": "removed", "highlight": "red", "note": "Retired"},
                    ],
                },
            )
            assert response.status_code == 200, response.text
            assert "role_modeling_test.xlsx" in response.headers.get("content-disposition", "")
            assert response.headers.get("content-type", "").startswith(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            workbook = load_workbook(BytesIO(response.content))
            sheet = workbook.active
            assert sheet.title == "New Model"
            assert sheet["A1"].value == "business role"
            assert sheet["B1"].value == "ruolo"
            assert sheet["A2"].value == "Finance AP"
            assert sheet["B2"].value == "SAP_AP"
            assert sheet["A3"].value == "Legacy"
            assert sheet["B3"].value == "LEGACY_X"
            assert sheet["A3"].fill.fgColor.rgb and "FDE9E7" in sheet["A3"].fill.fgColor.rgb.upper()
            assert sheet["A3"].font.color.rgb and "9C0006" in sheet["A3"].font.color.rgb.upper()
            """
        )

        with tempfile.TemporaryDirectory() as tmpdir:
            env = os.environ.copy()
            env["PYTHONPATH"] = str(backend_dir)
            env["JWT_SECRET"] = "role-modeling-xlsx-test-secret-32"
            python_bin = backend_dir / ".venv" / "bin" / "python"
            if not python_bin.exists():
                python_bin = Path(sys.executable)
            result = subprocess.run(
                [str(python_bin), "-c", script],
                cwd=tmpdir,
                env=env,
                text=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                check=False,
                timeout=90,
            )

        self.assertEqual(
            result.returncode,
            0,
            msg=f"stdout:\n{result.stdout}\nstderr:\n{result.stderr}",
        )

    def test_sod_matrix_export_generates_xlsx_with_severity_row_colors(self):
        backend_dir = Path(__file__).resolve().parent
        script = textwrap.dedent(
            """
            from io import BytesIO

            from fastapi.testclient import TestClient
            from openpyxl import load_workbook

            from app.server import app

            client = TestClient(app)
            login = client.post(
                "/api/auth/login",
                json={"domain": "example.internal", "username": "admin", "password": "admin123"},
            )
            assert login.status_code == 200, login.text
            token = login.json()["access_token"]
            headers = {"Authorization": f"Bearer {token}"}

            response = client.post(
                "/api/sod-matrix/export/xlsx",
                headers=headers,
                json={
                    "filename": "sod_matrix_test.xlsx",
                    "rows": [
                        {
                            "groupA": "SAP_AP",
                            "groupB": "SAP_PAY",
                            "users": 4,
                            "severity": "high",
                            "recommendation": "Separare autorizzazioni critiche",
                        },
                        {
                            "groupA": "SAP_MM",
                            "groupB": "SAP_FI",
                            "users": 2,
                            "severity": "medium",
                            "recommendation": "Rivedere assegnazioni",
                        },
                    ],
                },
            )
            assert response.status_code == 200, response.text
            assert "sod_matrix_test.xlsx" in response.headers.get("content-disposition", "")
            assert response.headers.get("content-type", "").startswith(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            workbook = load_workbook(BytesIO(response.content))
            sheet = workbook.active
            assert sheet.title == "SoD Matrix Alerts"
            assert [sheet["A1"].value, sheet["B1"].value, sheet["C1"].value, sheet["D1"].value, sheet["E1"].value] == [
                "Ruolo",
                "Ruolo in Conflitto",
                "Utenti",
                "Severità",
                "Raccomandazione",
            ]
            assert sheet["A2"].value == "SAP_AP"
            assert sheet["B2"].value == "SAP_PAY"
            assert sheet["C2"].value == 4
            assert sheet["D2"].value == "Alta"
            assert sheet["A2"].fill.fgColor.rgb and "FDE9E7" in sheet["A2"].fill.fgColor.rgb.upper()
            assert sheet["A3"].fill.fgColor.rgb and "FFF2CC" in sheet["A3"].fill.fgColor.rgb.upper()

            imported = client.post(
                "/api/sod-matrix/import/xlsx",
                headers=headers,
                files={
                    "file": (
                        "sod_matrix_test.xlsx",
                        response.content,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            assert imported.status_code == 200, imported.text
            imported_rows = imported.json()["rows"]
            assert imported_rows[0]["groupA"] == "SAP_AP"
            assert imported_rows[0]["groupB"] == "SAP_PAY"
            assert imported_rows[0]["users"] == 4
            assert imported_rows[0]["severity"] == "high"
            assert imported_rows[1]["severity"] == "medium"

            attestation_wb = load_workbook(BytesIO(response.content))
            attestation_ws = attestation_wb.active
            attestation_ws.delete_rows(1, attestation_ws.max_row)
            attestation_ws.append([
                "Status",
                "MemberType",
                "Ident_AttestationPolicy",
                "Member",
                "DisplayName",
                "CaseDisplayName",
                "FullPath",
                "ApprovedBy",
                "ApprovedOn",
                "CampaignDate",
                "NextReminder",
                "Approver",
                "DecisionType",
                "Comment",
                "Utente",
            ])
            for index in range(1, 13):
                attestation_ws.append([
                    "Denied" if index <= 6 else "Approved",
                    f"User {index} (USER{index})",
                    "ITA role membership attestation (Quarterly)",
                    "ARCADIA PERMISSION Ngas Media_User",
                    "ARCADIA PERMISSION Ngas Media_User",
                    f"ARCADIA PERMISSION Ngas Media_User - User {index}",
                    None,
                    "#LDS#Automatic approval." if index <= 6 else "Owner",
                    "17/04/2026",
                    "06/04/2026",
                    None,
                    None,
                    "Dismiss" if index <= 6 else "Grant",
                    "Requested system role assignment was successfully removed." if index <= 6 else "",
                    f"USER{index}",
                ])
                attestation_ws.append([
                    "Approved",
                    f"User {index} (USER{index})",
                    "ITA role membership attestation (Quarterly)",
                    "ARCADIA CRM BB ASSURANCE 150 Llama",
                    "ARCADIA CRM BB ASSURANCE 150 Llama",
                    f"ARCADIA CRM BB ASSURANCE 150 Llama - User {index}",
                    None,
                    "Owner",
                    "17/04/2026",
                    "06/04/2026",
                    None,
                    None,
                    "Grant",
                    "",
                    f"USER{index}",
                ])
            attestation_bytes = BytesIO()
            attestation_wb.save(attestation_bytes)
            attestation_bytes.seek(0)

            analyzed = client.post(
                "/api/sod-matrix/import/xlsx",
                headers=headers,
                files={
                    "file": (
                        "arcadia_attestation.xlsx",
                        attestation_bytes.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            assert analyzed.status_code == 200, analyzed.text
            analyzed_payload = analyzed.json()
            assert analyzed_payload["mode"] == "attestation_analysis"
            analyzed_rows = analyzed_payload["rows"]
            assert analyzed_rows
            assert analyzed_rows[0]["severity"] == "high"
            assert analyzed_rows[0]["denialUsers"] == 6
            assert "Verificare subito" in analyzed_rows[0]["recommendation"]
            assert "responsabile applicativo" in analyzed_rows[0]["recommendation"]
            assert "controllo compensativo" in analyzed_rows[0]["recommendation"]
            """
        )

        with tempfile.TemporaryDirectory() as tmpdir:
            env = os.environ.copy()
            env["PYTHONPATH"] = str(backend_dir)
            env["JWT_SECRET"] = "role-modeling-xlsx-test-secret-32"
            python_bin = backend_dir / ".venv" / "bin" / "python"
            if not python_bin.exists():
                python_bin = Path(sys.executable)
            result = subprocess.run(
                [str(python_bin), "-c", script],
                cwd=tmpdir,
                env=env,
                text=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                check=False,
                timeout=90,
            )

        self.assertEqual(
            result.returncode,
            0,
            msg=f"stdout:\n{result.stdout}\nstderr:\n{result.stderr}",
        )


if __name__ == "__main__":
    unittest.main()
