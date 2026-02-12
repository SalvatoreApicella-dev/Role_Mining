import os
import time
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Tuple

import jwt
import numpy as np
from fastapi import Depends, FastAPI, HTTPException, status
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from pydantic import BaseModel, Field
from sklearn.cluster import AgglomerativeClustering, MiniBatchKMeans
from sklearn.decomposition import TruncatedSVD
from fastapi import UploadFile, File, BackgroundTasks
import csv, io, re
try:
    from ldap3 import ALL, NTLM, SIMPLE, Connection, Server, Tls, NONE
    import ssl
except Exception:
    Connection = None  # type: ignore


APP_TITLE = "Role Mining API"
import secrets
# Use a persistent key for dev to avoid invalidating tokens on restart
JWT_SECRET = os.getenv("JWT_SECRET") or "dev_secret_key_persistent_change_in_prod"
APP_LOGIN_USER = os.getenv("APP_LOGIN_USER", "admin")
APP_LOGIN_PASS = os.getenv("APP_LOGIN_PASS", "admin123")
JWT_EXPIRE_MINUTES = int(os.getenv("JWT_EXPIRE_MINUTES", "240"))
MOCK_AD = os.getenv("MOCK_AD", "0") == "1"
LDAP_FETCH_ALL_ATTRIBUTES = os.getenv("LDAP_FETCH_ALL_ATTRIBUTES", "0") == "1"
LDAP_PAGE_SIZE = max(100, int(os.getenv("LDAP_PAGE_SIZE", "1000")))
LDAP_SEARCH_TIME_LIMIT = max(10, int(os.getenv("LDAP_SEARCH_TIME_LIMIT", "60")))
LDAP_EXTRA_ATTRIBUTES = [
    a.strip() for a in os.getenv("LDAP_EXTRA_ATTRIBUTES", "").split(",") if a.strip()
]
LDAP_BASE_ATTRIBUTES = [
    "sAMAccountName",
    "displayName",
    "memberOf",
    "department",
    "lastLogonTimestamp",
    "lastLogon",
    "employeeType",
    "distinguishedName",
    "mail",
    "userPrincipalName",
]



from openpyxl import load_workbook
from collections import defaultdict, Counter

# ML Engine import
from ml_engine import get_ml_engine, ACCOUNT_TYPES
ml_engine = get_ml_engine(data_dir="./ml_data")


# =============================================================================
# Response Cache (Solution 1: In-Memory TTL Cache)
# =============================================================================
class ResponseCache:
    """Simple in-memory cache with TTL (time-to-live) support."""
    
    def __init__(self):
        self._cache: Dict[str, tuple] = {}  # key -> (value, expire_time)
        self._hits = 0
        self._misses = 0
    
    def get(self, key: str) -> Optional[Any]:
        """Get cached value if exists and not expired."""
        if key in self._cache:
            value, expire_time = self._cache[key]
            if time.time() < expire_time:
                self._hits += 1
                return value
            else:
                del self._cache[key]
        self._misses += 1
        return None
    
    def set(self, key: str, value: Any, ttl_seconds: float = 30.0) -> None:
        """Store value with TTL."""
        self._cache[key] = (value, time.time() + ttl_seconds)
    
    def invalidate(self, key: str = None) -> None:
        """Invalidate specific key or all keys."""
        if key:
            self._cache.pop(key, None)
        else:
            self._cache.clear()

    def invalidate_prefix(self, prefix: str) -> None:
        """Invalidate all keys by prefix."""
        if not prefix:
            return
        keys = [k for k in self._cache.keys() if k.startswith(prefix)]
        for key in keys:
            self._cache.pop(key, None)
    
    def stats(self) -> Dict[str, Any]:
        """Return cache statistics."""
        total = self._hits + self._misses
        hit_rate = (self._hits / total * 100) if total > 0 else 0
        return {
            "hits": self._hits,
            "misses": self._misses,
            "hit_rate": f"{hit_rate:.1f}%",
            "cached_keys": len(self._cache)
        }

# Global cache instance
RESPONSE_CACHE = ResponseCache()

# Cache TTL settings (seconds)
CACHE_TTL_MINING = 60.0      # Mining results (invalidated on new mining)
CACHE_TTL_KPI = 60.0         # KPI data
CACHE_TTL_USERS = 30.0       # User list
CACHE_TTL_ROLES = 30.0       # Business roles

# Cache dell’ultima esecuzione (serve per drilldown)

def invalidate_hot_caches(*, users: bool = False, roles: bool = False, kpi: bool = False, mining: bool = False) -> None:
    if users:
        RESPONSE_CACHE.invalidate_prefix("users_")
        RESPONSE_CACHE.invalidate("ad_groups")
    if roles:
        RESPONSE_CACHE.invalidate("businessroles")
        RESPONSE_CACHE.invalidate_prefix("role_meta_")
    if kpi:
        RESPONSE_CACHE.invalidate("kpi")
    if mining:
        RESPONSE_CACHE.invalidate_prefix("rolemining_last_")

BROAD_MARKERS = ['all','tutti','tutte','full','global','everyone','any','anyone','everybody']

def _tokens(s: str) -> list[str]:
    s = (s or '').lower()
  # Treat underscore/dash like separators too, so VPN_ALL -> ["vpn","all"]
    s = s.replace('_', ' ').replace('-', ' ')
    s = re.sub(r'[^a-z0-9]', ' ', s)
    return [t for t in s.split() if t]

def _family_key(role_name: str) -> str:
    toks = _tokens(role_name)
    return toks[0] if toks else ""

def _is_broad(role_name: str) -> bool:
    toks = set(_tokens(role_name))
    return any(m in toks for m in BROAD_MARKERS)

def _matrix_user_roles(matrix: dict, username: str) -> set[str]:
    row = (matrix or {}).get(username) or {}
    return {r for r, v in row.items() if int(v) == 1}

def build_overprivileged_items(matrix: dict, top_pct: float = 10.0) -> dict:
    users = state.get("last_extract", {}).get("users") or []
    br_by_user = {u.get("username"): u.get("businessRole", "Unassigned") for u in users if u.get("username")}
    role_meta = state.get("role_meta", {}) or {}

    items = []
    for uname, row in (matrix or {}).items():
        actual = sorted([r for r, v in (row or {}).items() if int(v) == 1])

        br = br_by_user.get(uname, "Unassigned")
        expected = set((role_meta.get(br, {}) or {}).get("groups", []) or [])

        excess = sorted(set(actual) - expected)
        if not excess:
            continue  # SOLO over

        items.append({
        "username": uname,
        "groups": actual,          # tutti
        "overGroups": excess,      # solo eccesso
        "groupCount": len(excess), # numero eccessid
        
        "isOverprivileged": True,
        })
    items.sort(key=lambda r: r["groupCount"], reverse=True)
    return {"threshold": 1, "items": items}

import random

# In-memory cache for BRDB (not persisted to storage.json to keep it lean)
BRDB_CACHE: Dict[str, Dict[str, Any]] = {}

def generate_unique_color(existing_colors):
    """Genera un colore RRGGBB univoco rispetto a quelli esistenti"""
    existing = set(existing_colors)
    attempts = 0
    while attempts < 100:
        r = random.randint(80, 255)   # evita colori troppo scuri
        g = random.randint(80, 255)
        b = random.randint(80, 255)
        color = f"{r:02x}{g:02x}{b:02x}".upper()
        if color not in existing:
            return color
        attempts += 1
    return "6AA6FF"  # fallback

# Lista colori predefiniti (per i primi 12 ruoli)
PREDEFINED_COLORS = [
    "00B4FF", "FF9F1C", "71FFB2", "FF6B6B", "9B59B6", "3498DB",
    "F39C12", "1ABC9C", "E74C3C", "9B59B6", "3498DB", "F1C40F"
]

def predict_redundant(broad_role: str, specific_role: str, family: str, user_groups: set[str]) -> float:
  bt = set(_tokens(broad_role))
  st = set(_tokens(specific_role))
  if not bt or not st:
    return 0.0

  # Similarità nome ruolo (Jaccard su token)
  j_role = len(bt & st) / max(1, len(bt | st))

  # Broadness
  b_is_broad = 1.0 if _is_broad(broad_role) else 0.0

  # Overlap token specific vs gruppi reali utente
  st2 = [t for t in st if len(t) >= 3]
  hits = 0
  for g in user_groups:
    gl = (g or "").lower()
    if any(t in gl for t in st2):
      hits += 1
  g_overlap = hits / max(1, len(user_groups))

  p = 0.55 * j_role + 0.30 * b_is_broad + 0.15 * g_overlap
  return float(max(0.0, min(1.0, p)))



def build_ai_detection_items(matrix: dict) -> list[dict]:
    items = []
    for uname, row in (matrix or {}).items():
        roles = [r for r, v in (row or {}).items() if int(v) == 1]
        user_groups = _matrix_user_roles(matrix, uname)

        fam = defaultdict(list)
        for r in roles:
            k = _family_key(r)
            if k:
                fam[k].append(r)

        for family, rs in fam.items():
            broad = [r for r in rs if _is_broad(r)]
            specific = [r for r in rs if not _is_broad(r)]
            
            if not broad or not specific:
                continue

            redundant = []
            kept = set(specific)

            for b in broad:
                probs = [predict_redundant(b, s, family, user_groups) for s in specific]
                p = max(probs) if probs else 0.0
                if p >= 0.50:
                    redundant.append(b)

            # Fallback: se hai broad + specific ma l'algoritmo non è sicuro, 
            # mostrali comunque come "sospetti" invece di nasconderli
            if not redundant and broad:
                 redundant = list(broad)
            
            if redundant:
                items.append({
                    "username": uname,
                    "family": family,
                    "redundantRoles": sorted(set(redundant)),
                    "keptRoles": sorted(list(kept)),
                    "redundantCount": len(set(redundant)),
                })

    # Ordina per chi ne ha di più
    items.sort(key=lambda x: x["redundantCount"], reverse=True)
    return items


# ---------------------------------------------------------------------------
# Smart AI Detection (On-Demand) – Peer / Department / AccountType analysis
# ---------------------------------------------------------------------------
ADMIN_MARKERS = {'admin', 'administrator', 'superuser', 'root', 'owner',
                 'elevated', 'privileged', 'sudo', 'godmode'}
SERVICE_ACCOUNT_MARKERS = {"svc", "service", "serviceaccount", "service-account"}
RESTRICTED_ACCOUNT_TYPES = {
    'BlueCollar':  ADMIN_MARKERS,
    'External':    ADMIN_MARKERS | {'internal', 'staff', 'employee', 'hr',
                                     'payroll', 'finance', 'accounting'},
    'Contractor':  ADMIN_MARKERS | {'internal', 'staff'},
    # LLM guardrail: service/svc accounts should not carry administrative access.
    'Service':     ADMIN_MARKERS,
}
PEER_ANOMALY_THRESHOLD = 0.10   # flag if < 10% of peers have the group
DEPT_ANOMALY_THRESHOLD = 0.05   # flag if < 5% of dept members have group


def load_knowledge_base() -> dict:
    """Load the synthetic knowledge base (LLM-instructed rules)."""
    try:
        import json
        path = "backend/ml_data/knowledge_base.json"
        if not os.path.exists(path):
            return {}
        with open(path, "r") as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading KB: {e}")
        return {}


def _build_freq_tables(users: list) -> tuple:
    """Pre-compute frequency tables in a SINGLE pass over the user list.
    Returns (role_freq, dept_freq):
      role_freq  = { businessRole: { group: ratio_0_to_1 } }
      dept_freq  = { department:   { group: ratio_0_to_1 } }
    """
    from collections import Counter

    role_counts: dict[str, Counter] = {}   # role -> Counter of groups
    role_totals: dict[str, int] = {}       # role -> num users in role
    dept_counts: dict[str, Counter] = {}
    dept_totals: dict[str, int] = {}

    for u in (users or []):
        if u.get("excluded"):
            continue
        groups = u.get("groups") or []
        br = (u.get("businessRole") or "Unassigned").strip()
        dept = (u.get("department") or "Unknown").strip()

        role_totals[br] = role_totals.get(br, 0) + 1
        dept_totals[dept] = dept_totals.get(dept, 0) + 1

        if br not in role_counts:
            role_counts[br] = Counter()
        if dept not in dept_counts:
            dept_counts[dept] = Counter()

        for g in groups:
            role_counts[br][g] += 1
            dept_counts[dept][g] += 1

    # Convert counts to ratios
    role_freq: dict[str, dict[str, float]] = {}
    for br, ctr in role_counts.items():
        total = max(1, role_totals.get(br, 1))
        role_freq[br] = {g: cnt / total for g, cnt in ctr.items()}

    dept_freq: dict[str, dict[str, float]] = {}
    for dept, ctr in dept_counts.items():
        total = max(1, dept_totals.get(dept, 1))
        dept_freq[dept] = {g: cnt / total for g, cnt in ctr.items()}

    return role_freq, dept_freq


def _check_type_violation(group: str, account_type: str) -> str:
    """Return violation reason or empty string. O(1) per group."""
    if not account_type:
        return ""
    restricted = RESTRICTED_ACCOUNT_TYPES.get(account_type)
    if not restricted:
        return ""
    toks = set(_tokens(group))
    matched = toks & restricted
    if matched:
        return f"Policy: {account_type} should not have '{group}'"
    return ""


def _is_service_identity(u_data: dict) -> bool:
    acct_type = str(u_data.get("accountType") or "").strip().lower()
    if acct_type == "service":
        return True
    dn = str(u_data.get("displayName") or "").strip().lower()
    uname = str(u_data.get("username") or "").strip().lower()
    dn_toks = set(_tokens(dn))
    un_toks = set(_tokens(uname))
    if dn.startswith("svc") or uname.startswith("svc"):
        return True
    return bool((dn_toks | un_toks) & SERVICE_ACCOUNT_MARKERS)


def _is_admin_group_name(group: str) -> bool:
    toks = set(_tokens(group))
    return bool(toks & ADMIN_MARKERS)


def run_smart_ai_detection(users: list, matrix: dict) -> dict:
    """On-demand smart redundancy detection.
    Combines:
    1. Statistical signal (Peer / Dept freq)
    2. Knowledge Base rules (Redundancy, Policy, Norms)
    """
    role_freq, dept_freq = _build_freq_tables(users)
    kb = load_knowledge_base()
    
    kb_redundancy = kb.get("redundancy_rules", {})
    kb_policies = kb.get("account_type_policies", {})
    kb_role_defs = kb.get("role_definitions", {})
    kb_dept_norms = kb.get("department_norms", {})
    # Inject LLM guardrail for Service/SVC accounts.
    service_forbidden = set(kb_policies.get("Service", []) or [])
    service_forbidden.update(list(ADMIN_MARKERS))
    kb_policies["Service"] = sorted(service_forbidden)

    # Index users by username for O(1) lookup
    user_by_name: dict[str, dict] = {}
    for u in (users or []):
        if not u.get("excluded"):
            uname = u.get("username")
            if uname:
                user_by_name[uname] = u

    items: list[dict] = []
    total_anomalies = 0
    total_assignments = 0
    users_with_anomaly = 0

    for uname, row in (matrix or {}).items():
        groups = [g for g, v in (row or {}).items() if int(v) == 1]
        groups_set = set(groups)  # fast lookup
        total_assignments += len(groups)

        u_data = user_by_name.get(uname, {})
        br = (u_data.get("businessRole") or "Unassigned").strip()
        dept = (u_data.get("department") or "Unknown").strip()
        acct_type = (u_data.get("accountType") or "").strip()
        is_service_identity = _is_service_identity(u_data)

        peer_freqs = role_freq.get(br, {})
        dept_freqs = dept_freq.get(dept, {})

        # KB Norms for this user
        # Match "Engineer" in "Software Engineer III" etc.
        kb_role_norm = next((norm for k, norm in kb_role_defs.items() if k in br), [])
        kb_dept_norm = next((norm for k, norm in kb_dept_norms.items() if k in dept), [])

        anomalies: list[dict] = []
        for g in groups:
            reasons: list[str] = []
            confidence = 0.0

            # --- 1. KNOWLEDGE BASE CHECKS (High Confidence) ---
            
            # A) Known Redundancy Rule (Explicit and Heuristic)
            hierarchy_rules = kb.get("hierarchy_patterns", [])
            
            for other_g in groups:
                if other_g == g: continue
                
                # A1) Explicit KB Rule
                redundant_list = kb_redundancy.get(other_g, [])
                if g in redundant_list:
                    reasons.append(f"Redundant: Superceded by '{other_g}' (Explicit KB Rule)")
                    confidence = 1.0
                    break
                
                # A2) Heuristic Hierarchy Rule (Pattern-based)
                # e.g. Azure_1 vs Azure_All, App_Read vs App_Admin
                for pattern in hierarchy_rules:
                    root = pattern.get("root", "")
                    supersedes = pattern.get("supersedes", [])
                    
                    if root in other_g:
                        # Find the base name without the root suffix
                        base_other = other_g.replace(root, "")
                        for s in supersedes:
                            if s in g:
                                base_g = g.replace(s, "")
                                # If they share the same base name (e.g. "Azure"), it's likely a hierarchy violation
                                if base_other == base_g or base_other in g:
                                    reasons.append(f"Least Privilege: '{g}' is likely redundant given '{other_g}' ({root} hierarchy)")
                                    confidence = 0.9
                                    break
                        if reasons: break
                if reasons: break
            
            # B) Account Type Policy
            # Check against KB policies + hardcoded fallback
            forbidden = kb_policies.get(acct_type, [])
            if any(token in g for token in forbidden):
                 # re-check hardcoded function too just in case
                 pass

            # B2) Service/SVC + admin group guardrail (high confidence).
            if is_service_identity and _is_admin_group_name(g):
                reasons.append("LLM Policy: Service/SVC account with administrative group is high-risk")
                confidence = max(confidence, 0.99)
            
            # (Merged policy check logic)
            violation = _check_type_violation(g, acct_type)
            if violation:
                reasons.append(violation)
                confidence = max(confidence, 0.95)
            
            # C) KB Role/Dept Norms (finding out-of-pattern items)
            # If we generally know what this role has, and 'g' is NOT in it -> anomaly?
            # Careful: KB is generic, don't be too strict.
            # Only flag if it confirms a statistical anomaly.

            # --- 2. STATISTICAL CHECKS (Medium/Low Confidence) ---

            # Peer frequency check
            pf = peer_freqs.get(g, 0.0)
            is_stat_anomaly = False
            
            if pf < PEER_ANOMALY_THRESHOLD:
                reasons.append(f"Peer: only {pf * 100:.0f}% of '{br}' have this")
                confidence = max(confidence, 1.0 - pf)
                is_stat_anomaly = True

            # Department frequency check
            df = dept_freqs.get(g, 0.0)
            if df < DEPT_ANOMALY_THRESHOLD:
                # If KB says this dept SHOULD have it, ignore the anomaly
                if not any(k in g for k in kb_dept_norm):
                    reasons.append(f"Dept: only {df * 100:.0f}% of '{dept}' have this")
                    confidence = max(confidence, 1.0 - df)
                    is_stat_anomaly = True

            if reasons:
                anomalies.append({
                    "group": g,
                    "reasons": reasons,
                    "confidence": round(confidence, 2),
                    "peerFreq": round(pf, 4),
                    "deptFreq": round(df, 4),
                })


        if anomalies:
            users_with_anomaly += 1
            total_anomalies += len(anomalies)
            anomalies.sort(key=lambda a: a["confidence"], reverse=True)
            items.append({
                "username": uname,
                "displayName": u_data.get("displayName") or uname,
                "businessRole": br,
                "department": dept,
                "accountType": acct_type,
                "anomalyCount": len(anomalies),
                "anomalies": anomalies,
            })

    items.sort(key=lambda x: x["anomalyCount"], reverse=True)

    total_users = len(matrix or {})
    pct = (users_with_anomaly / max(1, total_users)) * 100.0
    print(f"[AI Detection] users_with_anomaly={users_with_anomaly}, total_users={total_users}, pct={pct:.2f}%", flush=True)

    return {
        "status": "ready",
        "items": items,
        "stats": {
            "aiDetection": round(pct, 2),
            "totalAnomalies": total_anomalies,
            "totalAssignments": total_assignments,
            "usersWithAnomaly": users_with_anomaly,
            "totalUsersScanned": len(matrix or {}),
        },
        "ts": datetime.now(timezone.utc).isoformat(),
    }


def build_cluster_quality_items(clusters: list, matrix: dict) -> list[dict]:
    """
    Drilldown “Cluster Quality”: ordina gli elementi che peggiorano la qualità del cluster.
    Assunzione minima: ogni cluster ha 'users' (lista username) e 'roleGroups' (lista gruppi del cluster).
    """
    out = []
    for c_idx, c in enumerate(clusters or []):
        users = c.get("users") or c.get("members") or []
        role_groups = set(c.get("roleGroups") or [])
        if not users or not role_groups:
            continue

        user_items = []
        for u in users:
            u_roles = _matrix_user_roles(matrix, u)
            missing = sorted(list(role_groups - u_roles))
            extra = sorted(list(u_roles - role_groups))
            denom = max(1, len(role_groups) + len(u_roles))
            distance = round((len(missing) + len(extra)) / denom, 4)  # più alto = più “sporco”

            user_items.append({
                "username": u,
                "distance": distance,
                "missingFromRole": missing,
                "extraVsRole": extra,
            })

        user_items.sort(key=lambda r: r["distance"], reverse=True)
        avg_distance = round(sum(x["distance"] for x in user_items) / max(1, len(user_items)), 4)

        out.append({
            "clusterIndex": c_idx,
            "clusterName": c.get("name") or f"Cluster {c_idx}",
            "avgDistance": avg_distance,
            "worstUsers": user_items[:50],  # top 50 driver
        })

    out.sort(key=lambda r: r["avgDistance"], reverse=True)
    return out




# ----------------------------
# Persistent Storage (replaces in-memory state)
# ----------------------------
from app.db.storage import get_store, init_default_state

# Initialize persistent storage
state = get_store()
init_default_state()
REQUIRED_DUPLICATE_ORDER = ["last_login", "groups_count", "dept_group_correlation", "has_department"]
MODEL_QUALITY_PRESETS: Dict[str, Dict[str, float]] = {
    "banking": {
        "role_entropy": 0.06,
        "template_coverage": 0.08,
        "noise_ratio": 0.09,
        "ambiguity": 0.07,
        "temporal_drift": 0.06,
        "matrix_density": 0.05,
        "orphan_weighted": 0.10,
        "overprivileged": 0.13,
        "stale_access": 0.12,
        "policy_violation": 0.14,
        "manual_override": 0.04,
        "generalization": 0.06,
    },
    "manufacturing": {
        "role_entropy": 0.08,
        "template_coverage": 0.11,
        "noise_ratio": 0.10,
        "ambiguity": 0.07,
        "temporal_drift": 0.09,
        "matrix_density": 0.06,
        "orphan_weighted": 0.09,
        "overprivileged": 0.10,
        "stale_access": 0.09,
        "policy_violation": 0.08,
        "manual_override": 0.06,
        "generalization": 0.07,
    },
    "retail": {
        "role_entropy": 0.08,
        "template_coverage": 0.10,
        "noise_ratio": 0.10,
        "ambiguity": 0.08,
        "temporal_drift": 0.09,
        "matrix_density": 0.07,
        "orphan_weighted": 0.08,
        "overprivileged": 0.10,
        "stale_access": 0.10,
        "policy_violation": 0.08,
        "manual_override": 0.05,
        "generalization": 0.07,
    },
}
state.setdefault("dq_rules", {})
state["dq_rules"]["duplicate_resolution_order"] = REQUIRED_DUPLICATE_ORDER.copy()
state.setdefault("dq_model_preset", "manufacturing")
state.setdefault("dq_model_weights", MODEL_QUALITY_PRESETS.get(state.get("dq_model_preset"), MODEL_QUALITY_PRESETS["manufacturing"]))


def get_active_model_weights() -> Dict[str, float]:
    preset = (state.get("dq_model_preset") or "manufacturing").strip().lower()
    base = MODEL_QUALITY_PRESETS.get(preset) or MODEL_QUALITY_PRESETS["manufacturing"]
    custom = state.get("dq_model_weights") or {}
    out = dict(base)
    for k, v in custom.items():
        try:
            out[str(k)] = float(v)
        except Exception:
            continue
    return out


def apply_business_roles(users: List[Dict[str, Any]]) -> None:
    """Add businessRole field to each user based on state mapping.
    
    CRITICAL: Only assigns from mapping if:
    1. User has no existing valid BR, OR
    2. Mapping has a valid (non-Unassigned) value for this user
    
    This preserves CSV-assigned BRs while allowing department-based assignment.
    """
    m = state.get("user_business_role", {})
    for u in users:
        existing_br = (u.get("businessRole") or "").strip()
        mapped_br = m.get(u.get("username"), "")
        
        # Priority: existing valid BR > mapped valid BR > "Unassigned"
        if existing_br and existing_br != "Unassigned":
            # Keep existing valid BR
            pass
        elif mapped_br and mapped_br != "Unassigned":
            # Use mapped BR if valid
            u["businessRole"] = mapped_br
        else:
            # No valid BR anywhere - mark as Unassigned
            if not existing_br:
                u["businessRole"] = "Unassigned"

def sync_roles_from_users(users: List[Dict[str, Any]]) -> int:
    """
    Registra in role_meta/business_roles tutti i BR presenti sugli utenti.
    Ritorna quanti BR nuovi sono stati creati.
    """
    role_meta = state.setdefault("role_meta", {})
    business_roles = state.setdefault("business_roles", set())
    created = 0

    for u in users or []:
        br = (u.get("businessRole") or "").strip()
        if not br or br == "Unassigned":
            continue
        if br not in role_meta:
            _ensure_role_registered(br)  # usa role_meta/business_roles
            created += 1
        business_roles.add(br)

    return created


# =============================================================================
# BRDB (NO AI): learning DB + inference engine
# =============================================================================
# (imports already at top of file)

BRDB_MIN_CONF = 0.70  # soglia per auto-assegnare

BRDB_STOP_TOKENS = {
    "grp", "group", "role", "access", "users", "user", "team", "dl",
    "sec", "security", "global", "all", "everyone"
}

def brdb_norm_group(g: str) -> str:
    return (g or "").strip()

def brdb_tokens(g: str) -> List[str]:
    s = (g or "").lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    toks = [t for t in s.split() if len(t) >= 2 and t not in BRDB_STOP_TOKENS]
    return toks[:20]

def brdb_inc_stat(stats: Dict[str, Dict[str, int]], key: str, br: str, inc: int) -> None:
    if not key or not br:
        return
    stats.setdefault(key, {})
    stats[key][br] = int(stats[key].get(br, 0)) + int(inc)

def brdb_rebuild() -> None:
    """
    Ricostruisce il DB interno usando:
    - user_business_role + last_extract.users(groups)
    - role_meta (template BR->groups) con peso maggiore
    """
    global BRDB_CACHE
    state.setdefault("brdb_group_stats", {})
    state.setdefault("brdb_token_stats", {})
    BRDB_CACHE = {}

    group_stats: Dict[str, Dict[str, int]] = {}
    token_stats: Dict[str, Dict[str, int]] = {}

    # 1) training da utenti assegnati
    user_br = state.get("user_business_role", {}) or {}
    users = (state.get("last_extract") or {}).get("users") or []
    for u in users:
        uname = u.get("username")
        br = (user_br.get(uname) or u.get("businessRole") or "").strip()
        if not br:
            continue
        for g in (u.get("groups") or []):
            g0 = brdb_norm_group(g)
            brdb_inc_stat(group_stats, g0, br, 1)
            for t in brdb_tokens(g0):
                brdb_inc_stat(token_stats, t, br, 1)

    # 2) training forte da template role_meta
    role_meta = state.get("role_meta") or {}
    for br, meta in role_meta.items():
        br = (br or "").strip()
        for g in (meta.get("groups") or []):
            g0 = brdb_norm_group(g)
            brdb_inc_stat(group_stats, g0, br, 4)     # peso alto
            for t in brdb_tokens(g0):
                brdb_inc_stat(token_stats, t, br, 2)

    state["brdb_group_stats"] = group_stats
    state["brdb_token_stats"] = token_stats
    BRDB_CACHE = {}
    state["brdb_ready"] = True

def brdb_ensure_ready() -> None:
    if not state.get("brdb_ready"):
        brdb_rebuild()

def brdb_infer_group(group: str) -> Dict[str, Any]:


    """
    Predice BR per un singolo gruppo.
    """
    brdb_ensure_ready()

    g0 = brdb_norm_group(group)
    if not g0:
        return {"role": "Unassigned", "confidence": 0.0, "evidence": {"reason": "empty_group"}}

    global BRDB_CACHE
    if g0 in BRDB_CACHE:
        return BRDB_CACHE[g0]

    group_stats = (state.get("brdb_group_stats") or {}).get(g0) or {}
    token_stats = state.get("brdb_token_stats") or {}

    scores = defaultdict(float)

    # segnale forte: gruppo già visto
    tot_g = sum(group_stats.values())
    if tot_g:
        for br, c in group_stats.items():
            scores[br] += 2.5 * (c / tot_g)

    # segnale debole: token del nome
    toks = brdb_tokens(g0)
    for t in toks:
        ts = token_stats.get(t) or {}
        tot_t = sum(ts.values())
        if not tot_t:
            continue
        for br, c in ts.items():
            scores[br] += 1.0 * (c / tot_t)

    if not scores:
        out = {"role": "Unassigned", "confidence": 0.0, "evidence": {"reason": "no_stats"}}
        BRDB_CACHE[g0] = out
        return out

    best_role, best_score = max(scores.items(), key=lambda x: x[1])
    sum_scores = sum(scores.values()) or 1.0
    conf = float(best_score / sum_scores)

    out = {
        "role": best_role,
        "confidence": round(conf, 3),
        "evidence": {
            "scoresTop": sorted(scores.items(), key=lambda x: x[1], reverse=True)[:5],
            "groupSeen": group_stats,
            "tokens": toks,
        },
        "ts": datetime.now(timezone.utc).isoformat(),
    }
    BRDB_CACHE[g0] = out
    return out

# (imports already at top of file)

DEPT_MINCONF = 0.80
DEPT_GROUP_SUPPORT = 0.60

def ensure_role_registered(role: str) -> None:
    _ensure_role_registered(role)

DEPT_MERGE_JACCARD = 0.55  # soglia similarità gruppi dept vs BR template


def classify_account(
    display_name: str,
    ou: str,
    employee_type: str,
    use_ml: bool = True,
    attributes: Optional[Dict[str, Any]] = None,
) -> str:
    """
    Classify account type using ML (if available) with rule-based fallback.
    
    Uses the ML engine for high-confidence predictions (>75%), otherwise
    falls back to the extended 12-type rule-based classification.
    """
    # Try ML-based classification first
    if use_ml:
        try:
            predicted_type, confidence, method = ml_engine.classify_account(
                display_name, ou, employee_type, confidence_threshold=0.75, attributes=attributes
            )
            # Custom rules have top priority and must be applied immediately.
            if method == "custom_rule":
                return predicted_type
            if method == "ml" and confidence >= 0.75:
                return predicted_type
        except Exception:
            pass  # Fall through to rules
    
    # Use ML engine's rule-based classification (extended 12 types)
    return ml_engine.classify_account_rules(display_name, ou, employee_type)


# =============================================================================
# BRDB Functions (Business Role Database) - Delegating to ML Engine
# =============================================================================

def brdb_rebuild():
    """Rebuild the Business Role Database from current user assignments."""
    users = state.get("last_extract", {}).get("users") or []
    ml_engine.brdb_rebuild(users)


def brdb_infer_group(group: str) -> dict:
    """Infer which role a group belongs to based on learned patterns."""
    return ml_engine.brdb_infer_group(group)


def brdb_learn_assignment(role: str, groups: list, weight: float = 1.0):
    """Record a confirmed role→groups assignment for learning."""
    ml_engine.brdb_learn_assignment(role, groups, weight)

def _jaccard(a: set[str], b: set[str]) -> float:
    if not a and not b:
        return 1.0
    return len(a & b) / max(1, len(a | b))

BR_ASSIGNMENT_RULE_WEIGHT = 7.5
BR_FIELD_RULE_WEIGHT = 8.5


def _br_assignment_rule_scores_for_user(user: Dict[str, Any]) -> Dict[str, float]:
    """
    Score business-role boosts from regex rules applied on group names.
    The higher the score, the stronger the influence in automatic assignment.
    """
    group_rules = list(state.get("br_assignment_pattern_rules") or [])
    field_rules = list(state.get("br_pattern_rules") or [])
    if not group_rules and not field_rules:
        return {}

    groups = [str(g).strip() for g in (user.get("groups") or []) if str(g).strip()]
    if not groups:
        return {}
    groups_lower = {g.lower() for g in groups}

    scores: Dict[str, float] = defaultdict(float)

    # Form 3: regex on group names.
    for rule in group_rules:
        business_role = (rule.get("business_role") or "").strip()
        regex = (rule.get("regex") or "").strip()
        legacy_role = (rule.get("role") or "").strip()
        if not business_role or not regex:
            continue

        # Backward compatibility: old rules might have an explicit group/role gate.
        if legacy_role and legacy_role.lower() not in groups_lower:
            continue

        match_count = 0
        try:
            for g in groups:
                if re.search(regex, g, re.IGNORECASE):
                    match_count += 1
        except re.error:
            continue

        if match_count <= 0:
            continue

        boost = BR_ASSIGNMENT_RULE_WEIGHT * (1.0 + 0.2 * (match_count - 1))
        scores[business_role] += boost

    # Form 2: business role + field + regex, evaluated on user field value.
    for rule in field_rules:
        business_role = (rule.get("business_role") or "").strip()
        field = (rule.get("field") or "").strip()
        regex = (rule.get("regex") or "").strip()
        if not business_role or not field or not regex:
            continue

        raw_val = user.get(field)
        if raw_val is None:
            continue
        if isinstance(raw_val, list):
            value = " ".join([str(x) for x in raw_val])
        else:
            value = str(raw_val)
        if not value.strip():
            continue

        try:
            if re.search(regex, value, re.IGNORECASE):
                scores[business_role] += BR_FIELD_RULE_WEIGHT
        except re.error:
            continue

    return dict(scores)


def apply_department_mapping(users: List[Dict[str, Any]], only_depts: Optional[set[str]] = None) -> None:
    # Prepara strutture coerenti
    state.setdefault("user_business_role", {})
    state.setdefault("role_meta", {})
    state.setdefault("business_roles", set())
    state.setdefault("dept_to_role", {})          # dept -> canonical BR
    state.setdefault("dept_role_analysis", {})    # dept -> evidence


    # Global rebuild is still useful for stats, but we can do it conditionally?
    # For now, keep it global as stats depend on all users.
    # We could optimize brdb_rebuild() too if needed.
    brdb_rebuild()

    targets = only_depts


    by_dept = defaultdict(list)
    for u in users or []:
        dept = (u.get("department") or "").strip()
        if dept:
            by_dept[dept].append(u)

    role_meta = state["role_meta"]
    user_br = state["user_business_role"]
    dept_to_role = state["dept_to_role"]
    dept_analysis = state["dept_role_analysis"]

    for dept, members in by_dept.items():
        if targets and dept not in targets:
            continue

        # 1) Baseline deterministica: BR = dept
        _ensure_role_registered(dept)

        # 2) Profilo gruppi frequenti del dipartimento
        n = max(1, len(members))
        cnt = defaultdict(int)
        for u in members:
            for g in (u.get("groups") or []):
                cnt[g] += 1
        dept_groups = {g for g, c in cnt.items() if (c / n) >= DEPT_GROUP_SUPPORT}

        # 3) Analisi: “votazione” ruolo macro dai gruppi utenti (BRDB)
        weights = defaultdict(float)
        rules_boost = defaultdict(float)
        for u in members:
            s = brdb_infer_groupset(u.get("groups") or [])
            r = (s.get("role") or "Unassigned").strip()
            c = float(s.get("confidence") or 0.0)
            if r and r != "Unassigned" and c > 0:
                weights[r] += c
            for rr, sc in _br_assignment_rule_scores_for_user(u).items():
                weights[rr] += float(sc)
                rules_boost[rr] += float(sc)

        chosen_role = dept
        evidence = {
            "dept": dept,
            "members": len(members),
            "deptGroups": sorted(dept_groups),
            "chosenRole": dept,
            "reason": "baseline",
        }

        if weights:
            best_role, best_w = max(weights.items(), key=lambda x: x[1])
            total = sum(weights.values()) or 1.0
            conf = best_w / total

            role_groups = set((role_meta.get(best_role, {}) or {}).get("groups") or [])
            sim = _jaccard(dept_groups, role_groups) if role_groups else 1.0  # se non ho template ancora, non blocco

            evidence.update({
                "bestRole": best_role,
                "confidence": round(conf, 3),
                "jaccard": round(sim, 3),
                "weightsTop": sorted(weights.items(), key=lambda x: x[1], reverse=True)[:5],
                "ruleBoostTop": sorted(rules_boost.items(), key=lambda x: x[1], reverse=True)[:5],
            })

            # Merge automatico (come mi hai chiesto), ma solo sopra soglie
            if best_role != "Unassigned" and conf >= DEPT_MINCONF and sim >= DEPT_MERGE_JACCARD:
                chosen_role = best_role
                _ensure_role_registered(chosen_role)
                evidence["chosenRole"] = chosen_role
                evidence["reason"] = "merged_by_analysis"

        # 4) Consolidamento template BR: aggiungo gruppi frequenti dept al ruolo scelto
        existing = set((role_meta.get(chosen_role, {}) or {}).get("groups") or [])
        role_meta[chosen_role]["groups"] = sorted(existing | dept_groups)

        # 5) Salvo mapping dept->BR e assegno utenti
        dept_to_role[dept] = chosen_role
        dept_analysis[dept] = evidence

        for u in members:
            uname = u.get("username")
            if uname:
                existing_br = (u.get("businessRole") or "").strip()
                # log("INFO", f"DEBUG_MAPPING: {uname} existing='{existing_br}' chosen='{chosen_role}'")

                if existing_br and existing_br != "Unassigned":
                    user_br[uname] = existing_br
                else:
                    user_rule_scores = _br_assignment_rule_scores_for_user(u)
                    if user_rule_scores:
                        boosted_role = max(user_rule_scores.items(), key=lambda x: x[1])[0]
                        _ensure_role_registered(boosted_role)
                        user_br[uname] = boosted_role
                    else:
                        user_br[uname] = chosen_role

    # Apply BR assignment rules also to users without department mapping.
    for u in (users or []):
        uname = u.get("username")
        if not uname:
            continue
        existing_br = (u.get("businessRole") or "").strip()
        mapped_br = (user_br.get(uname) or "").strip()
        if (existing_br and existing_br != "Unassigned") or (mapped_br and mapped_br != "Unassigned"):
            continue
        rule_scores = _br_assignment_rule_scores_for_user(u)
        if rule_scores:
            forced_role = max(rule_scores.items(), key=lambda x: x[1])[0]
            _ensure_role_registered(forced_role)
            user_br[uname] = forced_role

    # applica mapping sugli oggetti utente
    apply_business_roles(users)

def brdb_infer_groupset(groups: List[str]) -> Dict[str, Any]:
    """
    Predice BR per un insieme di gruppi (utente/riga CSV) aggregando le predizioni per gruppo.
    """
    groups = [brdb_norm_group(g) for g in (groups or []) if brdb_norm_group(g)]
    if not groups:
        return {"role": "Unassigned", "confidence": 0.0, "evidence": {"reason": "no_groups"}}

    weights = defaultdict(float)
    details = []
    for g in groups[:50]:
        s = brdb_infer_group(g)
        details.append({"group": g, **s})
        weights[s["role"]] += float(s.get("confidence", 0.0) or 0.0)

    best_role, best_w = max(weights.items(), key=lambda x: x[1])
    total = sum(weights.values()) or 1.0
    conf = float(best_w / total)

    return {
        "role": best_role,
        "confidence": round(conf, 3),
        "evidence": {"weights": dict(weights), "details": details[:25]},
    }

def brdb_learn_assignment(br: str, groups: List[str], weight: int = 5) -> None:
    """
    Aggiorna incrementale il DB interno quando hai una assegnazione 'vera' (CSV con BR o assegnazione manuale).
    """
    brdb_ensure_ready()

    br = (br or "").strip()
    if not br:
        return
    for g in (groups or []):
        g0 = brdb_norm_group(g)
        brdb_inc_stat(state["brdb_group_stats"], g0, br, weight)
        for t in brdb_tokens(g0):
            brdb_inc_stat(state["brdb_token_stats"], t, br, max(1, weight // 2))

    # invalida cache
    state["brdb_cache"] = {}


def _mk_candidate(
    *,
    source: str,
    candidate_id: str,
    display_name: str,
    business_role: str,
    roles: list[str],
    raw: str,
    department: str = "",
    last_login: Any = None,
) -> dict:
    return {
        "candidateId": candidate_id,
        "source": source,
        "displayName": (display_name or "").strip(),
        "businessRole": (business_role or "").strip(),
        "department": (department or "").strip(),
        "lastLogin": last_login,
        "roles": roles or [],
        "rawLine": raw or "",
    }

def rebuild_ingest_candidates() -> None:
    src = state.get("ingest_sources") or {}
    flat = []
    for _, items in src.items():
        flat.extend(items or [])
    state["ingest_candidates"] = flat

def apply_duplicate_displayname_resolution() -> None:
    """
    Applica la scelta: per ogni displayName duplicato, rende 'attivo' solo il candidato scelto.
    Effetto pratico: marca excluded=True sugli utenti non scelti (così spariscono da role mining e lista utenti).
    """
    choice = state.get("choice_by_displayName") or {}
    candidates = state.get("ingest_candidates") or []

    # build: displayName -> [candidate]
    by_dn = defaultdict(list)
    for c in candidates:
        dn = (c.get("displayName") or "").strip()
        if dn:
            by_dn[dn].append(c)

    # indicizza utenti attuali per displayName
    users = state.get("last_extract", {}).get("users") or []
    users_by_dn = defaultdict(list)
    for u in users:
        dn = (u.get("displayName") or "").strip()
        if dn:
            users_by_dn[dn].append(u)

    # reset excluded
    for u in users:
        if "excluded" in u:
            u["excluded"] = False

    # applica scelta sui duplicati
    for dn, rows in by_dn.items():
        if len(rows) <= 1:
            continue

        chosen_id = choice.get(dn) or rows[0].get("candidateId")  # default: prima
        chosen = next((r for r in rows if r.get("candidateId") == chosen_id), rows[0])

        # qui assumiamo: gli utenti "a sistema" sono identificati dal displayName (come nella tua richiesta)
        # => lasciamo attivo un solo user per quel displayName
        same_dn_users = users_by_dn.get(dn) or []
        if len(same_dn_users) <= 1:
            # se esiste un solo user a sistema, facciamo override dei campi con quelli del candidato scelto
            if same_dn_users:
                u0 = same_dn_users[0]
                u0["businessRole"] = chosen.get("businessRole") or u0.get("businessRole", "")
                u0["groups"] = sorted(set(chosen.get("roles") or u0.get("groups") or []))
            continue

        # se ci sono più user a sistema con lo stesso displayName, ne lasciamo attivo solo 1
        # (scelta semplice: il primo della lista resta attivo e viene “aggiornato” col candidato scelto)
        keep = same_dn_users[0]
        keep["businessRole"] = chosen.get("businessRole") or keep.get("businessRole", "")
        keep["groups"] = sorted(set(chosen.get("roles") or keep.get("groups") or []))

        for u in same_dn_users[1:]:
            u["excluded"] = True

    # aggiorna anche il mapping BR per coerenza UI
    apply_business_roles(users)


def _to_ts(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except Exception:
        pass
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).timestamp()
    except Exception:
        return 0.0


def _normalize_last_login(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, list):
        if not v:
            return None
        v = v[0]
    s = str(v).strip()
    if not s:
        return None

    # AD FILETIME (100ns intervals since 1601-01-01 UTC)
    try:
        iv = int(float(s))
        if iv > 116444736000000000:
            unix_ts = (iv - 116444736000000000) / 10000000.0
            dt = datetime.fromtimestamp(unix_ts, tz=timezone.utc)
            return dt.isoformat()
    except Exception:
        pass

    # Unix timestamp
    try:
        fv = float(s)
        if fv > 0 and fv < 4102444800:  # until year 2100
            dt = datetime.fromtimestamp(fv, tz=timezone.utc)
            return dt.isoformat()
    except Exception:
        pass

    # ISO-like string
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc).isoformat()
    except Exception:
        return s


def _build_dept_group_profile(rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, float]]:
    by_dept = defaultdict(lambda: defaultdict(int))
    dept_counts = defaultdict(int)
    for r in (rows or []):
        dept = (r.get("department") or "").strip()
        if not dept:
            continue
        dept_counts[dept] += 1
        for g in set(r.get("groups") or []):
            by_dept[dept][g] += 1

    out: Dict[str, Dict[str, float]] = {}
    for dept, gstats in by_dept.items():
        den = max(1, dept_counts.get(dept, 1))
        out[dept] = {g: (cnt / den) for g, cnt in gstats.items()}
    return out


def _score_duplicate_candidate(c: Dict[str, Any], dept_profile: Dict[str, Dict[str, float]]) -> Dict[str, Any]:
    dept = (c.get("department") or "").strip()
    groups = list(dict.fromkeys(c.get("groups") or []))
    has_dept = 1 if dept else 0
    groups_count = len(groups)
    last_login_ts = _to_ts(c.get("lastLogin"))

    corr = 0.0
    if dept and groups:
        probs = dept_profile.get(dept) or {}
        corr = sum(float(probs.get(g, 0.0)) for g in groups) / max(1, len(groups))

    order = REQUIRED_DUPLICATE_ORDER

    values = {
        "dept_group_correlation": round(corr, 6),
        "has_department": has_dept,
        "groups_count": groups_count,
        "last_login": last_login_ts,
    }
    rank = tuple(values.get(k, 0.0) for k in order)
    return {
        "rank": rank,
        "order": order,
        "reason": {
            "deptGroupCorrelation": round(corr, 4),
            "hasDepartment": bool(has_dept),
            "groupsCount": groups_count,
            "lastLoginTs": last_login_ts,
        },
    }


def _duplicate_resolution_items() -> List[Dict[str, Any]]:
    candidates = state.get("ingest_candidates") or []
    if not candidates:
        return []

    by_dn = defaultdict(list)
    for c in candidates:
        dn = (c.get("displayName") or "").strip()
        if dn:
            by_dn[dn.lower()].append(c)

    choice_raw = state.get("choice_by_displayName") or {}
    auto_raw = state.get("duplicate_autoselect") or {}
    choice = {(str(k).strip().lower()): v for k, v in choice_raw.items() if str(k).strip()}
    auto = {(str(k).strip().lower()): v for k, v in auto_raw.items() if str(k).strip()}
    items = []
    for dn_key, rows in by_dn.items():
        if len(rows) <= 1:
            continue
        chosen_id = choice.get(dn_key) or rows[0].get("candidateId")
        chosen = next((r for r in rows if r.get("candidateId") == chosen_id), rows[0])
        auto_id = ((auto.get(dn_key) or {}).get("candidateId")) or chosen_id
        display_name = (chosen.get("displayName") or rows[0].get("displayName") or "").strip()
        alternatives = [r for r in rows if r.get("candidateId") != chosen.get("candidateId")]
        items.append(
            {
                "displayName": display_name,
                "chosenCandidateId": chosen.get("candidateId"),
                "autoChosenCandidateId": auto_id,
                "chosen": chosen,
                "alternatives": alternatives,
                "count": len(rows),
                "autoReason": (auto.get(dn_key) or {}).get("reason"),
            }
        )

    items.sort(key=lambda x: x.get("displayName") or "")
    return items


def _record_duplicate_feedback(display_name: str, chosen_candidate_id: str, actor: str = "system") -> None:
    display_name = (display_name or "").strip()
    chosen_candidate_id = (chosen_candidate_id or "").strip()
    if not display_name or not chosen_candidate_id:
        return

    auto = state.get("duplicate_autoselect") or {}
    auto_item = auto.get(display_name) or auto.get(display_name.lower()) or {}
    auto_id = (auto_item.get("candidateId") or "").strip()
    if not auto_id or auto_id == chosen_candidate_id:
        return

    by_id = {}
    for c in (state.get("ingest_candidates") or []):
        cid = str(c.get("candidateId") or "").strip()
        if cid:
            by_id[cid] = c

    auto_c = by_id.get(auto_id) or {}
    chosen_c = by_id.get(chosen_candidate_id) or {}

    ev = {
        "ts": datetime.now(timezone.utc).isoformat(),
        "displayName": display_name,
        "actor": actor,
        "autoCandidateId": auto_id,
        "chosenCandidateId": chosen_candidate_id,
        "auto": {
            "department": auto_c.get("department"),
            "groupsCount": len(auto_c.get("roles") or []),
            "lastLogin": auto_c.get("lastLogin"),
        },
        "chosen": {
            "department": chosen_c.get("department"),
            "groupsCount": len(chosen_c.get("roles") or []),
            "lastLogin": chosen_c.get("lastLogin"),
        },
    }

    events = state.get("dq_feedback_events") or []
    events.append(ev)
    # keep recent history bounded
    state["dq_feedback_events"] = events[-500:]


def build_dq_rule_suggestions() -> List[Dict[str, Any]]:
    suggestions: List[Dict[str, Any]] = []
    rules = state.get("dq_rules") or {}
    order = list(rules.get("duplicate_resolution_order") or [])
    if not order:
        order = ["dept_group_correlation", "has_department", "groups_count", "last_login"]
    events = list(state.get("dq_feedback_events") or [])
    ingest = state.get("last_ingest_stats") or {}

    dep_better = 0
    grp_better = 0
    ll_better = 0
    considered = 0
    for ev in events:
        auto = ev.get("auto") or {}
        chosen = ev.get("chosen") or {}
        considered += 1
        if (not auto.get("department")) and (chosen.get("department")):
            dep_better += 1
        if int(chosen.get("groupsCount") or 0) > int(auto.get("groupsCount") or 0):
            grp_better += 1
        if _to_ts(chosen.get("lastLogin")) > _to_ts(auto.get("lastLogin")):
            ll_better += 1

    def _ratio(x: int, y: int) -> float:
        return (float(x) / float(y)) if y > 0 else 0.0

    def _idx(lst: List[str], key: str) -> int:
        try:
            return lst.index(key)
        except ValueError:
            return 999

    if considered >= 3 and _ratio(ll_better, considered) >= 0.60 and _idx(order, "last_login") > 1:
        new_order = [x for x in order if x != "last_login"]
        new_order.insert(1, "last_login")
        suggestions.append(
            {
                "ruleId": "dq-dup-priority-lastlogin",
                "title": "Alza priorita LastLogin nella deduplica",
                "description": "Gli override manuali scelgono spesso il candidato con ultimo accesso piu recente.",
                "confidence": round(_ratio(ll_better, considered), 2),
                "impact": {"manualOverridesAnalyzed": considered, "lastLoginPreferred": ll_better},
                "preview": {"duplicate_resolution_order": new_order},
                "current": {"duplicate_resolution_order": order},
                "alreadyApplied": new_order == order,
            }
        )

    if considered >= 3 and _ratio(grp_better, considered) >= 0.60 and _idx(order, "groups_count") > 1:
        new_order = [x for x in order if x != "groups_count"]
        new_order.insert(1, "groups_count")
        suggestions.append(
            {
                "ruleId": "dq-dup-priority-groups",
                "title": "Alza priorita numero gruppi nella deduplica",
                "description": "Gli override manuali privilegiano il candidato con maggiore copertura gruppi.",
                "confidence": round(_ratio(grp_better, considered), 2),
                "impact": {"manualOverridesAnalyzed": considered, "groupsPreferred": grp_better},
                "preview": {"duplicate_resolution_order": new_order},
                "current": {"duplicate_resolution_order": order},
                "alreadyApplied": new_order == order,
            }
        )

    if int(ingest.get("missingRoles") or 0) > 0 and not bool(rules.get("reject_empty_groups")):
        miss_roles = int(ingest.get("missingRoles") or 0)
        total_rows = max(1, int(ingest.get("rowsTotal") or 0))
        suggestions.append(
            {
                "ruleId": "dq-reject-empty-groups",
                "title": "Blocca righe senza gruppi",
                "description": "Scarta in import le righe con gruppi vuoti per ridurre utenti orfani e remediation manuale.",
                "confidence": round(min(1.0, miss_roles / total_rows), 2),
                "impact": {"missingRolesLastImport": miss_roles, "rowsTotal": total_rows},
                "preview": {"reject_empty_groups": True},
                "current": {"reject_empty_groups": bool(rules.get("reject_empty_groups"))},
                "alreadyApplied": False,
            }
        )

    suggestions.sort(key=lambda x: (x.get("alreadyApplied"), -(x.get("confidence") or 0)))
    return suggestions


def pick_best_user(cands: List[Dict[str, Any]]) -> Dict[str, Any]:
    # regola: preferisci lastLogin più recente (se presente), poi "completezza" (dept+br), poi più gruppi
    def score(u: Dict[str, Any]) -> tuple:
        last = u.get("lastLogin") or u.get("last_login") or ""
        has_dept = 1 if (u.get("department") or "").strip() else 0
        has_br = 1 if (u.get("businessRole") or "").strip() else 0
        ng = len(u.get("groups") or [])
        return (last, has_dept + has_br, ng)
    return sorted(cands, key=score, reverse=True)[0]

def filter_and_dedupe_connector_users(raw_users: List[Dict[str, Any]], source: str) -> List[Dict[str, Any]]:
    rejects = []
    stats = state.setdefault("last_ingest_stats", {})
    stats.update({
        "source": source,
        "rowsTotal": int(len(raw_users or [])),
        "rowsKept": 0,
        "duplicateDisplayName": 0,
        "missingDepartment": 0,
        "missingBusinessRole": 0,
        "missingDisplayName": 0,
        "missingUsername": 0,
        "ts": datetime.now(timezone.utc).isoformat(),
    })

    by_dn = defaultdict(list)
    for u in (raw_users or []):
        if not (u.get("username") or "").strip():
            stats["missingUsername"] += 1
        dn = (u.get("displayName") or "").strip()
        if dn:
            by_dn[dn].append(u)
        else:
            stats["missingDisplayName"] += 1
            rejects.append({"source": source, "reason": "Missing displayName", "user": u, "ts": stats["ts"]})

    chosen = []
    for dn, cands in by_dn.items():
        u = pick_best_user(cands)

        if len(cands) > 1:
            stats["duplicateDisplayName"] += (len(cands) - 1)

        # log degli altri duplicati
        for other in cands:
            if other is u:
                continue
            rejects.append({"source": source, "reason": f"Duplicate displayName '{dn}' (kept best)", "user": other, "ts": stats["ts"]})

        dept = (u.get("department") or "").strip()
        br = (u.get("businessRole") or "").strip()

        if not dept:
            stats["missingDepartment"] += 1
            rejects.append({"source": source, "reason": "Missing department", "user": u, "ts": stats["ts"]})
        count_missing_br = not str(source).lower().startswith("ad")

        if count_missing_br and not br:
            stats["missingBusinessRole"] += 1
            rejects.append({
                "source": source,
                "reason": "Missing businessRole",
                "user": u,
                "ts": stats.get("ts"),
            })

        chosen.append(u)

    stats["rowsKept"] = int(len(chosen))
    state["last_rejects"] = rejects
    return chosen


# (datetime import already at top of file)

def _merge_users_into_last_extract(new_users: list[dict], *, ou: str):
    state.setdefault("last_extract", {"ou": "", "users": [], "groups": [], "ts": None})
    base_users = state["last_extract"]["users"]

    by_username = {u.get("username"): u for u in base_users if u.get("username")}

    for nu in (new_users or []):
        uname = nu.get("username")
        if not uname:
            continue

        if uname in by_username:
            u = by_username[uname]
            # aggiorna displayName (se presente) e fai merge gruppi
            if nu.get("displayName"):
                u["displayName"] = nu["displayName"]
            if nu.get("department"):
                u["department"] = nu["department"]
            u["groups"] = sorted(set(nu.get("groups") or []))  # REPLACE da connettore

        else:
            base_users.append(nu)
            by_username[uname] = nu

    # riallinea campi derivati
    apply_business_roles(base_users)
    state["last_extract"]["ou"] = ou
    state["last_extract"]["groups"] = recompute_groups_from_users(base_users)
    state["last_extract"]["ts"] = datetime.now(timezone.utc).isoformat()


def replace_from_connector_pool(new_users: List[Dict[str, Any]], ou: str, source: str) -> Dict[str, int]:
    """
    Replace current extract snapshot with a full connector pool.
    Used by AD import to ensure all rules run on a complete fresh dataset.
    """
    state.setdefault("last_extract", {"ou": "", "users": [], "groups": [], "ts": None})
    previous_users = state["last_extract"].get("users") or []
    previous_groups = set(state["last_extract"].get("groups") or [])

    prev_by_username = {}
    prev_by_displayname = {}
    for u in previous_users:
        uname = str(u.get("username") or "").strip()
        if uname:
            prev_by_username[uname] = u
        dn = (u.get("displayName") or "").strip().lower()
        if dn:
            prev_by_displayname[dn] = u

    previous_group_members: Dict[str, set[str]] = defaultdict(set)
    for u in previous_users:
        uname0 = str(u.get("username") or "").strip()
        if not uname0:
            continue
        for g0 in (u.get("groups") or []):
            gname = str(g0 or "").strip()
            if gname:
                previous_group_members[gname].add(uname0)

    clean_users: List[Dict[str, Any]] = []
    new_users_count = 0
    updated_users_count = 0

    for u in (new_users or []):
        username = (u.get("username") or "").strip()
        if not username:
            continue

        groups = sorted(set(u.get("groups") or []))
        normalized = dict(u)
        normalized["username"] = username
        normalized["displayName"] = (u.get("displayName") or username).strip()
        normalized["groups"] = groups
        normalized["department"] = (u.get("department") or "").strip() or None
        normalized["businessRole"] = (u.get("businessRole") or "").strip() or None
        normalized["excluded"] = False
        clean_users.append(normalized)

        dn_key = normalized["displayName"].lower()
        prev = prev_by_username.get(username) or prev_by_displayname.get(dn_key)
        if not prev:
            new_users_count += 1
            continue

        changed = (
            str(prev.get("username") or "").strip() != normalized["username"]
            or str(prev.get("displayName") or "").strip() != normalized["displayName"]
            or sorted(set(prev.get("groups") or [])) != groups
            or (prev.get("department") or None) != normalized["department"]
            or (prev.get("businessRole") or None) != normalized["businessRole"]
            or (prev.get("accountType") or None) != (normalized.get("accountType") or None)
            or (prev.get("lastLogin") or None) != (normalized.get("lastLogin") or None)
        )
        if changed:
            updated_users_count += 1

    current_groups = recompute_groups_from_users(clean_users)

    current_group_members: Dict[str, set[str]] = defaultdict(set)
    for u in clean_users:
        uname1 = str(u.get("username") or "").strip()
        if not uname1:
            continue
        for g1 in (u.get("groups") or []):
            gname = str(g1 or "").strip()
            if gname:
                current_group_members[gname].add(uname1)

    common_groups = previous_groups & set(current_groups)
    updated_groups_count = sum(
        1 for g in common_groups
        if previous_group_members.get(g, set()) != current_group_members.get(g, set())
    )

    state["last_extract"] = {
        "ou": ou,
        "users": clean_users,
        "groups": current_groups,
        "ts": datetime.now(timezone.utc).isoformat(),
        "source": source,
    }
    state["mining_dirty"] = True

    return {
        "new_users": int(new_users_count),
        "updated_users": int(updated_users_count),
        "updated_by_displayname": int(updated_users_count),
        "new_groups": int(len(set(current_groups) - previous_groups)),
        "updated_groups": int(updated_groups_count),
    }


def merge_from_connector_by_displayname(new_users: List[Dict[str, Any]], ou: str, source: str) -> Dict[str, int]:
    """
    Non-destructive merge:
    - keep existing local users
    - update only users with same displayName
    - add AD users not found by displayName
    """
    state.setdefault("last_extract", {"ou": "", "users": [], "groups": [], "ts": None})
    base_users = state["last_extract"]["users"]
    previous_groups = set(state["last_extract"].get("groups") or [])

    previous_group_members: Dict[str, set[str]] = defaultdict(set)
    for u in (base_users or []):
        uname0 = str(u.get("username") or "").strip()
        if not uname0:
            continue
        for g0 in (u.get("groups") or []):
            gname = str(g0 or "").strip()
            if gname:
                previous_group_members[gname].add(uname0)

    by_displayname: Dict[str, Dict[str, Any]] = {}
    for u in base_users:
        dn = (u.get("displayName") or "").strip().lower()
        if dn and dn not in by_displayname:
            by_displayname[dn] = u

    new_users_count = 0
    updated_users_count = 0

    for u in (new_users or []):
        username = (u.get("username") or "").strip()
        display_name = (u.get("displayName") or username).strip()
        if not username or not display_name:
            continue

        dn_key = display_name.lower()
        existing_user = by_displayname.get(dn_key)
        normalized_groups = sorted(set(u.get("groups") or []))

        if existing_user:
            prev_username = str(existing_user.get("username") or "").strip()
            prev_display = str(existing_user.get("displayName") or "").strip()
            prev_department = str(existing_user.get("department") or "").strip()
            prev_business_role = str(existing_user.get("businessRole") or "").strip()
            prev_groups = sorted(set(existing_user.get("groups") or []))
            prev_account_type = existing_user.get("accountType")
            prev_last_login = existing_user.get("lastLogin")

            existing_user["username"] = username
            existing_user["displayName"] = display_name
            existing_user["groups"] = normalized_groups
            existing_user["department"] = (u.get("department") or "").strip() or None
            if (u.get("businessRole") or "").strip():
                existing_user["businessRole"] = (u.get("businessRole") or "").strip()
            if u.get("accountType") is not None:
                existing_user["accountType"] = u.get("accountType")
            if u.get("lastLogin") is not None:
                existing_user["lastLogin"] = u.get("lastLogin")
            for k in ("email", "upn", "employeeId", "manager", "statusAd", "statusHr", "attributes"):
                if u.get(k) is not None:
                    existing_user[k] = u.get(k)
            existing_user["excluded"] = False

            changed = (
                prev_username != str(existing_user.get("username") or "").strip()
                or prev_display != str(existing_user.get("displayName") or "").strip()
                or prev_department != str(existing_user.get("department") or "").strip()
                or prev_business_role != str(existing_user.get("businessRole") or "").strip()
                or prev_groups != sorted(set(existing_user.get("groups") or []))
                or prev_account_type != existing_user.get("accountType")
                or prev_last_login != existing_user.get("lastLogin")
            )
            if changed:
                updated_users_count += 1
            continue

        new_user = {
            "username": username,
            "displayName": display_name,
            "groups": normalized_groups,
            "department": (u.get("department") or "").strip() or None,
            "businessRole": (u.get("businessRole") or "").strip() or None,
            "excluded": False,
            "lastLogin": u.get("lastLogin"),
            "accountType": u.get("accountType"),
            "email": u.get("email"),
            "upn": u.get("upn"),
            "employeeId": u.get("employeeId"),
            "manager": u.get("manager"),
            "statusAd": u.get("statusAd"),
            "statusHr": u.get("statusHr"),
            "attributes": u.get("attributes"),
        }
        base_users.append(new_user)
        by_displayname[dn_key] = new_user
        new_users_count += 1

    current_groups = recompute_groups_from_users(base_users)
    state["last_extract"]["groups"] = current_groups
    state["last_extract"]["ts"] = datetime.now(timezone.utc).isoformat()
    state["last_extract"]["source"] = source
    if ou:
        state["last_extract"]["ou"] = ou
    state["mining_dirty"] = True

    current_group_members: Dict[str, set[str]] = defaultdict(set)
    for u in (base_users or []):
        uname1 = str(u.get("username") or "").strip()
        if not uname1:
            continue
        for g1 in (u.get("groups") or []):
            gname = str(g1 or "").strip()
            if gname:
                current_group_members[gname].add(uname1)

    common_groups = previous_groups & set(current_groups)
    updated_groups_count = sum(
        1 for g in common_groups
        if previous_group_members.get(g, set()) != current_group_members.get(g, set())
    )

    return {
        "new_users": int(new_users_count),
        "updated_users": int(updated_users_count),
        "new_groups": int(len(set(current_groups) - previous_groups)),
        "updated_groups": int(updated_groups_count),
    }


def merge_from_connector(new_users: List[Dict[str, Any]], ou: str, source: str) -> Dict[str, int]:
    """
    Merge new users from connector (AD/LDAP) into existing state.
    Matches by BOTH displayName AND username - if either matches, updates existing user.
    """
    # Ensure baseline exists (if first run)
    state.setdefault("last_extract", {"ou": "", "users": [], "groups": [], "ts": None})
    base_users = state["last_extract"]["users"]
    log("INFO", f"Merge start. Existing users in state: {len(base_users)}")

    # Build dual indexes for matching by BOTH username AND displayName
    by_username = {}
    by_displayname = {}
    for u in base_users:
        uname = u.get("username")
        if uname:
            by_username[uname] = u
        dn = (u.get("displayName") or "").strip().lower()
        if dn:
            by_displayname[dn] = u
    
    previous_groups = set(state["last_extract"].get("groups") or [])
    previous_group_members: Dict[str, set[str]] = defaultdict(set)
    for u in (base_users or []):
        uname0 = str(u.get("username") or "").strip()
        if not uname0:
            continue
        for g0 in (u.get("groups") or []):
            gname = str(g0 or "").strip()
            if gname:
                previous_group_members[gname].add(uname0)

    clean_new_users: List[Dict[str, Any]] = []
    updated_count = 0
    added_count = 0
    updated_groups_count = 0
    
    # Pre-clean new users 
    for u in (new_users or []):
        username = (u.get("username") or "").strip()
        if not username:
            continue
        clean_new_users.append({
            "username": username,
            "displayName": (u.get("displayName") or username).strip(),
            "groups": sorted(set(u.get("groups") or [])),
            "department": (u.get("department") or "").strip() or None,
            "businessRole": (u.get("businessRole") or "").strip() or None,
            "excluded": False,
        })

    for nu in clean_new_users:
        uname = nu["username"]
        dn_key = (nu.get("displayName") or "").strip().lower()
        
        # Merge policy:
        # 1) Match by username
        # 2) If not found, match by displayName and update existing record
        #    (required for connector updates with renamed/changed usernames)
        existing_user = by_username.get(uname)
        if existing_user is None:
            existing_user = by_displayname.get(dn_key)
        
        if existing_user:
            # REPLACE existing user with new data from import
            prev_username_for_user = str(existing_user.get("username") or "").strip()
            prev_display_for_user = str(existing_user.get("displayName") or "").strip()
            prev_department_for_user = str(existing_user.get("department") or "").strip()
            prev_business_role_for_user = str(existing_user.get("businessRole") or "").strip()
            prev_groups_for_user = sorted(set(existing_user.get("groups") or []))
            existing_user["displayName"] = nu["displayName"]
            existing_user["department"] = nu["department"]
            # REPLACE groups (not merge)
            existing_user["groups"] = nu.get("groups") or []
            if nu.get("businessRole"):
                existing_user["businessRole"] = nu["businessRole"]
            # Update username if it changed (displayName matched but username different)
            if existing_user.get("username") != uname:
                old_uname = existing_user.get("username")
                existing_user["username"] = uname
                # Update index
                if old_uname in by_username:
                    del by_username[old_uname]
                by_username[uname] = existing_user

            changed = (
                prev_username_for_user != str(existing_user.get("username") or "").strip()
                or prev_display_for_user != str(existing_user.get("displayName") or "").strip()
                or prev_department_for_user != str(existing_user.get("department") or "").strip()
                or prev_business_role_for_user != str(existing_user.get("businessRole") or "").strip()
                or prev_groups_for_user != sorted(set(existing_user.get("groups") or []))
            )
            if changed:
                updated_count += 1
        else:
            # New user - add to base
            base_users.append(nu)
            by_username[uname] = nu
            if dn_key:
                by_displayname[dn_key] = nu
            added_count += 1

    # Update metadata
    state["last_extract"]["ts"] = datetime.now(timezone.utc).isoformat()
    state["last_extract"]["source"] = source
    if ou:
        state["last_extract"]["ou"] = ou

    # Recompute global groups list
    current_groups = recompute_groups_from_users(base_users)
    state["last_extract"]["groups"] = current_groups

    current_group_members: Dict[str, set[str]] = defaultdict(set)
    for u in (base_users or []):
        uname1 = str(u.get("username") or "").strip()
        if not uname1:
            continue
        for g1 in (u.get("groups") or []):
            gname = str(g1 or "").strip()
            if gname:
                current_group_members[gname].add(uname1)

    common_groups = previous_groups & set(current_groups)
    updated_groups_count = sum(
        1 for g in common_groups
        if previous_group_members.get(g, set()) != current_group_members.get(g, set())
    )

    state["mining_dirty"] = True
    log("INFO", f"Merge complete. Updated: {updated_count}, Added: {added_count}, Total: {len(base_users)}")
    return {
        "new_users": int(added_count),
        "updated_users": int(updated_count),
        "new_groups": int(len(set(current_groups) - previous_groups)),
        "updated_groups": int(updated_groups_count),
    }

def rerun_auto_business_roles_after_connector(
    users: List[Dict[str, Any]],
    only_depts: Optional[set[str]] = None,
    preserve_existing_brs: bool = True,
) -> None:
    preserved_brs = {}
    if preserve_existing_brs:
        # Preserve existing valid Business Roles from user objects BEFORE resetting mapping
        # Useful for CSV imports where BR is explicitly curated.
        for u in (users or []):
            uname = u.get("username")
            br = (u.get("businessRole") or "").strip()
            if uname and br and br != "Unassigned":
                preserved_brs[uname] = br

    # Reset mapping (partial or full)
    if not only_depts:
        # Full rebuild. For AD import we can start clean and let department/rules reassign.
        state["user_business_role"] = preserved_brs.copy() if preserve_existing_brs else {}
    else:
        # Partial update: merge preserved into existing
        state.setdefault("user_business_role", {})
        if preserve_existing_brs:
            state["user_business_role"].update(preserved_brs)

    
    # Do NOT wipe existing business roles on the user objects themselves,
    # otherwise we lose the value we just imported from CSV/Connector.
    # for u in (users or []):
    #     u["businessRole"] = None


    # ricostruisci mapping usando dept + analisi merge
    apply_department_mapping(users, only_depts=only_depts)


    state["mining_dirty"] = True

# (numpy and typing imports already at top of file)

def compute_over_threshold(matrix: Dict[str, Dict[str, int]], pct: float = 0.10) -> Optional[int]:
    if not matrix:
        return None

    row_counts = np.array([int(sum(row.values())) for row in matrix.values()], dtype=np.int32)
    if row_counts.size == 0:
        return None

    k_top = int(np.ceil(pct * row_counts.size))
    k_top = max(1, k_top)

    thr = int(np.partition(row_counts, -k_top)[-k_top])
    return thr

def build_over_rows_only(matrix: Dict[str, Dict[str, int]], threshold: Optional[int]) -> List[Dict[str, Any]]:
    if not matrix or threshold is None:
        return []

    rows: List[Dict[str, Any]] = []
    for user, grants in matrix.items():
        groups = [g for g, v in (grants or {}).items() if int(v) == 1]
        n_groups = len(groups)

        if n_groups >= threshold:
            rows.append({
                "user": user,
                "nGroups": n_groups,
                "over": True,
                "groupsText": ", ".join(groups),
            })

    # default sort utile (visto che "over" è sempre True qui)
    rows.sort(key=lambda r: (-r["nGroups"], r["user"]))
    return rows


def log(level: str, message: str) -> None:
    state["logs"].insert(0, {"ts": datetime.now(timezone.utc).isoformat(), "level": level, "message": message})
    state["logs"] = state["logs"][:500]

def _json_safe_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    if isinstance(value, dict):
        out = {}
        for k, v in value.items():
            key = k.decode("utf-8", errors="replace") if isinstance(k, bytes) else str(k)
            out[key] = _json_safe_value(v)
        return out
    if isinstance(value, (list, tuple, set)):
        return [_json_safe_value(v) for v in value]
    return str(value)


def record_manual_user_change(
    *,
    actor: str,
    username: str,
    display_name: Optional[str],
    action: str,
    source: str,
    details: Optional[Dict[str, Any]] = None,
    persist: bool = True,
) -> None:
    events = state.setdefault("manual_user_changes", [])
    item = {
        "id": f"muc-{int(time.time()*1000)}-{len(events)+1}",
        "ts": datetime.now(timezone.utc).isoformat(),
        "actor": actor,
        "username": username,
        "displayName": display_name or username,
        "action": action,
        "source": source,
        "details": details or {},
    }
    events.append(item)
    trimmed = events[-2000:]
    if persist:
        state["manual_user_changes"] = trimmed
    else:
        events[:] = trimmed
    return item


def record_llm_learning_event(
    *,
    actor: str,
    source: str,
    signal_type: str,
    entity: str = "",
    details: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    events = state.setdefault("llm_learning_history", [])
    item = {
        "id": f"lle-{int(time.time()*1000)}-{len(events)+1}",
        "ts": datetime.now(timezone.utc).isoformat(),
        "actor": actor,
        "source": source,
        "signalType": signal_type,
        "entity": entity,
        "details": details or {},
    }
    events.append(item)
    state["llm_learning_history"] = events[-3000:]
    return item


def recalculate_assignments_background(trigger: str, actor: str) -> None:
    """
    Recompute BR assignments asynchronously after rule changes.
    Runs in background to keep API latency low.
    """
    try:
        users = state.get("last_extract", {}).get("users") or []
        if not users:
            return
        # Recompute account types so newly added pattern rules are applied to existing users.
        for u in users:
            dn = str(u.get("displayName") or u.get("username") or "")
            dept = str(u.get("department") or "")
            employee_type = str(u.get("employeeType") or u.get("employee_type") or "")
            try:
                new_type = classify_account(dn, dept, employee_type, use_ml=True, attributes=u)
                if new_type:
                    u["accountType"] = new_type
            except Exception:
                continue
        rerun_auto_business_roles_after_connector(users, only_depts=None)
        sync_roles_from_users(users)
        state["mining_dirty"] = True
        RESPONSE_CACHE.invalidate("businessroles")
        RESPONSE_CACHE.invalidate("kpi")
        state.save()
        log("INFO", f"Background assignment recalculation completed (trigger={trigger}, by={actor})")
    except Exception as exc:
        log("ERROR", f"Background assignment recalculation failed (trigger={trigger}, by={actor}): {exc}")


def refresh_ai_detection_background(trigger: str, actor: str) -> None:
    """
    Recompute mining (if dirty) and AI detection in background after imports.
    """
    try:
        ensure_last_mining(None)  # sync mining inside this background worker
        last = state.get("last_mining") or {}
        matrix = last.get("matrix") or {}
        if not matrix:
            return
        users = active_users(state.get("last_extract", {}).get("users") or [])
        result = run_smart_ai_detection(users, matrix)
        state["last_ai_detection"] = result
        RESPONSE_CACHE.invalidate("kpi")
        state.save()
        log("INFO", f"Background AI detection refresh completed (trigger={trigger}, by={actor})")
    except Exception as exc:
        log("ERROR", f"Background AI detection refresh failed (trigger={trigger}, by={actor}): {exc}")


def run_post_snapshot_logic_background(snapshot_ts: str, actor: str) -> None:
    """
    Run heavy post-import business logic after AD snapshot has been saved.
    If a newer snapshot exists, skip to avoid stale background work.
    """
    try:
        current_ts = str((state.get("last_extract") or {}).get("ts") or "")
        if not snapshot_ts or current_ts != str(snapshot_ts):
            log("INFO", f"Skip post-snapshot logic: stale snapshot (expected={snapshot_ts}, current={current_ts})")
            return

        users = (state.get("last_extract") or {}).get("users") or []
        rerun_auto_business_roles_after_connector(users, preserve_existing_brs=False)
        sync_roles_from_users(users)

        rebuild_ingest_candidates()
        apply_duplicate_displayname_resolution()
        invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)

        refresh_ai_detection_background("ad-import-post-snapshot", actor)
        log("INFO", f"Post-snapshot logic completed (snapshot_ts={snapshot_ts}, by={actor})")
    except Exception as exc:
        log("ERROR", f"Post-snapshot logic failed (snapshot_ts={snapshot_ts}, by={actor}): {exc}")


def run_post_csv_snapshot_logic_background(snapshot_ts: str, actor: str, touched_depts: List[str]) -> None:
    """
    Run heavy post-import business logic after CSV snapshot has been saved.
    If a newer snapshot exists, skip to avoid stale background work.
    """
    try:
        current_ts = str((state.get("last_extract") or {}).get("ts") or "")
        if not snapshot_ts or current_ts != str(snapshot_ts):
            log("INFO", f"Skip CSV post-snapshot logic: stale snapshot (expected={snapshot_ts}, current={current_ts})")
            return

        users = (state.get("last_extract") or {}).get("users") or []
        only_depts = {d for d in (touched_depts or []) if d}
        rerun_auto_business_roles_after_connector(
            users,
            only_depts=only_depts if only_depts else None,
            preserve_existing_brs=True,
        )
        sync_roles_from_users(users)

        rebuild_ingest_candidates()
        apply_duplicate_displayname_resolution()
        invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)

        refresh_ai_detection_background("csv-import-post-snapshot", actor)
        log("INFO", f"CSV post-snapshot logic completed (snapshot_ts={snapshot_ts}, by={actor})")
    except Exception as exc:
        log("ERROR", f"CSV post-snapshot logic failed (snapshot_ts={snapshot_ts}, by={actor}): {exc}")

def active_users(users: list[dict]) -> list[dict]:
    return [u for u in (users or []) if not u.get("excluded")]

def recompute_groups_from_users(users: list[dict]) -> list[str]:
    return sorted({g for u in active_users(users) for g in (u.get("groups") or [])})

# (_mk_candidate already defined at line 643)



def apply_choice_for_displayname(display_name: str,
                                 chosen_business_role: Optional[str],
                                 chosen_roles: Optional[List[str]]) -> None:
    users = state.get("last_extract", {}).get("users") or []

    same = [u for u in users if (u.get("displayName") or "").strip() == (display_name or "").strip()]
    if not same:
        return

    keep = same[0]
    keep["excluded"] = False

    if chosen_business_role:
        keep["businessRole"] = chosen_business_role
        state.setdefault("user_business_role", {})
        state["user_business_role"][keep["username"]] = chosen_business_role

    if chosen_roles is not None:
        keep["groups"] = sorted(set(chosen_roles))

    for u in same[1:]:
        u["excluded"] = True
        m = state.get("user_business_role") or {}
        if u.get("username") in m:
            del m[u["username"]]

    # state["last_extract"]["groups"] = recompute_groups_from_users(users)


# ----------------------------
# Models
# ----------------------------
class LoginRequest(BaseModel):
    username: str
    password: str


class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    username: str


class ConnectorConfig(BaseModel):
    server: str = Field(..., description="LDAP host/ip o 'mock'")
    bind_user: str = Field("", description="Utente bind (es: user@domain o DOMAIN\\user)")
    bind_password: str = Field("", description="Password bind")
    base_dn: str = Field("", description="Base DN (es: DC=example,DC=local)")
    auth: str = Field("SIMPLE", description="SIMPLE oppure NTLM")
    port: int = Field(389, description="LDAP Port")
    use_ssl: bool = Field(False, description="Use SSL/LDAPS")


class ExtractRequest(BaseModel):
    ou: str = Field("", description="OU DN (opzionale: fallback a base_dn connettore)")


class ExtractResponse(BaseModel):
    ou: str
    total_users: int
    total_groups: int
    snapshot_ready: bool = True
    processing_in_background: bool = False
    new_users: int = 0
    updated_users: int = 0
    updated_by_displayname: int = 0
    new_groups: int = 0
    updated_groups: int = 0
    users: List[Dict[str, Any]]
    groups: List[str]


class RoleMiningRequest(BaseModel):
    n_clusters: Optional[int] = Field(None, description="Se None, calcolato automaticamente")
    role_support: float = Field(0.5, ge=0.1, le=1.0, description="Soglia (0..1) per includere un gruppo nel ruolo del cluster")


class RoleMiningResponse(BaseModel):
    total_users: int
    total_groups: int
    n_clusters: int
    clusters: List[Dict[str, Any]]
    kpi: Dict[str, Any]


security = HTTPBearer(auto_error=False)


def create_access_token(username: str) -> str:
    exp = datetime.now(timezone.utc) + timedelta(minutes=JWT_EXPIRE_MINUTES)
    payload = {"sub": username, "exp": exp}
    return jwt.encode(payload, JWT_SECRET, algorithm="HS256")


def require_auth(creds: HTTPAuthorizationCredentials = Depends(security)) -> str:
    if creds is None or not creds.credentials:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Missing bearer token")
    try:
        payload = jwt.decode(creds.credentials, JWT_SECRET, algorithms=["HS256"])
        return payload["sub"]
    except Exception:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")


# ----------------------------
# LDAP / Mock AD
# ----------------------------
def mock_users() -> List[Dict[str, Any]]:
    return [
        {"username": "alice", "displayName": "Alice Rossi", "groups": ["HR", "AllEmployees", "Confluence"]},
        {"username": "bob", "displayName": "Bob Bianchi", "groups": ["HR", "AllEmployees", "Payroll"]},
        {"username": "carol", "displayName": "Carol Verdi", "groups": ["IT", "AllEmployees", "Azure", "GitLab"]},
        {"username": "dave", "displayName": "Dave Neri", "groups": ["IT", "AllEmployees", "Azuure"]},
        {"username": "erin", "displayName": "Erin Gialli", "groups": ["Sales", "AllEmployees", "CRM"]},
        {"username": "frank", "displayName": "Frank Blu", "groups": ["Sales", "AllEmployees", "CRM", "DiscountApproval"]},
    ]
def _mk_ldap_server(cfg: Dict[str, Any]):
    raw = (cfg.get("server") or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="Configura server LDAP")

    if Connection is None:
        raise HTTPException(status_code=500, detail="ldap3 non disponibile")

    raw_lower = raw.lower()
    scheme_ssl = raw_lower.startswith("ldaps://")
    host_port = raw.replace("ldaps://", "").replace("ldap://", "")
    host = host_port.split(":")[0].strip()
    server_port = None
    if ":" in host_port:
        try:
            server_port = int(host_port.rsplit(":", 1)[1].strip())
        except Exception:
            server_port = None

    cfg_use_ssl = cfg.get("use_ssl")
    use_ssl = bool(cfg_use_ssl) if cfg_use_ssl is not None else scheme_ssl

    cfg_port = cfg.get("port")
    port = None
    try:
        port = int(cfg_port) if cfg_port is not None else None
    except Exception:
        port = None
    if port is None:
        port = server_port
    if port is None:
        port = 636 if use_ssl else 389

    return {
        "server": Server(host, port=port, use_ssl=use_ssl, get_info=NONE, connect_timeout=15),
        "host": host,
        "port": port,
        "use_ssl": use_ssl,
    }


def _ldap_search_attrs() -> List[str]:
    if LDAP_FETCH_ALL_ATTRIBUTES:
        return ["*"]
    out = list(dict.fromkeys(LDAP_BASE_ATTRIBUTES + LDAP_EXTRA_ATTRIBUTES))
    return out or ["*"]


def _extract_group_cn(raw_dn: Any) -> str:
    dn_str = str(raw_dn or "").strip()
    if not dn_str:
        return ""
    for part in dn_str.split(","):
        p = part.strip()
        if p.upper().startswith("CN="):
            return p[3:]
    return dn_str





def extract_from_ldap(ou_dn: str) -> List[Dict[str, Any]]:
    if MOCK_AD or state.get("connector", {}).get("server") == "mock":
        # Ignora OU, ritorna dataset di test
        log("INFO", "Using MOCK AD extract")
        return mock_users()

    if Connection is None:
        raise HTTPException(status_code=500, detail="ldap3 library not installed or failed to import")

    cfg = state.get("connector", {})
    if not cfg.get("server") or not cfg.get("bind_user") or not cfg.get("bind_password"):
        raise HTTPException(status_code=400, detail="Configura server/bind_user/bind_password in Connettori")

    log("INFO", f"Connecting to LDAP server: {cfg['server']} as {cfg['bind_user']}")
    
    conn = None
    try:
        resolved = _mk_ldap_server(cfg)
        server = resolved["server"]
        host = resolved["host"]
        port = resolved["port"]
        use_ssl = resolved["use_ssl"]
        log("INFO", f"LDAP target resolved to {host}:{port} (use_ssl={use_ssl})")
        auth_mode = cfg.get("auth", "SIMPLE").upper()
        
        try:
            if auth_mode == "NTLM":
                conn = Connection(server, user=cfg["bind_user"], password=cfg["bind_password"], authentication=NTLM, auto_bind=True)
            else:
                conn = Connection(server, user=cfg["bind_user"], password=cfg["bind_password"], authentication=SIMPLE, auto_bind=True)
        except Exception as e:
            error_msg = str(e)
            log("ERROR", f"LDAP Connection failed: {error_msg}")
            if "timeout" in error_msg.lower():
                 raise HTTPException(status_code=504, detail=f"Timeout connettendosi al server LDAP {host}:{port} (use_ssl={use_ssl})")
            raise HTTPException(status_code=503, detail=f"Impossibile connettersi al server LDAP {host}:{port} (use_ssl={use_ssl}) - {error_msg}")

        # objectClass=user può includere account tecnici: in produzione aggiungere filtri più stretti
        search_filter = "(&(objectClass=user)(sAMAccountName=*))"
        attrs = _ldap_search_attrs()
        log(
            "INFO",
            f"Searching LDAP base='{ou_dn}' filter='{search_filter}' attrs={len(attrs)} page_size={LDAP_PAGE_SIZE}",
        )

        users: List[Dict[str, Any]] = []

        # Collect all available field names for UI
        available_fields = set()
        parsed_entries = 0

        if LDAP_PAGE_SIZE > 0:
            entries_iter = conn.extend.standard.paged_search(
                search_base=ou_dn,
                search_filter=search_filter,
                attributes=attrs,
                paged_size=LDAP_PAGE_SIZE,
                generator=True,
                time_limit=LDAP_SEARCH_TIME_LIMIT,
            )
        else:
            if not conn.search(
                search_base=ou_dn,
                search_filter=search_filter,
                attributes=attrs,
                time_limit=LDAP_SEARCH_TIME_LIMIT,
            ):
                result = conn.result or {}
                log("ERROR", f"LDAP search returned False. Result: {result}")
                result_desc = str(result.get("description") or "").lower()
                diagnostic = {
                    "message": f"LDAP search failed on base '{ou_dn}'",
                    "result": result,
                    "hints": [
                        "Verifica OU DN: deve esistere nel dominio LDAP (es. OU=Users,DC=example,DC=internal).",
                        f"Controlla coerenza con base_dn configurato: '{cfg.get('base_dn') or ''}'.",
                    ],
                }
                if "no such object" in result_desc or "nosuchobject" in result_desc:
                    raise HTTPException(status_code=400, detail=diagnostic)
                if "invalid dn syntax" in result_desc or "invaliddnsyntax" in result_desc:
                    raise HTTPException(status_code=400, detail=diagnostic)
                raise HTTPException(status_code=500, detail=diagnostic)
            entries_iter = (
                {"type": "searchResEntry", "attributes": e.entry_attributes_as_dict}
                for e in conn.entries
            )

        for entry in entries_iter:
            if entry.get("type") != "searchResEntry":
                continue
            parsed_entries += 1
            try:
                d = entry.get("attributes") or {}
                
                # Update available fields
                for k in d.keys():
                    available_fields.add(k)

                # Safe attribute extraction
                sAM = d.get("sAMAccountName") or []
                username = str(sAM[0]).strip() if sAM else ""
                
                disp = d.get("displayName") or []
                display = str(disp[0]).strip() if disp else ""
                
                member_of = d.get("memberOf") or []
                groups = []
                for dn in member_of:
                    cn = _extract_group_cn(dn)
                    if cn:
                        groups.append(cn)
                
                dept = d.get("department") or []
                department = str(dept[0]).strip() if dept else None
                
                llt = d.get("lastLogonTimestamp")
                ll = d.get("lastLogon")
                candidates = []
                n1 = _normalize_last_login(llt)
                n2 = _normalize_last_login(ll)
                if n1:
                    candidates.append(n1)
                if n2:
                    candidates.append(n2)
                if candidates:
                    last_login = max(candidates, key=_to_ts)
                else:
                    last_login = None

                et = d.get("employeeType")
                etype = str(et[0]).strip() if et else ""
                
                dn_full = d.get("distinguishedName")
                dn_str_val = str(dn_full[0]).strip() if dn_full else ""
                
                # Use department or parse OU from DN as fallback for classification
                ou_for_class = department or dn_str_val

                # Pass sanitized attributes to avoid bytes serialization errors in API responses.
                safe_attrs = _json_safe_value(d)
                account_type = classify_account(display or username, ou_for_class, etype, attributes=safe_attrs)

                if username:
                    users.append({
                        "username": username,
                        "displayName": display or username,
                        "groups": sorted(set(groups)),
                        "department": department,
                        "lastLogin": last_login,
                        "accountType": account_type,
                        "attributes": safe_attrs
                    })
            except Exception as entry_ex:
                # Log but continue processing other users
                log("WARNING", f"Error parsing LDAP entry: {entry_ex}. Entry: {str(entry)[:100]}...")
                continue
        
        # Update state with available fields
        state["ad_available_fields"] = sorted(list(available_fields))
        log("INFO", f"Updated available AD fields: {len(available_fields)} found.")

                
        log("INFO", f"LDAP extraction complete. Parsed entries={parsed_entries}, users={len(users)}.")
        return users

    except HTTPException:
        raise
    except Exception as e:
        log("ERROR", f"Unexpected error in extract_from_ldap: {e}")
        # traceback would be good here but keeping it simple
        raise HTTPException(status_code=500, detail=f"Errore interno durante estrazione LDAP: {str(e)}")
    finally:
        if conn:
            try:
                conn.unbind()
            except:
                pass


# ----------------------------
# Role Mining
# ----------------------------
def build_matrix(users: List[Dict[str, Any]]) -> Tuple[List[str], List[str], np.ndarray]:
    usernames = [u["username"] for u in users]
    all_groups = sorted({g for u in users for g in u["groups"]})
    g_index = {g: i for i, g in enumerate(all_groups)}

    X = np.zeros((len(users), len(all_groups)), dtype=np.int8)
    for i, u in enumerate(users):
        for g in u["groups"]:
            if g in g_index:
                X[i, g_index[g]] = 1
    return usernames, all_groups, X


def jaccard_distance_matrix(X: np.ndarray) -> np.ndarray:
    # D[i,j] = 1 - |A∩B|/|A∪B|
    n = X.shape[0]
    D = np.zeros((n, n), dtype=np.float32)
    for i in range(n):
        Ai = X[i]
        for j in range(i + 1, n):
            Aj = X[j]
            inter = int(np.logical_and(Ai, Aj).sum())
            union = int(np.logical_or(Ai, Aj).sum())
            dist = 0.0 if union == 0 else 1.0 - (inter / union)
            D[i, j] = dist
            D[j, i] = dist
    return D


def compute_purity(cluster_members_idx: List[int], X: np.ndarray) -> float:
    # Purity “semplice”: media della massima frequenza (per attributo) nel cluster
    if not cluster_members_idx:
        return 0.0
    sub = X[cluster_members_idx, :]
    # frequenza per colonna
    freq = sub.mean(axis=0)  # 0..1
    return float(freq.max())  # quanto un "gruppo dominante" spiega il cluster

# (_tokens, _family_key, _is_broad already defined at lines 96-109)
# (BROAD_MARKERS defined at line 94)

def compute_ai_detection(matrix: dict, users: list = None) -> dict:
    if users:
        # Use SMART logic (KB + Stats)
        res = run_smart_ai_detection(users, matrix)
        stats = res.get("stats", {})
        return {
            "aiDetection": stats.get("aiDetection", 0),
            "redundantAssignments": stats.get("totalAnomalies", 0),
            "totalAssignments": stats.get("totalAssignments", 0),
            "usersWithRedundancy": stats.get("usersWithAnomaly", 0),
            "redundantUsers": [], # Not needed for KPI summary
        }

    # --- Legacy Logic (Matrix only) ---
    total_assignments = 0
    redundant_assignments = 0

    users_with_redundancy = 0
    redundant_users = []

    for uname, row in (matrix or {}).items():
        roles = [r for r, v in (row or {}).items() if int(v) == 1]
        user_groups = _matrix_user_roles(matrix, uname)

        total_assignments += len(roles)

        fam = defaultdict(list)
        for r in roles:
            k = _family_key(r)
            if k:
                fam[k].append(r)

        has_redundancy = False

        for rs in fam.values():
            broad = [r for r in rs if _is_broad(r)]
            specific = [r for r in rs if not _is_broad(r)]
            if not broad or not specific:
                continue

            red = 0
            for b in broad:
                probs = [predict_redundant(b, s, _family_key(b), user_groups) for s in specific]
                p = max(probs) if probs else 0.0
                if p >= 0.50:
                    red += 1

            if red > 0:
                redundant_assignments += red
                has_redundancy = True

        if has_redundancy:
            users_with_redundancy += 1
            redundant_users.append(uname)

    total_users = len(matrix or {})
    pct = (users_with_redundancy / max(1, total_users) * 100.0)

    return {
        "aiDetection": round(pct, 2),
        "redundantAssignments": redundant_assignments,
        "totalAssignments": total_assignments,
        "usersWithRedundancy": users_with_redundancy,
        "redundantUsers": redundant_users,
    }


def compute_kpis(
    users: List[Dict[str, Any]],
    clusters: List[Dict[str, Any]],
    matrix: Dict[str, Dict[str, int]],
) -> Dict[str, Any]:
    import numpy as np

    total_users = len(users)
    if total_users == 0:
        return {
            "totalUsers": 0,
            "overprivilegedPct": 0,
            "clusterQuality": 0,
            "clusteringQuality": 0,
            "roleCoverage": 0,
            "aiDetection": 0,
            "redundantAssignments": 0,
            "totalAssignments": 0,
            "usersWithRedundancy": 0,
        }

    # --- Overprivileged: top 10% per numero gruppi calcolato dalla matrix ---
    row_counts = np.array(
        [int(sum((row or {}).values())) for row in (matrix or {}).values()],
        dtype=np.int32
    )
    if row_counts.size == 0:
        return {
            "totalUsers": total_users,
            "overprivilegedPct": 0,
            "clusterQuality": 0,
            "clusteringQuality": 0,
            "roleCoverage": 0,
            "aiDetection": 0,
            "redundantAssignments": 0,
            "totalAssignments": 0,
            "usersWithRedundancy": 0,
        }

    k_top = int(np.ceil(0.1 * row_counts.size))
    k_top = max(1, k_top)

    thr = int(np.partition(row_counts, -k_top)[-k_top])

    overpriv = float((row_counts >= thr).mean() * 100.0)

    # --- AI detection (redundant broad roles) ---
    ai = compute_ai_detection(matrix, users=users)

    # --- RoleCoverage: roleGroups copre i gruppi reali (da matrix) dei membri ---
    cov_vals: list[float] = []
    for c in (clusters or []):
        role_groups = set(c.get("roleGroups") or [])
        members = c.get("members") or []
        for uname in members:
            row = matrix.get(uname) or {}
            denom = int(sum((row or {}).values()))
            if denom <= 0:
                cov_vals.append(0.0)
            else:
                covered = sum(int(row.get(g, 0)) for g in role_groups)
                cov_vals.append(covered / denom)

    role_coverage = float((np.mean(cov_vals) if cov_vals else 0.0) * 100.0)

    # --- clusterQuality = data-quality score (ingest) ---
    ingest = state.get("last_ingest_stats") or {}
    total = int(ingest.get("rowsTotal") or 0)
    src = str(ingest.get("source") or "").lower()

    if total > 0:
        dup = int(ingest.get("duplicateDisplayName") or 0)
        miss_dept = int(ingest.get("missingDepartment") or 0)
        miss_br = int(ingest.get("missingBusinessRole") or 0)
        miss_dn = int(ingest.get("missingDisplayName") or 0)
        miss_user = int(ingest.get("missingUsername") or 0)

        if src.startswith("ad"):
            miss_br = 0
            dup = 0

        penalty = (
            1.00 * (dup / total) +
            0.70 * (miss_dept / total) +
            0.70 * (miss_br / total) +
            0.40 * (miss_dn / total) +
            0.40 * (miss_user / total)
        )
        penalty = min(1.0, penalty)
        cluster_quality = 100.0 * (1.0 - penalty)
    else:
        cluster_quality = 100.0

    # --- clusteringQuality = qualità clustering (purity media pesata) ---
    if clusters:
        total_sz = sum(int(c.get("size") or len(c.get("members") or [])) for c in clusters) or 1
        weighted = 0.0
        for c in clusters:
            sz = int(c.get("size") or len(c.get("members") or []))
            weighted += float(c.get("purity") or 0.0) * sz
        clustering_quality = (weighted / total_sz) * 100.0
    else:
        clustering_quality = 0.0

    # --- Model Quality (Enhanced Option B) ---
    groups_list = (state.get("last_extract") or {}).get("groups") or []
    if not groups_list and matrix:
        # Fallback to matrix keys if last_extract is empty
        all_groups = set()
        for row in matrix.values():
            if isinstance(row, list):
                all_groups.update(row)
            else:
                all_groups.update((row or {}).keys())
        groups_list = list(all_groups)
    
    mq = compute_model_quality(users, matrix, groups_list)

    # Ensure last_kpis is updated in state for drilldown fallback
    last_kpis = {
        "clusterQuality": round(cluster_quality, 2),
        "clusteringQuality": round(clustering_quality, 2),
        "modelQuality": mq.get("modelQuality", 0),
        "aiDetection": ai.get("score", 0),
        "roleCoverage": round(role_coverage, 2),
        "staleAccounts": mq.get("staleUsers", 0)
    }
    state["last_kpis"] = last_kpis

    return {
        "totalUsers": total_users,
        # "overprivilegedPct": round(overpriv, 2), # Removed/Deprecated
        "modelQuality": mq.get("modelQuality", 0),
        "orphanGroupsCount": mq.get("orphanGroups", 0),
        "overprivilegedCount": mq.get("overprivilegedUsers", 0),
        "zeroGroupCount": mq.get("zeroGroupUsers", 0),
        "staleAccountCount": mq.get("staleUsers", 0),

        "clusterQuality": round(float(cluster_quality), 2),
        "clusteringQuality": round(float(clustering_quality), 2),
        "aiDetection": ai.get("aiDetection", 0),
        "redundantAssignments": ai.get("redundantAssignments", 0),
        "totalAssignments": ai.get("totalAssignments", 0),
        "usersWithRedundancy": ai.get("usersWithRedundancy", 0),
        "roleCoverage": round(float(role_coverage), 2),
    }


def compute_model_quality(users: List[Dict[str, Any]], matrix: Dict[str, Dict[str, int]], groups: List[str]) -> Dict[str, Any]:
    import numpy as np
    from datetime import datetime, timezone
    import math

    total_users = len(users)
    total_groups = len(groups)

    if total_users == 0 or total_groups == 0:
        return {
            "modelQuality": 0,
            "orphanGroups": 0,
            "overprivilegedUsers": 0,
            "zeroGroupUsers": 0,
            "staleUsers": 0,
            "orphansList": [],
            "indicators": [],
            "policyViolations": [],
            "ambiguousUsers": [],
            "manualOverrideEvents": 0,
            "density": 0.0,
            "avgGeneralizationConfidence": 0.0,
        }

    def _active_groups(row: Any) -> set[str]:
        if row is None:
            return set()
        if isinstance(row, list):
            return {str(g) for g in row if str(g)}
        return {str(g) for g, val in (row or {}).items() if int(val) == 1}

    user_by_username = {str(u.get("username")): u for u in (users or []) if u.get("username")}
    user_groups_map: Dict[str, set[str]] = {}
    for uname in user_by_username.keys():
        user_groups_map[uname] = _active_groups(matrix.get(uname))
    for uname, row in (matrix or {}).items():
        if uname not in user_groups_map:
            user_groups_map[uname] = _active_groups(row)

    # 1) Orphan groups (weighted)
    group_counts = {g: 0 for g in groups}
    for active in user_groups_map.values():
        for g in active:
            if g in group_counts:
                group_counts[g] += 1
    orphans_list = [g for g, count in group_counts.items() if count == 0]
    n_orphans = len(orphans_list)
    critical_markers = ("admin", "all", "write", "prod", "root")
    weighted_orphans = 0.0
    for g in orphans_list:
        gl = str(g).lower()
        w = 2.0 if any(m in gl for m in critical_markers) else 1.0
        weighted_orphans += w
    orphan_weighted_pct = (weighted_orphans / max(1.0, total_groups * 2.0)) * 100.0

    # 2) Overprivileged concentration
    row_counts = np.array([len(gs) for gs in user_groups_map.values()], dtype=np.int32)
    if row_counts.size > 0:
        k_top = max(1, int(np.ceil(0.1 * row_counts.size)))
        thr = int(np.partition(row_counts, -k_top)[-k_top])
        n_over = int((row_counts >= thr).sum())
        over_pct = (n_over / total_users) * 100.0
    else:
        n_over, over_pct = 0, 0

    # 3) Users with Zero Groups
    n_zero = sum(1 for _, gs in user_groups_map.items() if len(gs) == 0)
    zero_pct = (n_zero / total_users) * 100.0 if total_users > 0 else 0

    # 4) Stale access quality (> 1 year)
    stale_count = 0
    now = datetime.now(timezone.utc)
    stale_list = []
    for u in (users or []):
        ll = u.get("lastLogin")
        if ll:
            try:
                clean_ll = str(ll).replace("Z", "+00:00")
                dt = datetime.fromisoformat(clean_ll)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                if (now - dt).days > 365:
                    stale_count += 1
                    stale_list.append({"username": u.get("username"), "displayName": u.get("displayName"), "lastLogon": ll})
            except Exception:
                pass
    stale_pct = (stale_count / total_users) * 100.0 if total_users > 0 else 0

    # 5) Role entropy per ruolo (eterogeneita interna)
    role_meta = state.get("role_meta") or {}
    user_br = state.get("user_business_role") or {}
    members_by_role = defaultdict(list)
    for u in (users or []):
        uname = u.get("username")
        if not uname:
            continue
        br = (u.get("businessRole") or user_br.get(uname) or "Unassigned").strip()
        members_by_role[br].append(uname)

    role_entropy_vals = []
    for _, members in members_by_role.items():
        n = len(members)
        if n <= 1:
            continue
        ent_vals = []
        for g in groups:
            present = sum(1 for uname in members if g in (user_groups_map.get(uname) or set()))
            p = present / max(1, n)
            if p <= 0.0 or p >= 1.0:
                continue
            h = -(p * math.log2(p) + (1 - p) * math.log2(1 - p))  # 0..1
            ent_vals.append(h)
        if ent_vals:
            role_entropy_vals.append(float(np.mean(ent_vals)))
    role_entropy_pct = float(np.mean(role_entropy_vals) * 100.0) if role_entropy_vals else 0.0

    # 6) Template coverage & 7) Noise ratio
    miss_template_ratios = []
    noise_ratios = []
    for u in (users or []):
        uname = u.get("username")
        if not uname:
            continue
        gs = user_groups_map.get(uname) or set()
        br = (u.get("businessRole") or user_br.get(uname) or "Unassigned").strip()
        tmpl = set((role_meta.get(br, {}) or {}).get("groups") or [])
        if tmpl:
            cov = len(gs & tmpl) / max(1, len(tmpl))
            miss_template_ratios.append(1.0 - cov)
        if gs:
            noise = len([g for g in gs if g not in tmpl]) / max(1, len(gs))
            noise_ratios.append(noise)
    template_coverage_penalty_pct = float(np.mean(miss_template_ratios) * 100.0) if miss_template_ratios else 0.0
    noise_ratio_pct = float(np.mean(noise_ratios) * 100.0) if noise_ratios else 0.0

    # 8) Ambiguita assegnazione ruolo (bassa confidence BRDB)
    ambiguous_users = []
    for u in (users or [])[:20000]:
        uname = u.get("username")
        if not uname:
            continue
        gs = list(user_groups_map.get(uname) or [])
        if not gs:
            continue
        s = brdb_infer_groupset(gs)
        conf = float(s.get("confidence") or 0.0)
        if conf < 0.55:
            ambiguous_users.append({"username": uname, "displayName": u.get("displayName"), "confidence": round(conf, 3)})
    ambiguity_pct = (len(ambiguous_users) / max(1, total_users)) * 100.0

    # 9) Drift temporale (utenti recenti ma bassa compatibilita template)
    drift_users = 0
    recent_users = 0
    for u in (users or []):
        uname = u.get("username")
        ll = u.get("lastLogin")
        if not uname or not ll:
            continue
        try:
            dt = datetime.fromisoformat(str(ll).replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
        except Exception:
            continue
        if (now - dt).days <= 90:
            recent_users += 1
            gs = user_groups_map.get(uname) or set()
            br = (u.get("businessRole") or user_br.get(uname) or "Unassigned").strip()
            tmpl = set((role_meta.get(br, {}) or {}).get("groups") or [])
            if gs and tmpl:
                fit = len(gs & tmpl) / max(1, len(gs | tmpl))
                if fit < 0.35:
                    drift_users += 1
    drift_pct = (drift_users / max(1, recent_users)) * 100.0 if recent_users else 0.0

    # 10) Matrix sparsity/overdensity
    total_assignments = int(sum(len(gs) for gs in user_groups_map.values()))
    density = total_assignments / max(1, total_users * total_groups)
    if density < 0.02:
        density_penalty_pct = min(100.0, ((0.02 - density) / 0.02) * 100.0)
    elif density > 0.35:
        density_penalty_pct = min(100.0, ((density - 0.35) / 0.35) * 100.0)
    else:
        density_penalty_pct = 0.0

    # 11) Policy violation rate (SoD-like)
    policy_violations = []
    for uname, gs in user_groups_map.items():
        by_prefix = defaultdict(set)
        for g in gs:
            parts = str(g).split("_")
            if len(parts) >= 3:
                pref = "_".join(parts[:2])
                by_prefix[pref].add(parts[-1].upper())
        bad = []
        for pref, suff in by_prefix.items():
            if "ALL" in suff and ("WRITE" in suff or "ADMIN" in suff):
                bad.append(pref)
        if bad:
            policy_violations.append({"username": uname, "conflicts": sorted(bad)})
    policy_violation_pct = (len(policy_violations) / max(1, total_users)) * 100.0

    # 12) Manual override dependency
    feedback_events = list(state.get("dq_feedback_events") or [])
    auto_resolved = int((state.get("last_ingest_stats") or {}).get("autoResolvedDuplicateUsers") or 0)
    manual_override_pct = min(100.0, (len(feedback_events) / max(1, auto_resolved)) * 100.0) if auto_resolved else 0.0

    # 13) Generalization score (BRDB confidence medio)
    confs = []
    for uname, gs in user_groups_map.items():
        if not gs:
            continue
        s = brdb_infer_groupset(list(gs))
        confs.append(float(s.get("confidence") or 0.0))
    avg_conf = float(np.mean(confs)) if confs else 0.0
    generalization_penalty_pct = (1.0 - avg_conf) * 100.0
    weights = get_active_model_weights()
    preset = (state.get("dq_model_preset") or "manufacturing").strip().lower()

    indicators = [
        {"id": "role_entropy", "label": "Role Entropy", "value": round(role_entropy_pct, 2), "penalty": round(role_entropy_pct, 2), "weight": float(weights.get("role_entropy", 0.08))},
        {"id": "template_coverage", "label": "Template Coverage Gap", "value": round(template_coverage_penalty_pct, 2), "penalty": round(template_coverage_penalty_pct, 2), "weight": float(weights.get("template_coverage", 0.10))},
        {"id": "noise_ratio", "label": "Noise Ratio", "value": round(noise_ratio_pct, 2), "penalty": round(noise_ratio_pct, 2), "weight": float(weights.get("noise_ratio", 0.10))},
        {"id": "ambiguity", "label": "Assignment Ambiguity", "value": round(ambiguity_pct, 2), "penalty": round(ambiguity_pct, 2), "weight": float(weights.get("ambiguity", 0.08))},
        {"id": "temporal_drift", "label": "Temporal Drift", "value": round(drift_pct, 2), "penalty": round(drift_pct, 2), "weight": float(weights.get("temporal_drift", 0.07))},
        {"id": "matrix_density", "label": "Matrix Density Risk", "value": round(density_penalty_pct, 2), "penalty": round(density_penalty_pct, 2), "weight": float(weights.get("matrix_density", 0.07))},
        {"id": "orphan_weighted", "label": "Weighted Orphans", "value": round(orphan_weighted_pct, 2), "penalty": round(orphan_weighted_pct, 2), "weight": float(weights.get("orphan_weighted", 0.09))},
        {"id": "overprivileged", "label": "Overprivileged Concentration", "value": round(over_pct, 2), "penalty": round(over_pct, 2), "weight": float(weights.get("overprivileged", 0.10))},
        {"id": "stale_access", "label": "Stale Access Quality", "value": round(stale_pct, 2), "penalty": round(stale_pct, 2), "weight": float(weights.get("stale_access", 0.10))},
        {"id": "policy_violation", "label": "Policy Violation Rate", "value": round(policy_violation_pct, 2), "penalty": round(policy_violation_pct, 2), "weight": float(weights.get("policy_violation", 0.08))},
        {"id": "manual_override", "label": "Manual Override Dependency", "value": round(manual_override_pct, 2), "penalty": round(manual_override_pct, 2), "weight": float(weights.get("manual_override", 0.07))},
        {"id": "generalization", "label": "Generalization Gap", "value": round(generalization_penalty_pct, 2), "penalty": round(generalization_penalty_pct, 2), "weight": float(weights.get("generalization", 0.07))},
    ]

    penalty = 0.0
    for i in indicators:
        contrib = float(i["weight"]) * float(i["penalty"])
        i["contribution"] = round(contrib, 2)
        penalty += contrib
    quality = max(0.0, 100.0 - penalty)

    return {
        "modelQuality": round(quality, 2),
        "orphanGroups": n_orphans,
        "overprivilegedUsers": n_over,
        "zeroGroupUsers": n_zero,
        "staleUsers": stale_count,
        "orphansList": orphans_list,
        "staleList": stale_list,
        "zeroList": [{"username": uname, "displayName": (user_by_username.get(uname) or {}).get("displayName"), "groupCount": 0} for uname, gs in user_groups_map.items() if len(gs) == 0],
        "overprivilegedList": [{"username": uname, "groupCount": len(gs)} for uname, gs in user_groups_map.items() if len(gs) >= (thr if row_counts.size > 0 else 999999)],
        "policyViolations": policy_violations,
        "ambiguousUsers": ambiguous_users,
        "manualOverrideEvents": len(feedback_events),
        "indicators": indicators,
        "density": round(density, 4),
        "avgGeneralizationConfidence": round(avg_conf, 3),
        "modelPreset": preset,
    }



def run_role_mining(
    users: List[Dict[str, Any]],
    n_clusters: Optional[int],
    role_support: float
) -> Dict[str, Any]:
    # usa solo utenti attivi
    users = [u for u in (users or []) if not u.get("excluded")]

    usernames, groups, X = build_matrix(users)

    # Colonne UI = gruppi snapshot (ultima estrazione) → NON dipendono dalle assegnazioni correnti
    snapshot_groups = ((state.get("last_extract") or {}).get("groups") or []).copy()
    snapshot_groups = [g for g in snapshot_groups if g]  # safety

    # Se lo snapshot è vuoto (es. mai estratto), fallback ai gruppi del build_matrix
    all_groups_ui = snapshot_groups if snapshot_groups else groups

    # dataset troppo piccolo
    if len(usernames) < 2 or len(groups) == 0:
        return {
            "clusters": [],
            "matrix": {},
            "kpi": compute_kpis(users, [], {}),
            "groups": all_groups_ui,
        }

    # scegli k
    auto_k = max(2, int(round(np.sqrt(len(usernames)))))
    if n_clusters:
        # Safety: never ask KMeans for more clusters than available samples
        k = max(2, min(int(n_clusters), len(usernames)))
    else:
        k = min(8, auto_k, len(usernames))

    # clustering su SVD + MiniBatchKMeans (O(N) complexity)
    # SVD reduction
    n_components = min(50, X.shape[1] - 1)
    if n_components > 1:
        svd = TruncatedSVD(n_components=n_components, random_state=42)
        X_reduced = svd.fit_transform(X)
    else:
        X_reduced = X

    # MiniBatchKMeans
    model = MiniBatchKMeans(n_clusters=k, random_state=42, batch_size=256, n_init="auto")
    labels = model.fit_predict(X_reduced)

    # clusters -> members/roleGroups/purity
    clusters: List[Dict[str, Any]] = []
    for cid in sorted(set(labels.tolist())):
        idx = [i for i, lab in enumerate(labels) if lab == cid]
        members = [usernames[i] for i in idx]

        sub = X[idx, :]
        freq = sub.mean(axis=0)  # 0..1
        role_groups = [groups[j] for j in range(len(groups)) if float(freq[j]) >= role_support]

        purity = compute_purity(idx, X)

        clusters.append({
            "clusterId": int(cid),
            "members": members,
            "roleGroups": role_groups,
            "purity": round(float(purity), 4),
            "size": len(members),
        })

    # matrix for UI (DEVE avere tutte le colonne di all_groups_ui)
    # mapping per i gruppi presenti in X
    gindex = {g: j for j, g in enumerate(groups)}

    matrix: Dict[str, Dict[str, int]] = {}
    for i, uname in enumerate(usernames):
        row: Dict[str, int] = {}
        for g in all_groups_ui:
            j = gindex.get(g)
            row[g] = int(X[i, j]) if j is not None else 0
        matrix[uname] = row

    kpi = compute_kpis(users, clusters, matrix)

    return {
        "clusters": clusters,
        "matrix": matrix,
        "kpi": kpi,
        "groups": all_groups_ui,
    }


def _mining_worker(n_clusters, role_support):
    ok = False
    try:
        users = active_users(state.get("last_extract", {}).get("users") or [])
        res = run_role_mining(users, n_clusters=n_clusters, role_support=role_support)

        # Build displayNames map (username -> displayName) for frontend optimization
        display_names = {}
        for u in users:
            uname = u.get("username")
            if uname:
                display_names[uname] = u.get("displayName") or uname

        state.update({
            "last_mining": {
                "clusters": res.get("clusters", []),
                "matrix": res.get("matrix", {}),
                "kpi": res.get("kpi", {}),
                "groups": res.get("groups", []),
                "displayNames": display_names,
                "ts": datetime.now(timezone.utc).isoformat(),
                "status": "ready"
            },
            "mining_dirty": False
        })
        ok = True
    except Exception as e:
        print(f"[Mining Worker] Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        state["mining_processing"] = False
        state["mining_status"] = "ready" if ok else "idle"

def ensure_last_mining(background_tasks: BackgroundTasks = None) -> None:
    """
    Ricalcola il clustering in background se dirty.
    """
    last_extract = state.get("last_extract") or {}
    last_mining = state.get("last_mining") or {}

    extract_ts = last_extract.get("ts")
    mining_ts = last_mining.get("ts")
    matrix = last_mining.get("matrix") or {}
    users = active_users(last_extract.get("users") or [])

    # No data loaded: do not trigger mining/polling loops.
    if not users:
        state["mining_processing"] = False
        state["mining_status"] = "idle"
        return

    stale = bool(extract_ts and (not mining_ts or str(extract_ts) > str(mining_ts)))

    # Recovery for stale persisted lock (e.g. crash/restart while mining_processing=True)
    if state.get("mining_processing"):
        started_at = state.get("mining_started_at")
        lock_stale = False
        if started_at:
            try:
                started_dt = datetime.fromisoformat(str(started_at).replace("Z", "+00:00"))
                if started_dt.tzinfo is None:
                    started_dt = started_dt.replace(tzinfo=timezone.utc)
                lock_stale = (datetime.now(timezone.utc) - started_dt).total_seconds() > 300
            except Exception:
                # Unknown timestamp format -> treat as stale to unblock system.
                lock_stale = True
        else:
            # No timestamp and no matrix means lock is likely stale
            lock_stale = not bool(matrix)

        if not lock_stale:
            return  # Already running

        state["mining_processing"] = False
        state["mining_status"] = "idle"

    if not state.get("mining_dirty") and matrix and not stale:
        return

    # Trigger background
    state["mining_processing"] = True
    state["mining_status"] = "running"
    state["mining_started_at"] = datetime.now(timezone.utc).isoformat()
    params = state.get("last_mining_params") or {}
    n_clusters = params.get("n_clusters", None)
    role_support = params.get("role_support", 0.6)
    
    if background_tasks:
        background_tasks.add_task(_mining_worker, n_clusters, role_support)
    else:
        # Fallback sync if no background_tasks provided (should verify calls)
        _mining_worker(n_clusters, role_support)



# ----------------------------
# FastAPI app
# ----------------------------
app = FastAPI(title=APP_TITLE)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://localhost:5173",
        "http://localhost:5174",  # dev
        "http://127.0.0.1:5173",
        "*",  # ← PER TEST (rimuovi in PROD)
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

from fastapi.middleware.gzip import GZipMiddleware
app.add_middleware(GZipMiddleware, minimum_size=1000)

@app.get("/api/kpi/drilldown")
def kpidrilldown_q(metric: str): #, username: str = Depends(require_auth)):
    # Passing None for background_tasks to avoid AttributeError on string
    return kpi_drilldown(metric, None)



@app.get("/api/health")
def health():
    return {"ok": True, "ts": int(time.time())}


class ToggleUserGroupRequest(BaseModel):
    username: str
    group: str
    enabled: bool  # True=assegna, False=rimuovi


@app.post("/api/users/groups/toggle")
def toggle_user_group(body: ToggleUserGroupRequest, username: str = Depends(require_auth)):
    # 1) trova utente nello stato
    users = (state.get("last_extract") or {}).get("users") or []
    uobj = next((u for u in users if u.get("username") == body.username), None)
    if not uobj:
        raise HTTPException(status_code=404, detail="User not found")

    # 2) valida gruppo (e impedisci di “resuscitare” gruppi non più nello snapshot AD/CSV)
    g = (body.group or "").strip()
    if not g:
        raise HTTPException(status_code=400, detail="Group empty")

    snapshot_groups = set(((state.get("last_extract") or {}).get("groups") or []))
    if snapshot_groups and g not in snapshot_groups:
        raise HTTPException(status_code=400, detail="Group not in last extract snapshot")

    # 3) toggle
    current = set(uobj.get("groups") or [])
    if body.enabled:
        current.add(g)
    else:
        current.discard(g)

    uobj["groups"] = sorted(current)
    record_manual_user_change(
        actor=username,
        username=body.username,
        display_name=uobj.get("displayName"),
        action="toggle-group",
        source="matrix",
        details={"group": g, "enabled": bool(body.enabled), "groupsCount": len(uobj.get("groups") or [])},
    )

    # IMPORTANTISSIMO: NON aggiornare state["last_extract"]["groups"] qui
    state["mining_dirty"] = True
    invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)
    return {"ok": True, "username": body.username, "group": g, "enabled": body.enabled}


@app.post("/api/auth/login", response_model=TokenResponse)
def login(body: LoginRequest):
    if body.username != APP_LOGIN_USER or body.password != APP_LOGIN_PASS:
        raise HTTPException(status_code=401, detail="Credenziali non valide")
    token = create_access_token(body.username)
    log("INFO", f"Login OK {body.username}")
    return TokenResponse(access_token=token, username=body.username)


@app.get("/api/me")
def me(username: str = Depends(require_auth)):
    return {"username": username}


@app.get("/api/config/connector", response_model=ConnectorConfig)
def get_connector(username: str = Depends(require_auth)):
    return ConnectorConfig(**state["connector"])


@app.post("/api/config/connector", response_model=ConnectorConfig)
def set_connector(cfg: ConnectorConfig, username: str = Depends(require_auth)):
    state["connector"] = cfg.model_dump()
    log("INFO", f"Connector config updated by {username} (server={cfg.server}, auth={cfg.auth})")
    return cfg


@app.post("/api/ad/extract", response_model=ExtractResponse)
def extract(req: ExtractRequest, background_tasks: BackgroundTasks, username: str = Depends(require_auth)):
    connector_cfg = state.get("connector") or {}
    ou_dn = (req.ou or "").strip() or (connector_cfg.get("base_dn") or "").strip()
    if not ou_dn:
        raise HTTPException(status_code=400, detail="OU/base_dn non valorizzato: imposta base_dn in Connettori o passa OU nella richiesta")

    users = extract_from_ldap(ou_dn)

    # 1) Costruisci candidati AD
    ad_candidates = []
    for u in users:
        ad_candidates.append(
            _mk_candidate(
                source="ad",
                candidate_id=f"ad:{u['username']}",
                display_name=u.get("displayName", ""),
                business_role=u.get("businessRole", ""),
                roles=(u.get("groups") or []),
                raw=f"AD:{u.get('username')}|{u.get('displayName')}|{','.join(u.get('groups') or [])}",
                department=u.get("department") or "",
                last_login=u.get("lastLogin"),
            )
        )

    # 2) Build full AD pool first, then merge into local DB:
    #    update only by same displayName; keep all other local users.
    users = filter_and_dedupe_connector_users(users, source="ad")
    merge_stats = merge_from_connector_by_displayname(users, ou=ou_dn, source="ad")
    
    # 3) Batch updates to state to minimize disk I/O
    updates = {
        "ingest_sources": {**state.get("ingest_sources", {}), "ad": ad_candidates},
        "mining_dirty": True
    }
    state.update(updates)
    invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)
    snapshot_ts = str((state.get("last_extract") or {}).get("ts") or "")
    background_tasks.add_task(run_post_snapshot_logic_background, snapshot_ts, username)

    return ExtractResponse(
        ou=state["last_extract"]["ou"],
        total_users=len(state["last_extract"]["users"]),
        total_groups=len(state["last_extract"]["groups"]),
        snapshot_ready=True,
        processing_in_background=True,
        new_users=merge_stats.get("new_users", 0),
        updated_users=merge_stats.get("updated_users", 0),
        updated_by_displayname=merge_stats.get("updated_by_displayname", 0),
        new_groups=merge_stats.get("new_groups", 0),
        updated_groups=merge_stats.get("updated_groups", 0),
        users=state["last_extract"]["users"],
        groups=state["last_extract"]["groups"],
    )


@app.get("/api/ad/extract/export-csv")
def export_last_extract_csv(username: str = Depends(require_auth)):
    users = (state.get("last_extract") or {}).get("users") or []
    if not users:
        raise HTTPException(status_code=400, detail="Nessuna estrazione disponibile")

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([
        "DisplayName",
        "Username",
        "Department",
        "BusinessRole",
        "Ruoli",
        "AccountType",
        "LastLogin",
        "Email",
        "UPN",
        "EmployeeId",
        "Manager",
        "StatusAD",
        "StatusHR",
    ])

    for u in users:
        writer.writerow([
            u.get("displayName") or "",
            u.get("username") or "",
            u.get("department") or "",
            u.get("businessRole") or "",
            ",".join(u.get("groups") or []),
            u.get("accountType") or "",
            u.get("lastLogin") or "",
            u.get("email") or "",
            u.get("upn") or "",
            u.get("employeeId") or "",
            u.get("manager") or "",
            u.get("statusAd") or "",
            u.get("statusHr") or "",
        ])

    output.seek(0)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"ad_extract_snapshot_{ts}.csv"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(output, media_type="text/csv; charset=utf-8", headers=headers)


@app.get("/api/users")
def list_users(q: str = "", type_q: str = "", limit: int = 100, offset: int = 0, sort_by: str = "", order: str = "asc", username: str = Depends(require_auth)):
    cache_key = f"users_{q}|{type_q}|{limit}|{offset}|{sort_by}|{order}"
    cached = RESPONSE_CACHE.get(cache_key)
    if cached:
        return cached

    users = active_users(state["last_extract"]["users"] or [])
    
    # Text Filter
    if q:
        ql = q.lower()
        users = [
            u for u in users
            if ql in (u.get("username") or "").lower()
            or ql in (u.get("displayName") or "").lower()
        ]
        
    # Type Filter
    if type_q:
        tql = type_q.lower()
        users = [
            u for u in users
            if tql in (u.get("accountType") or u.get("account_type") or "internal").lower()
        ]
    
    # Sorting support
    if sort_by:
        reverse = order.lower() == "desc"
        
        def get_sort_key(u):
            val = u.get(sort_by)
            # Fallback for accountType nuances
            if val is None and sort_by == "accountType":
                val = u.get("account_type")
            return (val or "").lower()
            
        users = sorted(users, key=get_sort_key, reverse=reverse)
    
    total = len(users)
    sliced = users[offset : offset + limit]
    result = {"total": total, "items": sliced, "limit": limit, "offset": offset}
    RESPONSE_CACHE.set(cache_key, result, CACHE_TTL_USERS)
    return result

@app.get("/api/users/{uname}")
def get_user(uname: str, username: str = Depends(require_auth)):
    users = state.get("last_extract", {}).get("users") or []
    u = next((x for x in users if x.get("username") == uname), None)
    if not u:
        raise HTTPException(status_code=404, detail="User not found")

    apply_business_roles(users)  # riallinea businessRole
    return {"user": u, "allGroups": recompute_groups_from_users(users)}


class UserUpdateRequest(BaseModel):
    groups: List[str] = []
    businessRole: Optional[str] = None
    accountType: Optional[str] = None

@app.post("/api/users/{uname}/update")
def update_user(uname: str, body: UserUpdateRequest, username: str = Depends(require_auth)):
    users = state.get("last_extract", {}).get("users") or []
    u = next((x for x in users if x.get("username") == uname), None)
    if not u:
        raise HTTPException(status_code=404, detail="User not found")

    prev_groups = sorted(set(u.get("groups") or []))
    prev_br = u.get("businessRole")
    prev_type = u.get("accountType")
    # Replace groups
    u["groups"] = sorted({g.strip() for g in (body.groups or []) if g and g.strip()})

    # (Opzionale) cambia anche BR qui, oppure continua a usare /api/businessroles/{role}/add
    if body.businessRole is not None:
        br = body.businessRole.strip()
        u["businessRole"] = br
        state.setdefault("user_business_role", {})
        if br and br != "Unassigned":
            state["user_business_role"][uname] = br
        else:
            # rimuovi mapping => torna Unassigned
            if uname in state["user_business_role"]:
                del state["user_business_role"][uname]

        # training “forte” come fai già in businessroles/roleadd
        try:
            brdb_learn_assignment(br, u.get("groups") or [], weight=10)
            record_llm_learning_event(
                actor=username,
                source="user-update",
                signal_type="brdb-assignment",
                entity=uname,
                details={"businessRole": br, "groupsCount": len(u.get("groups") or []), "weight": 10},
            )
        except Exception:
            pass

    if body.accountType:
        u["accountType"] = body.accountType.strip()

    # riallinea derivati
    apply_business_roles(users)
    # state["last_extract"]["groups"] = recompute_groups_from_users(users)
    state["mining_dirty"] = True
    record_manual_user_change(
        actor=username,
        username=uname,
        display_name=u.get("displayName"),
        action="update-user",
        source="user-detail",
        details={
            "groupsBefore": prev_groups,
            "groupsAfter": u.get("groups") or [],
            "businessRoleBefore": prev_br,
            "businessRoleAfter": u.get("businessRole"),
            "accountTypeBefore": prev_type,
            "accountTypeAfter": u.get("accountType"),
        },
    )

    invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)
    return {"ok": True, "user": u}


@app.get("/api/users/{uname}/peer-analysis")
def get_peer_analysis(uname: str, username: str = Depends(require_auth)):
    users = state.get("last_extract", {}).get("users") or []
    target = next((x for x in users if x.get("username") == uname), None)
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    
    br = target.get("businessRole")
    at = target.get("accountType") or "Internal"
    
    if not br or br == "Unassigned":
         return {"peersCount": 0, "anomalies": [], "suggestedGroups": []}

    # Find peers: same BR + same AccountType
    peers = [u for u in users if u.get("businessRole") == br and (u.get("accountType") or "Internal") == at]
    peers_count = len(peers)
    
    if peers_count < 2:
        return {"peersCount": peers_count, "anomalies": [], "suggestedGroups": []}
        
    # Calculate group frequencies
    grp_counts = defaultdict(int)
    for p in peers:
        for g in (p.get("groups") or []):
            grp_counts[g] += 1
            
    # Check target user's groups
    target_groups = set(target.get("groups") or [])
    anomalies = []
    
    for g in target_groups:
        freq = grp_counts[g] / peers_count
        if freq < 0.15:  # Threshold for anomaly (e.g., < 15% of peers have this group)
            anomalies.append({
                "group": g,
                "frequency": round(freq, 2),
                "count": grp_counts[g],
                "peers": peers_count
            })
    
    # Suggested groups: present in >=50% of peers but MISSING from this user
    suggested_groups = []
    for g, cnt in grp_counts.items():
        freq = cnt / peers_count
        if freq >= 0.50 and g not in target_groups:
            suggested_groups.append({
                "group": g,
                "frequency": round(freq, 2),
                "count": cnt,
                "peers": peers_count
            })
    suggested_groups.sort(key=lambda x: x["frequency"], reverse=True)
            
    return {
        "peersCount": peers_count,
        "anomalies": sorted(anomalies, key=lambda x: x["frequency"]),
        "suggestedGroups": suggested_groups,
    }


@app.post("/api/rolemining/run")
def rolemining_run(req: RoleMiningRequest, background_tasks: BackgroundTasks, username: str = Depends(require_auth)):
    users = active_users(state["last_extract"]["users"] or [])
    if not users:
        raise HTTPException(status_code=400, detail="Esegui prima AD Extract")

    state["last_mining_params"] = {"n_clusters": req.n_clusters, "role_support": req.role_support}
    state["mining_dirty"] = True
    
    # Avvia mining in background
    ensure_last_mining(background_tasks)

    return {"status": "started"}


@app.get("/api/rolemining/last")
def rolemining_last(background_tasks: BackgroundTasks, username: str = Depends(require_auth)):
    ensure_last_mining(background_tasks)
    
    # Check cache first
    status = state.get("mining_status", "idle")
    cache_key = f"rolemining_last_{status}"
    cached = RESPONSE_CACHE.get(cache_key)
    if cached:
        return cached
    
    last = state.get("last_mining") or {}
    
    # Optimize payload: Convert dense matrix to sparse (list of active groups)
    # This significantly reduces JSON size for large datasets (e.g. 5000 users)
    matrix_dense = last.get("matrix") or {}
    matrix_sparse = {}
    for u, row in matrix_dense.items():
        # Only include groups with value 1 (or truthy)
        matrix_sparse[u] = [g for g, v in row.items() if v]
    
    result = {
        **last,
        "matrix": matrix_sparse,
        "status": status
    }
    
    # Cache only if mining is not running (stable result)
    if status != "running":
        RESPONSE_CACHE.set(cache_key, result, CACHE_TTL_MINING)
    
    return result


@app.get("/api/kpi")
def kpi(background_tasks: BackgroundTasks, username: str = Depends(require_auth)):
    # Check cache first
    cache_key = "kpi"
    cached = RESPONSE_CACHE.get(cache_key)
    if cached:
        return cached
    
    ensure_last_mining(background_tasks)
    last = state.get("last_mining") or {}
    kpi_data = last.get("kpi") or {}

    # If modelQuality is missing from stored KPI (e.g. from older run), recompute it.
    if ("modelQuality" not in kpi_data or kpi_data.get("modelQuality") is None) and last.get("matrix"):
        kpi_data = compute_kpis(last.get("users", []), last.get("clusters", []), last.get("matrix", {}))
        last["kpi"] = kpi_data

    # Overlay latest Smart AI Detection stats if available (aligns KPI with internal page)
    last_ai = state.get("last_ai_detection") or {}
    stats = last_ai.get("stats")
    if stats and "aiDetection" in stats:
        # Use copy to avoid mutating persistent state unexpectedly, or modify if intended.
        # Safe to modify for display.
        kpi_data["aiDetection"] = stats["aiDetection"]
        kpi_data["redundantAssignments"] = stats.get("totalAnomalies", 0)
        kpi_data["totalAssignments"] = stats.get("totalAssignments", 0)
        kpi_data["usersWithRedundancy"] = stats.get("usersWithAnomaly", 0)

    if not kpi_data:
        # Frontend-safe fallback: return empty KPI instead of 400.
        # This prevents dashboard hard-fail when dataset is not loaded yet.
        kpi_data = {
            "totalUsers": 0,
            "modelQuality": 0,
            "orphanGroupsCount": 0,
            "overprivilegedCount": 0,
            "zeroGroupCount": 0,
            "staleAccountCount": 0,
            "clusterQuality": 0,
            "clusteringQuality": 0,
            "aiDetection": 0,
            "redundantAssignments": 0,
            "totalAssignments": 0,
            "usersWithRedundancy": 0,
            "roleCoverage": 0,
        }
    
    # Cache KPI data
    RESPONSE_CACHE.set(cache_key, kpi_data, CACHE_TTL_KPI)
    return kpi_data

@app.post("/api/kpi/clear-cache")
def clear_kpi_cache():
    RESPONSE_CACHE.invalidate("kpi")
    return {"status": "ok"}


@app.get("/api/cache/stats")
def cache_stats(username: str = Depends(require_auth)):
    """Return cache statistics for monitoring."""
    return RESPONSE_CACHE.stats()


@app.post("/api/ai-detection/run")
def ai_detection_run(background_tasks: BackgroundTasks, username: str = Depends(require_auth)):
    """On-demand smart AI detection. Computes peer/dept/type anomalies and caches the result."""
    ensure_last_mining(background_tasks)
    last = state.get("last_mining") or {}
    matrix = last.get("matrix") or {}
    if not matrix:
        raise HTTPException(status_code=400, detail="No mining data. Run role mining first.")

    users = active_users(state.get("last_extract", {}).get("users") or [])
    result = run_smart_ai_detection(users, matrix)
    state["last_ai_detection"] = result

    # Invalidate KPI cache so dashboard picks up the new aiDetection %
    RESPONSE_CACHE.invalidate("kpi")

    return result


@app.get("/api/ai-detection/last")
def ai_detection_last(username: str = Depends(require_auth)):
    """Return last cached AI detection results (fast read)."""
    cached = state.get("last_ai_detection")
    if not cached:
        return {"status": "not_run", "items": [], "stats": {}}
    return cached


@app.get("/api/kpi/drilldown/{metric}")
def kpi_drilldown(metric: str, background_tasks: BackgroundTasks): #, username: str = Depends(require_auth)):
    ensure_last_mining(background_tasks)
    last = state.get("last_mining") or {}
    matrix = last.get("matrix") or {}
    clusters = last.get("clusters") or []

    if not matrix:
        raise HTTPException(status_code=400, detail="Nessun risultato: dataset vuoto o role mining non eseguibile")

    if metric == "overprivileged":
        payload = build_overprivileged_items(matrix, top_pct=10.0)
        return {"metric": metric, **payload}

    if metric == "ai-detection":
        # Always use smart detection logic
        cached = state.get("last_ai_detection")
        if not cached or cached.get("status") != "ready":
             # If not cached or ready, compute it on the fly (ensure consistency)
             users = active_users(state.get("last_extract", {}).get("users") or [])
             cached = run_smart_ai_detection(users, matrix)
             state["last_ai_detection"] = cached
             # Also update stats in main KPI if needed, but for now just return consistent data
        
        return {"metric": metric, "items": cached.get("items", []), "stats": cached.get("stats", {})}

    if metric == "cluster-quality":
        # Cluster Quality in dashboard is Data/Ingest Quality. 
        ingest = state.get("last_ingest_stats") or {}
        last_extract = state.get("last_extract") or {}
        users = last_extract.get("users") or []
        
        # Fallback if state is empty (for consistency)
        if not ingest and not users:
             ingest = {"rowsTotal": 0, "duplicateDisplayName": 0, "missingDepartment": 0}

        # DEBUG: Inspect users
        if users:
            print(f"DEBUG: First user object: {users[0]}")
            missing_preview = [u for u in users if not (u.get("department") or "").strip()][:3]
            print(f"DEBUG: Missing Dept Preview (Raw): {missing_preview}")

        # Missing fields detection
        missing_dept = [{"username": u["username"], "displayName": u.get("displayName") or u.get("display_name") or u["username"]} for u in users if not (u.get("department") or "").strip()]
        missing_br = [{"username": u["username"], "displayName": u.get("displayName") or u.get("display_name") or u["username"]} for u in users if not (u.get("businessRole") or "").strip()]
        
        duplicate_items = _duplicate_resolution_items()
        if not duplicate_items:
            by_display_name = defaultdict(list)
            for u in users:
                dn = (u.get("displayName") or u.get("display_name") or "").strip()
                if dn:
                    by_display_name[dn].append(u)
            for dn, rows in by_display_name.items():
                if len(rows) <= 1:
                    continue
                chosen_user = rows[0]
                chosen = {
                    "candidateId": f"user:{chosen_user.get('username')}",
                    "source": "current",
                    "displayName": dn,
                    "department": chosen_user.get("department"),
                    "businessRole": chosen_user.get("businessRole"),
                    "roles": chosen_user.get("groups") or [],
                    "lastLogin": chosen_user.get("lastLogin"),
                }
                alternatives = []
                for alt in rows[1:]:
                    alternatives.append(
                        {
                            "candidateId": f"user:{alt.get('username')}",
                            "source": "current",
                            "displayName": dn,
                            "department": alt.get("department"),
                            "businessRole": alt.get("businessRole"),
                            "roles": alt.get("groups") or [],
                            "lastLogin": alt.get("lastLogin"),
                        }
                    )
                duplicate_items.append(
                    {
                        "displayName": dn,
                        "chosenCandidateId": chosen["candidateId"],
                        "autoChosenCandidateId": chosen["candidateId"],
                        "chosen": chosen,
                        "alternatives": alternatives,
                        "count": len(rows),
                        "autoReason": None,
                    }
                )

        candidate_dup_extra_rows = sum(max(0, int(x.get("count") or 0) - 1) for x in duplicate_items)

        # Also account for duplicate rejects from ingest filtering (metrics only)
        rejects = state.get("last_rejects") or []
        reject_dup_count = 0
        for r in rejects:
            reason = str((r or {}).get("reason") or "")
            if "Duplicate displayName" in reason:
                reject_dup_count += 1

        # Identity integrity metrics (focused on source/identity data quality)
        def _norm_txt(s: Any) -> str:
            return re.sub(r"[^a-z0-9]+", "", str(s or "").lower())

        def _is_valid_email(s: str) -> bool:
            return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", str(s or "").strip()))

        def _is_valid_upn(s: str) -> bool:
            v = str(s or "").strip().lower()
            return bool(v and "@" in v and _is_valid_email(v))

        def _is_valid_empid(s: str) -> bool:
            v = str(s or "").strip()
            return bool(re.match(r"^[A-Za-z0-9_-]{4,}$", v))

        now_utc = datetime.now(timezone.utc)
        known_usernames = {str(u.get("username") or "").strip().lower() for u in users if u.get("username")}
        known_emails = {str(u.get("email") or "").strip().lower() for u in users if u.get("email")}
        known_upns = {str(u.get("upn") or "").strip().lower() for u in users if u.get("upn")}

        invalid_identity_users = []
        invalid_lastlogon_users = []
        orphan_ref_users = []
        inactive_mismatch_users = []

        by_email = defaultdict(list)
        by_upn = defaultdict(list)
        by_empid = defaultdict(list)
        dept_variants = defaultdict(set)
        role_variants = defaultdict(set)
        dept_users_by_variant = defaultdict(list)
        role_users_by_variant = defaultdict(list)

        inactive_markers = {"inactive", "disabled", "terminated", "offboarded", "left"}
        active_markers = {"active", "enabled"}

        for u in users:
            uname = str(u.get("username") or "").strip()
            disp = str(u.get("displayName") or uname).strip()
            dept = str(u.get("department") or "").strip()
            br = str(u.get("businessRole") or "").strip()
            email = str(u.get("email") or "").strip().lower()
            upn = str(u.get("upn") or "").strip().lower()
            empid = str(u.get("employeeId") or "").strip()
            manager = str(u.get("manager") or "").strip()
            last_login = str(u.get("lastLogin") or "").strip()
            status_ad = str(u.get("statusAd") or "").strip().lower()
            status_hr = str(u.get("statusHr") or "").strip().lower()

            row_user = {"username": uname, "displayName": disp}

            if email:
                by_email[email].append(row_user)
            if upn:
                by_upn[upn].append(row_user)
            if empid:
                by_empid[empid].append(row_user)

            if dept:
                dnorm = _norm_txt(dept)
                if dnorm:
                    dept_variants[dnorm].add(dept)
                    dept_users_by_variant[(dnorm, dept)].append(row_user)
            if br:
                rnorm = _norm_txt(br)
                if rnorm:
                    role_variants[rnorm].add(br)
                    role_users_by_variant[(rnorm, br)].append(row_user)

            invalid_identity = False
            if email and not _is_valid_email(email):
                invalid_identity = True
            if upn and not _is_valid_upn(upn):
                invalid_identity = True
            if empid and not _is_valid_empid(empid):
                invalid_identity = True
            if invalid_identity:
                invalid_identity_users.append(row_user)

            if last_login:
                try:
                    dt = datetime.fromisoformat(last_login.replace("Z", "+00:00"))
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=timezone.utc)
                    if dt > now_utc + timedelta(days=1):
                        invalid_lastlogon_users.append(row_user)
                except Exception:
                    invalid_lastlogon_users.append(row_user)

            if manager:
                m = manager.lower()
                if (m not in known_usernames) and (m not in known_emails) and (m not in known_upns):
                    orphan_ref_users.append(row_user)

            if status_ad and status_hr:
                ad_inactive = any(k in status_ad for k in inactive_markers)
                hr_inactive = any(k in status_hr for k in inactive_markers)
                ad_active = any(k in status_ad for k in active_markers)
                hr_active = any(k in status_hr for k in active_markers)
                mismatch = (ad_inactive and hr_active) or (hr_inactive and ad_active)
                if mismatch:
                    inactive_mismatch_users.append(row_user)

        collision_users = []
        for bucket in (by_email, by_upn, by_empid):
            for _, vals in bucket.items():
                if len(vals) > 1:
                    collision_users.extend(vals)

        dept_drift_users = []
        for norm_k, variants in dept_variants.items():
            if len(variants) > 1:
                for v in variants:
                    dept_drift_users.extend(dept_users_by_variant.get((norm_k, v), []))

        role_drift_users = []
        for norm_k, variants in role_variants.items():
            if len(variants) > 1:
                for v in variants:
                    role_drift_users.extend(role_users_by_variant.get((norm_k, v), []))

        def _dedupe_users(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
            out = []
            seen = set()
            for r in rows:
                k = (str(r.get("username") or "").lower(), str(r.get("displayName") or "").lower())
                if k in seen:
                    continue
                seen.add(k)
                out.append(r)
            return out

        invalid_identity_users = _dedupe_users(invalid_identity_users)
        invalid_lastlogon_users = _dedupe_users(invalid_lastlogon_users)
        orphan_ref_users = _dedupe_users(orphan_ref_users)
        inactive_mismatch_users = _dedupe_users(inactive_mismatch_users)
        collision_users = _dedupe_users(collision_users)
        dept_drift_users = _dedupe_users(dept_drift_users)
        role_drift_users = _dedupe_users(role_drift_users)

        ingest_rejects = len(state.get("last_rejects") or [])
        import_reject_events = max(
            ingest_rejects,
            int(ingest.get("missingDisplayName") or 0)
            + int(ingest.get("missingRoles") or 0)
            + int(ingest.get("missingDepartment") or 0)
            + int(ingest.get("missingBusinessRole") or 0),
        )
        import_reject_rate = round((import_reject_events / max(1, len(users))) * 100.0, 2)

        identity_cases = [
            {"id": "invalid_identity_keys", "label": "Chiavi identita non valide", "count": len(invalid_identity_users), "users": invalid_identity_users},
            {"id": "identity_collisions", "label": "Collisioni identita", "count": len(collision_users), "users": collision_users},
            {"id": "invalid_lastlogon", "label": "LastLogon non valido", "count": len(invalid_lastlogon_users), "users": invalid_lastlogon_users},
            {"id": "department_vocab_drift", "label": "Deriva vocabolario dipartimento", "count": len(dept_drift_users), "users": dept_drift_users},
            {"id": "businessrole_vocab_drift", "label": "Deriva vocabolario business role", "count": len(role_drift_users), "users": role_drift_users},
            {"id": "orphan_references", "label": "Riferimenti orfani", "count": len(orphan_ref_users), "users": orphan_ref_users},
            {"id": "inactive_source_mismatch", "label": "Mismatch stato AD/HR", "count": len(inactive_mismatch_users), "users": inactive_mismatch_users},
            {"id": "import_reject_rate", "label": "Import reject rate", "count": import_reject_events, "rate": import_reject_rate, "users": []},
        ]
        identity_total = sum(int(c.get("count") or 0) for c in identity_cases)

        # Merging stats to reflect calculation on the actual displayed data
        stats = ingest.copy()
        stats["rowsTotal"] = len(users)
        # Keep ingest duplicate metric semantics (extra duplicate rows), but never hide detected duplicates.
        stats_dup = int(ingest.get("duplicateDisplayName") or 0)
        stats["duplicateDisplayName"] = max(stats_dup, candidate_dup_extra_rows, reject_dup_count)
        stats["missingDepartment"] = len(missing_dept)
        stats["missingBusinessRole"] = len(missing_br)
        stats["identityIntegrityIssues"] = identity_total

        return {
            "metric": "cluster-quality",
            "stats": stats,
            "items": [
                {"type": "Duplicates", "count": len(duplicate_items), "users": duplicate_items},
                {"type": "Missing Department", "count": len(missing_dept), "users": missing_dept},
                {"type": "Missing Business Role", "count": len(missing_br), "users": missing_br},
                {"type": "Identity Integrity", "count": identity_total, "cases": identity_cases, "users": []},
            ],
            "rejects": state.get("last_rejects") or []
        }

    if metric == "model-quality":
        users = active_users(state.get("last_extract", {}).get("users") or [])
        last_mining = state.get("last_mining") or {}
        matrix = last_mining.get("matrix") or {}
        # Get all unique groups from source of truth
        groups_list = (state.get("last_extract") or {}).get("groups") or []
        if not groups_list and matrix:
            first = next(iter(matrix.values()))
            groups_list = list(first if isinstance(first, list) else first.keys())

        mq = compute_model_quality(users, matrix, groups_list)

        orphans = [{"groupName": g, "userCount": 0} for g in mq.get("orphansList", [])]

        return {
            "metric": metric,
            "modelQuality": mq.get("modelQuality", 0),
            "groupsIssues": orphans,
            "staleAccounts": mq.get("staleList", []),
            "zeroGroupsUsers": mq.get("zeroList", []),
            "overprivilegedUsers": mq.get("overprivilegedList", []),
            "policyViolations": mq.get("policyViolations", []),
            "ambiguousUsers": mq.get("ambiguousUsers", []),
            "qualityIndicators": mq.get("indicators", []),
            "density": mq.get("density", 0),
            "avgGeneralizationConfidence": mq.get("avgGeneralizationConfidence", 0),
            "manualOverrideEvents": mq.get("manualOverrideEvents", 0),
            "modelPreset": mq.get("modelPreset", state.get("dq_model_preset") or "manufacturing"),
            "availableModelPresets": sorted(MODEL_QUALITY_PRESETS.keys()),
        }
    
    raise HTTPException(status_code=404, detail="Unknown metric")


# (FastAPI and typing imports already at top of file)

# helper: costruisce righe per UI (tutti gli utenti)
def build_overprivileged_rows(matrix: Dict[str, Dict[str, int]], threshold: Optional[int]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for user, grants in (matrix or {}).items():
        groups = [g for g, v in (grants or {}).items() if int(v) == 1]
        n_groups = len(groups)
        over = bool(threshold is not None and n_groups >= int(threshold))
        rows.append({
            "user": user,
            "nGroups": n_groups,
            "over": over,
            "groups": groups,              # lista (meglio della stringa)
            "groupsText": ", ".join(groups) # comodo se vuoi visualizzarla subito
        })

    # ordinamento default lato backend (opzionale): Over desc, poi #Gruppi desc, poi user asc
    rows.sort(key=lambda r: (not r["over"], -r["nGroups"], r["user"]))
    return rows

@app.get("/api/drilldown/overprivileged")
def drilldown_overprivileged(nclusters: int = 8, rolesupport: float = 0.1):
    last = state.get("last_mining") or {}
    matrix = last.get("matrix") or {}
    # Use existing kpi if available, don't re-run full mining on every drilldown hit
    kpi = last.get("kpi") or {}

    thr = compute_over_threshold(matrix, pct=0.10)     # calcolata su TUTTI gli utenti
    rows = build_over_rows_only(matrix, thr)           # ma ritorni SOLO gli over

    return {
        "rows": rows,
        # opzionale: la soglia non la mostri in UI, ma può tornare utile per debug
        "threshold": thr,
        "totalUsers": len(matrix or {}),
        "overUsers": len(rows),
    }



@app.get("/api/logs")
def logs(username: str = Depends(require_auth)):
    return {"total": len(state["logs"]), "items": state["logs"]}


class RoleAssignRequest(BaseModel):
    username: str


@app.get("/api/businessroles")
def businessroles(username: str = Depends(require_auth)):
    # Check cache first
    cache_key = "businessroles"
    cached = RESPONSE_CACHE.get(cache_key)
    if cached:
        return cached
    
    users = active_users(state["last_extract"]["users"] or [])

    apply_business_roles(users)
    roles = sorted({u.get("businessRole", "Unassigned") for u in users})
    extra = state.get("business_roles", set())
    roles = sorted(set(roles).union(set(extra)))
    roles_info = []
    for r in roles:
        members = [u for u in users if u.get("businessRole") == r]
        # Optimization: include meta (color, groups) directly
        meta = state.get("role_meta", {}).get(r, {})
        roles_info.append({
            "role": r,
            "count": len(members),
            "color": meta.get("color", "#6aa6ff"),
            "groups": meta.get("groups", [])
        })
    result = {
        "roles": roles_info,
        "assignments": {u["username"]: u.get("businessRole", "Unassigned") for u in users},
    }
    
    RESPONSE_CACHE.set(cache_key, result, CACHE_TTL_ROLES)
    return result

class RoleCreateRequest(BaseModel):
    role: str

# role -> meta (color + groups)
state.setdefault("role_meta", {
    "IT": {"color": "#00B4FF", "groups": ["Azure", "GitLab"]},
    "HR": {"color": "#FF9F1C", "groups": ["HR", "Payroll"]},
})
state.setdefault("business_roles", set(["IT", "HR"]))



class RoleColorRequest(BaseModel):
    color: str  # "#RRGGBB"

class RoleGroupRequest(BaseModel):
    group: str

@app.get("/api/ad/groups")
def ad_groups(username: str = Depends(require_auth)):
    cache_key = "ad_groups"
    cached = RESPONSE_CACHE.get(cache_key)
    if cached:
        return cached

    users = state["last_extract"].get("users") or []
    groups = recompute_groups_from_users(users)
    result = {"groups": groups}
    RESPONSE_CACHE.set(cache_key, result, CACHE_TTL_USERS)
    return result


@app.get("/api/businessroles/{role}/meta")
def businessrole_meta(role: str, username: str = Depends(require_auth)):
    # Check cache first
    cache_key = f"role_meta_{role}"
    cached = RESPONSE_CACHE.get(cache_key)
    if cached:
        return cached
    
    meta = state.get("role_meta", {}).get(role, {"color": "#ffffff", "groups": []})
    result = {"role": role, "color": meta.get("color", "#ffffff"), "groups": meta.get("groups", [])}
    
    RESPONSE_CACHE.set(cache_key, result, CACHE_TTL_ROLES)
    return result

@app.post("/api/businessroles/{role}/color")
def businessrole_set_color(role: str, body: RoleColorRequest, username: str = Depends(require_auth)):
    state.setdefault("role_meta", {})
    state["role_meta"].setdefault(role, {"color": "#ffffff", "groups": []})
    state["role_meta"][role]["color"] = body.color
    state.setdefault("business_roles", set()).add(role)
    log("INFO", f"Role color set: {role} -> {body.color} by {username}")
    invalidate_hot_caches(roles=True)
    return {"ok": True}

@app.post("/api/businessroles/{role}/groups/add")
def businessrole_add_group(role: str, body: RoleGroupRequest, username: str = Depends(require_auth)):
    state.setdefault("role_meta", {})
    state["role_meta"].setdefault(role, {"color": "#ffffff", "groups": []})
    gs = set(state["role_meta"][role].get("groups", []))
    gs.add(body.group)
    state["role_meta"][role]["groups"] = sorted(gs)
    state.setdefault("business_roles", set()).add(role)
    log("INFO", f"Role group add: {role} + {body.group} by {username}")
    invalidate_hot_caches(roles=True, kpi=True, mining=True)
    return {"ok": True}

# ----------------------------
# BRDB Suggestions for UI
# ----------------------------
class SuggestionPickRequest(BaseModel):
    group: str

def brdb_suggest_groups_for_role(role: str, *, limit: int = 50, min_conf: float = 0.60) -> List[Dict[str, Any]]:
    """
    Ritorna gruppi suggeriti da assegnare al Business Role (escludendo quelli già presenti in role_meta[role].groups).
    Richiede che tu abbia già le funzioni BRDB: brdb_rebuild(), brdb_infer_group().
    """
    role = (role or "").strip()
    if not role:
        return []

    # Assicura DB pronto (se usi brdb_ready/brdb_rebuild)
    try:
        brdb_rebuild()
    except Exception:
        # se non hai brdb_rebuild ancora incollato, evita crash dell'endpoint
        return []

    meta = (state.get("role_meta") or {}).get(role) or {}
    already = set(meta.get("groups") or [])

    # candidati: tutti i gruppi visti a sistema
    all_groups = (state.get("last_extract") or {}).get("groups") or []
    out = []
    for g in all_groups:
        g = (g or "").strip()
        if not g or g in already:
            continue

        s = brdb_infer_group(g)  # -> {"role": "...", "confidence": 0..1, "evidence": ...}
        if s.get("role") != role:
            continue
        if float(s.get("confidence") or 0.0) < float(min_conf):
            continue

        out.append({
            "group": g,
            "confidence": float(s.get("confidence") or 0.0),
            "evidence": s.get("evidence", {}),
        })

    out.sort(key=lambda x: x["confidence"], reverse=True)
    return out[: int(limit)]

@app.get("/api/businessroles/{role}/suggestions")
def businessrole_suggestions(role: str, limit: int = 50, min_conf: float = 0.60, username: str = Depends(require_auth)):
    """
    Endpoint per la sezione "AI Suggestion" della pagina Business Role.
    """
    items = brdb_suggest_groups_for_role(role, limit=limit, min_conf=min_conf)
    return {"role": role, "total": len(items), "items": items}

@app.post("/api/businessroles/{role}/suggestions/select")
def businessrole_suggestion_select(role: str, body: SuggestionPickRequest, username: str = Depends(require_auth)):
    """
    Pulsante "Select": assegna direttamente il gruppo suggerito al role_meta del BR.
    (È equivalente a chiamare /api/businessroles/{role}/groups/add, ma comodo per UI.)
    """
    g = (body.group or "").strip()
    if not g:
        raise HTTPException(status_code=400, detail="Group vuoto")

    # Riusa la stessa logica del groups/add
    state.setdefault("role_meta", {})
    state["role_meta"].setdefault(role, {"color": "#ffffff", "groups": []})
    gs = set(state["role_meta"][role].get("groups", []))
    gs.add(g)
    state["role_meta"][role]["groups"] = sorted(gs)
    state.setdefault("business_roles", set()).add(role)

    # training "forte" (se hai brdb_learn_assignment); altrimenti ignoralo
    try:
        brdb_learn_assignment(role, [g], weight=10)
        record_llm_learning_event(
            actor=username,
            source="businessroles-suggestion",
            signal_type="brdb-assignment",
            entity=role,
            details={"group": g, "weight": 10},
        )
    except Exception:
        pass

    state["mining_dirty"] = True
    log("INFO", f"Suggestion selected: {role} + {g} by {username}")
    invalidate_hot_caches(roles=True, kpi=True, mining=True)
    return {"ok": True, "role": role, "group": g}


@app.post("/api/businessroles/{role}/groups/remove")
def businessrole_remove_group(role: str, body: RoleGroupRequest, username: str = Depends(require_auth)):
    state.setdefault("role_meta", {})
    state["role_meta"].setdefault(role, {"color": "#ffffff", "groups": []})
    gs = [g for g in state["role_meta"][role].get("groups", []) if g != body.group]
    state["role_meta"][role]["groups"] = gs
    log("INFO", f"Role group remove: {role} - {body.group} by {username}")
    invalidate_hot_caches(roles=True, kpi=True, mining=True)
    return {"ok": True}


@app.post("/api/businessroles/recalculate/groups")
def businessroles_recalculate_groups(username: str = Depends(require_auth)):
    """
    Recompute group -> Business Role assignment from current users.
    Each group is assigned to the role where it appears most frequently.
    """
    users = active_users(state.get("last_extract", {}).get("users") or [])
    apply_business_roles(users)

    # group -> role -> count
    group_role_counts: Dict[str, Dict[str, int]] = {}
    for u in users:
        role = (u.get("businessRole") or "Unassigned").strip()
        for g in (u.get("groups") or []):
            if not g:
                continue
            bucket = group_role_counts.setdefault(g, {})
            bucket[role] = int(bucket.get(role, 0)) + 1

    # Build exclusive assignment: each group belongs to the role with max count.
    role_groups: Dict[str, set] = {}
    for g, counts in group_role_counts.items():
        best_role = None
        best_count = -1
        for role, cnt in counts.items():
            if cnt > best_count:
                best_role = role
                best_count = cnt
        if not best_role:
            continue
        role_groups.setdefault(best_role, set()).add(g)

    state.setdefault("role_meta", {})
    state.setdefault("business_roles", set())
    for role in set(state["business_roles"]).union(set(role_groups.keys())):
        _ensure_role_registered(role)

    for role in state["business_roles"]:
        assigned = sorted(role_groups.get(role, set()))
        state["role_meta"].setdefault(role, {"color": "#ffffff", "groups": []})
        state["role_meta"][role]["groups"] = assigned

    state["mining_dirty"] = True
    log("INFO", f"Business role groups recalculated by {username}")
    invalidate_hot_caches(roles=True, kpi=True, mining=True)
    return {
        "ok": True,
        "rolesUpdated": len(state["business_roles"]),
        "groupsAssigned": len(group_role_counts),
    }


@app.post("/api/businessroles/create")
def businessrole_create(body: RoleCreateRequest, username: str = Depends(require_auth)):
    role = body.role.strip()
    if not role:
        raise HTTPException(status_code=400, detail="Role vuoto")

    # Crea il ruolo senza assegnare utenti: basta “registrarlo”
    state.setdefault("business_roles", set())
    if isinstance(state["business_roles"], list):
        state["business_roles"] = set(state["business_roles"])
    state["business_roles"].add(role)

    log("INFO", f"Business role created: {role} by {username}")
    invalidate_hot_caches(roles=True)
    return {"ok": True, "role": role}

class ChooseCsvRowRequest(BaseModel):
    displayNameRaw: str
    rowId: str

@app.post("/api/csv/duplicates/choose")
def choose_csv_duplicate_row(body: ChooseCsvRowRequest, username: str = Depends(require_auth)):
    dn_raw = body.displayNameRaw
    row_id = body.rowId

    rows = state.get("last_csv_rows") or []
    rec = next((r for r in rows if r.get("rowId") == row_id and r.get("displayNameRaw") == dn_raw), None)
    if not rec:
        raise HTTPException(status_code=404, detail="Row not found")

    # salva scelta
    state.setdefault("csv_choice_by_dn", {})
    state["csv_choice_by_dn"][dn_raw] = row_id

    # applica override “a sistema” sull’utente (per displayName)
    users = state.get("last_extract", {}).get("users") or []
    dn_clean = rec.get("displayName") or ""
    br = rec.get("businessRole") or ""
    roles = rec.get("roles") or []

    uobj = next((u for u in users if (u.get("displayName") or "").strip() == dn_clean), None)
    if not uobj:
        raise HTTPException(status_code=400, detail="User not found in last_extract")

    dept = rec.get("department")
    if dept:
        uobj["department"] = dept

    uobj["groups"] = sorted(set(roles))
    if br:
        uobj["businessRole"] = br
        state.setdefault("user_business_role", {})
        state["user_business_role"][uobj["username"]] = br
    record_manual_user_change(
        actor=username,
        username=uobj.get("username"),
        display_name=uobj.get("displayName"),
        action="resolve-csv-duplicate",
        source="cluster-quality",
        details={"displayNameRaw": dn_raw, "chosenRowId": row_id},
    )

    log("INFO", f"CSV duplicate resolved: '{dn_raw}' -> rowId={row_id} by {username}")
    invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)
    return {"ok": True, "username": uobj["username"], "chosenRowId": row_id}


@app.get("/api/businessroles/{role}")
def businessrole_detail(role: str, username: str = Depends(require_auth)):
    users = active_users(state["last_extract"]["users"] or [])
    apply_business_roles(users)
    members = [u for u in users if u.get("businessRole") == role]
    return {"role": role, "users": members}

@app.get("/api/ingest/conflicts/duplicate-displayname")
def conflicts_duplicate_displayname(username: str = Depends(require_auth)):
    """
    Ritorna la lista dei displayName che hanno più di un candidato (conflitto).
    """
    items = _duplicate_resolution_items()
    for item in items:
        item["rows"] = [item.get("chosen")] + (item.get("alternatives") or [])
    items.sort(key=lambda x: len(x.get("rows") or []), reverse=True)
    return {"items": items}


class ChooseDuplicateRequest(BaseModel):
    displayName: str
    candidateId: str

@app.post("/api/ingest/conflicts/duplicate-displayname/choose")
def choose_duplicate(body: ChooseDuplicateRequest, username: str = Depends(require_auth)):
    state.setdefault("choice_by_displayName", {})
    state["choice_by_displayName"][body.displayName] = body.candidateId

    cand = next((c for c in state.get("ingest_candidates", [])
                 if c["candidateId"] == body.candidateId and c["displayName"] == body.displayName), None)
    if cand:
        apply_choice_for_displayname(
            display_name=body.displayName,
            chosen_business_role=cand.get("businessRole"),
            chosen_roles=cand.get("roles") or [],
        )
        users = state.get("last_extract", {}).get("users") or []
        uobj = next((u for u in users if (u.get("displayName") or "").strip() == body.displayName and not u.get("excluded")), None)
        if uobj:
            record_manual_user_change(
                actor=username,
                username=uobj.get("username"),
                display_name=uobj.get("displayName"),
                action="resolve-duplicate",
                source="cluster-quality",
                details={"candidateId": body.candidateId},
            )

    _record_duplicate_feedback(body.displayName, body.candidateId, actor=username)
    log("INFO", f"Duplicate resolved: {body.displayName} -> {body.candidateId} by {username}")
    invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)
    return {"ok": True}


@app.get("/api/data-quality/rules/suggestions")
def data_quality_rule_suggestions(username: str = Depends(require_auth)):
    return {
        "items": build_dq_rule_suggestions(),
        "activeRules": state.get("dq_rules") or {},
    }


@app.post("/api/data-quality/rules/suggestions/{rule_id}/apply")
def apply_data_quality_rule(rule_id: str, username: str = Depends(require_auth)):
    rid = (rule_id or "").strip()
    if not rid:
        raise HTTPException(status_code=400, detail="rule_id required")

    suggestions = {x.get("ruleId"): x for x in build_dq_rule_suggestions()}
    item = suggestions.get(rid)
    if not item:
        raise HTTPException(status_code=404, detail="Rule suggestion not found")

    rules = dict(state.get("dq_rules") or {})
    preview = item.get("preview") or {}
    for k, v in preview.items():
        rules[k] = v
    # Keep canonical duplicate ranking policy required by product.
    rules["duplicate_resolution_order"] = REQUIRED_DUPLICATE_ORDER.copy()
    state["dq_rules"] = rules

    log("INFO", f"DQ rule applied: {rid} by {username}")
    return {"ok": True, "ruleId": rid, "dqRules": rules}


@app.get("/api/data-quality/model/presets")
def data_quality_model_presets(username: str = Depends(require_auth)):
    active = (state.get("dq_model_preset") or "manufacturing").strip().lower()
    return {
        "activePreset": active,
        "availablePresets": sorted(MODEL_QUALITY_PRESETS.keys()),
        "weights": get_active_model_weights(),
    }


@app.post("/api/data-quality/model/presets/{preset}/apply")
def apply_data_quality_model_preset(preset: str, username: str = Depends(require_auth)):
    p = (preset or "").strip().lower()
    if p not in MODEL_QUALITY_PRESETS:
        raise HTTPException(status_code=404, detail="Unknown model preset")
    state["dq_model_preset"] = p
    state["dq_model_weights"] = dict(MODEL_QUALITY_PRESETS[p])
    RESPONSE_CACHE.invalidate("kpi")
    log("INFO", f"Model quality preset applied: {p} by {username}")
    return {"ok": True, "activePreset": p, "weights": state["dq_model_weights"]}



@app.get("/api/ingest/conflicts/{kind}")
def ingest_conflicts(kind: str, username: str = Depends(require_auth)):
    if kind == "duplicate-displayname":
        return conflicts_duplicate_displayname(username)
    raise HTTPException(status_code=404, detail="Unknown conflict kind")

class ChooseConflictRequest(BaseModel):
    kind: str                   # "duplicate-displayname"
    displayName: str
    candidateId: str

@app.post("/api/ingest/conflicts/choose")
def choose_conflict(body: ChooseConflictRequest, username: str = Depends(require_auth)):
    if body.kind != "duplicate-displayname":
        raise HTTPException(status_code=404, detail="Unknown conflict kind")

    state.setdefault("choice_by_displayName", {})
    state["choice_by_displayName"][body.displayName] = body.candidateId

    # Applica subito la scelta a “last_extract” (utente effettivo)

    return {"ok": True}


@app.post("/api/businessroles/{role}/add")
def businessrole_add(role: str, body: RoleAssignRequest, username: str = Depends(require_auth)):
    # assegna (e rimuove da eventuale ruolo precedente)
    m = state.get("user_business_role", {})
    m[body.username] = role
    state["user_business_role"] = m
    # aggiorna cache estrazione per riflettere subito la modifica in UI
    users = state["last_extract"]["users"] or []
    apply_business_roles(users)
        # BRDB training: quando assegni a mano è “ground truth”
    u = next((x for x in users if x.get("username") == body.username), None)
    if u:
        brdb_learn_assignment(role, u.get("groups") or [], weight=10)
        record_llm_learning_event(
            actor=username,
            source="businessroles-add-user",
            signal_type="brdb-assignment",
            entity=body.username,
            details={"businessRole": role, "groupsCount": len(u.get("groups") or []), "weight": 10},
        )
        record_manual_user_change(
            actor=username,
            username=body.username,
            display_name=u.get("displayName"),
            action="assign-business-role",
            source="business-roles",
            details={"businessRole": role},
        )

    log("INFO", f"Business role set: {body.username} -> {role} by {username}")
    invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)
    return {"ok": True, "username": body.username, "role": role}

def _slug_username(display_name: str) -> str:
    s = (display_name or "").strip().lower()
    s = re.sub(r"\s+", ".", s)
    s = re.sub(r"[^a-z0-9._-]", "", s)
    return s or "user"


# (DEPT_MINCONF already defined at line 479)

def _ensure_role_registered(role: str) -> None:
    role = (role or "").strip()
    if not role:
        return

    state.setdefault("role_meta", {})
    state.setdefault("business_roles", set())

    if role not in state["role_meta"]:
        r = random.randint(100, 255)
        g = random.randint(100, 255)
        b = random.randint(100, 255)
        color = f"{r:02x}{g:02x}{b:02x}".upper()
        if not color.startswith("#"):
            color = "#" + color
        state["role_meta"][role] = {"color": color, "groups": []}

    state["business_roles"].add(role)

def apply_department_mapping_if_missing(users: list[dict]) -> None:
    # usa BRDB già presente (inference sui gruppi)
    brdb_rebuild()

    by_dept = defaultdict(list)
    for u in (users or []):
        dept = (u.get("department") or "").strip()
        if dept:
            by_dept[dept].append(u)

    for dept, members in by_dept.items():
        weights = defaultdict(float)

        for u in members:
            s = brdb_infer_groupset(u.get("groups") or [])
            role = (s.get("role") or "Unassigned").strip()
            conf = float(s.get("confidence") or 0.0)
            if role and role != "Unassigned" and conf > 0:
                weights[role] += conf
            for rr, sc in _br_assignment_rule_scores_for_user(u).items():
                weights[rr] += float(sc)

        if weights:
            best_role, best_w = max(weights.items(), key=lambda x: x[1])
            total = sum(weights.values()) or 1.0
            dept_conf = best_w / total
        else:
            best_role, dept_conf = "Unassigned", 0.0

        chosen_role = best_role if (best_role != "Unassigned" and dept_conf >= DEPT_MINCONF) else dept
        _ensure_role_registered(chosen_role)

        # assegna SOLO se l'utente non ha già BR assegnato
        state.setdefault("user_business_role", {})
        for u in members:
            uname = u.get("username")
            if not uname:
                continue

            already = (state.get("user_business_role") or {}).get(uname)
            if already:
                continue

            # se per qualche motivo è già valorizzato sul record utente, non sovrascrivere
            if (u.get("businessRole") or "").strip():
                continue

            user_rule_scores = _br_assignment_rule_scores_for_user(u)
            if user_rule_scores:
                boosted_role = max(user_rule_scores.items(), key=lambda x: x[1])[0]
                _ensure_role_registered(boosted_role)
                state["user_business_role"][uname] = boosted_role
            else:
                state["user_business_role"][uname] = chosen_role

    # Also evaluate rule-based assignment for users without department.
    state.setdefault("user_business_role", {})
    for u in (users or []):
        uname = u.get("username")
        if not uname:
            continue
        already = (state.get("user_business_role") or {}).get(uname)
        if already:
            continue
        if (u.get("businessRole") or "").strip():
            continue
        rule_scores = _br_assignment_rule_scores_for_user(u)
        if rule_scores:
            forced_role = max(rule_scores.items(), key=lambda x: x[1])[0]
            _ensure_role_registered(forced_role)
            state["user_business_role"][uname] = forced_role



@app.post("/api/import/csv")
async def import_csv(file: UploadFile = File(...), background_tasks: BackgroundTasks = None, username: str = Depends(require_auth)):
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="File vuoto")

    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text), delimiter=";", skipinitialspace=True)

    if reader.fieldnames:
        reader.fieldnames = [h.strip() for h in reader.fieldnames if h is not None]

    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV senza header")

    norm_fields = { (h or "").strip().lower() for h in reader.fieldnames }
    if "displayname" not in norm_fields:
        raise HTTPException(
            status_code=400,
            detail=f"Headers richiesti: ['DisplayName']; trovati: {reader.fieldnames}",
        )

    def _norm_header(h: str) -> str:
        return (h or "").strip().lower()

    def _get_any(row_ci: dict, keys: list[str]) -> str:
        for k in keys:
            if k in row_ci and row_ci.get(k) is not None:
                return str(row_ci.get(k))
        return ""

    CSV_KEYS = {
        "displayName": ["displayname", "display name", "name", "utente", "user"],
        "username": ["username", "userprincipalname", "samaccountname", "login", "accountname"],
        "department": ["department", "dept", "dipartimento", "area", "funzione"],
        "businessRole": ["businessrole", "business role", "br", "ruolo business", "ruolo_business"],
        "roles": ["ruoli", "roles", "groups", "gruppi", "entitlements"],
        "accountType": ["accounttype", "account type", "tipo utente", "tipo_utente", "type"],
        "lastLogin": ["lastlogin", "last login", "last_logon", "lastlogon", "ultimo accesso", "ultimologin"],
        "email": ["email", "mail", "emailaddress", "posta"],
        "upn": ["upn", "user principal name", "userprincipalname"],
        "employeeId": ["employeeid", "employee id", "matricola", "badgeid"],
        "manager": ["manager", "owner", "responsabile"],
        "statusAd": ["statusad", "adstatus", "accountstatusad", "statoad"],
        "statusHr": ["statushr", "hrstatus", "accountstatushr", "statohr"],
        "orphanGroups": ["orphangroups", "orphan_groups", "cataloggroups", "catalog_groups", "groupscatalog"],
    }

    def _extract_csv_fields(row: dict) -> tuple:
        row_ci = {_norm_header(k): v for k, v in (row or {}).items()}
        dnraw = _get_any(row_ci, CSV_KEYS["displayName"])
        usernameraw = _get_any(row_ci, CSV_KEYS["username"])
        deptraw = _get_any(row_ci, CSV_KEYS["department"])
        brraw = _get_any(row_ci, CSV_KEYS["businessRole"])
        rolesraw = _get_any(row_ci, CSV_KEYS["roles"])
        type_raw = _get_any(row_ci, CSV_KEYS["accountType"])
        last_login_raw = _get_any(row_ci, CSV_KEYS["lastLogin"])
        email_raw = _get_any(row_ci, CSV_KEYS["email"])
        upn_raw = _get_any(row_ci, CSV_KEYS["upn"])
        employee_id_raw = _get_any(row_ci, CSV_KEYS["employeeId"])
        manager_raw = _get_any(row_ci, CSV_KEYS["manager"])
        status_ad_raw = _get_any(row_ci, CSV_KEYS["statusAd"])
        status_hr_raw = _get_any(row_ci, CSV_KEYS["statusHr"])
        orphan_groups_raw = _get_any(row_ci, CSV_KEYS["orphanGroups"])

        dn = (dnraw or "").strip()
        preferred_username = (usernameraw or "").strip()
        dept = (deptraw or "").strip()
        br = (brraw or "").strip()
        roles = (rolesraw or "").strip()
        acct_type = (type_raw or "WhiteCollar").strip()
        last_login = _normalize_last_login((last_login_raw or "").strip() or None)
        email = (email_raw or "").strip().lower()
        upn = (upn_raw or "").strip().lower()
        employee_id = (employee_id_raw or "").strip()
        manager = (manager_raw or "").strip()
        status_ad = (status_ad_raw or "").strip()
        status_hr = (status_hr_raw or "").strip()
        orphan_groups = [g.strip() for g in (orphan_groups_raw or "").split(",") if g and g.strip()]

        if (not dept) and dn and ("," in dn):
            dn, dept = [x.strip() for x in dn.split(",", 1)]

        if not br:
            br = dept

        return (
            dnraw,
            dn,
            preferred_username,
            dept,
            br,
            roles,
            acct_type,
            last_login,
            email,
            upn,
            employee_id,
            manager,
            status_ad,
            status_hr,
            orphan_groups,
        )

    state.setdefault("ingest_sources", {})
    state["ingest_sources"]["csv"] = []
    state["last_csv_rows"] = []
    state["csv_choice_by_dn"] = {}
    state["csv_rows_by_dn"] = defaultdict(list)

    csv_rows_total = 0
    csv_dup_dn_rows = 0
    csv_missing_displayname = 0
    csv_missing_department = 0
    csv_missing_businessrole = 0
    csv_missing_roles = 0
    csv_orphan_groups_catalog: set[str] = set()
    csv_rejects: List[Dict[str, Any]] = []
    reject_empty_groups = bool((state.get("dq_rules") or {}).get("reject_empty_groups"))

    existing_users_list = state.get("last_extract", {}).get("users", [])
    csv_candidates: List[Dict[str, Any]] = []
    state.setdefault("choice_by_displayName", {})
    state["duplicate_autoselect"] = {}

    for row in reader:
        csv_rows_total += 1
        row_id = f"csv:{csv_rows_total}"
        (
            dnraw,
            dn,
            preferred_username,
            dept,
            br,
            roles,
            acct_type,
            last_login,
            email,
            upn,
            employee_id,
            manager,
            status_ad,
            status_hr,
            orphan_groups,
        ) = _extract_csv_fields(row)
        for g in orphan_groups:
            csv_orphan_groups_catalog.add(g)

        if not dn:
            csv_missing_displayname += 1
            continue

        if not dept:
            csv_missing_department += 1
        if not br:
            csv_missing_businessrole += 1

        parsed_roles = [g.strip() for g in (roles or "").split(",") if g.strip()]
        if not parsed_roles:
            csv_missing_roles += 1
            if reject_empty_groups:
                csv_rejects.append(
                    {
                        "source": "csv",
                        "reason": "Missing groups (rejected by dq rule)",
                        "user": {"displayName": dn, "department": dept, "businessRole": br},
                        "ts": datetime.now(timezone.utc).isoformat(),
                    }
                )
                continue

        # Favor CSV provided type if it's not the default 'WhiteCollar', otherwise fallback to heuristic
        atype = classify_account(dn, dept, row.get("EmployeeType", ""))
        final_type = acct_type if acct_type not in ["", "WhiteCollar"] else atype

        user_payload = {
            "displayName": dn,
            "groups": parsed_roles,
            "department": dept or None,
            "businessRole": br or None,
            "excluded": False,
            "lastLogin": last_login,
            "accountType": final_type,
            "preferredUsername": preferred_username or None,
            "email": email or None,
            "upn": upn or None,
            "employeeId": employee_id or None,
            "manager": manager or None,
            "statusAd": status_ad or None,
            "statusHr": status_hr or None,
        }

        rec = {
            "rowId": row_id,
            "displayName": dn,
            "displayNameRaw": dnraw,
            "preferredUsername": preferred_username,
            "businessRole": br,
            "department": dept,
            "lastLogin": last_login,
            "roles": parsed_roles,
            "email": email,
            "upn": upn,
            "employeeId": employee_id,
            "manager": manager,
            "rawLine": f"{dnraw};{dept};{roles}",
        }
        state["last_csv_rows"].append(rec)
        state["csv_rows_by_dn"][dnraw].append(row_id)
        state["csv_choice_by_dn"].setdefault(dnraw, row_id)
        csv_candidates.append({"rowId": row_id, "rec": rec, "user": user_payload})

        candidate = _mk_candidate(
            source="csv",
            candidate_id=row_id,
            display_name=dn,
            business_role=br,
            roles=parsed_roles,
            raw=rec["rawLine"],
            department=dept,
            last_login=last_login,
        )
        state["ingest_sources"]["csv"].append(candidate)
        state["choice_by_displayName"].setdefault(candidate["displayName"], candidate["candidateId"])

    profile_rows = []
    for u in existing_users_list:
        profile_rows.append(
            {
                "department": (u.get("department") or "").strip(),
                "groups": list(u.get("groups") or []),
            }
        )
    for c in csv_candidates:
        profile_rows.append(
            {
                "department": (c["user"].get("department") or "").strip(),
                "groups": list(c["user"].get("groups") or []),
            }
        )
    dept_profile = _build_dept_group_profile(profile_rows)

    by_dn = defaultdict(list)
    for c in csv_candidates:
        dn_key = (c["user"].get("displayName") or "").strip().lower()
        if dn_key:
            by_dn[dn_key].append(c)

    csv_dup_dn_rows = 0
    for rows in by_dn.values():
        if len(rows) > 1:
            csv_dup_dn_rows += (len(rows) - 1)

    taken_usernames = {str(u.get("username") or "").strip() for u in existing_users_list if u.get("username")}
    new_users: List[Dict[str, Any]] = []
    for _, rows in by_dn.items():
        scored_rows = []
        for item in rows:
            score = _score_duplicate_candidate(item["user"], dept_profile)
            scored_rows.append({**item, "score": score})
        scored_rows.sort(key=lambda x: x["score"]["rank"], reverse=True)
        winner = scored_rows[0]
        winner_user = dict(winner["user"])

        preferred_uname = (winner_user.get("preferredUsername") or "").strip()
        base = _slug_username(preferred_uname or winner_user.get("displayName") or "")
        uname = base
        i = 2
        while uname in taken_usernames:
            uname = f"{base}{i}"
            i += 1
        taken_usernames.add(uname)
        winner_user["username"] = uname
        new_users.append(winner_user)

        if len(scored_rows) > 1:
            display_name = winner_user.get("displayName") or winner["rec"].get("displayName") or ""
            state["choice_by_displayName"][display_name] = winner["rowId"]
            state["duplicate_autoselect"][display_name] = {
                "candidateId": winner["rowId"],
                "reason": winner["score"]["reason"],
                "alternatives": [
                    {
                        "candidateId": s["rowId"],
                        "reason": s["score"]["reason"],
                    }
                    for s in scored_rows[1:]
                ],
            }
        else:
            display_name = winner_user.get("displayName") or winner["rec"].get("displayName") or ""
            state["choice_by_displayName"].setdefault(display_name, winner["rowId"])

    # Merge with existing users - match ONLY by displayName.
    # Keep all local users; update only same displayName; add others.
    existing_by_dn = {}
    for u in existing_users_list:
        dn = (u.get("displayName") or "").strip().lower()
        if dn and dn not in existing_by_dn:
            existing_by_dn[dn] = u

    added_users = 0
    updated_users = 0
    created_brs_count = 0

    for user in new_users:
        uname = user["username"]
        dn_key = (user.get("displayName") or "").strip().lower()
        
        # Match ONLY by displayName
        existing_user = existing_by_dn.get(dn_key)
        
        if existing_user:
            # REPLACE groups and other fields with new values from import
            existing_user["groups"] = user.get("groups") or []
            if user.get("businessRole"):
                existing_user["businessRole"] = user["businessRole"]
            if user.get("department"):
                existing_user["department"] = user["department"]
            if user.get("displayName"):
                existing_user["displayName"] = user["displayName"]
            if user.get("accountType"):
                existing_user["accountType"] = user["accountType"]
            if user.get("lastLogin"):
                existing_user["lastLogin"] = user["lastLogin"]
            for k in ("email", "upn", "employeeId", "manager", "statusAd", "statusHr"):
                if user.get(k) is not None:
                    existing_user[k] = user.get(k)
            
            # Update username if it changed
            if existing_user.get("username") != uname:
                existing_user["username"] = uname
            
            updated_users += 1
        else:
            # New user - add to indexes
            new_user = user.copy()
            existing_users_list.append(new_user)
            if dn_key:
                existing_by_dn[dn_key] = new_user
            added_users += 1

    merged_users = existing_users_list
    state["last_extract"]["users"] = merged_users
    computed_groups = set(recompute_groups_from_users(merged_users))
    computed_groups.update(csv_orphan_groups_catalog)
    state["last_extract"]["groups"] = sorted(computed_groups)
    state["last_extract"]["ou"] = "MERGED"
    state["last_extract"]["ts"] = time.time()

    # CRITICAL: Populate state["user_business_role"] with CSV Business Roles BEFORE auto-assignment
    # This ensures the preservation logic in apply_department_mapping has data to preserve
    state.setdefault("user_business_role", {})
    for u in merged_users:
        uname = u.get("username")
        br = (u.get("businessRole") or "").strip()
        if uname and br and br != "Unassigned":
            state["user_business_role"][uname] = br

    touched_depts = {u.get("department") for u in new_users if u.get("department")}
    csv_auto_resolved_duplicates = len(state.get("duplicate_autoselect") or {})

    state["last_ingest_stats"] = {
        "source": "csv",
        "rowsTotal": csv_rows_total,
        "rowsKept": len(merged_users),
        "duplicateDisplayName": csv_dup_dn_rows,
        "missingDepartment": csv_missing_department,
        "missingBusinessRole": csv_missing_businessrole,
        "missingDisplayName": csv_missing_displayname,
        "missingUsername": 0,
        "missingRoles": csv_missing_roles,
        "orphanGroupsCatalog": len(csv_orphan_groups_catalog),
        "autoResolvedDuplicateUsers": csv_auto_resolved_duplicates,
        "ts": datetime.now(timezone.utc).isoformat(),
    }

    state["last_csv_stats"] = {
        "csvRowsTotal": csv_rows_total,
        "csvRowsMissingBR": csv_missing_businessrole,
        "csvDuplicateDisplayNameRows": csv_dup_dn_rows,
        "by": username,
        "ts": datetime.now(timezone.utc).isoformat(),
    }

    state["last_rejects"] = csv_rejects
    state["mining_dirty"] = True
    invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)
    snapshot_ts = str((state.get("last_extract") or {}).get("ts") or "")
    touched_depts_list = sorted([d for d in touched_depts if d])
    if background_tasks:
        background_tasks.add_task(
            run_post_csv_snapshot_logic_background,
            snapshot_ts,
            username,
            touched_depts_list,
        )
    else:
        run_post_csv_snapshot_logic_background(snapshot_ts, username, touched_depts_list)

    return {
        "ok": True,
        "snapshotReady": True,
        "processingInBackground": True,
        "addedUsers": added_users,
        "updatedUsers": updated_users,
        "updatedByDisplayName": updated_users,
        "totalUsers": len(merged_users),
        "rowsTotal": csv_rows_total,
        "csvDuplicateDisplayNameRows": csv_dup_dn_rows,
        "autoResolvedDuplicateUsers": csv_auto_resolved_duplicates,
        "newBusinessRoles": 0,
    }



def apply_all_choices_to_last_extract() -> None:
    rebuild_ingest_candidates()
    for dn, cid in (state.get("choice_by_displayName") or {}).items():
        cand = next(
            (c for c in (state.get("ingest_candidates") or [])
             if c.get("candidateId") == cid and c.get("displayName") == dn),
            None
        )
        if cand:
            apply_choice_for_displayname(
                display_name=dn,
                chosen_business_role=cand.get("businessRole"),
                chosen_roles=cand.get("roles") or [],
            )

    # riallinea lista gruppi (tiene conto degli excluded)
    users = state.get("last_extract", {}).get("users") or []
    state["last_extract"]["groups"] = recompute_groups_from_users(users)



# (_slug_username already defined at line 2299)
# (datetime import already at top of file)

def applyimportrow(displayname: str, businessrole: str, ruoli: str, department: str = ""):
    displayname = (displayname or "").strip()
    businessrole = (businessrole or "").strip()
    ruoli = (ruoli or "").strip()
    department = (department or "").strip()

    if not displayname:
        return {"skipped": True}

    # Username stabile
    uname = _slug_username(displayname)
    
    # Gruppi
    groups = [g.strip() for g in (ruoli or "").split(",") if g.strip()]
    
    # User object
    user = {
        "username": uname,
        "displayName": displayname,
        "groups": groups,
        "department": department or None,
        "businessRole": businessrole or None,
        "excluded": False,
    }
    
    # Merge logic (simplified for single row)
    last_extract = state.setdefault("last_extract", {"users": [], "groups": [], "ou": "IMPORT", "ts": None})
    users = last_extract.get("users", [])
    
    existing = next((u for u in users if u.get("username") == uname), None)
    created_user = False
    created_role = False
    added_groups = 0
    
    if existing:
        old_groups = set(existing.get("groups") or [])
        new_groups = set(groups)
        added_groups = len(new_groups - old_groups)
        existing["groups"] = sorted(old_groups | new_groups)
        if businessrole:
            existing["businessRole"] = businessrole
        if department:
            existing["department"] = department
    else:
        users.append(user)
        created_user = True
        added_groups = len(groups)
        
    if businessrole:
        business_roles = state.setdefault("business_roles", set())
        if businessrole not in business_roles:
            _ensure_role_registered(businessrole)
            created_role = True

    state["mining_dirty"] = True
    return {
        "ok": True,
        "created_user": created_user,
        "created_role": created_role,
        "added_groups": added_groups
    }


@app.post("/api/import/xlsx")
async def import_xlsx(file: UploadFile = File(...), username: str = Depends(require_auth)):
    raw = await file.read()
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active  # primo foglio

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return {"ok": False, "error": "Empty Excel"}

    # mappa colonne (richiede intestazioni esatte)
    cols = {str(v).strip(): i for i, v in enumerate(header) if v is not None}
    required = ["DisplayName", "BusinessRole", "Ruoli"]
    if any(c not in cols for c in required):
        return {"ok": False, "error": f"Missing headers: {required}", "found": list(cols.keys())}

    created_users = 0
    created_roles = 0
    assigned_users = 0
    added_groups_total = 0

    for r in rows:
        dn = r[cols["DisplayName"]] if cols["DisplayName"] < len(r) else ""
        br = r[cols["BusinessRole"]] if cols["BusinessRole"] < len(r) else ""
        ru = r[cols["Ruoli"]] if cols["Ruoli"] < len(r) else ""

        out = applyimportrow(str(dn or ""), str(br or ""), str(ru or ""))
        if out.get("skipped"):
            continue
        created_users += 1 if out.get("created_user") else 0
        created_roles += 1 if out.get("created_role") else 0
        assigned_users += 1
        added_groups_total += int(out.get("added_groups") or 0)

    return {
        "ok": True,
        "created_users": created_users,
        "created_roles": created_roles,
        "assigned_users": assigned_users,
        "added_groups": added_groups_total
    }


# =============================================================================
# ML ENGINE API ENDPOINTS
# =============================================================================

@app.get("/api/config/ad-fields")
def get_ad_fields(username: str = Depends(require_auth)):
    """Return list of all AD fields found during last import."""
    fields = state.get("ad_available_fields", [])
    # Ensure default fields are always present
    defaults = {"displayName", "department", "title", "employeeType", "company", "manager", "mail"}
    combined = sorted(list(set(fields) | defaults))
    return {"fields": combined}

@app.get("/api/brdb/status")
def brdb_status_api(username: str = Depends(require_auth)):
    """
    Ritorna lo stato del calcolo BRDB (background).
    """
    return {
        "calculated": state.get("brdb_calculated", False),
        "min_confidence": state.get("brdb_min_confidence", BRDB_MIN_CONF),
        "last_update": state.get("brdb_last_update")
    }

@app.get("/api/ml/status")
def ml_status(username: str = Depends(require_auth)):
    """Return ML engine status and metrics."""
    return ml_engine.get_status()


@app.post("/api/ml/train")
def ml_train(username: str = Depends(require_auth)):
    """Trigger ML model training from accumulated data."""
    # Build training data from existing users
    users = state.get("last_extract", {}).get("users") or []
    training_data = []
    
    for u in users:
        if u.get("accountType"):
            training_data.append({
                "display_name": u.get("displayName", ""),
                "ou": u.get("department", ""),
                "employee_type": "",  # Not always available
                "account_type": u.get("accountType", "Internal"),
            })
    
    # Also include corrections/confirmations
    result = ml_engine.retrain_from_history()
    if not result.get("success") and training_data:
        result = ml_engine.train_classifier(training_data)
    timeline = state.setdefault("ai_training_timeline", [])
    timeline.append({
        "id": f"run-{int(time.time()*1000)}",
        "ts": datetime.now(timezone.utc).isoformat(),
        "triggeredBy": username,
        "datasetSize": len(training_data),
        "modelName": "account-type-classifier",
        "status": "success" if result.get("success") else "failed",
        "metrics": {
            "accuracy": float(result.get("accuracy") or 0.0),
            "f1": float(result.get("f1") or 0.0),
            "precision": float(result.get("precision") or 0.0),
            "recall": float(result.get("recall") or 0.0),
        },
    })
    record_llm_learning_event(
        actor=username,
        source="ml-train",
        signal_type="model-train",
        entity="account-type-classifier",
        details={
            "datasetSize": len(training_data),
            "success": bool(result.get("success")),
            "accuracy": float(result.get("accuracy") or 0.0),
        },
    )

    return result


@app.post("/api/ml/rebuild-brdb")
def ml_rebuild_brdb(username: str = Depends(require_auth)):
    """Force rebuild of Business Role Database."""
    brdb_rebuild()
    return {"ok": True, "message": "BRDB rebuilt", **ml_engine.get_status()["brdb"]}


class AccountTypeConfirmRequest(BaseModel):
    confirmed_type: str


@app.post("/api/users/{uname}/confirm-type")
def confirm_account_type(uname: str, body: AccountTypeConfirmRequest, username: str = Depends(require_auth)):
    """Record a type confirmation (trains the ML model)."""
    users = state.get("last_extract", {}).get("users") or []
    target = next((x for x in users if x.get("username") == uname), None)
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Record confirmation
    ml_engine.record_confirmation(
        username=uname,
        display_name=target.get("displayName", ""),
        ou=target.get("department", ""),
        employee_type="",
        confirmed_type=body.confirmed_type
    )
    
    # Update user type
    target["accountType"] = body.confirmed_type
    record_llm_learning_event(
        actor=username,
        source="confirm-account-type",
        signal_type="supervised-label",
        entity=uname,
        details={"confirmed_type": body.confirmed_type},
    )
    record_manual_user_change(
        actor=username,
        username=uname,
        display_name=target.get("displayName"),
        action="confirm-account-type",
        source="ai-training",
        details={"accountType": body.confirmed_type},
    )
    
    return {"ok": True, "user": uname, "type": body.confirmed_type}


@app.post("/api/users/{uname}/correct-type")
def correct_account_type(uname: str, body: AccountTypeConfirmRequest, username: str = Depends(require_auth)):
    """Record a type correction (trains the ML model)."""
    users = state.get("last_extract", {}).get("users") or []
    target = next((x for x in users if x.get("username") == uname), None)
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    
    old_type = target.get("accountType", "Internal")
    
    # Record correction
    ml_engine.record_correction(
        username=uname,
        display_name=target.get("displayName", ""),
        ou=target.get("department", ""),
        employee_type="",
        old_type=old_type,
        new_type=body.confirmed_type
    )
    
    # Update user type
    target["accountType"] = body.confirmed_type
    record_llm_learning_event(
        actor=username,
        source="correct-account-type",
        signal_type="supervised-correction",
        entity=uname,
        details={"old_type": old_type, "new_type": body.confirmed_type},
    )
    record_manual_user_change(
        actor=username,
        username=uname,
        display_name=target.get("displayName"),
        action="correct-account-type",
        source="ai-training",
        details={"from": old_type, "to": body.confirmed_type},
    )
    
    return {"ok": True, "user": uname, "old_type": old_type, "new_type": body.confirmed_type}


@app.get("/api/ml/account-types")
def get_account_types(username: str = Depends(require_auth)):
    """Return the list of supported account types."""
    return {"types": ACCOUNT_TYPES}


# =============================================================================
# AI LAB API ENDPOINTS (Data Drift, Timeline, A/B, Fairness, Synthetic, Feedback)
# =============================================================================

class TimelineRunRequest(BaseModel):
    model_name: str = "account-type-classifier"
    note: Optional[str] = None


class AbPlaygroundRequest(BaseModel):
    model_a: str = "baseline-v1"
    model_b: str = "candidate-v2"
    sample_size: int = 400


class SyntheticGenerateRequest(BaseModel):
    count: int = 30
    scenario: str = "mixed"
    persist: bool = True


class FeedbackEventRequest(BaseModel):
    username: str
    predicted_type: str
    corrected_type: str
    confidence: float = Field(default=0.0, ge=0.0, le=1.0)
    note: Optional[str] = None


def _ai_lab_users() -> List[Dict[str, Any]]:
    return active_users(state.get("last_extract", {}).get("users") or [])


def _safe_dist(values: List[str]) -> Dict[str, float]:
    total = max(1, len(values))
    cnt = Counter([str(v or "unknown").strip().lower() for v in values])
    return {k: v / total for k, v in cnt.items()}


def _psi(base: Dict[str, float], current: Dict[str, float]) -> float:
    eps = 1e-6
    keys = set(base.keys()) | set(current.keys())
    out = 0.0
    for k in keys:
        b = max(eps, float(base.get(k, 0.0)))
        c = max(eps, float(current.get(k, 0.0)))
        out += (c - b) * np.log(c / b)
    return float(out)


def _default_timeline_entry() -> Dict[str, Any]:
    st = ml_engine.get_status() or {}
    return {
        "id": f"seed-{int(time.time()*1000)}",
        "ts": datetime.now(timezone.utc).isoformat(),
        "triggeredBy": "system",
        "datasetSize": int((st.get("training_data") or {}).get("total_samples") or 0),
        "modelName": "account-type-classifier",
        "status": "success",
        "metrics": {
            "accuracy": float((st.get("model") or {}).get("accuracy") or 0.0),
            "f1": float((st.get("model") or {}).get("f1") or 0.0),
            "precision": float((st.get("model") or {}).get("precision") or 0.0),
            "recall": float((st.get("model") or {}).get("recall") or 0.0),
        },
        "note": "Snapshot iniziale",
    }


@app.get("/api/ai-lab/drift")
def ai_lab_drift(username: str = Depends(require_auth)):
    users = _ai_lab_users()
    if len(users) < 10:
        return {"summary": {"population": len(users), "psi_account_type": 0, "psi_department": 0, "mean_groups_delta": 0}, "signals": []}

    sorted_users = sorted(users, key=lambda u: str(u.get("username") or ""))
    split = max(3, int(len(sorted_users) * 0.35))
    baseline = sorted_users[:split]
    current = sorted_users[split:]

    base_type = _safe_dist([u.get("accountType") for u in baseline])
    cur_type = _safe_dist([u.get("accountType") for u in current])
    base_dept = _safe_dist([u.get("department") for u in baseline])
    cur_dept = _safe_dist([u.get("department") for u in current])

    base_groups = np.mean([len(u.get("groups") or []) for u in baseline]) if baseline else 0.0
    cur_groups = np.mean([len(u.get("groups") or []) for u in current]) if current else 0.0
    delta_groups = float(cur_groups - base_groups)

    psi_type = _psi(base_type, cur_type)
    psi_dept = _psi(base_dept, cur_dept)

    signals = [
        {
            "id": "account-type-distribution",
            "label": "Distribuzione Account Type",
            "baseline": base_type,
            "current": cur_type,
            "psi": round(psi_type, 4),
            "severity": "high" if psi_type >= 0.25 else ("medium" if psi_type >= 0.1 else "low"),
        },
        {
            "id": "department-distribution",
            "label": "Distribuzione Department",
            "baseline": base_dept,
            "current": cur_dept,
            "psi": round(psi_dept, 4),
            "severity": "high" if psi_dept >= 0.25 else ("medium" if psi_dept >= 0.1 else "low"),
        },
        {
            "id": "mean-groups-per-user",
            "label": "Gruppi medi per utente",
            "baseline": round(base_groups, 3),
            "current": round(cur_groups, 3),
            "delta": round(delta_groups, 3),
            "severity": "high" if abs(delta_groups) >= 2.5 else ("medium" if abs(delta_groups) >= 1.0 else "low"),
        },
    ]
    return {
        "summary": {
            "population": len(users),
            "psi_account_type": round(psi_type, 4),
            "psi_department": round(psi_dept, 4),
            "mean_groups_delta": round(delta_groups, 3),
        },
        "signals": signals,
    }


@app.get("/api/ai-lab/training-timeline")
def ai_lab_timeline(username: str = Depends(require_auth)):
    timeline = state.setdefault("ai_training_timeline", [])
    if not timeline:
        timeline.append(_default_timeline_entry())
    timeline.sort(key=lambda x: str(x.get("ts") or ""), reverse=True)
    learning = sorted(state.get("llm_learning_history") or [], key=lambda x: str(x.get("ts") or ""), reverse=True)
    return {"items": timeline[:60], "learningHistory": learning[:500]}


@app.post("/api/ai-lab/training-timeline/run")
def ai_lab_timeline_run(body: TimelineRunRequest, username: str = Depends(require_auth)):
    st = ml_engine.get_status() or {}
    acc = float((st.get("model") or {}).get("accuracy") or 0.78)
    jitter = (hash(f"{body.model_name}:{time.time_ns()}") % 200 - 100) / 10000.0
    acc2 = max(0.0, min(1.0, acc + jitter))
    f1 = max(0.0, min(1.0, acc2 - 0.02))
    precision = max(0.0, min(1.0, acc2 - 0.01))
    recall = max(0.0, min(1.0, acc2 - 0.015))
    users = _ai_lab_users()
    item = {
        "id": f"run-{int(time.time()*1000)}",
        "ts": datetime.now(timezone.utc).isoformat(),
        "triggeredBy": username,
        "datasetSize": len(users),
        "modelName": body.model_name,
        "status": "success",
        "metrics": {
            "accuracy": round(acc2, 4),
            "f1": round(f1, 4),
            "precision": round(precision, 4),
            "recall": round(recall, 4),
        },
        "note": body.note or "Run simulato da AI Lab",
    }
    timeline = state.setdefault("ai_training_timeline", [])
    timeline.append(item)
    timeline.sort(key=lambda x: str(x.get("ts") or ""), reverse=True)
    return {"ok": True, "item": item}


def _simulate_quality_for_model(model_name: str, users: List[Dict[str, Any]], sample_size: int) -> Dict[str, Any]:
    pool = users[:max(1, min(len(users), sample_size))]
    if not pool:
        return {"accuracy": 0.0, "precision": 0.0, "recall": 0.0, "f1": 0.0, "size": 0}
    base = 0.79
    if "candidate" in model_name.lower() or "v2" in model_name.lower():
        base += 0.02
    if "xgboost" in model_name.lower():
        base += 0.015
    if "legacy" in model_name.lower():
        base -= 0.03
    penalties = 0.0
    for u in pool:
        if not (u.get("department") or "").strip():
            penalties += 0.002
        if len(u.get("groups") or []) == 0:
            penalties += 0.002
    acc = max(0.45, min(0.98, base - penalties / max(1, len(pool))))
    prec = max(0.0, min(1.0, acc - 0.01))
    rec = max(0.0, min(1.0, acc - 0.015))
    f1 = (2 * prec * rec / (prec + rec)) if (prec + rec) > 0 else 0.0
    return {
        "accuracy": round(acc, 4),
        "precision": round(prec, 4),
        "recall": round(rec, 4),
        "f1": round(f1, 4),
        "size": len(pool),
    }


def _guess_csv_delimiter(text: str) -> str:
    header = (text.splitlines() or [""])[0]
    if header.count(";") >= header.count(","):
        return ";"
    return ","


def _ab_norm_header(h: str) -> str:
    return (h or "").strip().lower().replace("_", "").replace(" ", "")


def _ab_get(row_ci: Dict[str, Any], keys: List[str]) -> str:
    for k in keys:
        v = row_ci.get(_ab_norm_header(k))
        if v is not None:
            return str(v)
    return ""


def _build_matrix_from_users(users: List[Dict[str, Any]]) -> Tuple[Dict[str, Dict[str, int]], List[str]]:
    all_groups = sorted({g for u in users for g in (u.get("groups") or [])})
    matrix: Dict[str, Dict[str, int]] = {}
    for u in users:
        uname = str(u.get("username") or "")
        ug = set(u.get("groups") or [])
        matrix[uname] = {g: (1 if g in ug else 0) for g in all_groups}
    return matrix, all_groups


def _parse_users_from_csv_bytes(raw: bytes, dataset_name: str) -> Dict[str, Any]:
    text = raw.decode("utf-8-sig", errors="replace")
    delim = _guess_csv_delimiter(text)
    reader = csv.DictReader(io.StringIO(text), delimiter=delim, skipinitialspace=True)
    rows = list(reader)

    KEY = {
        "displayName": ["displayname", "display name", "name"],
        "username": ["username", "userprincipalname", "samaccountname", "login"],
        "department": ["department", "dept", "dipartimento"],
        "businessRole": ["businessrole", "business role", "br"],
        "roles": ["ruoli", "roles", "groups", "gruppi", "entitlements"],
        "accountType": ["accounttype", "account type", "type"],
        "lastLogin": ["lastlogon", "lastlogin", "last login", "last_logon"],
    }

    users: List[Dict[str, Any]] = []
    seen_usernames = set()
    missing_display = 0
    missing_department = 0
    missing_business_role = 0
    zero_groups = 0
    invalid_last_login = 0

    by_dn = defaultdict(list)
    for i, row in enumerate(rows, start=1):
        row_ci = {_ab_norm_header(k): v for k, v in (row or {}).items()}
        dn = (_ab_get(row_ci, KEY["displayName"]) or "").strip()
        if not dn:
            missing_display += 1
            continue
        uname = (_ab_get(row_ci, KEY["username"]) or "").strip()
        dept = (_ab_get(row_ci, KEY["department"]) or "").strip()
        br = (_ab_get(row_ci, KEY["businessRole"]) or "").strip()
        roles_raw = (_ab_get(row_ci, KEY["roles"]) or "").strip()
        atype = (_ab_get(row_ci, KEY["accountType"]) or "").strip() or "Internal"
        ll_raw = (_ab_get(row_ci, KEY["lastLogin"]) or "").strip()
        ll = _normalize_last_login(ll_raw or None)
        if ll_raw and not ll:
            invalid_last_login += 1

        groups = [g.strip() for g in roles_raw.split(",") if g and g.strip()]
        if not groups:
            zero_groups += 1
        if not dept:
            missing_department += 1
        if not br:
            missing_business_role += 1
            br = dept or "Unassigned"

        if not uname:
            uname = _slug_username(dn)
        base = uname
        n = 2
        while uname in seen_usernames:
            uname = f"{base}{n}"
            n += 1
        seen_usernames.add(uname)

        u = {
            "username": uname,
            "displayName": dn,
            "department": dept,
            "businessRole": br,
            "groups": groups,
            "accountType": atype or classify_account(dn, dept, ""),
            "lastLogin": ll,
            "excluded": False,
        }
        users.append(u)
        by_dn[dn.strip().lower()].append(u)

    duplicate_displayname_rows = sum(max(0, len(v) - 1) for v in by_dn.values())
    return {
        "name": dataset_name,
        "users": users,
        "rowsTotal": len(rows),
        "missingDisplayName": missing_display,
        "missingDepartment": missing_department,
        "missingBusinessRole": missing_business_role,
        "zeroGroupsUsers": zero_groups,
        "invalidLastLogin": invalid_last_login,
        "duplicateDisplayNameRows": duplicate_displayname_rows,
    }


def _compute_dataset_scores_from_users(dataset: Dict[str, Any]) -> Dict[str, Any]:
    users = dataset.get("users") or []
    total = max(1, len(users))
    matrix, groups = _build_matrix_from_users(users)
    mq = compute_model_quality(users, matrix, groups)
    ai = run_smart_ai_detection(users, matrix)

    metrics = {
        "rows_total": len(users),
        "duplicate_displayname_rows": int(dataset.get("duplicateDisplayNameRows") or 0),
        "missing_department": int(dataset.get("missingDepartment") or 0),
        "missing_business_role": int(dataset.get("missingBusinessRole") or 0),
        "zero_groups_users": int(dataset.get("zeroGroupsUsers") or 0),
        "invalid_last_login": int(dataset.get("invalidLastLogin") or 0),
        "stale_users": int(len(mq.get("staleList") or [])),
        "overprivileged_users": int(len(mq.get("overprivilegedList") or [])),
        "policy_violations": int(len(mq.get("policyViolations") or [])),
        "ambiguous_users": int(len(mq.get("ambiguousUsers") or [])),
        "ai_anomaly_users": int((ai.get("stats") or {}).get("usersWithAnomaly") or 0),
        "model_quality_score": float(mq.get("modelQuality") or 0.0),
    }

    # weighted score [0..100], where higher is better
    rates = {
        "dup": metrics["duplicate_displayname_rows"] / total,
        "miss_dept": metrics["missing_department"] / total,
        "miss_br": metrics["missing_business_role"] / total,
        "zero": metrics["zero_groups_users"] / total,
        "invalid_ll": metrics["invalid_last_login"] / total,
        "stale": metrics["stale_users"] / total,
        "over": metrics["overprivileged_users"] / total,
        "policy": metrics["policy_violations"] / total,
        "amb": metrics["ambiguous_users"] / total,
        "ai": metrics["ai_anomaly_users"] / total,
    }
    score = 100.0 * (
        0.33 * (metrics["model_quality_score"] / 100.0) +
        0.09 * (1.0 - min(1.0, rates["dup"])) +
        0.08 * (1.0 - min(1.0, rates["miss_dept"])) +
        0.08 * (1.0 - min(1.0, rates["miss_br"])) +
        0.08 * (1.0 - min(1.0, rates["zero"])) +
        0.07 * (1.0 - min(1.0, rates["invalid_ll"])) +
        0.07 * (1.0 - min(1.0, rates["stale"])) +
        0.07 * (1.0 - min(1.0, rates["over"])) +
        0.06 * (1.0 - min(1.0, rates["policy"])) +
        0.07 * (1.0 - min(1.0, rates["amb"])) +
        0.10 * (1.0 - min(1.0, rates["ai"]))
    )
    score = max(0.0, min(100.0, score))

    return {
        "name": dataset.get("name"),
        "score": round(score, 2),
        "metrics": metrics,
    }


@app.post("/api/ai-lab/ab-playground/compare")
def ai_lab_ab_compare(body: AbPlaygroundRequest, username: str = Depends(require_auth)):
    users = _ai_lab_users()
    sample_size = max(50, min(int(body.sample_size or 400), max(50, len(users))))
    a = _simulate_quality_for_model(body.model_a, users, sample_size)
    b = _simulate_quality_for_model(body.model_b, users, sample_size)
    delta = {k: round(float(b.get(k, 0.0)) - float(a.get(k, 0.0)), 4) for k in ("accuracy", "precision", "recall", "f1")}
    return {
        "sampleSize": sample_size,
        "modelA": {"name": body.model_a, **a},
        "modelB": {"name": body.model_b, **b},
        "delta": delta,
        "winner": body.model_b if delta.get("f1", 0.0) > 0 else body.model_a,
    }


@app.post("/api/ai-lab/ab-playground/upload-compare")
async def ai_lab_ab_compare_upload(
    file_a: UploadFile = File(...),
    file_b: UploadFile = File(...),
    username: str = Depends(require_auth),
):
    raw_a = await file_a.read()
    raw_b = await file_b.read()
    if not raw_a or not raw_b:
        raise HTTPException(status_code=400, detail="Entrambi i file CSV sono obbligatori.")

    parsed_a = _parse_users_from_csv_bytes(raw_a, file_a.filename or "csv_a")
    parsed_b = _parse_users_from_csv_bytes(raw_b, file_b.filename or "csv_b")
    scored_a = _compute_dataset_scores_from_users(parsed_a)
    scored_b = _compute_dataset_scores_from_users(parsed_b)

    meta = {
        "model_quality_score": {"higher_is_better": True, "label": "Model Quality Score"},
        "rows_total": {"higher_is_better": True, "label": "Rows Total"},
        "duplicate_displayname_rows": {"higher_is_better": False, "label": "Duplicate DisplayName Rows"},
        "missing_department": {"higher_is_better": False, "label": "Missing Department"},
        "missing_business_role": {"higher_is_better": False, "label": "Missing Business Role"},
        "zero_groups_users": {"higher_is_better": False, "label": "Zero Groups Users"},
        "invalid_last_login": {"higher_is_better": False, "label": "Invalid LastLogin"},
        "stale_users": {"higher_is_better": False, "label": "Stale Users"},
        "overprivileged_users": {"higher_is_better": False, "label": "Overprivileged Users"},
        "policy_violations": {"higher_is_better": False, "label": "Policy Violations"},
        "ambiguous_users": {"higher_is_better": False, "label": "Ambiguous Users"},
        "ai_anomaly_users": {"higher_is_better": False, "label": "AI Anomaly Users"},
    }

    compare_rows = []
    improved = 0
    worsened = 0
    unchanged = 0
    for k, m in meta.items():
        a_val = float(scored_a["metrics"].get(k, 0))
        b_val = float(scored_b["metrics"].get(k, 0))
        diff = round(b_val - a_val, 4)
        hib = bool(m["higher_is_better"])
        if diff == 0:
            trend = "unchanged"
            unchanged += 1
        else:
            better = diff > 0 if hib else diff < 0
            trend = "improved" if better else "worsened"
            if better:
                improved += 1
            else:
                worsened += 1
        compare_rows.append({
            "metric": k,
            "label": m["label"],
            "a": a_val,
            "b": b_val,
            "diff": diff,
            "trend": trend,
        })

    score_diff = round(float(scored_b["score"]) - float(scored_a["score"]), 2)
    winner = scored_b["name"] if score_diff > 0 else (scored_a["name"] if score_diff < 0 else "tie")
    return {
        "datasetA": scored_a,
        "datasetB": scored_b,
        "scoreDiff": score_diff,
        "winner": winner,
        "comparison": compare_rows,
        "summary": {"improved": improved, "worsened": worsened, "unchanged": unchanged},
    }


@app.get("/api/ai-lab/fairness")
def ai_lab_fairness(username: str = Depends(require_auth)):
    users = _ai_lab_users()
    if not users:
        return {"overallErrorRate": 0.0, "byDepartment": [], "byAccountType": []}

    def _err(u: Dict[str, Any]) -> int:
        risk = 0
        if len(u.get("groups") or []) == 0:
            risk += 1
        if not (u.get("businessRole") or "").strip():
            risk += 1
        if str(u.get("accountType") or "").lower() in {"external", "contractor"} and len(u.get("groups") or []) > 7:
            risk += 1
        return 1 if risk > 0 else 0

    overall = np.mean([_err(u) for u in users]) if users else 0.0

    def _aggregate(key_name: str) -> List[Dict[str, Any]]:
        buckets = defaultdict(list)
        for u in users:
            k = str(u.get(key_name) or "Unknown").strip() or "Unknown"
            buckets[k].append(u)
        out = []
        for k, arr in buckets.items():
            err_rate = float(np.mean([_err(x) for x in arr])) if arr else 0.0
            out.append({
                "group": k,
                "size": len(arr),
                "errorRate": round(err_rate, 4),
                "gapVsOverall": round(err_rate - overall, 4),
            })
        out.sort(key=lambda x: abs(x["gapVsOverall"]), reverse=True)
        return out

    return {
        "overallErrorRate": round(float(overall), 4),
        "byDepartment": _aggregate("department")[:20],
        "byAccountType": _aggregate("accountType")[:20],
    }


@app.get("/api/ai-lab/synthetic")
def ai_lab_synthetic(username: str = Depends(require_auth)):
    templates = [
        {"id": "svc_admin_overlap", "label": "Service con gruppi admin", "risk": "high"},
        {"id": "missing_identity_keys", "label": "Chiavi identita mancanti", "risk": "medium"},
        {"id": "future_last_login", "label": "LastLogin futura", "risk": "medium"},
        {"id": "department_role_mismatch", "label": "Dipartimento/BusinessRole incoerenti", "risk": "high"},
        {"id": "ghost_manager", "label": "Manager non esistente", "risk": "high"},
    ]
    return {"templates": templates, "lastGenerated": state.get("ai_synthetic_cases", [])[:100]}


@app.post("/api/ai-lab/synthetic/generate")
def ai_lab_synthetic_generate(body: SyntheticGenerateRequest, username: str = Depends(require_auth)):
    users = _ai_lab_users()
    templates = ["svc_admin_overlap", "missing_identity_keys", "future_last_login", "department_role_mismatch", "ghost_manager"]
    count = max(1, min(int(body.count or 30), 300))
    generated = []
    for i in range(count):
        t = body.scenario if body.scenario != "mixed" else templates[i % len(templates)]
        ref = users[i % len(users)] if users else {}
        generated.append({
            "id": f"syn-{int(time.time()*1000)}-{i}",
            "template": t,
            "displayName": f"syn_{(ref.get('displayName') or 'utente').replace(' ', '_').lower()}_{i+1}",
            "department": ref.get("department") or "Unknown",
            "businessRole": ref.get("businessRole") or "Unassigned",
            "groups": (ref.get("groups") or [])[:4],
            "severity": "high" if t in {"svc_admin_overlap", "department_role_mismatch", "ghost_manager"} else "medium",
        })
    if body.persist:
        state["ai_synthetic_cases"] = generated
    return {"ok": True, "count": len(generated), "items": generated}


@app.get("/api/ai-lab/feedback")
def ai_lab_feedback(username: str = Depends(require_auth)):
    events = state.get("ai_feedback_events") or []
    total = len(events)
    recent = sorted(events, key=lambda x: str(x.get("ts") or ""), reverse=True)[:200]
    by_corrected = Counter([str(e.get("corrected_type") or "Unknown") for e in events])
    history = list(state.get("manual_user_changes") or [])
    history = sorted(history, key=lambda x: str(x.get("ts") or ""), reverse=True)[:500]
    return {
        "total": total,
        "byCorrectedType": [{"type": k, "count": v} for k, v in by_corrected.most_common()],
        "items": recent,
        "history": history,
    }


@app.post("/api/ai-lab/feedback")
def ai_lab_feedback_add(body: FeedbackEventRequest, background_tasks: BackgroundTasks, username: str = Depends(require_auth)):
    users = state.get("last_extract", {}).get("users") or []
    target = next((u for u in users if str(u.get("username") or "") == body.username), None)
    if target:
        target["accountType"] = body.corrected_type
    event = {
        "id": f"fb-{int(time.time()*1000)}",
        "ts": datetime.now(timezone.utc).isoformat(),
        "author": username,
        "username": body.username,
        "predicted_type": body.predicted_type,
        "corrected_type": body.corrected_type,
        "confidence": float(body.confidence),
        "note": body.note or "",
    }
    events = state.setdefault("ai_feedback_events", [])
    events.append(event)
    events[:] = events[-2000:]
    dq_ev = state.setdefault("dq_feedback_events", [])
    dq_ev.append({
        "kind": "ai-feedback",
        "username": body.username,
        "predicted": body.predicted_type,
        "corrected": body.corrected_type,
        "ts": event["ts"],
    })
    dq_ev[:] = dq_ev[-3000:]
    record_llm_learning_event(
        actor=username,
        source="ai-lab-feedback",
        signal_type="supervised-correction",
        entity=body.username,
        details={
            "predicted_type": body.predicted_type,
            "corrected_type": body.corrected_type,
            "confidence": float(body.confidence),
        },
    )
    manual_event = record_manual_user_change(
        actor=username,
        username=body.username,
        display_name=target.get("displayName") if target else body.username,
        action="feedback-correction",
        source="ai-lab-feedback",
        details={
            "predicted_type": body.predicted_type,
            "corrected_type": body.corrected_type,
            "confidence": float(body.confidence),
            "note": body.note or "",
        },
        persist=False,
    )
    # Persist once in background to reduce response latency on large storage files.
    background_tasks.add_task(state.save)
    return {"ok": True, "event": event, "manualEvent": manual_event}


# =============================================================================
# PATTERN RULES API ENDPOINTS
# =============================================================================

class PatternRuleRequest(BaseModel):
    account_type: str
    field: str
    regex: str


class BrPatternRuleRequest(BaseModel):
    business_role: str
    field: str
    regex: str


class BrAssignmentPatternRuleRequest(BaseModel):
    business_role: str
    role: Optional[str] = None  # legacy optional gate on exact group name
    regex: str


@app.get("/api/ml/patterns")
def get_patterns(username: str = Depends(require_auth)):
    """Return all classification patterns (static + custom)."""
    return ml_engine.get_patterns()


@app.post("/api/ml/patterns")
def add_pattern(body: PatternRuleRequest, background_tasks: BackgroundTasks, username: str = Depends(require_auth)):
    """Add a new custom regex pattern rule for account classification."""
    result = ml_engine.add_pattern(body.account_type, body.field, body.regex)
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error", "Failed"))
    log("INFO", f"Pattern added: {body.account_type}/{body.field}/{body.regex} by {username}")
    record_llm_learning_event(
        actor=username,
        source="pattern-rules",
        signal_type="rule-added",
        entity=body.account_type,
        details={"field": body.field, "regex": body.regex},
    )
    background_tasks.add_task(recalculate_assignments_background, "account-pattern-rule-added", username)
    return result


@app.delete("/api/ml/patterns/{index}")
def delete_pattern_endpoint(index: int, background_tasks: BackgroundTasks, username: str = Depends(require_auth)):
    """Delete a custom pattern by index."""
    result = ml_engine.delete_pattern(index)
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error", "Failed"))
    log("INFO", f"Pattern deleted: index={index} by {username}")
    record_llm_learning_event(
        actor=username,
        source="pattern-rules",
        signal_type="rule-deleted",
        entity=str(index),
        details={},
    )
    background_tasks.add_task(recalculate_assignments_background, "account-pattern-rule-deleted", username)
    return result


@app.get("/api/ml/br-patterns")
def get_br_patterns(username: str = Depends(require_auth)):
    rules = state.setdefault("br_pattern_rules", [])
    return {"custom": list(rules)}


@app.post("/api/ml/br-patterns")
def add_br_pattern(body: BrPatternRuleRequest, background_tasks: BackgroundTasks, username: str = Depends(require_auth)):
    business_role = (body.business_role or "").strip()
    field = (body.field or "").strip()
    regex = (body.regex or "").strip()
    if not business_role:
        raise HTTPException(status_code=400, detail="Business Role is required")
    if not field:
        raise HTTPException(status_code=400, detail="Field is required")
    if not regex:
        raise HTTPException(status_code=400, detail="Regex is required")
    try:
        re.compile(regex)
    except re.error as exc:
        raise HTTPException(status_code=400, detail=f"Invalid regex: {exc}") from exc

    rules = state.setdefault("br_pattern_rules", [])
    rule = {
        "business_role": business_role,
        "field": field,
        "regex": regex,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    rules.append(rule)
    state["br_pattern_rules"] = rules
    state.save()
    log("INFO", f"BR pattern added: {business_role}/{field}/{regex} by {username}")
    record_llm_learning_event(
        actor=username,
        source="pattern-rules",
        signal_type="br-rule-added",
        entity=business_role,
        details={"field": field, "regex": regex},
    )
    background_tasks.add_task(recalculate_assignments_background, "br-pattern-rule-added", username)
    return {"success": True, "rule": rule, "total_custom_rules": len(rules)}


@app.delete("/api/ml/br-patterns/{index}")
def delete_br_pattern(index: int, background_tasks: BackgroundTasks, username: str = Depends(require_auth)):
    rules = list(state.setdefault("br_pattern_rules", []))
    if index < 0 or index >= len(rules):
        raise HTTPException(status_code=400, detail="Invalid index")
    removed = rules.pop(index)
    state["br_pattern_rules"] = rules
    state.save()
    log("INFO", f"BR pattern deleted: index={index} by {username}")
    record_llm_learning_event(
        actor=username,
        source="pattern-rules",
        signal_type="br-rule-deleted",
        entity=str(index),
        details={},
    )
    background_tasks.add_task(recalculate_assignments_background, "br-pattern-rule-deleted", username)
    return {"success": True, "removed": removed, "total_custom_rules": len(rules)}


@app.get("/api/ml/br-assignment-patterns")
def get_br_assignment_patterns(username: str = Depends(require_auth)):
    rules = state.setdefault("br_assignment_pattern_rules", [])
    return {"custom": list(rules)}


@app.post("/api/ml/br-assignment-patterns")
def add_br_assignment_pattern(body: BrAssignmentPatternRuleRequest, background_tasks: BackgroundTasks, username: str = Depends(require_auth)):
    business_role = (body.business_role or "").strip()
    regex = (body.regex or "").strip()
    if not business_role:
        raise HTTPException(status_code=400, detail="Business Role is required")
    if not regex:
        raise HTTPException(status_code=400, detail="Regex is required")
    try:
        re.compile(regex)
    except re.error as exc:
        raise HTTPException(status_code=400, detail=f"Invalid regex: {exc}") from exc

    rules = state.setdefault("br_assignment_pattern_rules", [])
    rule = {
        "business_role": business_role,
        "regex": regex,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    legacy_role = (body.role or "").strip()
    if legacy_role:
        rule["role"] = legacy_role
    rules.append(rule)
    state["br_assignment_pattern_rules"] = rules
    state.save()
    log("INFO", f"BR assignment pattern added: {business_role}/{regex} by {username}")
    record_llm_learning_event(
        actor=username,
        source="pattern-rules",
        signal_type="br-assignment-rule-added",
        entity=business_role,
        details={"regex": regex, "legacyRole": legacy_role},
    )
    background_tasks.add_task(recalculate_assignments_background, "br-assignment-rule-added", username)
    return {"success": True, "rule": rule, "total_custom_rules": len(rules)}


@app.delete("/api/ml/br-assignment-patterns/{index}")
def delete_br_assignment_pattern(index: int, background_tasks: BackgroundTasks, username: str = Depends(require_auth)):
    rules = list(state.setdefault("br_assignment_pattern_rules", []))
    if index < 0 or index >= len(rules):
        raise HTTPException(status_code=400, detail="Invalid index")
    removed = rules.pop(index)
    state["br_assignment_pattern_rules"] = rules
    state.save()
    log("INFO", f"BR assignment pattern deleted: index={index} by {username}")
    record_llm_learning_event(
        actor=username,
        source="pattern-rules",
        signal_type="br-assignment-rule-deleted",
        entity=str(index),
        details={},
    )
    background_tasks.add_task(recalculate_assignments_background, "br-assignment-rule-deleted", username)
    return {"success": True, "removed": removed, "total_custom_rules": len(rules)}


# (peer-analysis endpoint defined at line 1846)


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=False)
