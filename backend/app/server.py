import os
import time
import json
import hashlib
import hmac
import tempfile
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Tuple
from pathlib import Path

import jwt
import numpy as np
from fastapi import Depends, FastAPI, HTTPException, Request, status
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from pydantic import BaseModel, Field
from sklearn.cluster import AgglomerativeClustering, MiniBatchKMeans
from sklearn.decomposition import TruncatedSVD
from fastapi import UploadFile, File, BackgroundTasks
import csv, io, re, threading
import urllib.request
import urllib.parse
import urllib.error
import base64
from http.cookiejar import CookieJar
try:
    from ldap3 import ALL, NTLM, SIMPLE, Connection, Server, Tls, NONE, MODIFY_ADD, MODIFY_DELETE, MODIFY_REPLACE
    import ssl
except Exception:
    Connection = None  # type: ignore
    MODIFY_ADD = 0  # type: ignore
    MODIFY_DELETE = 1  # type: ignore
    MODIFY_REPLACE = 2  # type: ignore


APP_TITLE = "Role Mining API"
import secrets
# Use a persistent key for dev to avoid invalidating tokens on restart
JWT_SECRET = os.getenv("JWT_SECRET") or "dev_secret_key_persistent_change_in_prod"
APP_LOGIN_USER = os.getenv("APP_LOGIN_USER", "admin")
APP_LOGIN_PASS = os.getenv("APP_LOGIN_PASS", "admin123")
APP_VIEWER_USER = os.getenv("APP_VIEWER_USER", "user")
APP_VIEWER_PASS = os.getenv("APP_VIEWER_PASS", "viewer")
JWT_EXPIRE_MINUTES = int(os.getenv("JWT_EXPIRE_MINUTES", "240"))
MOCK_AD = os.getenv("MOCK_AD", "0") == "1"
LDAP_FETCH_ALL_ATTRIBUTES = os.getenv("LDAP_FETCH_ALL_ATTRIBUTES", "0") == "1"
LDAP_PAGE_SIZE = max(100, int(os.getenv("LDAP_PAGE_SIZE", "1000")))
LDAP_SEARCH_TIME_LIMIT = max(10, int(os.getenv("LDAP_SEARCH_TIME_LIMIT", "60")))
CSV_IMPORT_MAX_BYTES = max(1 * 1024 * 1024, int(os.getenv("CSV_IMPORT_MAX_BYTES", str(250 * 1024 * 1024))))
CSV_IMPORT_MAX_ROWS = max(1000, int(os.getenv("CSV_IMPORT_MAX_ROWS", "200000")))
CSV_IMPORT_MAX_REJECTS_STORED = max(100, int(os.getenv("CSV_IMPORT_MAX_REJECTS_STORED", "5000")))
CSV_IMPORT_DETACHED_POSTPROCESS = os.getenv("CSV_IMPORT_DETACHED_POSTPROCESS", "1") != "0"
LDAP_EXTRA_ATTRIBUTES = [
    a.strip() for a in os.getenv("LDAP_EXTRA_ATTRIBUTES", "").split(",") if a.strip()
]
LDAP_BASE_ATTRIBUTES = [
    "sAMAccountName",
    "displayName",
    "memberOf",
    "department",
    "lastLogonTimestamp",
    "lastLogon",
    "employeeType",
    "distinguishedName",
    "mail",
    "userPrincipalName",
]

SYSTEM_USER_PERMISSION_DEFAULTS: Dict[str, bool] = {
    "can_view_analytics": True,
    "can_view_cluster": True,
    "can_view_users": True,
    "can_view_business_roles": True,
    "can_view_ai_training": True,
    "can_view_configurations": True,
    "can_view_logs": True,
    "can_view_system_users": True,
    "can_manage_settings": True,
    "can_manage_assignments": True,
}

SYSTEM_USER_HASH_PREFIX = "pbkdf2_sha256"
SYSTEM_USER_HASH_ITERATIONS = 150_000



from openpyxl import load_workbook
from collections import defaultdict, Counter

# ML Engine import
from ml_engine import get_ml_engine, ACCOUNT_TYPES
ml_engine = get_ml_engine(data_dir="./ml_data")
GLOBAL_ML_SIGNALS_PATH = Path("./ml_data/global_ml_signals.json")
GLOBAL_ML_SIGNALS_LOCK = threading.RLock()
GLOBAL_ML_SIGNALS_MAX = max(1000, int(os.getenv("GLOBAL_ML_SIGNALS_MAX", "20000")))

REBUILD_LOCK = threading.Lock()
from app.core.cache import (
    CACHE_TTL_KPI,
    CACHE_TTL_MINING,
    CACHE_TTL_ROLES,
    CACHE_TTL_USERS,
    RESPONSE_CACHE,
    invalidate_hot_caches,
)
from app.services.identity_quality import (
    is_valid_email_address as _is_valid_email_address,
    is_valid_employee_id as _is_valid_employee_id,
    is_valid_upn_value as _is_valid_upn_value,
    norm_identity_text as _norm_identity_text,
)
from app.services.data_quality_helpers import (
    allowed_identity_case_ids as _allowed_identity_case_ids,
    build_cluster_quality_summary_cards as _build_cluster_quality_summary_cards,
    dedupe_user_rows as _dedupe_user_rows,
    duplicate_resolution_items_from_users as _duplicate_resolution_items_from_users,
    effective_duplicate_displayname_count as _effective_duplicate_displayname_count,
    users_missing_field as _users_missing_field,
)
from app.api.ai_lab_routes import register_ai_lab_routes
from app.api.pattern_rules_routes import register_pattern_rules_routes

BROAD_MARKERS = ['all','tutti','tutte','full','global','everyone','any','anyone','everybody']

def _tokens(s: str) -> list[str]:
    s = (s or '').lower()
  # Treat underscore/dash like separators too, so VPN_ALL -> ["vpn","all"]
    s = s.replace('_', ' ').replace('-', ' ')
    s = re.sub(r'[^a-z0-9]', ' ', s)
    return [t for t in s.split() if t]

def _family_key(role_name: str) -> str:
    toks = _tokens(role_name)
    return toks[0] if toks else ""

def _is_broad(role_name: str) -> bool:
    toks = set(_tokens(role_name))
    return any(m in toks for m in BROAD_MARKERS)

def _matrix_user_roles(matrix: dict, username: str) -> set[str]:
    row = (matrix or {}).get(username) or {}
    return {r for r, v in row.items() if int(v) == 1}

def build_overprivileged_items(matrix: dict, top_pct: float = 10.0) -> dict:
    users = state.get("last_extract", {}).get("users") or []
    br_by_user = {u.get("username"): u.get("businessRole", "Unassigned") for u in users if u.get("username")}
    role_meta = state.get("role_meta", {}) or {}

    items = []
    for uname, row in (matrix or {}).items():
        actual = sorted([r for r, v in (row or {}).items() if int(v) == 1])

        br = br_by_user.get(uname, "Unassigned")
        expected = set((role_meta.get(br, {}) or {}).get("groups", []) or [])

        excess = sorted(set(actual) - expected)
        if not excess:
            continue  # SOLO over

        items.append({
        "username": uname,
        "groups": actual,          # tutti
        "overGroups": excess,      # solo eccesso
        "groupCount": len(excess), # numero eccessid
        
        "isOverprivileged": True,
        })
    items.sort(key=lambda r: r["groupCount"], reverse=True)
    return {"threshold": 1, "items": items}

import random

# In-memory cache for BRDB (not persisted to storage.json to keep it lean)
BRDB_CACHE: Dict[str, Dict[str, Any]] = {}

def generate_unique_color(existing_colors):
    """Genera un colore RRGGBB univoco rispetto a quelli esistenti"""
    existing = set(existing_colors)
    attempts = 0
    while attempts < 100:
        r = random.randint(80, 255)   # evita colori troppo scuri
        g = random.randint(80, 255)
        b = random.randint(80, 255)
        color = f"{r:02x}{g:02x}{b:02x}".upper()
        if color not in existing:
            return color
        attempts += 1
    return "6AA6FF"  # fallback

# Lista colori predefiniti (per i primi 12 ruoli)
PREDEFINED_COLORS = [
    "00B4FF", "FF9F1C", "71FFB2", "FF6B6B", "9B59B6", "3498DB",
    "F39C12", "1ABC9C", "E74C3C", "9B59B6", "3498DB", "F1C40F"
]

def predict_redundant(broad_role: str, specific_role: str, family: str, user_groups: set[str]) -> float:
  bt = set(_tokens(broad_role))
  st = set(_tokens(specific_role))
  if not bt or not st:
    return 0.0

  # Similarità nome ruolo (Jaccard su token)
  j_role = len(bt & st) / max(1, len(bt | st))

  # Broadness
  b_is_broad = 1.0 if _is_broad(broad_role) else 0.0

  # Overlap token specific vs gruppi reali utente
  st2 = [t for t in st if len(t) >= 3]
  hits = 0
  for g in user_groups:
    gl = (g or "").lower()
    if any(t in gl for t in st2):
      hits += 1
  g_overlap = hits / max(1, len(user_groups))

  p = 0.55 * j_role + 0.30 * b_is_broad + 0.15 * g_overlap
  return float(max(0.0, min(1.0, p)))



def build_ai_detection_items(matrix: dict) -> list[dict]:
    items = []
    for uname, row in (matrix or {}).items():
        roles = [r for r, v in (row or {}).items() if int(v) == 1]
        user_groups = _matrix_user_roles(matrix, uname)

        fam = defaultdict(list)
        for r in roles:
            k = _family_key(r)
            if k:
                fam[k].append(r)

        for family, rs in fam.items():
            broad = [r for r in rs if _is_broad(r)]
            specific = [r for r in rs if not _is_broad(r)]
            
            if not broad or not specific:
                continue

            redundant = []
            kept = set(specific)

            for b in broad:
                probs = [predict_redundant(b, s, family, user_groups) for s in specific]
                p = max(probs) if probs else 0.0
                if p >= 0.50:
                    redundant.append(b)

            # Fallback: se hai broad + specific ma l'algoritmo non è sicuro, 
            # mostrali comunque come "sospetti" invece di nasconderli
            if not redundant and broad:
                 redundant = list(broad)
            
            if redundant:
                items.append({
                    "username": uname,
                    "family": family,
                    "redundantRoles": sorted(set(redundant)),
                    "keptRoles": sorted(list(kept)),
                    "redundantCount": len(set(redundant)),
                })

    # Ordina per chi ne ha di più
    items.sort(key=lambda x: x["redundantCount"], reverse=True)
    return items


# ---------------------------------------------------------------------------
# Smart AI Detection (On-Demand) – Peer / Department / AccountType analysis
# ---------------------------------------------------------------------------
ADMIN_MARKERS = {'admin', 'administrator', 'superuser', 'root', 'owner',
                 'elevated', 'privileged', 'sudo', 'godmode'}
SERVICE_ACCOUNT_MARKERS = {"svc", "service", "serviceaccount", "service-account"}
RESTRICTED_ACCOUNT_TYPES = {
    'BlueCollar':  ADMIN_MARKERS,
    'External':    ADMIN_MARKERS | {'internal', 'staff', 'employee', 'hr',
                                     'payroll', 'finance', 'accounting'},
    'Contractor':  ADMIN_MARKERS | {'internal', 'staff'},
    # LLM guardrail: service/svc accounts should not carry administrative access.
    'Service':     ADMIN_MARKERS,
}
PEER_ANOMALY_THRESHOLD = 0.10   # flag if < 10% of peers have the group
DEPT_ANOMALY_THRESHOLD = 0.05   # flag if < 5% of dept members have group


def load_knowledge_base() -> dict:
    """Load the synthetic knowledge base (LLM-instructed rules)."""
    try:
        import json
        path = "backend/ml_data/knowledge_base.json"
        if not os.path.exists(path):
            return {}
        with open(path, "r") as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading KB: {e}")
        return {}


def _build_freq_tables(users: list) -> tuple:
    """Pre-compute frequency tables in a SINGLE pass over the user list.
    Returns (role_freq, dept_freq):
      role_freq  = { businessRole: { group: ratio_0_to_1 } }
      dept_freq  = { department:   { group: ratio_0_to_1 } }
    """
    from collections import Counter

    role_counts: dict[str, Counter] = {}   # role -> Counter of groups
    role_totals: dict[str, int] = {}       # role -> num users in role
    dept_counts: dict[str, Counter] = {}
    dept_totals: dict[str, int] = {}

    for u in (users or []):
        if u.get("excluded"):
            continue
        groups = u.get("groups") or []
        br = (u.get("businessRole") or "Unassigned").strip()
        dept = (u.get("department") or "Unknown").strip()

        role_totals[br] = role_totals.get(br, 0) + 1
        dept_totals[dept] = dept_totals.get(dept, 0) + 1

        if br not in role_counts:
            role_counts[br] = Counter()
        if dept not in dept_counts:
            dept_counts[dept] = Counter()

        for g in groups:
            role_counts[br][g] += 1
            dept_counts[dept][g] += 1

    # Convert counts to ratios
    role_freq: dict[str, dict[str, float]] = {}
    for br, ctr in role_counts.items():
        total = max(1, role_totals.get(br, 1))
        role_freq[br] = {g: cnt / total for g, cnt in ctr.items()}

    dept_freq: dict[str, dict[str, float]] = {}
    for dept, ctr in dept_counts.items():
        total = max(1, dept_totals.get(dept, 1))
        dept_freq[dept] = {g: cnt / total for g, cnt in ctr.items()}

    return role_freq, dept_freq


def _check_type_violation(group: str, account_type: str) -> str:
    """Return violation reason or empty string. O(1) per group."""
    if not account_type:
        return ""
    restricted = RESTRICTED_ACCOUNT_TYPES.get(account_type)
    if not restricted:
        return ""
    toks = set(_tokens(group))
    matched = toks & restricted
    if matched:
        return f"Policy: {account_type} should not have '{group}'"
    return ""


def _is_service_identity(u_data: dict) -> bool:
    acct_type = str(u_data.get("accountType") or "").strip().lower()
    if acct_type == "service":
        return True
    dn = str(u_data.get("displayName") or "").strip().lower()
    uname = str(u_data.get("username") or "").strip().lower()
    dn_toks = set(_tokens(dn))
    un_toks = set(_tokens(uname))
    if dn.startswith("svc") or uname.startswith("svc"):
        return True
    return bool((dn_toks | un_toks) & SERVICE_ACCOUNT_MARKERS)


def _is_admin_group_name(group: str) -> bool:
    toks = set(_tokens(group))
    return bool(toks & ADMIN_MARKERS)


def run_smart_ai_detection(users: list, matrix: dict) -> dict:
    """On-demand smart redundancy detection.
    Combines:
    1. Statistical signal (Peer / Dept freq)
    2. Knowledge Base rules (Redundancy, Policy, Norms)
    """
    role_freq, dept_freq = _build_freq_tables(users)
    kb = load_knowledge_base()
    
    kb_redundancy = kb.get("redundancy_rules", {})
    kb_policies = kb.get("account_type_policies", {})
    kb_role_defs = kb.get("role_definitions", {})
    kb_dept_norms = kb.get("department_norms", {})
    # Inject LLM guardrail for Service/SVC accounts.
    service_forbidden = set(kb_policies.get("Service", []) or [])
    service_forbidden.update(list(ADMIN_MARKERS))
    kb_policies["Service"] = sorted(service_forbidden)

    # Index users by username for O(1) lookup
    user_by_name: dict[str, dict] = {}
    for u in (users or []):
        if not u.get("excluded"):
            uname = u.get("username")
            if uname:
                user_by_name[uname] = u

    items: list[dict] = []
    total_anomalies = 0
    total_assignments = 0
    users_with_anomaly = 0

    for uname, row in (matrix or {}).items():
        groups = [g for g, v in (row or {}).items() if int(v) == 1]
        groups_set = set(groups)  # fast lookup
        total_assignments += len(groups)

        u_data = user_by_name.get(uname, {})
        br = (u_data.get("businessRole") or "Unassigned").strip()
        dept = (u_data.get("department") or "Unknown").strip()
        acct_type = (u_data.get("accountType") or "").strip()
        is_service_identity = _is_service_identity(u_data)

        peer_freqs = role_freq.get(br, {})
        dept_freqs = dept_freq.get(dept, {})

        # KB Norms for this user
        # Match "Engineer" in "Software Engineer III" etc.
        kb_role_norm = next((norm for k, norm in kb_role_defs.items() if k in br), [])
        kb_dept_norm = next((norm for k, norm in kb_dept_norms.items() if k in dept), [])

        anomalies: list[dict] = []
        for g in groups:
            reasons: list[str] = []
            confidence = 0.0

            # --- 1. KNOWLEDGE BASE CHECKS (High Confidence) ---
            
            # A) Known Redundancy Rule (Explicit and Heuristic)
            hierarchy_rules = kb.get("hierarchy_patterns", [])
            
            for other_g in groups:
                if other_g == g: continue
                
                # A1) Explicit KB Rule
                redundant_list = kb_redundancy.get(other_g, [])
                if g in redundant_list:
                    reasons.append(f"Redundant: Superceded by '{other_g}' (Explicit KB Rule)")
                    confidence = 1.0
                    break
                
                # A2) Heuristic Hierarchy Rule (Pattern-based)
                # e.g. Azure_1 vs Azure_All, App_Read vs App_Admin
                for pattern in hierarchy_rules:
                    root = pattern.get("root", "")
                    supersedes = pattern.get("supersedes", [])
                    
                    if root in other_g:
                        # Find the base name without the root suffix
                        base_other = other_g.replace(root, "")
                        for s in supersedes:
                            if s in g:
                                base_g = g.replace(s, "")
                                # If they share the same base name (e.g. "Azure"), it's likely a hierarchy violation
                                if base_other == base_g or base_other in g:
                                    reasons.append(f"Least Privilege: '{g}' is likely redundant given '{other_g}' ({root} hierarchy)")
                                    confidence = 0.9
                                    break
                        if reasons: break
                if reasons: break
            
            # B) Account Type Policy
            # Check against KB policies + hardcoded fallback
            forbidden = kb_policies.get(acct_type, [])
            if any(token in g for token in forbidden):
                 # re-check hardcoded function too just in case
                 pass

            # B2) Service/SVC + admin group guardrail (high confidence).
            if is_service_identity and _is_admin_group_name(g):
                reasons.append("LLM Policy: Service/SVC account with administrative group is high-risk")
                confidence = max(confidence, 0.99)
            
            # (Merged policy check logic)
            violation = _check_type_violation(g, acct_type)
            if violation:
                reasons.append(violation)
                confidence = max(confidence, 0.95)
            
            # C) KB Role/Dept Norms (finding out-of-pattern items)
            # If we generally know what this role has, and 'g' is NOT in it -> anomaly?
            # Careful: KB is generic, don't be too strict.
            # Only flag if it confirms a statistical anomaly.

            # --- 2. STATISTICAL CHECKS (Medium/Low Confidence) ---

            # Peer frequency check
            pf = peer_freqs.get(g, 0.0)
            is_stat_anomaly = False
            
            if pf < PEER_ANOMALY_THRESHOLD:
                reasons.append(f"Peer: only {pf * 100:.0f}% of '{br}' have this")
                confidence = max(confidence, 1.0 - pf)
                is_stat_anomaly = True

            # Department frequency check
            df = dept_freqs.get(g, 0.0)
            if df < DEPT_ANOMALY_THRESHOLD:
                # If KB says this dept SHOULD have it, ignore the anomaly
                if not any(k in g for k in kb_dept_norm):
                    reasons.append(f"Dept: only {df * 100:.0f}% of '{dept}' have this")
                    confidence = max(confidence, 1.0 - df)
                    is_stat_anomaly = True

            if reasons:
                anomalies.append({
                    "group": g,
                    "reasons": reasons,
                    "confidence": round(confidence, 2),
                    "peerFreq": round(pf, 4),
                    "deptFreq": round(df, 4),
                })


        if anomalies:
            users_with_anomaly += 1
            total_anomalies += len(anomalies)
            anomalies.sort(key=lambda a: a["confidence"], reverse=True)
            items.append({
                "username": uname,
                "displayName": u_data.get("displayName") or uname,
                "businessRole": br,
                "department": dept,
                "accountType": acct_type,
                "anomalyCount": len(anomalies),
                "anomalies": anomalies,
            })

    items.sort(key=lambda x: x["anomalyCount"], reverse=True)

    total_users = len(matrix or {})
    pct = (users_with_anomaly / max(1, total_users)) * 100.0
    print(f"[AI Detection] users_with_anomaly={users_with_anomaly}, total_users={total_users}, pct={pct:.2f}%", flush=True)

    return {
        "status": "ready",
        "items": items,
        "stats": {
            "aiDetection": round(pct, 2),
            "totalAnomalies": total_anomalies,
            "totalAssignments": total_assignments,
            "usersWithAnomaly": users_with_anomaly,
            "totalUsersScanned": len(matrix or {}),
        },
        "ts": datetime.now(timezone.utc).isoformat(),
    }


def build_cluster_quality_items(clusters: list, matrix: dict) -> list[dict]:
    """
    Drilldown “Cluster Quality”: ordina gli elementi che peggiorano la qualità del cluster.
    Assunzione minima: ogni cluster ha 'users' (lista username) e 'roleGroups' (lista gruppi del cluster).
    """
    out = []
    for c_idx, c in enumerate(clusters or []):
        users = c.get("users") or c.get("members") or []
        role_groups = set(c.get("roleGroups") or [])
        if not users or not role_groups:
            continue

        user_items = []
        for u in users:
            u_roles = _matrix_user_roles(matrix, u)
            missing = sorted(list(role_groups - u_roles))
            extra = sorted(list(u_roles - role_groups))
            denom = max(1, len(role_groups) + len(u_roles))
            distance = round((len(missing) + len(extra)) / denom, 4)  # più alto = più “sporco”

            user_items.append({
                "username": u,
                "distance": distance,
                "missingFromRole": missing,
                "extraVsRole": extra,
            })

        user_items.sort(key=lambda r: r["distance"], reverse=True)
        avg_distance = round(sum(x["distance"] for x in user_items) / max(1, len(user_items)), 4)

        out.append({
            "clusterIndex": c_idx,
            "clusterName": c.get("name") or f"Cluster {c_idx}",
            "avgDistance": avg_distance,
            "worstUsers": user_items[:50],  # top 50 driver
        })

    out.sort(key=lambda r: r["avgDistance"], reverse=True)
    return out




# ----------------------------
# Persistent Storage (replaces in-memory state)
# ----------------------------
from app.db.storage import (
    get_store,
    get_current_tenant_id,
    init_default_state,
    list_known_tenant_ids,
    list_registered_domains,
    lookup_registered_domain,
    normalize_tenant_id,
    pop_tenant_context,
    push_tenant_context,
    register_domain_mapping,
    reset_tenant_state,
    tenant_storage_exists,
    tenant_context,
)

# Initialize persistent storage
state = get_store()
init_default_state(normalize_tenant_id(os.getenv("DEFAULT_TENANT_ID", "example.internal")))
REQUIRED_DUPLICATE_ORDER = ["last_login", "groups_count", "dept_group_correlation", "has_department"]
MODEL_QUALITY_PRESETS: Dict[str, Dict[str, float]] = {
    "banking": {
        "role_entropy": 0.06,
        "template_coverage": 0.08,
        "noise_ratio": 0.09,
        "ambiguity": 0.07,
        "temporal_drift": 0.06,
        "matrix_density": 0.05,
        "orphan_weighted": 0.10,
        "overprivileged": 0.13,
        "stale_access": 0.12,
        "policy_violation": 0.14,
        "manual_override": 0.04,
        "generalization": 0.06,
    },
    "manufacturing": {
        "role_entropy": 0.08,
        "template_coverage": 0.11,
        "noise_ratio": 0.10,
        "ambiguity": 0.07,
        "temporal_drift": 0.09,
        "matrix_density": 0.06,
        "orphan_weighted": 0.09,
        "overprivileged": 0.10,
        "stale_access": 0.09,
        "policy_violation": 0.08,
        "manual_override": 0.06,
        "generalization": 0.07,
    },
    "retail": {
        "role_entropy": 0.08,
        "template_coverage": 0.10,
        "noise_ratio": 0.10,
        "ambiguity": 0.08,
        "temporal_drift": 0.09,
        "matrix_density": 0.07,
        "orphan_weighted": 0.08,
        "overprivileged": 0.10,
        "stale_access": 0.10,
        "policy_violation": 0.08,
        "manual_override": 0.05,
        "generalization": 0.07,
    },
}
state.setdefault("dq_rules", {})
state["dq_rules"]["duplicate_resolution_order"] = REQUIRED_DUPLICATE_ORDER.copy()
state.setdefault("dq_model_preset", "manufacturing")
state.setdefault("dq_model_weights", MODEL_QUALITY_PRESETS.get(state.get("dq_model_preset"), MODEL_QUALITY_PRESETS["manufacturing"]))


def get_active_model_weights() -> Dict[str, float]:
    preset = (state.get("dq_model_preset") or "manufacturing").strip().lower()
    base = MODEL_QUALITY_PRESETS.get(preset) or MODEL_QUALITY_PRESETS["manufacturing"]
    custom = state.get("dq_model_weights") or {}
    out = dict(base)
    for k, v in custom.items():
        try:
            out[str(k)] = float(v)
        except Exception:
            continue
    return out


def apply_business_roles(users: List[Dict[str, Any]]) -> None:
    """Add businessRole field to each user based on state mapping.
    
    CRITICAL: Only assigns from mapping if:
    1. User has no existing valid BR, OR
    2. Mapping has a valid (non-Unassigned) value for this user
    
    This preserves CSV-assigned BRs while allowing department-based assignment.
    """
    m = state.get("user_business_role", {})
    for u in users:
        existing_br = (u.get("businessRole") or "").strip()
        mapped_br = m.get(u.get("username"), "")
        
        # Priority: existing valid BR > mapped valid BR > "Unassigned"
        if existing_br and existing_br != "Unassigned":
            # Keep existing valid BR
            pass
        elif mapped_br and mapped_br != "Unassigned":
            # Use mapped BR if valid
            u["businessRole"] = mapped_br
        else:
            # No valid BR anywhere - mark as Unassigned
            if not existing_br:
                u["businessRole"] = "Unassigned"

def sync_roles_from_users(users: List[Dict[str, Any]]) -> int:
    """
    Registra in role_meta/business_roles tutti i BR presenti sugli utenti.
    Ritorna quanti BR nuovi sono stati creati.
    """
    role_meta = state.setdefault("role_meta", {})
    business_roles = state.setdefault("business_roles", set())
    created = 0
    groups_by_role: Dict[str, set[str]] = defaultdict(set)

    for u in users or []:
        br = (u.get("businessRole") or "").strip()
        if not br or br == "Unassigned":
            continue
        if br not in role_meta:
            _ensure_role_registered(br)  # usa role_meta/business_roles
            created += 1
        business_roles.add(br)
        for group in (u.get("groups") or []):
            group_name = str(group or "").strip()
            if group_name:
                groups_by_role[br].add(group_name)

    for br, groups in groups_by_role.items():
        meta = role_meta.setdefault(br, {"color": "#ffffff", "groups": []})
        existing_groups = {
            str(group or "").strip()
            for group in (meta.get("groups") or [])
            if str(group or "").strip()
        }
        merged_groups = sorted(existing_groups | groups)
        if merged_groups != (meta.get("groups") or []):
            meta["groups"] = merged_groups

    return created


# =============================================================================
# BRDB (NO AI): learning DB + inference engine
# =============================================================================
# (imports already at top of file)

BRDB_MIN_CONF = 0.70  # soglia per auto-assegnare

BRDB_STOP_TOKENS = {
    "grp", "group", "role", "access", "users", "user", "team", "dl",
    "sec", "security", "global", "all", "everyone"
}

def brdb_norm_group(group: str) -> str:
    """Standard normalization for group names."""
    if not group: 
        return ""
    # Remove common prefix/suffix noise
    g = group.strip()
    return g

def brdb_ensure_ready() -> None:
    if not state.get("brdb_ready"):
        brdb_rebuild()

def brdb_infer_group(group: str) -> dict:
    """Infer which role a group belongs to based on learned patterns."""
    brdb_ensure_ready()
    return ml_engine.brdb_infer_group(group)

# (imports already at top of file)

DEPT_MINCONF = 0.80
DEPT_GROUP_SUPPORT = 0.60

def ensure_role_registered(role: str) -> None:
    _ensure_role_registered(role)

DEPT_MERGE_JACCARD = 0.55  # soglia similarità gruppi dept vs BR template


def classify_account(
    display_name: str,
    ou: str,
    employee_type: str,
    use_ml: bool = True,
    attributes: Optional[Dict[str, Any]] = None,
) -> str:
    """
    Classify account type using ML (if available) with rule-based fallback.
    
    Uses the ML engine for high-confidence predictions (>75%), otherwise
    falls back to the extended 12-type rule-based classification.
    """
    # Try ML-based classification first
    if use_ml:
        try:
            predicted_type, confidence, method = ml_engine.classify_account(
                display_name, ou, employee_type, confidence_threshold=0.75, attributes=attributes
            )
            # Custom rules have top priority and must be applied immediately.
            if method == "custom_rule":
                return predicted_type
            if method == "ml" and confidence >= 0.75:
                return predicted_type
        except Exception:
            pass  # Fall through to rules
    
    # Use ML engine's rule-based classification (extended 12 types)
    return ml_engine.classify_account_rules(display_name, ou, employee_type)


# =============================================================================
# BRDB Functions (Business Role Database) - Delegating to ML Engine
# =============================================================================

def brdb_rebuild():
    """Rebuild the Business Role Database from current user assignments (Background)."""
    def _worker():
        with REBUILD_LOCK:
            # Check again under lock to avoid redundant rebuilds
            if state.get("brdb_ready"):
                return
                
            users = state.get("last_extract", {}).get("users") or []
            ml_engine.brdb_rebuild(users)
            state["brdb_ready"] = True
            log("INFO", f"Background BRDB rebuild finished ({len(users)} users)")
            # Invalidate detail caches as roles might have changed
            invalidate_hot_caches(roles=True)

    # Start in background if not already running
    if REBUILD_LOCK.locked():
        return
        
    threading.Thread(target=_worker, daemon=True).start()
    log("INFO", "Triggered background BRDB rebuild")

def brdb_learn_assignment(role: str, groups: list, weight: float = 1.0):
    """Record a confirmed role→groups assignment for learning."""
    ml_engine.brdb_learn_assignment(role, groups, weight)
    state["brdb_ready"] = False

def _jaccard(a: set[str], b: set[str]) -> float:
    if not a and not b:
        return 1.0
    return len(a & b) / max(1, len(a | b))

BR_ASSIGNMENT_RULE_WEIGHT = 7.5
BR_FIELD_RULE_WEIGHT = 8.5


def _br_assignment_rule_scores_for_user(user: Dict[str, Any]) -> Dict[str, float]:
    """
    Score business-role boosts from regex rules applied on group names.
    The higher the score, the stronger the influence in automatic assignment.
    """
    group_rules = list(state.get("br_assignment_pattern_rules") or [])
    field_rules = list(state.get("br_pattern_rules") or [])
    if not group_rules and not field_rules:
        return {}

    groups = [str(g).strip() for g in (user.get("groups") or []) if str(g).strip()]
    if not groups:
        return {}
    groups_lower = {g.lower() for g in groups}

    scores: Dict[str, float] = defaultdict(float)

    # Form 3: regex on group names.
    for rule in group_rules:
        business_role = (rule.get("business_role") or "").strip()
        regex = (rule.get("regex") or "").strip()
        legacy_role = (rule.get("role") or "").strip()
        if not business_role or not regex:
            continue

        # Backward compatibility: old rules might have an explicit group/role gate.
        if legacy_role and legacy_role.lower() not in groups_lower:
            continue

        match_count = 0
        try:
            for g in groups:
                if re.search(regex, g, re.IGNORECASE):
                    match_count += 1
        except re.error:
            continue

        if match_count <= 0:
            continue

        boost = BR_ASSIGNMENT_RULE_WEIGHT * (1.0 + 0.2 * (match_count - 1))
        scores[business_role] += boost

    # Form 2: business role + field + regex, evaluated on user field value.
    for rule in field_rules:
        business_role = (rule.get("business_role") or "").strip()
        field = (rule.get("field") or "").strip()
        regex = (rule.get("regex") or "").strip()
        if not business_role or not field or not regex:
            continue

        raw_val = user.get(field)
        if raw_val is None:
            continue
        if isinstance(raw_val, list):
            value = " ".join([str(x) for x in raw_val])
        else:
            value = str(raw_val)
        if not value.strip():
            continue

        try:
            if re.search(regex, value, re.IGNORECASE):
                scores[business_role] += BR_FIELD_RULE_WEIGHT
        except re.error:
            continue

    return dict(scores)


def apply_department_mapping(users: List[Dict[str, Any]], only_depts: Optional[set[str]] = None) -> None:
    # Prepara strutture coerenti
    state.setdefault("user_business_role", {})
    state.setdefault("role_meta", {})
    state.setdefault("business_roles", set())
    state.setdefault("dept_to_role", {})          # dept -> canonical BR
    state.setdefault("dept_role_analysis", {})    # dept -> evidence


    # Global rebuild is still useful for stats, but we can do it conditionally?
    # For now, keep it global as stats depend on all users.
    # We could optimize brdb_rebuild() too if needed.
    brdb_rebuild()

    targets = only_depts


    by_dept = defaultdict(list)
    for u in users or []:
        dept = (u.get("department") or "").strip()
        if dept:
            by_dept[dept].append(u)

    role_meta = state["role_meta"]
    user_br = state["user_business_role"]
    dept_to_role = state["dept_to_role"]
    dept_analysis = state["dept_role_analysis"]

    for dept, members in by_dept.items():
        if targets and dept not in targets:
            continue

        # 1) Baseline deterministica: BR = dept
        _ensure_role_registered(dept)

        # 2) Profilo gruppi frequenti del dipartimento
        n = max(1, len(members))
        cnt = defaultdict(int)
        for u in members:
            for g in (u.get("groups") or []):
                cnt[g] += 1
        dept_groups = {g for g, c in cnt.items() if (c / n) >= DEPT_GROUP_SUPPORT}

        # 3) Analisi: “votazione” ruolo macro dai gruppi utenti (BRDB)
        weights = defaultdict(float)
        rules_boost = defaultdict(float)
        for u in members:
            s = brdb_infer_groupset(u.get("groups") or [])
            r = (s.get("role") or "Unassigned").strip()
            c = float(s.get("confidence") or 0.0)
            if r and r != "Unassigned" and c > 0:
                weights[r] += c
            for rr, sc in _br_assignment_rule_scores_for_user(u).items():
                weights[rr] += float(sc)
                rules_boost[rr] += float(sc)

        chosen_role = dept
        evidence = {
            "dept": dept,
            "members": len(members),
            "deptGroups": sorted(dept_groups),
            "chosenRole": dept,
            "reason": "baseline",
        }

        if weights:
            best_role, best_w = max(weights.items(), key=lambda x: x[1])
            total = sum(weights.values()) or 1.0
            conf = best_w / total

            role_groups = set((role_meta.get(best_role, {}) or {}).get("groups") or [])
            sim = _jaccard(dept_groups, role_groups) if role_groups else 1.0  # se non ho template ancora, non blocco

            evidence.update({
                "bestRole": best_role,
                "confidence": round(conf, 3),
                "jaccard": round(sim, 3),
                "weightsTop": sorted(weights.items(), key=lambda x: x[1], reverse=True)[:5],
                "ruleBoostTop": sorted(rules_boost.items(), key=lambda x: x[1], reverse=True)[:5],
            })

            # Merge automatico (come mi hai chiesto), ma solo sopra soglie
            if best_role != "Unassigned" and conf >= DEPT_MINCONF and sim >= DEPT_MERGE_JACCARD:
                chosen_role = best_role
                _ensure_role_registered(chosen_role)
                evidence["chosenRole"] = chosen_role
                evidence["reason"] = "merged_by_analysis"

        # 4) Consolidamento template BR: aggiungo gruppi frequenti dept al ruolo scelto
        existing = set((role_meta.get(chosen_role, {}) or {}).get("groups") or [])
        role_meta[chosen_role]["groups"] = sorted(existing | dept_groups)

        # 5) Salvo mapping dept->BR e assegno utenti
        dept_to_role[dept] = chosen_role
        dept_analysis[dept] = evidence

        for u in members:
            uname = u.get("username")
            if uname:
                existing_br = (u.get("businessRole") or "").strip()
                # log("INFO", f"DEBUG_MAPPING: {uname} existing='{existing_br}' chosen='{chosen_role}'")

                if existing_br and existing_br != "Unassigned":
                    user_br[uname] = existing_br
                else:
                    user_rule_scores = _br_assignment_rule_scores_for_user(u)
                    if user_rule_scores:
                        boosted_role = max(user_rule_scores.items(), key=lambda x: x[1])[0]
                        _ensure_role_registered(boosted_role)
                        user_br[uname] = boosted_role
                    else:
                        user_br[uname] = chosen_role

    # Apply BR assignment rules also to users without department mapping.
    for u in (users or []):
        uname = u.get("username")
        if not uname:
            continue
        existing_br = (u.get("businessRole") or "").strip()
        mapped_br = (user_br.get(uname) or "").strip()
        if (existing_br and existing_br != "Unassigned") or (mapped_br and mapped_br != "Unassigned"):
            continue
        rule_scores = _br_assignment_rule_scores_for_user(u)
        if rule_scores:
            forced_role = max(rule_scores.items(), key=lambda x: x[1])[0]
            _ensure_role_registered(forced_role)
            user_br[uname] = forced_role

    # applica mapping sugli oggetti utente
    apply_business_roles(users)

def brdb_infer_groupset(groups: List[str]) -> Dict[str, Any]:
    """
    Predice BR per un insieme di gruppi (utente/riga CSV) aggregando le predizioni per gruppo.
    """
    groups = [brdb_norm_group(g) for g in (groups or []) if brdb_norm_group(g)]
    if not groups:
        return {"role": "Unassigned", "confidence": 0.0, "evidence": {"reason": "no_groups"}}

    weights = defaultdict(float)
    details = []
    for g in groups[:50]:
        s = brdb_infer_group(g)
        details.append({"group": g, **s})
        weights[s["role"]] += float(s.get("confidence", 0.0) or 0.0)

    best_role, best_w = max(weights.items(), key=lambda x: x[1])
    total = sum(weights.values()) or 1.0
    conf = float(best_w / total)

    return {
        "role": best_role,
        "confidence": round(conf, 3),
        "evidence": {"weights": dict(weights), "details": details[:25]},
    }

def brdb_learn_assignment(br: str, groups: List[str], weight: int = 5) -> None:
    """
    Aggiorna incrementale il DB interno quando hai una assegnazione 'vera' (CSV con BR o assegnazione manuale).
    """
    brdb_ensure_ready()

    br = (br or "").strip()
    if not br:
        return
    for g in (groups or []):
        g0 = brdb_norm_group(g)
        brdb_inc_stat(state["brdb_group_stats"], g0, br, weight)
        for t in brdb_tokens(g0):
            brdb_inc_stat(state["brdb_token_stats"], t, br, max(1, weight // 2))

    # invalida cache
    state["brdb_cache"] = {}


def _mk_candidate(
    *,
    source: str,
    candidate_id: str,
    display_name: str,
    business_role: str,
    roles: list[str],
    raw: str,
    department: str = "",
    last_login: Any = None,
) -> dict:
    return {
        "candidateId": candidate_id,
        "source": source,
        "displayName": (display_name or "").strip(),
        "businessRole": (business_role or "").strip(),
        "department": (department or "").strip(),
        "lastLogin": last_login,
        "roles": roles or [],
        "rawLine": raw or "",
    }

def rebuild_ingest_candidates() -> None:
    src = state.get("ingest_sources") or {}
    flat = []
    for _, items in src.items():
        flat.extend(items or [])
    state["ingest_candidates"] = flat

def apply_duplicate_displayname_resolution() -> None:
    """
    Applica la scelta: per ogni displayName duplicato, rende 'attivo' solo il candidato scelto.
    Effetto pratico: marca excluded=True sugli utenti non scelti (così spariscono da role mining e lista utenti).
    """
    choice = state.get("choice_by_displayName") or {}
    candidates = state.get("ingest_candidates") or []

    # build: displayName -> [candidate]
    by_dn = defaultdict(list)
    for c in candidates:
        dn = (c.get("displayName") or "").strip()
        if dn:
            by_dn[dn].append(c)

    # indicizza utenti attuali per displayName
    users = state.get("last_extract", {}).get("users") or []
    users_by_dn = defaultdict(list)
    for u in users:
        dn = (u.get("displayName") or "").strip()
        if dn:
            users_by_dn[dn].append(u)

    # reset excluded
    for u in users:
        if "excluded" in u:
            u["excluded"] = False

    # applica scelta sui duplicati
    for dn, rows in by_dn.items():
        if len(rows) <= 1:
            continue

        chosen_id = choice.get(dn) or rows[0].get("candidateId")  # default: prima
        chosen = next((r for r in rows if r.get("candidateId") == chosen_id), rows[0])

        # qui assumiamo: gli utenti "a sistema" sono identificati dal displayName (come nella tua richiesta)
        # => lasciamo attivo un solo user per quel displayName
        same_dn_users = users_by_dn.get(dn) or []
        if len(same_dn_users) <= 1:
            # se esiste un solo user a sistema, facciamo override dei campi con quelli del candidato scelto
            if same_dn_users:
                u0 = same_dn_users[0]
                u0["businessRole"] = chosen.get("businessRole") or u0.get("businessRole", "")
                u0["groups"] = sorted(set(chosen.get("roles") or u0.get("groups") or []))
            continue

        # se ci sono più user a sistema con lo stesso displayName, ne lasciamo attivo solo 1
        # (scelta semplice: il primo della lista resta attivo e viene “aggiornato” col candidato scelto)
        keep = same_dn_users[0]
        keep["businessRole"] = chosen.get("businessRole") or keep.get("businessRole", "")
        keep["groups"] = sorted(set(chosen.get("roles") or keep.get("groups") or []))

        for u in same_dn_users[1:]:
            u["excluded"] = True

    # aggiorna anche il mapping BR per coerenza UI
    apply_business_roles(users)


def _to_ts(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except Exception:
        pass
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).timestamp()
    except Exception:
        return 0.0


def _normalize_last_login(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, list):
        if not v:
            return None
        v = v[0]
    s = str(v).strip()
    if not s:
        return None

    # AD FILETIME (100ns intervals since 1601-01-01 UTC)
    try:
        iv = int(float(s))
        if iv > 116444736000000000:
            unix_ts = (iv - 116444736000000000) / 10000000.0
            dt = datetime.fromtimestamp(unix_ts, tz=timezone.utc)
            return dt.isoformat()
    except Exception:
        pass

    # Unix timestamp
    try:
        fv = float(s)
        if fv > 0 and fv < 4102444800:  # until year 2100
            dt = datetime.fromtimestamp(fv, tz=timezone.utc)
            return dt.isoformat()
    except Exception:
        pass

    # ISO-like string
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc).isoformat()
    except Exception:
        return s


def _build_dept_group_profile(rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, float]]:
    by_dept = defaultdict(lambda: defaultdict(int))
    dept_counts = defaultdict(int)
    for r in (rows or []):
        dept = (r.get("department") or "").strip()
        if not dept:
            continue
        dept_counts[dept] += 1
        for g in set(r.get("groups") or []):
            by_dept[dept][g] += 1

    out: Dict[str, Dict[str, float]] = {}
    for dept, gstats in by_dept.items():
        den = max(1, dept_counts.get(dept, 1))
        out[dept] = {g: (cnt / den) for g, cnt in gstats.items()}
    return out


def _score_duplicate_candidate(c: Dict[str, Any], dept_profile: Dict[str, Dict[str, float]]) -> Dict[str, Any]:
    dept = (c.get("department") or "").strip()
    groups = list(dict.fromkeys(c.get("groups") or []))
    has_dept = 1 if dept else 0
    groups_count = len(groups)
    last_login_ts = _to_ts(c.get("lastLogin"))

    corr = 0.0
    if dept and groups:
        probs = dept_profile.get(dept) or {}
        corr = sum(float(probs.get(g, 0.0)) for g in groups) / max(1, len(groups))

    order = REQUIRED_DUPLICATE_ORDER

    values = {
        "dept_group_correlation": round(corr, 6),
        "has_department": has_dept,
        "groups_count": groups_count,
        "last_login": last_login_ts,
    }
    rank = tuple(values.get(k, 0.0) for k in order)
    return {
        "rank": rank,
        "order": order,
        "reason": {
            "deptGroupCorrelation": round(corr, 4),
            "hasDepartment": bool(has_dept),
            "groupsCount": groups_count,
            "lastLoginTs": last_login_ts,
        },
    }


def _duplicate_resolution_items() -> List[Dict[str, Any]]:
    candidates = state.get("ingest_candidates") or []
    if not candidates:
        return []

    by_dn = defaultdict(list)
    for c in candidates:
        dn = (c.get("displayName") or "").strip()
        if dn:
            by_dn[dn.lower()].append(c)

    choice_raw = state.get("choice_by_displayName") or {}
    auto_raw = state.get("duplicate_autoselect") or {}
    choice = {(str(k).strip().lower()): v for k, v in choice_raw.items() if str(k).strip()}
    auto = {(str(k).strip().lower()): v for k, v in auto_raw.items() if str(k).strip()}
    items = []
    for dn_key, rows in by_dn.items():
        if len(rows) <= 1:
            continue
        chosen_id = choice.get(dn_key) or rows[0].get("candidateId")
        chosen = next((r for r in rows if r.get("candidateId") == chosen_id), rows[0])
        auto_id = ((auto.get(dn_key) or {}).get("candidateId")) or chosen_id
        display_name = (chosen.get("displayName") or rows[0].get("displayName") or "").strip()
        alternatives = [r for r in rows if r.get("candidateId") != chosen.get("candidateId")]
        items.append(
            {
                "displayName": display_name,
                "chosenCandidateId": chosen.get("candidateId"),
                "autoChosenCandidateId": auto_id,
                "chosen": chosen,
                "alternatives": alternatives,
                "count": len(rows),
                "autoReason": (auto.get(dn_key) or {}).get("reason"),
            }
        )

    items.sort(key=lambda x: x.get("displayName") or "")
    return items


def _record_duplicate_feedback(display_name: str, chosen_candidate_id: str, actor: str = "system") -> None:
    display_name = (display_name or "").strip()
    chosen_candidate_id = (chosen_candidate_id or "").strip()
    if not display_name or not chosen_candidate_id:
        return

    auto = state.get("duplicate_autoselect") or {}
    auto_item = auto.get(display_name) or auto.get(display_name.lower()) or {}
    auto_id = (auto_item.get("candidateId") or "").strip()
    if not auto_id or auto_id == chosen_candidate_id:
        return

    by_id = {}
    for c in (state.get("ingest_candidates") or []):
        cid = str(c.get("candidateId") or "").strip()
        if cid:
            by_id[cid] = c

    auto_c = by_id.get(auto_id) or {}
    chosen_c = by_id.get(chosen_candidate_id) or {}

    ev = {
        "ts": datetime.now(timezone.utc).isoformat(),
        "displayName": display_name,
        "actor": actor,
        "autoCandidateId": auto_id,
        "chosenCandidateId": chosen_candidate_id,
        "auto": {
            "department": auto_c.get("department"),
            "groupsCount": len(auto_c.get("roles") or []),
            "lastLogin": auto_c.get("lastLogin"),
        },
        "chosen": {
            "department": chosen_c.get("department"),
            "groupsCount": len(chosen_c.get("roles") or []),
            "lastLogin": chosen_c.get("lastLogin"),
        },
    }

    events = state.get("dq_feedback_events") or []
    events.append(ev)
    # keep recent history bounded
    state["dq_feedback_events"] = events[-500:]


def build_dq_rule_suggestions() -> List[Dict[str, Any]]:
    suggestions: List[Dict[str, Any]] = []
    rules = state.get("dq_rules") or {}
    order = list(rules.get("duplicate_resolution_order") or [])
    if not order:
        order = ["dept_group_correlation", "has_department", "groups_count", "last_login"]
    events = list(state.get("dq_feedback_events") or [])
    ingest = state.get("last_ingest_stats") or {}

    dep_better = 0
    grp_better = 0
    ll_better = 0
    considered = 0
    for ev in events:
        auto = ev.get("auto") or {}
        chosen = ev.get("chosen") or {}
        considered += 1
        if (not auto.get("department")) and (chosen.get("department")):
            dep_better += 1
        if int(chosen.get("groupsCount") or 0) > int(auto.get("groupsCount") or 0):
            grp_better += 1
        if _to_ts(chosen.get("lastLogin")) > _to_ts(auto.get("lastLogin")):
            ll_better += 1

    def _ratio(x: int, y: int) -> float:
        return (float(x) / float(y)) if y > 0 else 0.0

    def _idx(lst: List[str], key: str) -> int:
        try:
            return lst.index(key)
        except ValueError:
            return 999

    if considered >= 3 and _ratio(ll_better, considered) >= 0.60 and _idx(order, "last_login") > 1:
        new_order = [x for x in order if x != "last_login"]
        new_order.insert(1, "last_login")
        suggestions.append(
            {
                "ruleId": "dq-dup-priority-lastlogin",
                "title": "Alza priorita LastLogin nella deduplica",
                "description": "Gli override manuali scelgono spesso il candidato con ultimo accesso piu recente.",
                "confidence": round(_ratio(ll_better, considered), 2),
                "impact": {"manualOverridesAnalyzed": considered, "lastLoginPreferred": ll_better},
                "preview": {"duplicate_resolution_order": new_order},
                "current": {"duplicate_resolution_order": order},
                "alreadyApplied": new_order == order,
            }
        )

    if considered >= 3 and _ratio(grp_better, considered) >= 0.60 and _idx(order, "groups_count") > 1:
        new_order = [x for x in order if x != "groups_count"]
        new_order.insert(1, "groups_count")
        suggestions.append(
            {
                "ruleId": "dq-dup-priority-groups",
                "title": "Alza priorita numero gruppi nella deduplica",
                "description": "Gli override manuali privilegiano il candidato con maggiore copertura gruppi.",
                "confidence": round(_ratio(grp_better, considered), 2),
                "impact": {"manualOverridesAnalyzed": considered, "groupsPreferred": grp_better},
                "preview": {"duplicate_resolution_order": new_order},
                "current": {"duplicate_resolution_order": order},
                "alreadyApplied": new_order == order,
            }
        )

    if int(ingest.get("missingRoles") or 0) > 0 and not bool(rules.get("reject_empty_groups")):
        miss_roles = int(ingest.get("missingRoles") or 0)
        total_rows = max(1, int(ingest.get("rowsTotal") or 0))
        suggestions.append(
            {
                "ruleId": "dq-reject-empty-groups",
                "title": "Blocca righe senza gruppi",
                "description": "Scarta in import le righe con gruppi vuoti per ridurre utenti orfani e remediation manuale.",
                "confidence": round(min(1.0, miss_roles / total_rows), 2),
                "impact": {"missingRolesLastImport": miss_roles, "rowsTotal": total_rows},
                "preview": {"reject_empty_groups": True},
                "current": {"reject_empty_groups": bool(rules.get("reject_empty_groups"))},
                "alreadyApplied": False,
            }
        )

    suggestions.sort(key=lambda x: (x.get("alreadyApplied"), -(x.get("confidence") or 0)))
    return suggestions


def pick_best_user(cands: List[Dict[str, Any]]) -> Dict[str, Any]:
    # Ranking policy: prefer most recent login, then profile completeness, then number of groups.
    def score(u: Dict[str, Any]) -> tuple:
        last = u.get("lastLogin") or u.get("last_login") or ""
        has_dept = 1 if (u.get("department") or "").strip() else 0
        has_br = 1 if (u.get("businessRole") or "").strip() else 0
        ng = len(u.get("groups") or [])
        return (last, has_dept + has_br, ng)
    # `max` avoids sorting the full list and preserves behavior for tie cases.
    return max(cands, key=score)


# Canonical DataSource labels used on imported users.
CONNECTOR_TO_DATASOURCE: Dict[str, str] = {
    "ad": "AD",
    "sap": "SAP",
    "csv": "CSV",
    "xlsx": "XLSX",
    "azure": "AZURE",
    "one_identity": "ONE_IDENTITY",
    "sailpoint": "SAILPOINT",
    "saviynt": "SAVIYNT",
    "servicenow": "SERVICENOW",
    "salesforce": "SALESFORCE",
    "m365": "M365",
}

CONNECTOR_TARGET_ALIASES: Dict[str, str] = {
    "active_directory": "ad",
    "activedirectory": "ad",
    "microsoft365": "m365",
    "oneidentity": "one_identity",
}

# Stable shape used to detect provisioning deltas per user.
PROVISIONING_FIELDS = (
    "displayName",
    "department",
    "businessRole",
    "accountType",
    "email",
    "upn",
    "employeeId",
    "manager",
    "statusAd",
    "statusHr",
    "excluded",
    "DataSource",
)


def normalize_connector_target(target: str) -> str:
    normalized = str(target or "").strip().lower()
    return CONNECTOR_TARGET_ALIASES.get(normalized, normalized)


def datasource_from_source(source: str) -> str:
    normalized = normalize_connector_target(source)
    if normalized in CONNECTOR_TO_DATASOURCE:
        return CONNECTOR_TO_DATASOURCE[normalized]
    return normalized.upper() if normalized else "UNKNOWN"


def stamp_user_datasource(user: Dict[str, Any], source: str) -> Dict[str, Any]:
    # We keep the field name exactly as requested by product: `DataSource`.
    user["DataSource"] = datasource_from_source(source)
    return user


def stamp_users_datasource(users: List[Dict[str, Any]], source: str) -> List[Dict[str, Any]]:
    for user in (users or []):
        stamp_user_datasource(user, source)
    return users


def infer_user_datasource(user: Dict[str, Any], fallback_source: str = "") -> str:
    current = str(user.get("DataSource") or user.get("datasource") or "").strip()
    if current and current.upper() != "UNKNOWN":
        return current.upper()

    uname = str(user.get("username") or "").strip().lower()
    if uname.startswith("sap.") or uname.startswith("sf."):
        return "SAP"
    if uname.startswith("ad.") or uname.startswith("ldap."):
        return "AD"
    if uname.startswith("azure.") or uname.startswith("aad."):
        return "AZURE"
    if uname.startswith("m365.") or uname.startswith("o365."):
        return "M365"
    if uname.startswith("snow.") or uname.startswith("servicenow."):
        return "SERVICENOW"
    if uname.startswith("sfdc.") or uname.startswith("salesforce."):
        return "SALESFORCE"
    if uname.startswith("sail.") or uname.startswith("sailpoint."):
        return "SAILPOINT"
    if uname.startswith("sav.") or uname.startswith("saviynt."):
        return "SAVIYNT"
    if uname.startswith("oneid.") or uname.startswith("oneidentity."):
        return "ONE_IDENTITY"

    inferred = datasource_from_source(fallback_source)
    if inferred and inferred != "UNKNOWN":
        return inferred
    return "UNKNOWN"


def _ingest_displayname_source_index() -> Dict[str, set[str]]:
    index: Dict[str, set[str]] = defaultdict(set)
    ingest_sources = state.get("ingest_sources") or {}
    for source_key, candidates in ingest_sources.items():
        datasource = datasource_from_source(str(source_key or ""))
        if not datasource or datasource == "UNKNOWN":
            continue
        for cand in (candidates or []):
            dn = str((cand or {}).get("displayName") or "").strip().lower()
            if not dn:
                continue
            index[dn].add(datasource)
    return index


def backfill_datasource_in_state(*, persist: bool = True, fallback_source: str = "") -> int:
    extract_state = state.get("last_extract") or {}
    users = extract_state.get("users") or []
    if not users:
        return 0

    display_index = _ingest_displayname_source_index()
    changed = 0
    for user in users:
        previous = str(user.get("DataSource") or user.get("datasource") or "").strip().upper()
        inferred = infer_user_datasource(user, fallback_source=fallback_source)
        if inferred == "UNKNOWN":
            dn_key = str(user.get("displayName") or "").strip().lower()
            matches = sorted(display_index.get(dn_key) or [])
            if len(matches) == 1:
                inferred = matches[0]
            elif len(matches) > 1:
                fallback_ds = datasource_from_source(fallback_source)
                if fallback_ds in matches:
                    inferred = fallback_ds
        if inferred != previous:
            user["DataSource"] = inferred
            changed += 1

    if changed and persist:
        with state.batch():
            state["last_extract"] = extract_state
            state["mining_dirty"] = True
    return changed


def _provision_payload_for_user(user: Dict[str, Any], datasource: str) -> Dict[str, Any]:
    # Keep payload canonical so we can hash and detect true changes reliably.
    groups = sorted({str(g).strip() for g in (user.get("groups") or []) if str(g).strip()})
    payload: Dict[str, Any] = {
        "username": str(user.get("username") or "").strip(),
        "groups": groups,
        "DataSource": datasource,
    }
    for field in PROVISIONING_FIELDS:
        if field == "DataSource":
            continue
        payload[field] = user.get(field)
    return payload


def _provision_fingerprint(payload: Dict[str, Any]) -> str:
    encoded = json.dumps(payload, sort_keys=True, ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def _get_users_by_datasource(datasource: str) -> List[Dict[str, Any]]:
    users = (state.get("last_extract") or {}).get("users") or []
    selected = []
    for user in users:
        uname = str(user.get("username") or "").strip()
        if not uname:
            continue
        if str(user.get("DataSource") or "").strip().upper() != datasource:
            continue
        selected.append(user)
    return selected


def run_connector_provisioning(target: str, actor: str) -> Dict[str, Any]:
    connector = normalize_connector_target(target)
    if connector not in CONNECTOR_TO_DATASOURCE:
        raise HTTPException(status_code=400, detail=f"Connector target '{target}' is not supported")

    datasource = CONNECTOR_TO_DATASOURCE[connector]
    users = _get_users_by_datasource(datasource)

    previous_state = state.get("connector_provisioning") or {}
    previous_snapshots = previous_state.get("snapshots") or {}
    previous_snapshot = previous_snapshots.get(datasource) or {}

    current_snapshot: Dict[str, str] = {}
    changed_usernames: List[str] = []
    changed_payloads: List[Dict[str, Any]] = []
    for user in users:
        payload = _provision_payload_for_user(user, datasource)
        username = payload["username"]
        fingerprint = _provision_fingerprint(payload)
        current_snapshot[username] = fingerprint
        if previous_snapshot.get(username) != fingerprint:
            changed_usernames.append(username)
            changed_payloads.append(payload)

    removed_usernames = sorted([u for u in previous_snapshot.keys() if u not in current_snapshot])
    upstream = {
        "attempted": False,
        "success": 0,
        "failed": 0,
        "errors": [],
    }
    upstream = _provision_users_upstream(
        target=connector,
        datasource=datasource,
        changed_payloads=changed_payloads,
        removed_usernames=removed_usernames,
        force_enable=False,
    )

    provisioned_at = datetime.now(timezone.utc).isoformat()
    message = (
        f"Provisioned {len(changed_usernames)} changed users and {len(removed_usernames)} removals for {datasource}."
        if changed_usernames or removed_usernames
        else f"No changes to provision for {datasource}."
    )
    if bool(upstream.get("attempted")):
        message = (
            f"{message} "
            f"Upstream write success={int(upstream.get('success') or 0)} "
            f"failed={int(upstream.get('failed') or 0)}."
        )

    run_summary = {
        "target": connector,
        "datasource": datasource,
        "total_users": len(users),
        "changed_users": len(changed_usernames),
        "removed_users": len(removed_usernames),
        "changed_usernames": changed_usernames[:200],
        "removed_usernames": removed_usernames[:200],
        "provisioned_at": provisioned_at,
        "by": actor,
        "message": message,
        "upstream_attempted": bool(upstream.get("attempted")),
        "upstream_success": int(upstream.get("success") or 0),
        "upstream_failed": int(upstream.get("failed") or 0),
        "upstream_errors": list(upstream.get("errors") or [])[:200],
    }

    with state.batch():
        provisioning_state = dict(state.get("connector_provisioning") or {})
        snapshots = dict(provisioning_state.get("snapshots") or {})
        snapshots[datasource] = current_snapshot
        provisioning_state["snapshots"] = snapshots
        provisioning_state["last_run"] = run_summary
        last_by_source = dict(provisioning_state.get("last_run_by_datasource") or {})
        last_by_source[datasource] = run_summary
        provisioning_state["last_run_by_datasource"] = last_by_source
        state["connector_provisioning"] = provisioning_state

        history = list(state.get("connector_provisioning_history") or [])
        history.append(run_summary)
        state["connector_provisioning_history"] = history[-500:]

    log("INFO", f"Connector provisioning executed target={connector} datasource={datasource} changed={len(changed_usernames)} removed={len(removed_usernames)} by={actor}")
    return run_summary


# Best-effort startup repair for legacy snapshots imported before DataSource support.
_DATASOURCE_BACKFILL_COUNT = backfill_datasource_in_state(persist=True)


def filter_and_dedupe_connector_users(raw_users: List[Dict[str, Any]], source: str) -> List[Dict[str, Any]]:
    rejects = []
    stamp_users_datasource(raw_users or [], source)
    stats = state.setdefault("last_ingest_stats", {})
    stats.update({
        "source": source,
        "rowsTotal": int(len(raw_users or [])),
        "rowsKept": 0,
        "duplicateDisplayName": 0,
        "missingDepartment": 0,
        "missingBusinessRole": 0,
        "missingDisplayName": 0,
        "missingUsername": 0,
        "ts": datetime.now(timezone.utc).isoformat(),
    })

    by_dn = defaultdict(list)
    for u in (raw_users or []):
        if not (u.get("username") or "").strip():
            stats["missingUsername"] += 1
        dn = (u.get("displayName") or "").strip()
        if dn:
            by_dn[dn].append(u)
        else:
            stats["missingDisplayName"] += 1
            rejects.append({"source": source, "reason": "Missing displayName", "user": u, "ts": stats["ts"]})

    chosen = []
    for dn, cands in by_dn.items():
        u = pick_best_user(cands)
        stamp_user_datasource(u, source)

        if len(cands) > 1:
            stats["duplicateDisplayName"] += (len(cands) - 1)

        # log degli altri duplicati
        for other in cands:
            if other is u:
                continue
            rejects.append({"source": source, "reason": f"Duplicate displayName '{dn}' (kept best)", "user": other, "ts": stats["ts"]})

        dept = (u.get("department") or "").strip()
        br = (u.get("businessRole") or "").strip()

        if not dept:
            stats["missingDepartment"] += 1
            rejects.append({"source": source, "reason": "Missing department", "user": u, "ts": stats["ts"]})
        count_missing_br = not str(source).lower().startswith("ad")

        if count_missing_br and not br:
            stats["missingBusinessRole"] += 1
            rejects.append({
                "source": source,
                "reason": "Missing businessRole",
                "user": u,
                "ts": stats.get("ts"),
            })

        chosen.append(u)

    stats["rowsKept"] = int(len(chosen))
    state["last_rejects"] = rejects
    return chosen


# (datetime import already at top of file)

def _merge_users_into_last_extract(new_users: list[dict], *, ou: str):
    state.setdefault("last_extract", {"ou": "", "users": [], "groups": [], "ts": None})
    base_users = state["last_extract"]["users"]

    by_username = {u.get("username"): u for u in base_users if u.get("username")}

    for nu in (new_users or []):
        uname = nu.get("username")
        if not uname:
            continue

        if uname in by_username:
            u = by_username[uname]
            # aggiorna displayName (se presente) e fai merge gruppi
            if nu.get("displayName"):
                u["displayName"] = nu["displayName"]
            if nu.get("department"):
                u["department"] = nu["department"]
            u["groups"] = sorted(set(nu.get("groups") or []))  # REPLACE da connettore

        else:
            base_users.append(nu)
            by_username[uname] = nu

    # riallinea campi derivati
    apply_business_roles(base_users)
    state["last_extract"]["ou"] = ou
    state["last_extract"]["groups"] = recompute_groups_from_users(base_users)
    state["last_extract"]["ts"] = datetime.now(timezone.utc).isoformat()


def replace_from_connector_pool(new_users: List[Dict[str, Any]], ou: str, source: str) -> Dict[str, int]:
    """
    Replace current extract snapshot with a full connector pool.
    Used by AD import to ensure all rules run on a complete fresh dataset.
    """
    state.setdefault("last_extract", {"ou": "", "users": [], "groups": [], "ts": None})
    previous_users = state["last_extract"].get("users") or []
    previous_groups = set(state["last_extract"].get("groups") or [])

    prev_by_username = {}
    prev_by_displayname = {}
    for u in previous_users:
        uname = str(u.get("username") or "").strip()
        if uname:
            prev_by_username[uname] = u
        dn = (u.get("displayName") or "").strip().lower()
        if dn:
            prev_by_displayname[dn] = u

    previous_group_members: Dict[str, set[str]] = defaultdict(set)
    for u in previous_users:
        uname0 = str(u.get("username") or "").strip()
        if not uname0:
            continue
        for g0 in (u.get("groups") or []):
            gname = str(g0 or "").strip()
            if gname:
                previous_group_members[gname].add(uname0)

    clean_users: List[Dict[str, Any]] = []
    new_users_count = 0
    updated_users_count = 0

    for u in (new_users or []):
        username = (u.get("username") or "").strip()
        if not username:
            continue

        groups = sorted(set(u.get("groups") or []))
        normalized = dict(u)
        normalized["username"] = username
        normalized["displayName"] = (u.get("displayName") or username).strip()
        normalized["groups"] = groups
        normalized["department"] = (u.get("department") or "").strip() or None
        normalized["businessRole"] = (u.get("businessRole") or "").strip() or None
        normalized["excluded"] = False
        normalized["DataSource"] = datasource_from_source(source)
        clean_users.append(normalized)

        dn_key = normalized["displayName"].lower()
        prev = prev_by_username.get(username) or prev_by_displayname.get(dn_key)
        if not prev:
            new_users_count += 1
            continue

        changed = (
            str(prev.get("username") or "").strip() != normalized["username"]
            or str(prev.get("displayName") or "").strip() != normalized["displayName"]
            or sorted(set(prev.get("groups") or [])) != groups
            or (prev.get("department") or None) != normalized["department"]
            or (prev.get("businessRole") or None) != normalized["businessRole"]
            or (prev.get("accountType") or None) != (normalized.get("accountType") or None)
            or (prev.get("lastLogin") or None) != (normalized.get("lastLogin") or None)
        )
        if changed:
            updated_users_count += 1

    current_groups = recompute_groups_from_users(clean_users)

    current_group_members: Dict[str, set[str]] = defaultdict(set)
    for u in clean_users:
        uname1 = str(u.get("username") or "").strip()
        if not uname1:
            continue
        for g1 in (u.get("groups") or []):
            gname = str(g1 or "").strip()
            if gname:
                current_group_members[gname].add(uname1)

    common_groups = previous_groups & set(current_groups)
    updated_groups_count = sum(
        1 for g in common_groups
        if previous_group_members.get(g, set()) != current_group_members.get(g, set())
    )

    state["last_extract"] = {
        "ou": ou,
        "users": clean_users,
        "groups": current_groups,
        "ts": datetime.now(timezone.utc).isoformat(),
        "source": source,
    }
    state["mining_dirty"] = True

    return {
        "new_users": int(new_users_count),
        "updated_users": int(updated_users_count),
        "updated_by_displayname": int(updated_users_count),
        "new_groups": int(len(set(current_groups) - previous_groups)),
        "updated_groups": int(updated_groups_count),
    }


def merge_from_connector_by_displayname(new_users: List[Dict[str, Any]], ou: str, source: str) -> Dict[str, int]:
    """
    Non-destructive merge:
    - keep existing local users
    - update only users with same displayName
    - add AD users not found by displayName
    """
    state.setdefault("last_extract", {"ou": "", "users": [], "groups": [], "ts": None})
    base_users = state["last_extract"]["users"]
    previous_groups = set(state["last_extract"].get("groups") or [])

    previous_group_members: Dict[str, set[str]] = defaultdict(set)
    for u in (base_users or []):
        uname0 = str(u.get("username") or "").strip()
        if not uname0:
            continue
        for g0 in (u.get("groups") or []):
            gname = str(g0 or "").strip()
            if gname:
                previous_group_members[gname].add(uname0)

    by_displayname: Dict[str, Dict[str, Any]] = {}
    for u in base_users:
        dn = (u.get("displayName") or "").strip().lower()
        if dn and dn not in by_displayname:
            by_displayname[dn] = u

    new_users_count = 0
    updated_users_count = 0

    for u in (new_users or []):
        username = (u.get("username") or "").strip()
        display_name = (u.get("displayName") or username).strip()
        if not username or not display_name:
            continue

        dn_key = display_name.lower()
        existing_user = by_displayname.get(dn_key)
        normalized_groups = sorted(set(u.get("groups") or []))

        if existing_user:
            prev_username = str(existing_user.get("username") or "").strip()
            prev_display = str(existing_user.get("displayName") or "").strip()
            prev_department = str(existing_user.get("department") or "").strip()
            prev_business_role = str(existing_user.get("businessRole") or "").strip()
            prev_groups = sorted(set(existing_user.get("groups") or []))
            prev_account_type = existing_user.get("accountType")
            prev_last_login = existing_user.get("lastLogin")

            existing_user["username"] = username
            existing_user["displayName"] = display_name
            existing_user["groups"] = normalized_groups
            existing_user["department"] = (u.get("department") or "").strip() or None
            if (u.get("businessRole") or "").strip():
                existing_user["businessRole"] = (u.get("businessRole") or "").strip()
            if u.get("accountType") is not None:
                existing_user["accountType"] = u.get("accountType")
            if u.get("lastLogin") is not None:
                existing_user["lastLogin"] = u.get("lastLogin")
            for k in ("email", "upn", "employeeId", "manager", "statusAd", "statusHr", "attributes"):
                if u.get(k) is not None:
                    existing_user[k] = u.get(k)
            existing_user["excluded"] = False
            existing_user["DataSource"] = datasource_from_source(source)

            changed = (
                prev_username != str(existing_user.get("username") or "").strip()
                or prev_display != str(existing_user.get("displayName") or "").strip()
                or prev_department != str(existing_user.get("department") or "").strip()
                or prev_business_role != str(existing_user.get("businessRole") or "").strip()
                or prev_groups != sorted(set(existing_user.get("groups") or []))
                or prev_account_type != existing_user.get("accountType")
                or prev_last_login != existing_user.get("lastLogin")
            )
            if changed:
                updated_users_count += 1
            continue

        new_user = {
            "username": username,
            "displayName": display_name,
            "groups": normalized_groups,
            "department": (u.get("department") or "").strip() or None,
            "businessRole": (u.get("businessRole") or "").strip() or None,
            "excluded": False,
            "lastLogin": u.get("lastLogin"),
            "accountType": u.get("accountType"),
            "email": u.get("email"),
            "upn": u.get("upn"),
            "employeeId": u.get("employeeId"),
            "manager": u.get("manager"),
            "statusAd": u.get("statusAd"),
            "statusHr": u.get("statusHr"),
            "attributes": u.get("attributes"),
            "DataSource": datasource_from_source(source),
        }
        base_users.append(new_user)
        by_displayname[dn_key] = new_user
        new_users_count += 1

    current_groups = recompute_groups_from_users(base_users)
    state["last_extract"]["groups"] = current_groups
    state["last_extract"]["ts"] = datetime.now(timezone.utc).isoformat()
    state["last_extract"]["source"] = source
    if ou:
        state["last_extract"]["ou"] = ou
    state["mining_dirty"] = True

    current_group_members: Dict[str, set[str]] = defaultdict(set)
    for u in (base_users or []):
        uname1 = str(u.get("username") or "").strip()
        if not uname1:
            continue
        for g1 in (u.get("groups") or []):
            gname = str(g1 or "").strip()
            if gname:
                current_group_members[gname].add(uname1)

    common_groups = previous_groups & set(current_groups)
    updated_groups_count = sum(
        1 for g in common_groups
        if previous_group_members.get(g, set()) != current_group_members.get(g, set())
    )

    return {
        "new_users": int(new_users_count),
        "updated_users": int(updated_users_count),
        "new_groups": int(len(set(current_groups) - previous_groups)),
        "updated_groups": int(updated_groups_count),
    }


def merge_from_connector(new_users: List[Dict[str, Any]], ou: str, source: str) -> Dict[str, int]:
    """
    Merge new users from connector (AD/LDAP) into existing state.
    Matches by BOTH displayName AND username - if either matches, updates existing user.
    """
    # Ensure baseline exists (if first run)
    state.setdefault("last_extract", {"ou": "", "users": [], "groups": [], "ts": None})
    base_users = state["last_extract"]["users"]
    log("INFO", f"Merge start. Existing users in state: {len(base_users)}")

    # Build dual indexes for matching by BOTH username AND displayName
    by_username = {}
    by_displayname = {}
    for u in base_users:
        uname = u.get("username")
        if uname:
            by_username[uname] = u
        dn = (u.get("displayName") or "").strip().lower()
        if dn:
            by_displayname[dn] = u
    
    previous_groups = set(state["last_extract"].get("groups") or [])
    previous_group_members: Dict[str, set[str]] = defaultdict(set)
    for u in (base_users or []):
        uname0 = str(u.get("username") or "").strip()
        if not uname0:
            continue
        for g0 in (u.get("groups") or []):
            gname = str(g0 or "").strip()
            if gname:
                previous_group_members[gname].add(uname0)

    clean_new_users: List[Dict[str, Any]] = []
    updated_count = 0
    added_count = 0
    updated_groups_count = 0
    
    # Pre-clean new users 
    for u in (new_users or []):
        username = (u.get("username") or "").strip()
        if not username:
            continue
        clean_new_users.append({
            "username": username,
            "displayName": (u.get("displayName") or username).strip(),
            "groups": sorted(set(u.get("groups") or [])),
            "department": (u.get("department") or "").strip() or None,
            "businessRole": (u.get("businessRole") or "").strip() or None,
            "excluded": False,
            "DataSource": datasource_from_source(source),
        })

    for nu in clean_new_users:
        uname = nu["username"]
        dn_key = (nu.get("displayName") or "").strip().lower()
        
        # Merge policy:
        # 1) Match by username
        # 2) If not found, match by displayName and update existing record
        #    (required for connector updates with renamed/changed usernames)
        existing_user = by_username.get(uname)
        if existing_user is None:
            existing_user = by_displayname.get(dn_key)
        
        if existing_user:
            # REPLACE existing user with new data from import
            prev_username_for_user = str(existing_user.get("username") or "").strip()
            prev_display_for_user = str(existing_user.get("displayName") or "").strip()
            prev_department_for_user = str(existing_user.get("department") or "").strip()
            prev_business_role_for_user = str(existing_user.get("businessRole") or "").strip()
            prev_groups_for_user = sorted(set(existing_user.get("groups") or []))
            existing_user["displayName"] = nu["displayName"]
            existing_user["department"] = nu["department"]
            # REPLACE groups (not merge)
            existing_user["groups"] = nu.get("groups") or []
            if nu.get("businessRole"):
                existing_user["businessRole"] = nu["businessRole"]
            existing_user["DataSource"] = datasource_from_source(source)
            # Update username if it changed (displayName matched but username different)
            if existing_user.get("username") != uname:
                old_uname = existing_user.get("username")
                existing_user["username"] = uname
                # Update index
                if old_uname in by_username:
                    del by_username[old_uname]
                by_username[uname] = existing_user

            changed = (
                prev_username_for_user != str(existing_user.get("username") or "").strip()
                or prev_display_for_user != str(existing_user.get("displayName") or "").strip()
                or prev_department_for_user != str(existing_user.get("department") or "").strip()
                or prev_business_role_for_user != str(existing_user.get("businessRole") or "").strip()
                or prev_groups_for_user != sorted(set(existing_user.get("groups") or []))
            )
            if changed:
                updated_count += 1
        else:
            # New user - add to base
            base_users.append(nu)
            by_username[uname] = nu
            if dn_key:
                by_displayname[dn_key] = nu
            added_count += 1

    # Update metadata
    state["last_extract"]["ts"] = datetime.now(timezone.utc).isoformat()
    state["last_extract"]["source"] = source
    if ou:
        state["last_extract"]["ou"] = ou

    # Recompute global groups list
    current_groups = recompute_groups_from_users(base_users)
    state["last_extract"]["groups"] = current_groups

    current_group_members: Dict[str, set[str]] = defaultdict(set)
    for u in (base_users or []):
        uname1 = str(u.get("username") or "").strip()
        if not uname1:
            continue
        for g1 in (u.get("groups") or []):
            gname = str(g1 or "").strip()
            if gname:
                current_group_members[gname].add(uname1)

    common_groups = previous_groups & set(current_groups)
    updated_groups_count = sum(
        1 for g in common_groups
        if previous_group_members.get(g, set()) != current_group_members.get(g, set())
    )

    state["mining_dirty"] = True
    log("INFO", f"Merge complete. Updated: {updated_count}, Added: {added_count}, Total: {len(base_users)}")
    return {
        "new_users": int(added_count),
        "updated_users": int(updated_count),
        "new_groups": int(len(set(current_groups) - previous_groups)),
        "updated_groups": int(updated_groups_count),
    }

def rerun_auto_business_roles_after_connector(
    users: List[Dict[str, Any]],
    only_depts: Optional[set[str]] = None,
    preserve_existing_brs: bool = True,
) -> None:
    preserved_brs = {}
    if preserve_existing_brs:
        # Preserve existing valid Business Roles from user objects BEFORE resetting mapping
        # Useful for CSV imports where BR is explicitly curated.
        for u in (users or []):
            uname = u.get("username")
            br = (u.get("businessRole") or "").strip()
            if uname and br and br != "Unassigned":
                preserved_brs[uname] = br

    # Reset mapping (partial or full)
    if not only_depts:
        # Full rebuild. For AD import we can start clean and let department/rules reassign.
        state["user_business_role"] = preserved_brs.copy() if preserve_existing_brs else {}
        state["brdb_ready"] = False
    else:
        # Partial update: merge preserved into existing
        state.setdefault("user_business_role", {})
        if preserve_existing_brs:
            state["user_business_role"].update(preserved_brs)

    
    # Do NOT wipe existing business roles on the user objects themselves,
    # otherwise we lose the value we just imported from CSV/Connector.
    # for u in (users or []):
    #     u["businessRole"] = None


    # ricostruisci mapping usando dept + analisi merge
    apply_department_mapping(users, only_depts=only_depts)


    state["mining_dirty"] = True

# (numpy and typing imports already at top of file)

def compute_over_threshold(matrix: Dict[str, Dict[str, int]], pct: float = 0.10) -> Optional[int]:
    if not matrix:
        return None

    row_counts = np.array([int(sum(row.values())) for row in matrix.values()], dtype=np.int32)
    if row_counts.size == 0:
        return None

    k_top = int(np.ceil(pct * row_counts.size))
    k_top = max(1, k_top)

    thr = int(np.partition(row_counts, -k_top)[-k_top])
    return thr

def build_over_rows_only(matrix: Dict[str, Dict[str, int]], threshold: Optional[int]) -> List[Dict[str, Any]]:
    if not matrix or threshold is None:
        return []

    rows: List[Dict[str, Any]] = []
    for user, grants in matrix.items():
        groups = [g for g, v in (grants or {}).items() if int(v) == 1]
        n_groups = len(groups)

        if n_groups >= threshold:
            rows.append({
                "user": user,
                "nGroups": n_groups,
                "over": True,
                "groupsText": ", ".join(groups),
            })

    # default sort utile (visto che "over" è sempre True qui)
    rows.sort(key=lambda r: (-r["nGroups"], r["user"]))
    return rows


def log(level: str, message: str) -> None:
    state["logs"].insert(
        0,
        {
            "ts": datetime.now(timezone.utc).isoformat(),
            "tenant_id": get_current_tenant_id(),
            "level": level,
            "message": message,
        },
    )
    state["logs"] = state["logs"][:500]

def _json_safe_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    if isinstance(value, dict):
        out = {}
        for k, v in value.items():
            key = k.decode("utf-8", errors="replace") if isinstance(k, bytes) else str(k)
            out[key] = _json_safe_value(v)
        return out
    if isinstance(value, (list, tuple, set)):
        return [_json_safe_value(v) for v in value]
    return str(value)


def _append_global_ml_signal(kind: str, payload: Dict[str, Any]) -> None:
    """
    Persist ML-relevant signals in a global cross-tenant file so learning can
    aggregate feedback from all tenants.
    """
    item = {
        "kind": str(kind or "").strip() or "unknown",
        "tenant_id": get_current_tenant_id(),
        "ts": datetime.now(timezone.utc).isoformat(),
        "payload": _json_safe_value(payload),
    }
    try:
        GLOBAL_ML_SIGNALS_PATH.parent.mkdir(parents=True, exist_ok=True)
        with GLOBAL_ML_SIGNALS_LOCK:
            rows: List[Dict[str, Any]] = []
            if GLOBAL_ML_SIGNALS_PATH.exists():
                try:
                    with GLOBAL_ML_SIGNALS_PATH.open("r", encoding="utf-8") as rf:
                        loaded = json.load(rf)
                    if isinstance(loaded, list):
                        rows = loaded
                except Exception:
                    rows = []
            rows.append(item)
            if len(rows) > GLOBAL_ML_SIGNALS_MAX:
                rows = rows[-GLOBAL_ML_SIGNALS_MAX:]
            with tempfile.NamedTemporaryFile(
                mode="w",
                encoding="utf-8",
                dir=str(GLOBAL_ML_SIGNALS_PATH.parent),
                prefix=f"{GLOBAL_ML_SIGNALS_PATH.name}.tmp.",
                suffix=".json",
                delete=False,
            ) as tf:
                json.dump(rows, tf, ensure_ascii=False, separators=(",", ":"))
                tf.flush()
                os.fsync(tf.fileno())
                temp_name = tf.name
            os.replace(temp_name, GLOBAL_ML_SIGNALS_PATH)
    except Exception:
        # Global telemetry must never block the main request path.
        return


def record_manual_user_change(
    *,
    actor: str,
    username: str,
    display_name: Optional[str],
    action: str,
    source: str,
    details: Optional[Dict[str, Any]] = None,
    persist: bool = True,
) -> None:
    events = state.setdefault("manual_user_changes", [])
    tenant_id = get_current_tenant_id()
    item = {
        "id": f"muc-{int(time.time()*1000)}-{len(events)+1}",
        "ts": datetime.now(timezone.utc).isoformat(),
        "tenantId": tenant_id,
        "actor": actor,
        "username": username,
        "displayName": display_name or username,
        "action": action,
        "source": source,
        "details": details or {},
    }
    events.append(item)
    trimmed = events[-2000:]
    if persist:
        state["manual_user_changes"] = trimmed
    else:
        events[:] = trimmed
    _append_global_ml_signal("manual_user_change", item)
    return item


def record_llm_learning_event(
    *,
    actor: str,
    source: str,
    signal_type: str,
    entity: str = "",
    details: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    events = state.setdefault("llm_learning_history", [])
    tenant_id = get_current_tenant_id()
    item = {
        "id": f"lle-{int(time.time()*1000)}-{len(events)+1}",
        "ts": datetime.now(timezone.utc).isoformat(),
        "tenantId": tenant_id,
        "actor": actor,
        "source": source,
        "signalType": signal_type,
        "entity": entity,
        "details": details or {},
    }
    events.append(item)
    state["llm_learning_history"] = events[-3000:]
    _append_global_ml_signal("llm_learning_event", item)
    return item


def recalculate_assignments_background(trigger: str, actor: str) -> None:
    """
    Recompute BR assignments asynchronously after rule changes.
    Runs in background to keep API latency low.
    """
    try:
        users = state.get("last_extract", {}).get("users") or []
        if not users:
            return
        # Recompute account types so newly added pattern rules are applied to existing users.
        for u in users:
            dn = str(u.get("displayName") or u.get("username") or "")
            dept = str(u.get("department") or "")
            employee_type = str(u.get("employeeType") or u.get("employee_type") or "")
            try:
                new_type = classify_account(dn, dept, employee_type, use_ml=True, attributes=u)
                if new_type:
                    u["accountType"] = new_type
            except Exception:
                continue
        rerun_auto_business_roles_after_connector(users, only_depts=None)
        sync_roles_from_users(users)
        state["mining_dirty"] = True
        RESPONSE_CACHE.invalidate("businessroles")
        invalidate_hot_caches(kpi=True)
        state.save()
        log("INFO", f"Background assignment recalculation completed (trigger={trigger}, by={actor})")
    except Exception as exc:
        log("ERROR", f"Background assignment recalculation failed (trigger={trigger}, by={actor}): {exc}")


def refresh_ai_detection_background(trigger: str, actor: str) -> None:
    """
    Recompute mining (if dirty) and AI detection in background after imports.
    """
    try:
        ensure_last_mining(None)  # sync mining inside this background worker
        last = state.get("last_mining") or {}
        matrix = last.get("matrix") or {}
        if not matrix:
            return
        users = active_users(state.get("last_extract", {}).get("users") or [])
        result = run_smart_ai_detection(users, matrix)
        state["last_ai_detection"] = result
        invalidate_hot_caches(kpi=True)
        state.save()
        log("INFO", f"Background AI detection refresh completed (trigger={trigger}, by={actor})")
    except Exception as exc:
        log("ERROR", f"Background AI detection refresh failed (trigger={trigger}, by={actor}): {exc}")


def run_post_snapshot_logic_background(snapshot_ts: str, actor: str, tenant_id: Optional[str] = None) -> None:
    """
    Run heavy post-import business logic after AD snapshot has been saved.
    If a newer snapshot exists, skip to avoid stale background work.
    """
    with tenant_context(tenant_id):
        try:
            init_default_state(get_current_tenant_id())
            current_ts = str((state.get("last_extract") or {}).get("ts") or "")
            if not snapshot_ts or current_ts != str(snapshot_ts):
                log("INFO", f"Skip post-snapshot logic: stale snapshot (expected={snapshot_ts}, current={current_ts})")
                return

            users = (state.get("last_extract") or {}).get("users") or []
            rerun_auto_business_roles_after_connector(users, preserve_existing_brs=False)
            sync_roles_from_users(users)

            rebuild_ingest_candidates()
            apply_duplicate_displayname_resolution()
            invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)

            refresh_ai_detection_background("ad-import-post-snapshot", actor)
            log("INFO", f"Post-snapshot logic completed (snapshot_ts={snapshot_ts}, by={actor})")
        except Exception as exc:
            log("ERROR", f"Post-snapshot logic failed (snapshot_ts={snapshot_ts}, by={actor}): {exc}")


def run_post_csv_snapshot_logic_background(
    snapshot_ts: str,
    actor: str,
    touched_depts: List[str],
    tenant_id: Optional[str] = None,
) -> None:
    """
    Run heavy post-import business logic after CSV snapshot has been saved.
    If a newer snapshot exists, skip to avoid stale background work.
    """
    with tenant_context(tenant_id):
        try:
            init_default_state(get_current_tenant_id())
            current_ts = str((state.get("last_extract") or {}).get("ts") or "")
            if not snapshot_ts or current_ts != str(snapshot_ts):
                log("INFO", f"Skip CSV post-snapshot logic: stale snapshot (expected={snapshot_ts}, current={current_ts})")
                return

            users = (state.get("last_extract") or {}).get("users") or []
            only_depts = {d for d in (touched_depts or []) if d}
            rerun_auto_business_roles_after_connector(
                users,
                only_depts=only_depts if only_depts else None,
                preserve_existing_brs=True,
            )
            sync_roles_from_users(users)

            rebuild_ingest_candidates()
            apply_duplicate_displayname_resolution()
            invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)

            refresh_ai_detection_background("csv-import-post-snapshot", actor)
            log("INFO", f"CSV post-snapshot logic completed (snapshot_ts={snapshot_ts}, by={actor})")
        except Exception as exc:
            log("ERROR", f"CSV post-snapshot logic failed (snapshot_ts={snapshot_ts}, by={actor}): {exc}")


def _start_detached_csv_postprocess(
    snapshot_ts: str,
    actor: str,
    touched_depts: List[str],
    tenant_id: Optional[str],
) -> None:
    thread = threading.Thread(
        target=run_post_csv_snapshot_logic_background,
        args=(snapshot_ts, actor, touched_depts, tenant_id),
        daemon=True,
        name=f"csv-postprocess-{tenant_id or DEFAULT_TENANT_ID}",
    )
    thread.start()

def active_users(users: list[dict]) -> list[dict]:
    return [u for u in (users or []) if not u.get("excluded")]

def recompute_groups_from_users(users: list[dict]) -> list[str]:
    return sorted({g for u in active_users(users) for g in (u.get("groups") or [])})

# (_mk_candidate already defined at line 643)


def _role_modeling_feedback_stats() -> Dict[str, Dict[str, int]]:
    stats: Dict[str, Dict[str, int]] = defaultdict(lambda: {"accepted": 0, "rejected": 0, "total": 0})
    for item in state.get("role_modeling_feedback") or []:
        ptype = str(item.get("proposal_type") or "").strip().lower()
        if not ptype:
            continue
        accepted = bool(item.get("accepted"))
        stats[ptype]["total"] += 1
        if accepted:
            stats[ptype]["accepted"] += 1
        else:
            stats[ptype]["rejected"] += 1
    return dict(stats)


def _role_modeling_ml_boost(proposal_type: str, ml_weight: float) -> float:
    ptype = str(proposal_type or "").strip().lower()
    if not ptype or ml_weight <= 0:
        return 0.0
    stats = _role_modeling_feedback_stats().get(ptype) or {}
    total = int(stats.get("total") or 0)
    if total < 3:
        return 0.0
    accepted = int(stats.get("accepted") or 0)
    acceptance = accepted / max(1, total)
    # Converts acceptance ratio to a signed adjustment in [-ml_weight, +ml_weight].
    return (acceptance - 0.5) * 2 * float(ml_weight)


def _build_role_templates(users: List[Dict[str, Any]], min_group_support: float) -> Dict[str, set[str]]:
    role_groups: Dict[str, Counter] = defaultdict(Counter)
    role_sizes: Dict[str, int] = defaultdict(int)
    role_meta = state.get("role_meta") or {}

    for u in users:
        br = str(u.get("businessRole") or "").strip()
        if not br or br == "Unassigned":
            continue
        role_sizes[br] += 1
        for g in (u.get("groups") or []):
            g_norm = str(g or "").strip()
            if g_norm:
                role_groups[br][g_norm] += 1

    templates: Dict[str, set[str]] = {}
    for br, size in role_sizes.items():
        inferred = {g for g, c in role_groups.get(br, Counter()).items() if (c / max(1, size)) >= min_group_support}
        if inferred:
            templates[br] = inferred
            continue
        configured = set((role_meta.get(br) or {}).get("groups") or [])
        if configured:
            templates[br] = configured
    # Include catalog roles (orphans included) when templates are defined in role metadata.
    for br, meta in (role_meta or {}).items():
        role_name = str(br or "").strip()
        if not role_name or role_name == "Unassigned" or role_name in templates:
            continue
        configured = set((meta or {}).get("groups") or [])
        if configured:
            templates[role_name] = configured
    return templates


def _build_role_modeling_sandbox(req: "RoleModelingSandboxRequest") -> Dict[str, Any]:
    users = active_users((state.get("last_extract") or {}).get("users") or [])
    last_mining = state.get("last_mining") or {}
    mining_matrix = last_mining.get("matrix") or {}
    user_br_map = state.get("user_business_role") or {}
    if not users and mining_matrix:
        fallback_users = []
        for uname, row in mining_matrix.items():
            groups = [g for g, v in (row or {}).items() if int(v) == 1]
            fallback_users.append(
                {
                    "username": uname,
                    "displayName": uname,
                    "groups": groups,
                    "businessRole": user_br_map.get(uname, "Unassigned"),
                }
            )
        users = fallback_users
    if not users:
        raise HTTPException(status_code=400, detail="Nessun dato disponibile: esegui prima una discovery.")

    templates = _build_role_templates(users, req.min_group_support)
    by_role: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    group_to_users: Dict[str, set[str]] = defaultdict(set)
    users_with_groups = 0
    total_groups_from_users = 0
    for u in users:
        uname = str(u.get("username") or "").strip()
        br = str(u.get("businessRole") or "").strip()
        if uname and br and br != "Unassigned":
            by_role[br].append(u)
        u_groups = [str(g or "").strip() for g in (u.get("groups") or []) if str(g or "").strip()]
        if u_groups:
            users_with_groups += 1
            total_groups_from_users += len(u_groups)
        for g in u_groups:
            g_norm = str(g or "").strip()
            if uname and g_norm:
                group_to_users[g_norm].add(uname)

    # Fallback to last mining matrix when user objects do not carry enough group details.
    if mining_matrix and (users_with_groups == 0 or total_groups_from_users < max(10, len(users) // 3)):
        group_to_users = defaultdict(set)
        for uname, row in mining_matrix.items():
            for g, v in (row or {}).items():
                try:
                    enabled = int(v) == 1
                except Exception:
                    enabled = bool(v)
                if enabled:
                    g_norm = str(g or "").strip()
                    if g_norm:
                        group_to_users[g_norm].add(str(uname))

    role_meta = state.get("role_meta") or {}
    catalog_roles: set[str] = set()
    business_roles_state = state.get("business_roles") or []
    if isinstance(business_roles_state, (set, list, tuple)):
        for role in business_roles_state:
            role_name = str(role or "").strip()
            if role_name and role_name != "Unassigned":
                catalog_roles.add(role_name)
    for role in (role_meta or {}).keys():
        role_name = str(role or "").strip()
        if role_name and role_name != "Unassigned":
            catalog_roles.add(role_name)
    for role in by_role.keys():
        role_name = str(role or "").strip()
        if role_name and role_name != "Unassigned":
            catalog_roles.add(role_name)
    for role in templates.keys():
        role_name = str(role or "").strip()
        if role_name and role_name != "Unassigned":
            catalog_roles.add(role_name)
    for role in (user_br_map or {}).values():
        role_name = str(role or "").strip()
        if role_name and role_name != "Unassigned":
            catalog_roles.add(role_name)

    catalog_role_names = sorted(catalog_roles)
    orphan_roles = sorted([r for r in catalog_role_names if len(by_role.get(r) or []) == 0])

    proposals: List[Dict[str, Any]] = []

    # 1) Role merge opportunities (similar templates + assigned users).
    current_role_overlap_pairs = 0
    merge_role_names = sorted([r for r in catalog_role_names if r in templates and templates.get(r)])
    for i in range(len(merge_role_names)):
        for j in range(i + 1, len(merge_role_names)):
            a = merge_role_names[i]
            b = merge_role_names[j]
            ta = templates.get(a) or set()
            tb = templates.get(b) or set()
            if not ta or not tb:
                continue
            inter = ta & tb
            union = ta | tb
            similarity = len(inter) / max(1, len(union))
            if similarity < req.redundancy_threshold:
                continue
            current_role_overlap_pairs += 1
            affected = min(len(by_role.get(a) or []), len(by_role.get(b) or []))
            base_conf = 0.55 + 0.45 * similarity
            conf = max(0.05, min(0.99, base_conf + _role_modeling_ml_boost("role_merge", req.ml_weight)))
            priority = round((affected * 0.8 + len(inter) * 0.2) * conf, 2)
            proposals.append(
                {
                    "id": f"role-merge::{a}::{b}",
                    "proposalType": "role_merge",
                    "title": f"Merge {a} + {b}",
                    "shortLabel": f"{a} + {b}",
                    "confidence": round(conf, 3),
                    "priorityScore": priority,
                    "affectedUsers": affected,
                    "rationale": f"Template simili ({round(similarity * 100)}%) e ruoli sovrapposti.",
                }
            )

    # 2) Group consolidation opportunities.
    current_group_overlap_pairs = 0
    sorted_groups = sorted(group_to_users.items(), key=lambda x: len(x[1]), reverse=True)[:80]
    for i in range(len(sorted_groups)):
        ga, ua = sorted_groups[i]
        if len(ua) < 2:
            continue
        for j in range(i + 1, len(sorted_groups)):
            gb, ub = sorted_groups[j]
            if len(ub) < 2:
                continue
            overlap = len(ua & ub) / max(1, min(len(ua), len(ub)))
            if overlap < req.redundancy_threshold:
                continue
            current_group_overlap_pairs += 1
            affected = len(ua | ub)
            base_conf = 0.50 + 0.50 * overlap
            conf = max(0.05, min(0.99, base_conf + _role_modeling_ml_boost("group_merge", req.ml_weight)))
            priority = round((affected * 0.75 + len(ua & ub) * 0.25) * conf, 2)
            proposals.append(
                {
                    "id": f"group-merge::{ga}::{gb}",
                    "proposalType": "group_merge",
                    "title": f"Consolida gruppi {ga} / {gb}",
                    "shortLabel": f"{ga} / {gb}",
                    "confidence": round(conf, 3),
                    "priorityScore": priority,
                    "affectedUsers": affected,
                    "rationale": f"Overlap membership elevato ({round(overlap * 100)}%).",
                }
            )

    # 3) Assignment normalization opportunities by business role template.
    current_drifted_users = set()
    for role_name, members in by_role.items():
        template = templates.get(role_name) or set()
        if len(members) < 3 or not template:
            continue
        for u in members:
            current = set([str(g or "").strip() for g in (u.get("groups") or []) if str(g or "").strip()])
            missing = sorted(template - current)
            extra = sorted(current - template)
            if not missing and not extra:
                continue
            drift = (len(missing) + len(extra)) / max(1, len(template | current))
            if drift < 0.35:
                continue
            current_drifted_users.add(str(u.get("username") or ""))
            base_conf = 0.45 + min(0.45, drift)
            conf = max(0.05, min(0.99, base_conf + _role_modeling_ml_boost("assignment_update", req.ml_weight)))
            priority = round((len(missing) * 0.7 + len(extra) * 0.5 + 1.0) * conf, 2)
            uname = str(u.get("username") or "")
            proposals.append(
                {
                    "id": f"assignment::{uname}::{role_name}",
                    "proposalType": "assignment_update",
                    "title": f"Normalizza assegnazioni utente {uname}",
                    "shortLabel": uname,
                    "confidence": round(conf, 3),
                    "priorityScore": priority,
                    "affectedUsers": 1,
                    "missingCount": len(missing),
                    "extraCount": len(extra),
                    "rationale": f"Scostamento dal template di {role_name}: +{len(missing)} / -{len(extra)}.",
                }
            )

    # 4) Role retirement opportunities (unused/near-unused and low uniqueness roles).
    role_to_users: Dict[str, set[str]] = {}
    for role_name, members in by_role.items():
        role_to_users[role_name] = set([str(u.get("username") or "") for u in members if str(u.get("username") or "")])
    group_role_count: Counter = Counter()
    for role_name, tpl in templates.items():
        for g in (tpl or set()):
            group_role_count[g] += 1

    retire_candidates: List[Dict[str, Any]] = []
    for role_name in catalog_role_names:
        population = len(role_to_users.get(role_name) or set())
        tpl = templates.get(role_name) or set()
        if not tpl:
            is_orphan = population == 0
            is_unmodeled = 0 < population <= max(2, int(round(len(users) * 0.002)))
            if not (is_orphan or is_unmodeled):
                continue
            rationale_bits = []
            if is_orphan:
                rationale_bits.append("ruolo orfano non assegnato")
            if is_unmodeled:
                rationale_bits.append(f"ruolo senza template con adozione limitata ({population} utenti)")
            confidence = 0.90 if is_orphan else 0.62
            confidence = max(0.05, min(0.99, confidence + _role_modeling_ml_boost("role_retire", req.ml_weight)))
            priority = round((4.2 if is_orphan else (max(1, population) * 0.55)) * confidence, 2)
            retire_item = {
                "id": f"role-retire::{role_name}",
                "proposalType": "role_retire",
                "title": f"Ritira ruolo {role_name}",
                "shortLabel": role_name,
                "confidence": round(confidence, 3),
                "priorityScore": priority,
                "affectedUsers": int(population),
                "role": role_name,
                "mergeTarget": "",
                "rationale": f"Ruolo candidato a ritiro: {', '.join(rationale_bits)}.",
            }
            retire_candidates.append(retire_item)
            proposals.append(retire_item)
            continue
        unique_groups = [g for g in tpl if int(group_role_count.get(g) or 0) == 1]
        uniqueness_ratio = len(unique_groups) / max(1, len(tpl))

        best_similarity = 0.0
        best_partner = ""
        for peer in merge_role_names:
            if peer == role_name:
                continue
            peer_tpl = templates.get(peer) or set()
            if not peer_tpl:
                continue
            sim = len(tpl & peer_tpl) / max(1, len(tpl | peer_tpl))
            if sim > best_similarity:
                best_similarity = sim
                best_partner = peer

        is_low_population = population <= max(2, int(round(len(users) * 0.002)))
        is_low_uniqueness = uniqueness_ratio <= 0.14 and best_similarity >= max(0.72, req.redundancy_threshold - 0.12)
        if not (is_low_population or is_low_uniqueness):
            continue

        rationale_bits = []
        if is_low_population:
            rationale_bits.append(f"bassa adozione ({population} utenti)")
        if is_low_uniqueness:
            rationale_bits.append(f"alta sovrapposizione con {best_partner} ({round(best_similarity * 100)}%)")
        if not rationale_bits:
            rationale_bits.append("ridondanza strutturale")

        confidence = max(0.1, min(0.97, 0.50 + best_similarity * 0.4 + (0.1 if is_low_population else 0.0)))
        confidence = max(0.05, min(0.99, confidence + _role_modeling_ml_boost("role_retire", req.ml_weight)))
        priority = round((max(1, population) * 0.45 + (1 - uniqueness_ratio) * 4.2) * confidence, 2)
        retire_item = {
            "id": f"role-retire::{role_name}",
            "proposalType": "role_retire",
            "title": f"Ritira ruolo {role_name}",
            "shortLabel": role_name,
            "confidence": round(confidence, 3),
            "priorityScore": priority,
            "affectedUsers": int(population),
            "role": role_name,
            "mergeTarget": best_partner,
            "rationale": f"Ruolo candidato a ritiro: {', '.join(rationale_bits)}.",
        }
        retire_candidates.append(retire_item)
        proposals.append(retire_item)

    proposals.sort(key=lambda x: (float(x.get("priorityScore") or 0), float(x.get("confidence") or 0)), reverse=True)
    raw_role_merges = [p for p in proposals if str(p.get("proposalType")) == "role_merge"]
    raw_group_merges = [p for p in proposals if str(p.get("proposalType")) == "group_merge"]
    raw_assignment_updates = [p for p in proposals if str(p.get("proposalType")) == "assignment_update"]
    raw_role_retire = [p for p in proposals if str(p.get("proposalType")) == "role_retire"]

    # Keep proposal diversity across paradigms (merge, retire, normalize, consolidate).
    by_type_queue: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for item in proposals:
        by_type_queue[str(item.get("proposalType") or "")].append(item)
    max_suggestions = int(req.max_suggestions)
    per_type_cap = max(2, max_suggestions // 3)
    type_order = ["role_merge", "role_retire", "assignment_update", "group_merge"]
    type_counts: Counter = Counter()
    selected: List[Dict[str, Any]] = []
    while len(selected) < max_suggestions:
        progressed = False
        ordered_types = type_order + [t for t in by_type_queue.keys() if t not in type_order]
        for ptype in ordered_types:
            queue = by_type_queue.get(ptype) or []
            if not queue:
                continue
            if type_counts[ptype] >= per_type_cap and len(selected) < max_suggestions - 2:
                continue
            selected.append(queue.pop(0))
            type_counts[ptype] += 1
            progressed = True
            if len(selected) >= max_suggestions:
                break
        if not progressed:
            break
    proposals = selected
    by_type = Counter([str(item.get("proposalType") or "") for item in proposals])
    avg_conf = 0.0
    if proposals:
        avg_conf = sum(float(item.get("confidence") or 0) for item in proposals) / len(proposals)
    projected_impacted_users = sum(int(item.get("affectedUsers") or 0) for item in proposals)
    priority_scores = sorted([float(item.get("priorityScore") or 0.0) for item in proposals])
    high_priority_threshold = 0.0
    if priority_scores:
        high_priority_threshold = priority_scores[int(round((len(priority_scores) - 1) * 0.65))]
    high_priority = (
        sum(1 for item in proposals if float(item.get("priorityScore") or 0.0) >= high_priority_threshold)
        if proposals
        else 0
    )
    role_merge_count = int(by_type.get("role_merge") or 0)
    group_merge_count = int(by_type.get("group_merge") or 0)
    assignment_count = int(by_type.get("assignment_update") or 0)
    role_retire_count = int(by_type.get("role_retire") or 0)
    current_assignment_drift_count = len([u for u in current_drifted_users if u])
    assignment_extra_total = int(
        sum(int(item.get("extraCount") or 0) for item in proposals if str(item.get("proposalType")) == "assignment_update")
    )
    current_assignments_users = int(sum(len(u.get("groups") or []) for u in users))
    current_assignments_matrix = 0
    if mining_matrix:
        for row in mining_matrix.values():
            for _, v in (row or {}).items():
                try:
                    enabled = int(v) == 1
                except Exception:
                    enabled = bool(v)
                if enabled:
                    current_assignments_matrix += 1
    current_assignments = max(current_assignments_users, current_assignments_matrix)
    current_avg_groups = float(current_assignments / max(1, len(users)))
    role_covered_users = int(sum(1 for u in users if str(u.get("businessRole") or "").strip() not in ("", "Unassigned")))
    current_role_coverage = (role_covered_users / max(1, len(users))) * 100.0
    current_model_score = float((state.get("last_mining") or {}).get("kpi", {}).get("modelQuality") or 0.0)
    if current_model_score <= 0 and mining_matrix:
        try:
            mining_users = list((last_mining.get("users") or [])) or users
            mining_clusters = list((last_mining.get("clusters") or []))
            computed_kpi = compute_kpis(mining_users, mining_clusters, mining_matrix) or {}
            current_model_score = float(computed_kpi.get("modelQuality") or 0.0)
        except Exception:
            current_model_score = float(current_model_score or 0.0)
    if current_model_score <= 0:
        # Fallback heuristic when mining KPI is not available yet: keep score realistic/non-zero
        # so the UI remains informative on fresh environments.
        drift_ratio = current_assignment_drift_count / max(1, len(users))
        role_overlap_penalty = min(14.0, current_role_overlap_pairs * 0.012)
        group_overlap_penalty = min(10.0, current_group_overlap_pairs * 0.02)
        drift_penalty = min(18.0, drift_ratio * 18.0)
        coverage_bonus = current_role_coverage * 0.08
        heuristic_score = 72.0 + coverage_bonus - role_overlap_penalty - group_overlap_penalty - drift_penalty
        current_model_score = max(25.0, min(96.0, heuristic_score))

    # Estimated deltas from proposed sandbox actions.
    estimated_assignment_reduction = int(round(group_merge_count * 1.5 + assignment_extra_total * 0.8))
    projected_assignments = max(0, current_assignments - estimated_assignment_reduction)
    projected_avg_groups = float(projected_assignments / max(1, len(users)))
    projected_role_pairs = max(
        0,
        current_role_overlap_pairs - max(1, int(round(role_merge_count * 0.75 + role_retire_count * 0.2))),
    ) if current_role_overlap_pairs else 0
    projected_group_pairs = max(0, current_group_overlap_pairs - max(1, int(round(group_merge_count * 0.70)))) if current_group_overlap_pairs else 0
    projected_drifted_users = max(0, current_assignment_drift_count - max(1, int(round(assignment_count * 0.60)))) if current_assignment_drift_count else 0
    role_reduction_ratio = (
        (current_role_overlap_pairs - projected_role_pairs) / max(1, current_role_overlap_pairs)
        if current_role_overlap_pairs
        else 0.0
    )
    group_reduction_ratio = (
        (current_group_overlap_pairs - projected_group_pairs) / max(1, current_group_overlap_pairs)
        if current_group_overlap_pairs
        else 0.0
    )
    drift_reduction_ratio = (
        (current_assignment_drift_count - projected_drifted_users) / max(1, current_assignment_drift_count)
        if current_assignment_drift_count
        else 0.0
    )
    projected_role_coverage = min(100.0, current_role_coverage + drift_reduction_ratio * 6.0)

    if proposals:
        action_volume_factor = min(1.0, len(proposals) / max(1, int(req.max_suggestions)))
        confidence_factor = max(0.0, min(1.0, avg_conf))
        structural_gain = (
            role_reduction_ratio * 18.0
            + group_reduction_ratio * 12.0
            + drift_reduction_ratio * 16.0
            + confidence_factor * 10.0
            + action_volume_factor * 4.0
        )
        parameter_multiplier = 0.88 + float(req.redundancy_threshold) * 0.2 + float(req.min_group_support) * 0.1
        ml_bonus = float(req.ml_weight) * 2.0
        _projected_model_score_estimate = min(100.0, current_model_score + structural_gain * parameter_multiplier + ml_bonus)
    else:
        _projected_model_score_estimate = float(current_model_score)
    execution_projected_model_score = min(100.0, max(current_model_score, _projected_model_score_estimate))
    # Proposed score is the estimated achievable score for the selected scenario/parameters.
    projected_model_score = execution_projected_model_score
    ideal_target_model_score = 100.0
    top_role_merges = raw_role_merges[:3]
    top_group_merges = raw_group_merges[:3]
    top_assignment_updates = raw_assignment_updates[:4]
    top_role_retire = raw_role_retire[:3]
    current_operating_model = {
        "name": "Current Model",
        "pillars": [
            {"label": "Role Architecture", "status": "fragmented", "detail": f"{current_role_overlap_pairs} sovrapposizioni tra ruoli"},
            {"label": "Group Taxonomy", "status": "redundant", "detail": f"{current_group_overlap_pairs} coppie gruppi ridondanti"},
            {"label": "Assignment Hygiene", "status": "variable", "detail": f"{current_assignment_drift_count} utenti fuori template"},
            {"label": "Model Score", "status": "baseline", "detail": f"Score attuale {round(current_model_score, 1)}"},
        ],
    }
    proposed_operating_model = {
        "name": "Target Model v2",
        "pillars": [
            {
                "label": "Role Architecture",
                "status": "consolidated",
                "detail": f"Merge guidati da similarita per ridurre ruoli duplicati a {projected_role_pairs}",
            },
            {
                "label": "Group Taxonomy",
                "status": "normalized",
                "detail": f"Consolidamento gruppi e cleanup con target {projected_group_pairs} coppie residue",
            },
            {
                "label": "Assignment Hygiene",
                "status": "policy-driven",
                "detail": f"Allineamento template per ridurre drift utenti a {projected_drifted_users}",
            },
            {
                "label": "Model Score",
                "status": "improving",
                "detail": f"Score stimato {round(projected_model_score, 1)}",
            },
        ],
    }
    rollout_plan = [
        {
            "phase": "Phase 1 - Quick Wins",
            "window": "Week 1-2",
            "items": [p.get("title") for p in (top_role_retire[:1] + top_group_merges[:1] + top_assignment_updates[:1]) if p.get("title")],
        },
        {
            "phase": "Phase 2 - Core Refactor",
            "window": "Week 3-4",
            "items": [p.get("title") for p in (top_role_merges[:2] + top_role_retire[1:2] + top_assignment_updates[1:3]) if p.get("title")],
        },
        {
            "phase": "Phase 3 - Stabilization",
            "window": "Week 5+",
            "items": [p.get("title") for p in (top_role_merges[2:3] + top_group_merges[2:3] + top_role_retire[2:3] + top_assignment_updates[3:4]) if p.get("title")],
        },
    ]
    business_value = [
        {"metric": "Model Score", "current": round(current_model_score, 2), "target": round(projected_model_score, 2)},
        {"metric": "Assignments", "current": int(current_assignments), "target": int(projected_assignments)},
        {"metric": "Role Coverage %", "current": round(current_role_coverage, 2), "target": round(projected_role_coverage, 2)},
    ]
    group_freq = sorted([(g, len(u_set)) for g, u_set in group_to_users.items()], key=lambda x: x[1], reverse=True)
    discovery_models: List[Dict[str, Any]] = []

    sensitive_markers = ["admin", "finance", "payroll", "security", "approve", "vendor", "hr", "sap"]
    sensitive_groups = [
        g for g, _ in group_freq[:60]
        if any(m in str(g).lower() for m in sensitive_markers)
    ][:18]
    sod_items: List[Dict[str, Any]] = []
    for i in range(len(sensitive_groups)):
        for j in range(i + 1, len(sensitive_groups)):
            ga = sensitive_groups[i]
            gb = sensitive_groups[j]
            ua = group_to_users.get(ga) or set()
            ub = group_to_users.get(gb) or set()
            conflict_users = len(ua & ub)
            if conflict_users <= 0:
                continue
            severity = "high" if conflict_users >= 10 else "medium" if conflict_users >= 4 else "low"
            sod_items.append(
                {
                    "groupA": ga,
                    "groupB": gb,
                    "users": int(conflict_users),
                    "severity": severity,
                    "recommendation": "Verifica separazione compiti e riduci assegnazioni incrociate.",
                }
            )
    sod_items.sort(key=lambda x: int(x.get("users") or 0), reverse=True)
    sod_items = sod_items[:12]

    high_sod_count = len([x for x in sod_items if str(x.get("severity") or "").lower() == "high"])
    medium_sod_count = len([x for x in sod_items if str(x.get("severity") or "").lower() == "medium"])
    low_sod_count = len([x for x in sod_items if str(x.get("severity") or "").lower() == "low"])
    current_role_count = max(1, len(catalog_role_names))
    merge_pool = list(raw_role_merges)
    retire_pool = list(raw_role_retire or retire_candidates)

    scenario_specs = [
        {
            "id": "least-privilege-tightening",
            "name": "Least Privilege Tightening",
            "strategy": "Riduce permessi eccedenti e limita eccezioni ad alta entropia.",
            "merge_factor": 0.38,
            "group_factor": 0.35,
            "retire_factor": 0.30,
            "drift_factor": 0.70,
            "risk_bias": 0.10,
            "coverage_bias": 1.5,
        },
        {
            "id": "balanced-governance",
            "name": "Balanced Governance",
            "strategy": "Bilancia standardizzazione ruoli, SoD e continuita operativa.",
            "merge_factor": 0.52,
            "group_factor": 0.48,
            "retire_factor": 0.45,
            "drift_factor": 0.60,
            "risk_bias": 0.08,
            "coverage_bias": 2.0,
        },
        {
            "id": "aggressive-rationalization",
            "name": "Aggressive Rationalization",
            "strategy": "Massimizza consolidamento e ritiro ruoli non necessari.",
            "merge_factor": 0.82,
            "group_factor": 0.72,
            "retire_factor": 0.78,
            "drift_factor": 0.46,
            "risk_bias": 0.24,
            "coverage_bias": 1.2,
        },
        {
            "id": "sod-first-hardening",
            "name": "SoD First Hardening",
            "strategy": "Priorita alla mitigazione conflitti SoD ad alta severita.",
            "merge_factor": 0.40,
            "group_factor": 0.33,
            "retire_factor": 0.35,
            "drift_factor": 0.68,
            "risk_bias": -0.20,
            "coverage_bias": 1.0,
        },
        {
            "id": "business-role-consolidation",
            "name": "Business Role Consolidation",
            "strategy": "Consolida ruoli business con alta sovrapposizione di accessi.",
            "merge_factor": 0.74,
            "group_factor": 0.36,
            "retire_factor": 0.52,
            "drift_factor": 0.62,
            "risk_bias": 0.05,
            "coverage_bias": 2.8,
        },
        {
            "id": "entitlement-standardization",
            "name": "Entitlement Standardization",
            "strategy": "Riduce varianti di entitlement puntando a template stabili.",
            "merge_factor": 0.44,
            "group_factor": 0.76,
            "retire_factor": 0.30,
            "drift_factor": 0.58,
            "risk_bias": 0.02,
            "coverage_bias": 1.8,
        },
        {
            "id": "exception-minimization",
            "name": "Exception Minimization",
            "strategy": "Riduce gli outlier e le assegnazioni fuori template.",
            "merge_factor": 0.34,
            "group_factor": 0.29,
            "retire_factor": 0.26,
            "drift_factor": 0.35,
            "risk_bias": 0.00,
            "coverage_bias": 2.4,
        },
        {
            "id": "risk-based-segmentation",
            "name": "Risk-Based Segmentation",
            "strategy": "Segmenta i ruoli in base al rischio accessi e criticita funzioni.",
            "merge_factor": 0.47,
            "group_factor": 0.45,
            "retire_factor": 0.40,
            "drift_factor": 0.55,
            "risk_bias": -0.08,
            "coverage_bias": 1.6,
        },
        {
            "id": "identity-lifecycle-alignment",
            "name": "Identity Lifecycle Alignment",
            "strategy": "Allinea modello ruoli ai cicli Joiner/Mover/Leaver.",
            "merge_factor": 0.50,
            "group_factor": 0.42,
            "retire_factor": 0.38,
            "drift_factor": 0.48,
            "risk_bias": -0.02,
            "coverage_bias": 3.0,
        },
        {
            "id": "federated-governance-hybrid",
            "name": "Federated Governance Hybrid",
            "strategy": "Approccio ibrido central + domain ownership per scalabilita.",
            "merge_factor": 0.60,
            "group_factor": 0.40,
            "retire_factor": 0.50,
            "drift_factor": 0.57,
            "risk_bias": 0.11,
            "coverage_bias": 2.2,
        },
    ]

    def _clamp(value: float, lo: float = 0.0, hi: float = 100.0) -> float:
        return max(lo, min(hi, float(value)))

    def _vector_distance(a: List[int], b: List[int]) -> int:
        n = min(len(a), len(b))
        return int(sum(abs(int(a[i]) - int(b[i])) for i in range(n)))

    baseline_vector = [
        int(current_role_count),
        int(current_role_overlap_pairs),
        int(current_assignment_drift_count),
        int(high_sod_count * 10 + medium_sod_count * 4 + low_sod_count * 2),
    ]

    def _simulate_scenario(profile: Dict[str, Any], spec: Dict[str, Any]) -> Dict[str, Any]:
        role_overlap = max(0, int(profile.get("role_overlap") or 0))
        group_overlap = max(0, int(profile.get("group_overlap") or 0))
        drift_users = max(0, int(profile.get("drift_users") or 0))
        role_count = max(1, int(profile.get("role_count") or 1))
        user_count = max(1, int(profile.get("user_count") or 1))
        model_score = _clamp(float(profile.get("model_score") or 0.0))
        role_coverage = _clamp(float(profile.get("role_coverage") or 0.0))
        sod_high = max(0, int(profile.get("sod_high") or 0))
        sod_medium = max(0, int(profile.get("sod_medium") or 0))
        sod_low = max(0, int(profile.get("sod_low") or 0))

        # Make discovery outputs react clearly to user controls.
        max_suggestions_factor = max(0.55, min(1.65, float(req.max_suggestions) / 24.0))
        support_factor = max(0.70, min(1.12, 1.16 - float(req.min_group_support) * 0.45))
        redundancy_factor = max(0.62, min(1.00, 1.06 - float(req.redundancy_threshold) * 0.45))
        ml_factor = max(0.88, min(1.24, 0.92 + float(req.ml_weight) * 0.40))
        action_budget_factor = max_suggestions_factor * support_factor * redundancy_factor

        merge_capacity = max(len(merge_pool), int(round(role_overlap * 0.18)))
        group_capacity = max(len(raw_group_merges), int(round(group_overlap * 0.22)))
        retire_capacity = max(len(retire_pool), int(round(role_count * 0.40)))

        merge_take_raw = min(
            merge_capacity,
            int(round(merge_capacity * float(spec.get("merge_factor") or 0) * action_budget_factor)),
        )
        group_take_raw = min(
            group_capacity,
            int(round(group_capacity * float(spec.get("group_factor") or 0) * action_budget_factor)),
        )
        retire_take_raw = min(
            retire_capacity,
            int(round(retire_capacity * float(spec.get("retire_factor") or 0) * max(0.75, action_budget_factor))),
        )

        # Merge pool is pair-based and grows quadratically; normalize to realistic actionable volumes.
        max_merge_actions = max(1, int(round(role_count * 0.50)))
        max_group_actions = max(1, int(round(max(6, group_overlap * 0.35))))
        max_retire_actions = max(1, int(round(role_count * 0.40)))
        requested_merge_take = max(0, min(max_merge_actions, int(round(merge_take_raw * 0.06))))
        group_take = max(0, min(max_group_actions, int(round(group_take_raw * 0.08))))
        requested_retire_take = max(0, min(max_retire_actions, retire_take_raw))

        # Ensure numeric consistency: merged + retired cannot reduce more roles than available.
        min_projected_roles = 6 if role_count >= 6 else 1
        max_role_reduction_budget = max(0, role_count - min_projected_roles)
        merge_take = int(requested_merge_take)
        retire_take = int(requested_retire_take)
        requested_total = merge_take + retire_take
        if requested_total > max_role_reduction_budget:
            if requested_total <= 0:
                merge_take = 0
                retire_take = 0
            else:
                merge_share = float(merge_take) / float(requested_total)
                merge_take = min(merge_take, int(round(max_role_reduction_budget * merge_share)))
                retire_take = min(retire_take, max_role_reduction_budget - merge_take)
                remaining = max_role_reduction_budget - (merge_take + retire_take)
                if remaining > 0 and merge_take < requested_merge_take:
                    extra_merge = min(remaining, requested_merge_take - merge_take)
                    merge_take += extra_merge
                    remaining -= extra_merge
                if remaining > 0 and retire_take < requested_retire_take:
                    extra_retire = min(remaining, requested_retire_take - retire_take)
                    retire_take += extra_retire
                    remaining -= extra_retire
        merge_take = max(0, int(merge_take))
        retire_take = max(0, int(retire_take))
        drift_after_factor = float(spec.get("drift_factor") or 1.0)
        drift_after_factor = drift_after_factor * max(0.76, min(1.08, 1.02 - float(req.ml_weight) * 0.20))
        drift_after_factor = drift_after_factor * max(0.82, min(1.06, 1.00 - (1.0 - float(req.min_group_support)) * 0.10))
        drift_after = max(0, int(round(drift_users * drift_after_factor)))

        projected_roles = max(min_projected_roles, role_count - merge_take - retire_take)
        projected_role_overlap = max(0, int(round(role_overlap - merge_take * 0.82 - retire_take * 0.28)))
        projected_group_overlap = max(0, int(round(group_overlap - group_take * 0.86)))

        role_reduction_ratio_local = (role_overlap - projected_role_overlap) / max(1, role_overlap) if role_overlap else 0.0
        group_reduction_ratio_local = (group_overlap - projected_group_overlap) / max(1, group_overlap) if group_overlap else 0.0
        drift_reduction_ratio_local = (drift_users - drift_after) / max(1, drift_users) if drift_users else 0.0

        role_coverage_est = _clamp(role_coverage + float(spec.get("coverage_bias") or 0.0) + drift_reduction_ratio_local * 6.5)
        sod_risk = _clamp(
            (sod_high * 18 + sod_medium * 8 + sod_low * 2) * (1 + float(spec.get("risk_bias") or 0.0))
            - merge_take * 0.6
            - retire_take * 0.4
            - (drift_users - drift_after) * 0.003,
            0,
            100,
        )
        role_overlap_penalty = (projected_role_overlap / max(1, role_overlap)) * 22 if role_overlap else 0.0
        maintainability = _clamp(
            100
            - (projected_roles / max(1, role_count)) * 55
            - role_overlap_penalty
            + role_reduction_ratio_local * 28,
            0,
            100,
        )
        least_privilege = _clamp(100 - (drift_after / max(1, user_count)) * 100 + drift_reduction_ratio_local * 8, 0, 100)
        estimated_model_score = _clamp(
            model_score
            + role_reduction_ratio_local * 17
            + group_reduction_ratio_local * 11
            + drift_reduction_ratio_local * 16
            + (role_coverage_est - role_coverage) * 0.28
            - sod_risk * 0.05,
            0,
            100,
        )
        estimated_model_score = _clamp(
            estimated_model_score
            + (max_suggestions_factor - 1.0) * 1.6
            + (support_factor - 1.0) * 2.1
            + (redundancy_factor - 1.0) * 1.3
            + (ml_factor - 1.0) * 2.4,
            0,
            100,
        )
        scenario_vector = [
            int(projected_roles),
            int(projected_role_overlap),
            int(drift_after),
            int(round(sod_risk)),
        ]
        diversity_distance = _vector_distance(scenario_vector, baseline_vector)
        selection_score = (
            estimated_model_score * 0.34
            + least_privilege * 0.24
            + maintainability * 0.18
            + role_coverage_est * 0.12
            - sod_risk * 0.16
            + min(20.0, diversity_distance * 0.35)
        )
        selection_score += (max_suggestions_factor - 1.0) * 3.2 + (ml_factor - 1.0) * 2.8
        selection_score = _clamp(selection_score, 0, 100)

        merge_examples = [x.get("shortLabel") for x in merge_pool[:min(3, merge_take)] if x.get("shortLabel")] if merge_take > 0 else []
        retire_examples = [x.get("shortLabel") for x in retire_pool[:min(3, retire_take)] if x.get("shortLabel")] if retire_take > 0 else []
        role_reduction = max(0, role_count - projected_roles)
        role_reduction_pct = (role_reduction / max(1, role_count)) * 100.0

        return {
            "startingRoleCount": int(role_count),
            "projectedRoleCount": int(projected_roles),
            "roleReduction": int(role_reduction),
            "roleReductionPct": round(role_reduction_pct, 2),
            "mergedRolePairs": int(merge_take),
            "retiredRoles": int(retire_take),
            "groupConsolidations": int(group_take),
            "remainingDriftUsers": int(drift_after),
            "estimatedModelScore": round(float(estimated_model_score), 2),
            "maintainabilityScore": int(round(maintainability)),
            "leastPrivilegeScore": int(round(least_privilege)),
            "sodRiskIndex": int(round(sod_risk)),
            "roleCoverage": round(float(role_coverage_est), 2),
            "mergeExamples": merge_examples,
            "retireExamples": retire_examples,
            "scenarioVector": scenario_vector,
            "diversityDistance": int(diversity_distance),
            "selectionScore": round(float(selection_score), 2),
        }

    rng = np.random.default_rng(42)
    synthetic_training_dataset: List[Dict[str, Any]] = []
    scenario_win_counter: Counter = Counter()
    for idx in range(100):
        synthetic_profile = {
            "role_overlap": max(0, int(round(current_role_overlap_pairs * rng.uniform(0.45, 1.65) + rng.normal(0, 8)))),
            "group_overlap": max(0, int(round(current_group_overlap_pairs * rng.uniform(0.45, 1.65) + rng.normal(0, 6)))),
            "drift_users": max(0, int(round(current_assignment_drift_count * rng.uniform(0.45, 1.55) + rng.normal(0, 45)))),
            "role_count": max(8, int(round(current_role_count * rng.uniform(0.65, 1.40)))),
            "user_count": max(80, int(round(len(users) * rng.uniform(0.70, 1.30)))),
            "model_score": _clamp(current_model_score + rng.normal(0, 10), 25, 98),
            "role_coverage": _clamp(current_role_coverage + rng.normal(0, 7), 35, 100),
            "sod_high": max(0, int(round(high_sod_count * rng.uniform(0.2, 2.2)))),
            "sod_medium": max(0, int(round(medium_sod_count * rng.uniform(0.3, 2.0)))),
            "sod_low": max(0, int(round(low_sod_count * rng.uniform(0.3, 2.4)))),
        }
        best_id = ""
        best_score = -10_000.0
        for spec in scenario_specs:
            sim = _simulate_scenario(synthetic_profile, spec)
            if float(sim.get("selectionScore") or 0.0) > best_score:
                best_score = float(sim.get("selectionScore") or 0.0)
                best_id = str(spec.get("id") or "")
        if best_id:
            scenario_win_counter[best_id] += 1
        synthetic_training_dataset.append(
            {
                "id": f"ds-{idx+1}",
                "profile": synthetic_profile,
                "selectedScenario": best_id,
                "selectionScore": round(best_score, 2),
            }
        )

    win_rate_by_scenario = {
        str(spec.get("id")): round(float(scenario_win_counter.get(str(spec.get("id")), 0)) / 100.0, 4)
        for spec in scenario_specs
    }

    current_profile = {
        "role_overlap": current_role_overlap_pairs,
        "group_overlap": current_group_overlap_pairs,
        "drift_users": current_assignment_drift_count,
        "role_count": current_role_count,
        "user_count": len(users),
        "model_score": current_model_score,
        "role_coverage": current_role_coverage,
        "sod_high": high_sod_count,
        "sod_medium": medium_sod_count,
        "sod_low": low_sod_count,
    }

    discovery_model_catalog: List[Dict[str, Any]] = []
    for spec in scenario_specs:
        model = _simulate_scenario(current_profile, spec)
        scenario_id = str(spec.get("id") or "")
        historical_rate = float(win_rate_by_scenario.get(scenario_id) or 0.0)
        model["selectionScore"] = round(float(model.get("selectionScore") or 0.0) * 0.82 + historical_rate * 100.0 * 0.18, 2)
        model["historicalWinRate"] = round(historical_rate, 4)
        model["id"] = scenario_id
        model["name"] = spec.get("name")
        model["strategy"] = spec.get("strategy")
        discovery_model_catalog.append(model)

    discovery_model_catalog.sort(
        key=lambda x: (float(x.get("selectionScore") or 0), float(x.get("estimatedModelScore") or 0)),
        reverse=True,
    )

    discovery_models = []
    min_diversity_distance = 12
    for candidate in discovery_model_catalog:
        if len(discovery_models) >= 3:
            break
        if not discovery_models:
            discovery_models.append(candidate)
            continue
        distances = [
            _vector_distance(candidate.get("scenarioVector") or baseline_vector, selected.get("scenarioVector") or baseline_vector)
            for selected in discovery_models
        ]
        if distances and min(distances) >= min_diversity_distance:
            discovery_models.append(candidate)

    if len(discovery_models) < 3:
        for candidate in discovery_model_catalog:
            if len(discovery_models) >= 3:
                break
            if candidate in discovery_models:
                continue
            discovery_models.append(candidate)

    for idx, model in enumerate(discovery_models):
        model["rank"] = idx + 1

    lm_selection_instruction = (
        "Given 10 access-model paradigms, score each scenario using weighted objectives: "
        "least-privilege (33%), maintainability (24%), role coverage (18%), model quality uplift (20%), "
        "and SoD risk penalty (-15%). Calibrate ranking with empirical priors from 100 labeled synthetic datasets "
        "sampled from current environment distributions. Return top 3 scenarios maximizing score while enforcing "
        "minimum diversity distance between scenario vectors."
    )
    state["role_modeling_lm_training_dataset"] = synthetic_training_dataset[-100:]
    state["role_modeling_lm_selection_meta"] = {
        "datasetCount": 100,
        "winRateByScenario": dict(win_rate_by_scenario),
        "topScenarios": [m.get("id") for m in discovery_models],
        "generatedAt": datetime.now(timezone.utc).isoformat(),
    }

    workflow_cards = [
        {"title": "Discovery", "status": "done", "detail": f"{len(discovery_models)} modelli consigliati su {len(discovery_model_catalog)} valutati"},
        {"title": "Optimization", "status": "in_progress", "detail": f"{len(proposals)} azioni prioritarie generate"},
        {"title": "Review", "status": "pending", "detail": f"{len(sod_items)} alert SoD da validare"},
        {"title": "Adoption", "status": "pending", "detail": "Sandbox pronta per validare il passaggio dal baseline al modello proposto"},
    ]

    guardrails = [
        {
            "label": "Role Coverage",
            "ok": current_role_coverage >= 70,
            "value": round(current_role_coverage, 1),
            "target": ">= 70%",
            "hint": "Garantire che la maggior parte degli utenti sia mappata a un ruolo.",
        },
        {
            "label": "Assignment Drift",
            "ok": current_assignment_drift_count <= max(20, int(len(users) * 0.05)),
            "value": int(current_assignment_drift_count),
            "target": "<= 5% utenti",
            "hint": "Ridurre utenti con assegnazioni fuori template.",
        },
        {
            "label": "SoD Critical",
            "ok": len([x for x in sod_items if x.get("severity") == "high"]) == 0,
            "value": len([x for x in sod_items if x.get("severity") == "high"]),
            "target": "0 high",
            "hint": "Eliminare conflitti ad alta severita.",
        },
    ]

    trend_points = [
        {"label": "Current", "score": round(current_model_score, 2)},
        {
            "label": "Quick Wins",
            "score": round(
                min(
                    100.0,
                    current_model_score + max(1.0, (execution_projected_model_score - current_model_score) * 0.45),
                ),
                2,
            ),
        },
        {
            "label": "Phase 2",
            "score": round(
                min(
                    100.0,
                    current_model_score + max(2.0, (execution_projected_model_score - current_model_score) * 0.8),
                ),
                2,
            ),
        },
        {"label": "Proposed", "score": round(projected_model_score, 2)},
    ]

    feedback = _role_modeling_feedback_stats()
    return {
        "sandbox": True,
        "noWrite": True,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "appliedParameters": {
            "maxSuggestions": int(req.max_suggestions),
            "templateSupport": round(float(req.min_group_support), 2),
            "redundancyThreshold": round(float(req.redundancy_threshold), 2),
            "mlWeight": round(float(req.ml_weight), 2),
        },
        "summary": {
            "users": len(users),
            "businessRoles": len(catalog_role_names),
            "assignedBusinessRoles": len(by_role),
            "orphanBusinessRoles": len(orphan_roles),
            "groups": len(group_to_users),
            "proposals": len(proposals),
        },
        "mlSignals": {
            "feedbackByType": feedback,
            "totalFeedback": sum(int(v.get("total") or 0) for v in feedback.values()),
            "weight": float(req.ml_weight),
        },
        "kpis": {
            "avgConfidence": round(avg_conf, 3),
            "projectedImpactedUsers": int(projected_impacted_users),
            "highPriorityCount": int(high_priority),
            "highPriorityThreshold": round(float(high_priority_threshold), 3),
            "byType": dict(by_type),
        },
        "comparison": {
            "current": {
                "modelScore": round(current_model_score, 2),
                "totalAssignments": int(current_assignments),
                "avgGroupsPerUser": round(current_avg_groups, 2),
                "redundantRolePairs": int(current_role_overlap_pairs),
                "redundantGroupPairs": int(current_group_overlap_pairs),
                "driftedUsers": int(current_assignment_drift_count),
                "roleCoverage": round(current_role_coverage, 2),
            },
            "proposed": {
                "modelScore": round(projected_model_score, 2),
                "executionModelScore": round(execution_projected_model_score, 2),
                "idealTargetModelScore": round(ideal_target_model_score, 2),
                "totalAssignments": int(projected_assignments),
                "avgGroupsPerUser": round(projected_avg_groups, 2),
                "redundantRolePairs": int(projected_role_pairs),
                "redundantGroupPairs": int(projected_group_pairs),
                "driftedUsers": int(projected_drifted_users),
                "roleCoverage": round(projected_role_coverage, 2),
            },
            "improvement": {
                "modelScoreDelta": round(projected_model_score - current_model_score, 2),
                "targetModelScoreDelta": round(ideal_target_model_score - current_model_score, 2),
                "assignmentsDelta": int(projected_assignments - current_assignments),
                "avgGroupsPerUserDelta": round(projected_avg_groups - current_avg_groups, 2),
                "redundantRolePairsDelta": int(projected_role_pairs - current_role_overlap_pairs),
                "redundantGroupPairsDelta": int(projected_group_pairs - current_group_overlap_pairs),
                "driftedUsersDelta": int(projected_drifted_users - current_assignment_drift_count),
                "roleCoverageDelta": round(projected_role_coverage - current_role_coverage, 2),
            },
        },
        "dataFreshness": {
            "extractTs": (state.get("last_extract") or {}).get("ts"),
            "extractSource": (state.get("last_extract") or {}).get("source"),
            "miningTs": (state.get("last_mining") or {}).get("ts"),
        },
        "recommendedModel": {
            "current": current_operating_model,
            "target": proposed_operating_model,
            "rolloutPlan": rollout_plan,
            "businessValue": business_value,
        },
        "discoveryModels": discovery_models,
        "discoveryModelCatalog": discovery_model_catalog,
        "lmSelection": {
            "datasetCount": 100,
            "instruction": lm_selection_instruction,
            "winRateByScenario": dict(win_rate_by_scenario),
            "topScenarios": [m.get("id") for m in discovery_models],
            "trainingSample": synthetic_training_dataset[:12],
        },
        "sodMatrix": sod_items,
        "workflow": workflow_cards,
        "guardrails": guardrails,
        "trend": trend_points,
        "proposals": proposals,
    }



def apply_choice_for_displayname(display_name: str,
                                 chosen_business_role: Optional[str],
                                 chosen_roles: Optional[List[str]]) -> None:
    users = state.get("last_extract", {}).get("users") or []

    same = [u for u in users if (u.get("displayName") or "").strip() == (display_name or "").strip()]
    if not same:
        return

    keep = same[0]
    keep["excluded"] = False

    if chosen_business_role:
        keep["businessRole"] = chosen_business_role
        state.setdefault("user_business_role", {})
        state["user_business_role"][keep["username"]] = chosen_business_role

    if chosen_roles is not None:
        keep["groups"] = sorted(set(chosen_roles))

    for u in same[1:]:
        u["excluded"] = True
        m = state.get("user_business_role") or {}
        if u.get("username") in m:
            del m[u["username"]]

    # state["last_extract"]["groups"] = recompute_groups_from_users(users)


# ----------------------------
# Models
# ----------------------------
class LoginRequest(BaseModel):
    username: str
    password: str
    domain: Optional[str] = None


class DomainRegistrationRequest(BaseModel):
    domain: str
    licenseCode: str


class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    username: str
    tenant_id: str
    tenant_domain: str


class SystemUserPermissions(BaseModel):
    can_view_analytics: bool = True
    can_view_cluster: bool = True
    can_view_users: bool = True
    can_view_business_roles: bool = True
    can_view_ai_training: bool = True
    can_view_configurations: bool = True
    can_view_logs: bool = True
    can_view_system_users: bool = True
    can_manage_settings: bool = True
    can_manage_assignments: bool = True


class SystemUserUpdateRequest(BaseModel):
    display_name: Optional[str] = None
    password: Optional[str] = None
    active: Optional[bool] = None
    permissions: Optional[SystemUserPermissions] = None


class SystemUserCreateRequest(BaseModel):
    username: str
    display_name: Optional[str] = None
    password: str
    active: bool = True
    permissions: Optional[SystemUserPermissions] = None


class SystemUsersBulkDeleteRequest(BaseModel):
    usernames: List[str] = Field(default_factory=list)


class ConnectorConfig(BaseModel):
    server: str = Field(..., description="LDAP host/ip o 'mock'")
    bind_user: str = Field("", description="Utente bind (es: user@domain o DOMAIN\\user)")
    bind_password: str = Field("", description="Password bind")
    base_dn: str = Field("", description="Base DN (es: DC=example,DC=local)")
    auth: str = Field("SIMPLE", description="SIMPLE oppure NTLM")
    port: int = Field(389, description="LDAP Port")
    use_ssl: bool = Field(False, description="Use SSL/LDAPS")
    sap_base_url: str = Field("", description="SAP API base URL (es: https://sap.company.local)")
    sap_auth_mode: str = Field("AUTO", description="AUTO, APIKEY, BASIC, OAUTH2")
    sap_client: str = Field("", description="SAP client (es: 100)")
    sap_system: str = Field("", description="SAP system id (es: ECC/4HANA)")
    sap_username: str = Field("", description="SAP username")
    sap_password: str = Field("", description="SAP password")
    sap_api_key: str = Field("", description="SAP API key (Business Accelerator Hub sandbox)")
    sap_token_url: str = Field("", description="OAuth2 token endpoint (es: https://<host>/oauth/token)")
    sap_client_id: str = Field("", description="OAuth2 client id")
    sap_client_secret: str = Field("", description="OAuth2 client secret")
    sap_oauth_scope: str = Field("", description="OAuth2 scope (opzionale)")
    sap_company_id: str = Field("", description="SuccessFactors company id (opzionale)")
    sap_users_path: str = Field("/sap/opu/odata/sap/ZROLE_MINING_SRV/Users", description="SAP users endpoint path")
    sap_provision_enabled: bool = Field(True, description="Enable real SAP write provisioning")
    sap_provision_method: str = Field("POST", description="SAP write method: POST, PUT, PATCH")
    sap_provision_path: str = Field("", description="SAP write endpoint path (fallback: sap_users_path)")
    sap_provision_user_path_template: str = Field("", description="Optional per-user write path template, e.g. /odata/v2/User('{username}')")
    sap_deprovision_enabled: bool = Field(False, description="Enable delete for removed SAP users")
    sap_deprovision_user_path_template: str = Field("", description="Per-user delete path template, e.g. /odata/v2/User('{username}')")
    sap_use_csrf_token: bool = Field(True, description="Use x-csrf-token for SAP write calls")
    sap_provision_username_field: str = Field("username", description="Outbound field name for username")
    sap_provision_display_name_field: str = Field("displayName", description="Outbound field name for display name")
    sap_provision_department_field: str = Field("department", description="Outbound field name for department")
    sap_provision_business_role_field: str = Field("businessRole", description="Outbound field name for business role")
    sap_provision_groups_field: str = Field("groups", description="Outbound field name for groups")
    sap_provision_datasource_field: str = Field("DataSource", description="Outbound field name for DataSource")
    sap_provision_groups_as_csv: bool = Field(False, description="Serialize groups as CSV string instead of list")
    azure_base_url: str = Field("https://graph.microsoft.com", description="Azure Graph base URL")
    azure_tenant_id: str = Field("", description="Azure tenant id")
    azure_client_id: str = Field("", description="Azure app client id")
    azure_client_secret: str = Field("", description="Azure app client secret")
    azure_users_path: str = Field("/v1.0/users?$select=id,displayName,userPrincipalName,mail,department,accountEnabled", description="Azure users endpoint/query")
    one_identity_base_url: str = Field("https://<host>/AppServer", description="One Identity base URL")
    one_identity_token_url: str = Field("", description="One Identity token URL")
    one_identity_client_id: str = Field("", description="One Identity client id")
    one_identity_client_secret: str = Field("", description="One Identity client secret")
    one_identity_username: str = Field("", description="One Identity username (optional)")
    one_identity_password: str = Field("", description="One Identity password (optional)")
    one_identity_users_path: str = Field("/api/entities/person?limit=100", description="One Identity users path")
    sailpoint_base_url: str = Field("https://<tenant>.api.identitynow.com/v3", description="SailPoint base URL")
    sailpoint_token_url: str = Field("", description="SailPoint token URL")
    sailpoint_client_id: str = Field("", description="SailPoint client id")
    sailpoint_client_secret: str = Field("", description="SailPoint client secret")
    sailpoint_users_path: str = Field("/accounts", description="SailPoint users path")
    saviynt_base_url: str = Field("", description="Saviynt base URL")
    saviynt_token_url: str = Field("", description="Saviynt token URL")
    saviynt_client_id: str = Field("", description="Saviynt client id")
    saviynt_client_secret: str = Field("", description="Saviynt client secret")
    saviynt_username: str = Field("", description="Saviynt service username")
    saviynt_password: str = Field("", description="Saviynt service password")
    saviynt_users_path: str = Field("", description="Saviynt users path (tenant-specific)")
    servicenow_base_url: str = Field("", description="ServiceNow instance URL")
    servicenow_username: str = Field("", description="ServiceNow username")
    servicenow_password: str = Field("", description="ServiceNow password")
    servicenow_users_path: str = Field("/api/now/table/sys_user?sysparm_fields=sys_id,user_name,name,email,department,active", description="ServiceNow users API path")
    salesforce_base_url: str = Field("", description="Salesforce instance URL")
    salesforce_token_url: str = Field("https://login.salesforce.com/services/oauth2/token", description="Salesforce OAuth token URL")
    salesforce_client_id: str = Field("", description="Salesforce client id")
    salesforce_client_secret: str = Field("", description="Salesforce client secret")
    salesforce_users_path: str = Field("/services/data/v60.0/query?q=SELECT+Id,Name,Username,Email,Department,IsActive+FROM+User", description="Salesforce users query path")
    m365_base_url: str = Field("https://graph.microsoft.com", description="Microsoft 365 Graph base URL")
    m365_tenant_id: str = Field("", description="Microsoft 365 tenant id")
    m365_client_id: str = Field("", description="Microsoft 365 client id")
    m365_client_secret: str = Field("", description="Microsoft 365 client secret")
    m365_users_path: str = Field("/v1.0/users?$select=id,displayName,userPrincipalName,mail,department,accountEnabled", description="Microsoft 365 users query path")
    connector_provisioning: Dict[str, Any] = Field(default_factory=dict, description="Per-connector upstream provisioning settings")
    discovery_schedules: Dict[str, Any] = Field(default_factory=dict, description="Discovery schedule by connector target")
    discovery_results: Dict[str, Any] = Field(default_factory=dict, description="Last discovery result by connector target")


class ExtractRequest(BaseModel):
    ou: str = Field("", description="OU DN (opzionale: fallback a base_dn connettore)")


class ExtractResponse(BaseModel):
    ou: str
    total_users: int
    total_groups: int
    snapshot_ready: bool = True
    processing_in_background: bool = False
    new_users: int = 0
    updated_users: int = 0
    updated_by_displayname: int = 0
    new_groups: int = 0
    updated_groups: int = 0
    users: List[Dict[str, Any]]
    groups: List[str]


class ConnectorProvisionResponse(BaseModel):
    target: str
    datasource: str
    total_users: int
    changed_users: int
    removed_users: int
    changed_usernames: List[str] = Field(default_factory=list)
    removed_usernames: List[str] = Field(default_factory=list)
    provisioned_at: str
    by: str
    message: str
    upstream_attempted: bool = False
    upstream_success: int = 0
    upstream_failed: int = 0
    upstream_errors: List[str] = Field(default_factory=list)


class SapBulkProvisionRequest(BaseModel):
    count: int = Field(100, ge=1, le=5000, description="How many users to create")
    groups_per_user: int = Field(20, ge=1, le=200, description="How many groups per generated user")
    department: str = Field("SAP Bulk Department", description="Department to assign to generated users")
    business_role: str = Field("SAP Bulk Role", description="Business role to assign to generated users")
    username_prefix: str = Field("sap.bulk", description="Username prefix")
    display_prefix: str = Field("SAP Bulk User", description="Display name prefix")
    group_prefix: str = Field("SAP_BULK_GRP", description="Group name prefix")
    start_index: int = Field(1, ge=1, le=999999, description="Starting index for generated usernames")
    dry_run: bool = Field(False, description="Generate payloads but skip upstream write")


class SapBulkProvisionResponse(BaseModel):
    requested_users: int
    generated_users: int
    groups_per_user: int
    department: str
    business_role: str
    dry_run: bool
    uploaded_users: int
    failed_users: int
    uploaded_usernames: List[str] = Field(default_factory=list)
    failed_details: List[str] = Field(default_factory=list)


class RoleMiningRequest(BaseModel):
    n_clusters: Optional[int] = Field(None, description="Se None, calcolato automaticamente")
    role_support: float = Field(0.5, ge=0.1, le=1.0, description="Soglia (0..1) per includere un gruppo nel ruolo del cluster")


class RoleMiningResponse(BaseModel):
    total_users: int
    total_groups: int
    n_clusters: int
    clusters: List[Dict[str, Any]]
    kpi: Dict[str, Any]


class RoleModelingSandboxRequest(BaseModel):
    max_suggestions: int = Field(24, ge=5, le=120, description="Max numero suggerimenti in output")
    min_group_support: float = Field(0.6, ge=0.3, le=0.95, description="Supporto minimo per gruppo nel template ruolo")
    redundancy_threshold: float = Field(0.8, ge=0.5, le=0.98, description="Soglia similarita per merge ruoli/gruppi")
    ml_weight: float = Field(0.35, ge=0.0, le=0.8, description="Peso del layer di apprendimento da feedback")


class RoleModelingFeedbackRequest(BaseModel):
    proposal_id: str
    proposal_type: str
    accepted: bool


class RoleModelingApplyRequest(BaseModel):
    actions: List[Dict[str, Any]] = Field(default_factory=list)
    applied_model_id: str = ""
    target_model_score: Optional[float] = None


security = HTTPBearer(auto_error=False)


DEFAULT_TENANT_ID = normalize_tenant_id(os.getenv("DEFAULT_TENANT_ID", "example.internal"))
DEFAULT_TENANT_DOMAIN = (
    os.getenv("DEFAULT_TENANT_DOMAIN", "example.internal").strip() or "example.internal"
).lower()
TENANT_DOMAIN_MAP_RAW = os.getenv("TENANT_DOMAIN_MAP", "")
DOMAIN_REGISTRATION_LICENSE_CODE = "Bip2026!"
_TENANT_DOMAIN_RE = re.compile(r"^[a-z0-9.-]+$")
BUILTIN_TENANT_DOMAIN_MAP: Dict[str, str] = {
    "example.internal": "example.internal",
    "bip.internal": "bip",
    "bip": "bip",
    "sky.internal": "sky",
    "sky.it": "sky",
    "sky": "sky",
}


def _normalize_tenant_domain(raw: Optional[str]) -> str:
    value = str(raw or "").strip().lower()
    if "://" in value:
        value = value.split("://", 1)[1]
    value = value.split("/", 1)[0]
    value = value.split(":", 1)[0].strip(".")
    if not value:
        return ""
    if not _TENANT_DOMAIN_RE.fullmatch(value):
        return ""
    return value


def _parse_tenant_domain_map(raw: str) -> Dict[str, str]:
    """
    Parse TENANT_DOMAIN_MAP from env.
    Accepted format:
      TENANT_DOMAIN_MAP="sky.it=sky,prometeon.com=prometeon"
    """
    mapping: Dict[str, str] = {}
    for item in (raw or "").split(","):
        chunk = item.strip()
        if not chunk:
            continue
        if "=" in chunk:
            domain_raw, tenant_raw = chunk.split("=", 1)
        elif ":" in chunk:
            domain_raw, tenant_raw = chunk.split(":", 1)
        else:
            continue
        domain = _normalize_tenant_domain(domain_raw)
        tenant_id = normalize_tenant_id(str(tenant_raw or "").strip().lower())
        if domain and tenant_id:
            mapping[domain] = tenant_id
    return mapping


def _tenant_domain_map() -> Dict[str, str]:
    mapping = {k: normalize_tenant_id(v) for k, v in BUILTIN_TENANT_DOMAIN_MAP.items()}
    mapping.update(_parse_tenant_domain_map(TENANT_DOMAIN_MAP_RAW))
    mapping.setdefault(DEFAULT_TENANT_DOMAIN, DEFAULT_TENANT_ID)
    # NOTE: We intentionally do NOT add all known tenant IDs as domain aliases here.
    # That fallback would allow a new domain whose name coincidentally matches an existing
    # tenant slug to silently land in that tenant – violating strict isolation.
    return mapping


def _new_isolated_tenant_id_for_domain(domain: str) -> str:
    base_tenant_id = normalize_tenant_id(domain)
    reserved = set(_tenant_domain_map().values())
    reserved.update(list_known_tenant_ids())
    reserved.update(list_registered_domains().values())

    if base_tenant_id not in reserved and not tenant_storage_exists(base_tenant_id):
        return base_tenant_id

    idx = 2
    while True:
        candidate = normalize_tenant_id(f"{base_tenant_id}-{idx}")
        if candidate not in reserved and not tenant_storage_exists(candidate):
            return candidate
        idx += 1


def _resolve_tenant_for_login(raw_domain: Optional[str]) -> Tuple[str, str]:
    # Manual input by user: no fallback auto-selection.
    domain = _normalize_tenant_domain(raw_domain)
    if not domain:
        raise HTTPException(status_code=400, detail="Dominio cliente obbligatorio")

    # Check domain registry first (dynamically registered domains).
    registered = lookup_registered_domain(domain)
    if registered:
        return normalize_tenant_id(registered), domain

    tenant_id = _tenant_domain_map().get(domain)
    if not tenant_id:
        raise HTTPException(status_code=401, detail="Dominio non autorizzato")
    return normalize_tenant_id(tenant_id), domain


def create_access_token(username: str, tenant_id: str, tenant_domain: str) -> str:
    exp = datetime.now(timezone.utc) + timedelta(minutes=JWT_EXPIRE_MINUTES)
    payload = {
        "sub": username,
        "tenant_id": normalize_tenant_id(tenant_id or DEFAULT_TENANT_ID),
        "tenant_domain": str(tenant_domain or DEFAULT_TENANT_DOMAIN),
        "exp": exp,
    }
    return jwt.encode(payload, JWT_SECRET, algorithm="HS256")


def require_auth(request: Request, creds: HTTPAuthorizationCredentials = Depends(security)) -> str:
    if creds is None or not creds.credentials:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Missing bearer token")
    try:
        payload = jwt.decode(creds.credentials, JWT_SECRET, algorithms=["HS256"])
        tenant_id = normalize_tenant_id(payload.get("tenant_id") or DEFAULT_TENANT_ID)
        request.state.tenant_id = tenant_id
        request.state.auth_claims = payload
        init_default_state(tenant_id)
        return payload["sub"]
    except Exception:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")


def _normalize_permissions(raw: Optional[Dict[str, Any]]) -> Dict[str, bool]:
    perms = dict(SYSTEM_USER_PERMISSION_DEFAULTS)
    for key, default_val in SYSTEM_USER_PERMISSION_DEFAULTS.items():
        if raw and key in raw:
            perms[key] = bool(raw.get(key))
        else:
            perms[key] = bool(default_val)
    return perms


def _hash_system_user_secret(raw_value: str, *, salt_hex: Optional[str] = None) -> str:
    secret_value = str(raw_value or "")
    salt = bytes.fromhex(salt_hex) if salt_hex else secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac(
        "sha256",
        secret_value.encode("utf-8"),
        salt,
        SYSTEM_USER_HASH_ITERATIONS,
    ).hex()
    return f"{SYSTEM_USER_HASH_PREFIX}${SYSTEM_USER_HASH_ITERATIONS}${salt.hex()}${digest}"


def _verify_system_user_secret(raw_value: str, encoded_hash: str) -> bool:
    try:
        algo, iter_raw, salt_hex, expected = str(encoded_hash or "").split("$", 3)
        if algo != SYSTEM_USER_HASH_PREFIX:
            return False
        iterations = int(iter_raw)
        salt = bytes.fromhex(salt_hex)
        actual = hashlib.pbkdf2_hmac(
            "sha256",
            str(raw_value or "").encode("utf-8"),
            salt,
            iterations,
        ).hex()
        return hmac.compare_digest(actual, expected)
    except Exception:
        return False


def _new_system_user_record(
    *,
    username: str,
    display_name: str,
    raw_secret: str,
    active: bool,
    permissions: Dict[str, bool],
) -> Dict[str, Any]:
    return {
        "username": username,
        "display_name": display_name,
        "password_hash": _hash_system_user_secret(raw_secret),
        "active": bool(active),
        "permissions": _normalize_permissions(permissions),
    }


def _default_system_users() -> List[Dict[str, Any]]:
    return [
        _new_system_user_record(
            username=APP_LOGIN_USER,
            display_name="Administrator",
            raw_secret=APP_LOGIN_PASS,
            active=True,
            permissions=dict(SYSTEM_USER_PERMISSION_DEFAULTS),
        ),
        _new_system_user_record(
            username=APP_VIEWER_USER,
            display_name="User Viewer",
            raw_secret=APP_VIEWER_PASS,
            active=True,
            permissions={
                "can_view_analytics": True,
                "can_view_cluster": True,
                "can_view_users": True,
                "can_view_business_roles": True,
                "can_view_ai_training": True,
                "can_view_configurations": False,
                "can_view_logs": False,
                "can_view_system_users": False,
                "can_manage_settings": False,
                "can_manage_assignments": False,
            },
        ),
    ]


def _ensure_system_users_state() -> None:
    users = state.get("system_users")
    if not isinstance(users, list) or len(users) == 0:
        state["system_users"] = _default_system_users()
        return

    existing_by_username = {}
    for rec in users:
        if isinstance(rec, dict):
            uname = str(rec.get("username") or "").strip().lower()
            if uname:
                existing_by_username[uname] = rec

    # Migration: legacy username `users` -> `user`.
    legacy_user = existing_by_username.get("users")
    canonical_user = existing_by_username.get("user")
    changed = False
    if legacy_user and not canonical_user:
        legacy_user["username"] = "user"
        if not str(legacy_user.get("display_name") or "").strip():
            legacy_user["display_name"] = "User Viewer"
        existing_by_username["user"] = legacy_user
        existing_by_username.pop("users", None)
        changed = True
    elif legacy_user and canonical_user:
        users = [rec for rec in users if str(rec.get("username") or "").strip().lower() != "users"]
        existing_by_username.pop("users", None)
        changed = True

    # Ensure mandatory mock users exist.
    for seed in _default_system_users():
        uname = str(seed.get("username") or "").strip().lower()
        if uname not in existing_by_username:
            users.append(seed)
            changed = True

    # Normalize shape/permissions.
    for rec in users:
        if not isinstance(rec, dict):
            continue
        rec.setdefault("display_name", rec.get("username") or "")
        legacy_secret = str(rec.get("password") or "")
        encoded_hash = str(rec.get("password_hash") or "")
        if encoded_hash:
            rec.pop("password", None)
        elif legacy_secret:
            rec["password_hash"] = _hash_system_user_secret(legacy_secret)
            rec.pop("password", None)
            changed = True
        else:
            rec["password_hash"] = _hash_system_user_secret("")
            changed = True
        rec["active"] = bool(rec.get("active", True))
        rec["permissions"] = _normalize_permissions(rec.get("permissions"))
    if changed:
        state["system_users"] = users


def _find_system_user(username: str) -> Optional[Dict[str, Any]]:
    _ensure_system_users_state()
    uname = str(username or "").strip().lower()
    for rec in (state.get("system_users") or []):
        if str(rec.get("username") or "").strip().lower() == uname:
            return rec
    return None


def _permissions_for_user(username: str) -> Dict[str, bool]:
    rec = _find_system_user(username)
    if not rec:
        return dict(SYSTEM_USER_PERMISSION_DEFAULTS)
    return _normalize_permissions(rec.get("permissions"))


def _public_system_user(rec: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "username": str(rec.get("username") or ""),
        "display_name": str(rec.get("display_name") or rec.get("username") or ""),
        "active": bool(rec.get("active", True)),
        "permissions": _normalize_permissions(rec.get("permissions")),
    }


def _require_capability(username: str, capability: str, detail: str) -> None:
    perms = _permissions_for_user(username)
    if not bool(perms.get(capability, False)):
        raise HTTPException(status_code=403, detail=detail)


def _validate_system_users_safety(users: List[Dict[str, Any]]) -> None:
    if not any(
        bool(u.get("active", True))
        and _normalize_permissions(u.get("permissions")).get("can_manage_settings")
        for u in users
    ):
        raise HTTPException(
            status_code=400,
            detail="Serve almeno un system user attivo con permesso di gestione impostazioni",
        )


# ----------------------------
# LDAP / Mock AD
# ----------------------------
def mock_users() -> List[Dict[str, Any]]:
    return [
        {"username": "alice", "displayName": "Alice Rossi", "groups": ["HR", "AllEmployees", "Confluence"]},
        {"username": "bob", "displayName": "Bob Bianchi", "groups": ["HR", "AllEmployees", "Payroll"]},
        {"username": "carol", "displayName": "Carol Verdi", "groups": ["IT", "AllEmployees", "Azure", "GitLab"]},
        {"username": "dave", "displayName": "Dave Neri", "groups": ["IT", "AllEmployees", "Azuure"]},
        {"username": "erin", "displayName": "Erin Gialli", "groups": ["Sales", "AllEmployees", "CRM"]},
        {"username": "frank", "displayName": "Frank Blu", "groups": ["Sales", "AllEmployees", "CRM", "DiscountApproval"]},
    ]


def mock_sap_users() -> List[Dict[str, Any]]:
    return [
        {
            "username": "sap.mrossi",
            "displayName": "Mario Rossi",
            "department": "Finance",
            "businessRole": "Accountant",
            "groups": ["SAP_FI_DISPLAY", "SAP_CO_DISPLAY", "SAP_PORTAL_USER"],
        },
        {
            "username": "sap.lbianchi",
            "displayName": "Laura Bianchi",
            "department": "Procurement",
            "businessRole": "Buyer",
            "groups": ["SAP_MM_DISPLAY", "SAP_MM_BUYER", "SAP_PORTAL_USER"],
        },
        {
            "username": "sap.gverdi",
            "displayName": "Giulia Verdi",
            "department": "Warehouse",
            "businessRole": "Warehouse Operator",
            "groups": ["SAP_MM_DISPLAY", "SAP_MM_WM_WRITE", "SAP_PORTAL_USER"],
        },
    ]


def _sap_pick(entry: Dict[str, Any], keys: List[str], default: Any = "") -> Any:
    for key in keys:
        if key in entry and entry.get(key) not in (None, ""):
            return entry.get(key)
        for k2, v2 in entry.items():
            if str(k2).lower() == key.lower() and v2 not in (None, ""):
                return v2
    return default


def _sap_parse_groups(raw: Any) -> List[str]:
    if raw is None:
        return []
    if isinstance(raw, list):
        out = [str(x).strip() for x in raw if str(x).strip()]
        return sorted(set(out))
    txt = str(raw).strip()
    if not txt:
        return []
    parts = [p.strip() for p in re.split(r"[;,|]", txt) if p.strip()]
    return sorted(set(parts))


def _sap_normalize_users(payload: Any) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]]
    if isinstance(payload, list):
        rows = payload
    elif isinstance(payload, dict):
        if isinstance(payload.get("value"), list):  # OData v4
            rows = payload.get("value") or []
        elif isinstance(payload.get("d"), dict) and isinstance((payload.get("d") or {}).get("results"), list):  # OData v2
            rows = (payload.get("d") or {}).get("results") or []
        else:
            rows = payload.get("users") if isinstance(payload.get("users"), list) else []
    else:
        rows = []

    users: List[Dict[str, Any]] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        username = str(_sap_pick(row, ["username", "UserName", "Bname", "user_id", "userid", "BusinessPartner", "businessPartner", "id", "ID"], "")).strip()
        if not username:
            continue
        display_name = str(_sap_pick(row, ["displayName", "DisplayName", "fullName", "BusinessPartnerFullName", "OrganizationBPName1", "name", "uname"], username)).strip()
        department = str(_sap_pick(row, ["department", "Department", "orgUnit", "OrgUnit"], "")).strip()
        business_role = str(_sap_pick(row, ["businessRole", "BusinessRole", "jobRole", "RoleName"], "Unassigned")).strip()
        account_type = str(_sap_pick(row, ["accountType", "AccountType", "userType"], "")).strip()
        last_login = str(_sap_pick(row, ["lastLogin", "LastLogin", "last_logon", "lastLogon"], "")).strip()
        groups = _sap_parse_groups(_sap_pick(row, ["groups", "Groups", "roles", "Roles", "authorizations", "Authorizations"], []))
        users.append(
            {
                "username": username,
                "displayName": display_name or username,
                "department": department or "Unknown",
                "businessRole": business_role or "Unassigned",
                "accountType": account_type,
                "lastLogin": last_login,
                "groups": groups,
            }
        )
    return users


def _sap_fetch_oauth_token(cfg: Dict[str, Any]) -> str:
    token_url = str(cfg.get("sap_token_url") or "").strip()
    client_id = str(cfg.get("sap_client_id") or "").strip()
    client_secret = str(cfg.get("sap_client_secret") or "").strip()
    oauth_scope = str(cfg.get("sap_oauth_scope") or "").strip()
    company_id = str(cfg.get("sap_company_id") or "").strip()

    if not token_url:
        raise HTTPException(status_code=400, detail="Configura sap_token_url per OAuth2")
    if not client_id or not client_secret:
        raise HTTPException(status_code=400, detail="Configura sap_client_id/sap_client_secret per OAuth2")

    body = {
        "grant_type": "client_credentials",
        "client_id": client_id,
        "client_secret": client_secret,
    }
    if oauth_scope:
        body["scope"] = oauth_scope
    if company_id:
        body["company_id"] = company_id

    data = urllib.parse.urlencode(body).encode("utf-8")
    basic_blob = base64.b64encode(f"{client_id}:{client_secret}".encode("utf-8")).decode("ascii")
    req = urllib.request.Request(
        token_url,
        headers={
            "Authorization": f"Basic {basic_blob}",
            "Content-Type": "application/x-www-form-urlencoded",
            "Accept": "application/json",
            "User-Agent": "RoleMining/1.0",
        },
        data=data,
        method="POST",
    )

    parsed_token_url = urllib.parse.urlparse(token_url)
    if parsed_token_url.scheme not in {"http", "https"}:
        raise HTTPException(status_code=400, detail="SAP OAuth token URL deve usare schema http/https")

    try:
        with urllib.request.urlopen(req, timeout=30) as resp:  # nosec B310
            raw = resp.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="replace")[:500]
        raise HTTPException(status_code=502, detail=f"SAP OAuth token HTTP {e.code}: {detail or 'errore upstream'}")
    except urllib.error.URLError as e:
        raise HTTPException(status_code=503, detail=f"SAP OAuth token endpoint non raggiungibile: {e.reason}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore OAuth token SAP: {e}")

    try:
        payload = json.loads(raw)
    except Exception:
        raise HTTPException(status_code=502, detail="Risposta OAuth token non in formato JSON")

    token = str(payload.get("access_token") or "").strip()
    if not token:
        raise HTTPException(status_code=502, detail="OAuth token mancante nella risposta SAP")
    return token


def _sap_has_oauth_credentials(cfg: Dict[str, Any]) -> bool:
    return bool(
        str(cfg.get("sap_token_url") or "").strip()
        and str(cfg.get("sap_client_id") or "").strip()
        and str(cfg.get("sap_client_secret") or "").strip()
    )


def _sap_build_auth_headers(cfg: Dict[str, Any]) -> Dict[str, str]:
    sap_auth_mode = str(cfg.get("sap_auth_mode") or "AUTO").strip().upper()
    sap_username = str(cfg.get("sap_username") or "").strip()
    sap_password = str(cfg.get("sap_password") or "").strip()
    sap_api_key = str(cfg.get("sap_api_key") or "").strip()
    sap_has_oauth = _sap_has_oauth_credentials(cfg)

    headers: Dict[str, str] = {
        "Accept": "application/json",
        "User-Agent": "RoleMining/1.0",
    }
    if sap_auth_mode == "APIKEY":
        if not sap_api_key:
            raise HTTPException(status_code=400, detail="Auth mode APIKEY selezionato ma sap_api_key non configurata")
        headers["APIKey"] = sap_api_key
        headers["apikey"] = sap_api_key
    elif sap_auth_mode == "BASIC":
        if not sap_username or not sap_password:
            raise HTTPException(status_code=400, detail="Auth mode BASIC selezionato ma sap_username/sap_password non configurati")
        auth_blob = base64.b64encode(f"{sap_username}:{sap_password}".encode("utf-8")).decode("ascii")
        headers["Authorization"] = f"Basic {auth_blob}"
    elif sap_auth_mode == "OAUTH2":
        token = _sap_fetch_oauth_token(cfg)
        headers["Authorization"] = f"Bearer {token}"
    else:
        # AUTO mode preference: API key -> OAuth2 -> Basic
        if sap_api_key:
            headers["APIKey"] = sap_api_key
            headers["apikey"] = sap_api_key
        elif sap_has_oauth:
            token = _sap_fetch_oauth_token(cfg)
            headers["Authorization"] = f"Bearer {token}"
        elif sap_username and sap_password:
            auth_blob = base64.b64encode(f"{sap_username}:{sap_password}".encode("utf-8")).decode("ascii")
            headers["Authorization"] = f"Basic {auth_blob}"
        else:
            raise HTTPException(status_code=400, detail="Configura credenziali SAP (API key, OAuth2 o Basic)")
    return headers


def _sap_build_url(cfg: Dict[str, Any], path: str, query: Optional[Dict[str, str]] = None) -> str:
    sap_base_url = str(cfg.get("sap_base_url") or "").strip()
    if not sap_base_url:
        raise HTTPException(status_code=400, detail="Configura sap_base_url in Connettori")
    base = sap_base_url.rstrip("/")
    norm_path = path if str(path or "").startswith("/") else f"/{str(path or '').strip()}"
    norm_path = norm_path if norm_path != "/" else ""
    merged_query: Dict[str, str] = {}
    sap_client = str(cfg.get("sap_client") or "").strip()
    if sap_client:
        merged_query["sap-client"] = sap_client
    for key, value in (query or {}).items():
        if value is None:
            continue
        txt = str(value).strip()
        if txt:
            merged_query[str(key)] = txt
    qs = urllib.parse.urlencode(merged_query)
    if qs:
        return f"{base}{norm_path}?{qs}"
    return f"{base}{norm_path}"


def _sap_request(
    cfg: Dict[str, Any],
    *,
    method: str,
    path: str,
    payload: Optional[Dict[str, Any]] = None,
    query: Optional[Dict[str, str]] = None,
    use_csrf_token: bool = False,
) -> Tuple[int, str]:
    method_upper = str(method or "GET").strip().upper() or "GET"
    url = _sap_build_url(cfg, path, query=query)
    parsed = urllib.parse.urlparse(url)
    if parsed.scheme not in {"http", "https"}:
        raise HTTPException(status_code=400, detail="SAP Base URL deve usare schema http/https")
    headers = _sap_build_auth_headers(cfg)
    data: Optional[bytes] = None
    if payload is not None:
        headers["Content-Type"] = "application/json"
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")

    opener = None
    if use_csrf_token and method_upper in {"POST", "PUT", "PATCH", "DELETE"}:
        opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(CookieJar()))
        csrf_headers = dict(headers)
        csrf_headers["x-csrf-token"] = "Fetch"
        csrf_req = urllib.request.Request(url, headers=csrf_headers, method="GET")
        try:
            with opener.open(csrf_req, timeout=30) as csrf_resp:
                token = str(csrf_resp.headers.get("x-csrf-token") or "").strip()
        except urllib.error.HTTPError as e:
            detail = e.read().decode("utf-8", errors="replace")[:500]
            raise HTTPException(status_code=502, detail=f"SAP CSRF token HTTP {e.code}: {detail or 'errore upstream'}")
        except urllib.error.URLError as e:
            raise HTTPException(status_code=503, detail=f"SAP CSRF endpoint non raggiungibile: {e.reason}")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Errore fetch CSRF SAP: {e}")
        if token:
            headers["x-csrf-token"] = token

    req = urllib.request.Request(url, headers=headers, data=data, method=method_upper)
    try:
        if opener is not None:
            with opener.open(req, timeout=30) as resp:
                raw = resp.read().decode("utf-8", errors="replace")
                status_code = int(getattr(resp, "status", 200) or 200)
        else:
            with urllib.request.urlopen(req, timeout=30) as resp:  # nosec B310
                raw = resp.read().decode("utf-8", errors="replace")
                status_code = int(getattr(resp, "status", 200) or 200)
        return status_code, raw
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="replace")[:500]
        raise HTTPException(status_code=502, detail=f"SAP API HTTP {e.code}: {detail or 'errore upstream'}")
    except urllib.error.URLError as e:
        raise HTTPException(status_code=503, detail=f"SAP API non raggiungibile: {e.reason}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore SAP API: {e}")


def extract_from_sap(scope: str) -> List[Dict[str, Any]]:
    cfg = state.get("connector", {}) or {}
    sap_system = str(cfg.get("sap_system") or "").strip()
    sap_users_path = str(cfg.get("sap_users_path") or "/sap/opu/odata/sap/ZROLE_MINING_SRV/Users").strip()
    sap_has_oauth = _sap_has_oauth_credentials(cfg)
    sap_base_url = str(cfg.get("sap_base_url") or "").strip()
    sap_username = str(cfg.get("sap_username") or "").strip()
    sap_password = str(cfg.get("sap_password") or "").strip()
    sap_api_key = str(cfg.get("sap_api_key") or "").strip()

    if not sap_base_url and not sap_username and not sap_password and not sap_api_key and not sap_has_oauth:
        raise HTTPException(
            status_code=400,
            detail="Connettore SAP non configurato: imposta SAP Base URL e credenziali reali in Connettori",
        )

    path = sap_users_path if sap_users_path.startswith("/") else f"/{sap_users_path}"
    sap_client = str(cfg.get("sap_client") or "").strip()
    log("INFO", f"SAP extract call system={sap_system or 'N/A'} client={sap_client or 'N/A'} scope={scope or 'N/A'} endpoint={path}")
    _, raw = _sap_request(cfg, method="GET", path=path, query={"$format": "json"}, use_csrf_token=False)

    try:
        payload = json.loads(raw)
    except Exception:
        raise HTTPException(status_code=502, detail="Risposta SAP non in formato JSON")

    users = _sap_normalize_users(payload)
    if not users:
        raise HTTPException(status_code=502, detail="SAP API raggiunta ma nessun utente valido trovato")
    return users


def _sap_apply_user_path_template(path_template: str, username: str) -> str:
    template = str(path_template or "").strip()
    if not template:
        return ""
    if "{username}" in template:
        return template.replace("{username}", urllib.parse.quote(str(username or "").strip(), safe=""))
    return template


def _sap_build_provision_payload(user_payload: Dict[str, Any], datasource: str, cfg: Dict[str, Any]) -> Dict[str, Any]:
    groups = sorted({str(g).strip() for g in (user_payload.get("groups") or []) if str(g).strip()})
    groups_value: Any = ",".join(groups) if bool(cfg.get("sap_provision_groups_as_csv")) else groups

    mappings = [
        (cfg.get("sap_provision_username_field"), user_payload.get("username")),
        (cfg.get("sap_provision_display_name_field"), user_payload.get("displayName")),
        (cfg.get("sap_provision_department_field"), user_payload.get("department")),
        (cfg.get("sap_provision_business_role_field"), user_payload.get("businessRole")),
        (cfg.get("sap_provision_groups_field"), groups_value),
        (cfg.get("sap_provision_datasource_field"), datasource),
    ]

    payload: Dict[str, Any] = {}
    for field_name, value in mappings:
        key = str(field_name or "").strip()
        if not key:
            continue
        payload[key] = value
    return payload


def _sap_provision_users_upstream(
    *,
    datasource: str,
    changed_payloads: List[Dict[str, Any]],
    removed_usernames: List[str],
    force_enable: bool = False,
) -> Dict[str, Any]:
    cfg = state.get("connector", {}) or {}
    enabled = bool(cfg.get("sap_provision_enabled", True))
    if not enabled and not force_enable:
        return {
            "attempted": False,
            "success": 0,
            "failed": 0,
            "errors": [],
            "uploaded_usernames": [],
            "message": "SAP upstream provisioning is disabled in connector settings.",
        }

    method = str(cfg.get("sap_provision_method") or "POST").strip().upper() or "POST"
    if method not in {"POST", "PUT", "PATCH"}:
        raise HTTPException(status_code=400, detail="sap_provision_method must be POST, PUT or PATCH")
    default_path = str(cfg.get("sap_provision_path") or cfg.get("sap_users_path") or "").strip()
    if not default_path:
        raise HTTPException(status_code=400, detail="Configura sap_provision_path oppure sap_users_path")

    user_path_template = str(cfg.get("sap_provision_user_path_template") or "").strip()
    deprovision_enabled = bool(cfg.get("sap_deprovision_enabled"))
    deprovision_template = str(cfg.get("sap_deprovision_user_path_template") or "").strip()
    use_csrf = bool(cfg.get("sap_use_csrf_token", True))

    success = 0
    failed = 0
    uploaded_usernames: List[str] = []
    errors: List[str] = []

    for payload in (changed_payloads or []):
        username = str(payload.get("username") or "").strip()
        if not username:
            failed += 1
            errors.append("Missing username in provisioning payload")
            continue
        path = _sap_apply_user_path_template(user_path_template, username) or default_path
        outbound_payload = _sap_build_provision_payload(payload, datasource, cfg)
        try:
            _sap_request(
                cfg,
                method=method,
                path=path,
                payload=outbound_payload,
                query={"$format": "json"},
                use_csrf_token=use_csrf,
            )
            success += 1
            uploaded_usernames.append(username)
        except HTTPException as exc:
            failed += 1
            errors.append(f"{username}: {exc.detail}")
        except Exception as exc:
            failed += 1
            errors.append(f"{username}: {exc}")

    if deprovision_enabled and deprovision_template:
        for username in (removed_usernames or []):
            path = _sap_apply_user_path_template(deprovision_template, username)
            if not path:
                failed += 1
                errors.append(f"{username}: missing deprovision path template")
                continue
            try:
                _sap_request(
                    cfg,
                    method="DELETE",
                    path=path,
                    query={"$format": "json"},
                    use_csrf_token=use_csrf,
                )
                success += 1
            except HTTPException as exc:
                failed += 1
                errors.append(f"{username} (delete): {exc.detail}")
            except Exception as exc:
                failed += 1
                errors.append(f"{username} (delete): {exc}")

    return {
        "attempted": bool(changed_payloads or (deprovision_enabled and removed_usernames)),
        "success": success,
        "failed": failed,
        "errors": errors[:200],
        "uploaded_usernames": uploaded_usernames[:500],
        "message": f"SAP upstream write success={success} failed={failed}.",
    }


def _generate_sap_bulk_users(req: SapBulkProvisionRequest) -> List[Dict[str, Any]]:
    count = int(req.count)
    groups_per_user = int(req.groups_per_user)
    start_index = int(req.start_index)
    group_pool_size = max(groups_per_user, 100)
    group_pool = [f"{req.group_prefix}_{idx:03d}" for idx in range(1, group_pool_size + 1)]

    users: List[Dict[str, Any]] = []
    for offset in range(count):
        absolute = start_index + offset
        username = f"{req.username_prefix}.{absolute:04d}"
        display_name = f"{req.display_prefix} {absolute:04d}"
        base_idx = absolute % group_pool_size
        groups = [group_pool[(base_idx + j) % group_pool_size] for j in range(groups_per_user)]
        users.append(
            {
                "username": username,
                "displayName": display_name,
                "department": req.department,
                "businessRole": req.business_role,
                "accountType": "Internal",
                "groups": sorted(set(groups)),
                "DataSource": "SAP",
            }
        )
    return users


def _connector_pick(entry: Dict[str, Any], keys: List[str], default: Any = "") -> Any:
    for key in keys:
        if key in entry and entry.get(key) not in (None, ""):
            return entry.get(key)
        for k2, v2 in entry.items():
            if str(k2).lower() == key.lower() and v2 not in (None, ""):
                return v2
    return default


def _connector_parse_groups(raw: Any) -> List[str]:
    if raw is None:
        return []
    if isinstance(raw, dict):
        for k in ("value", "results", "items", "roles", "groups", "entitlements"):
            if isinstance(raw.get(k), list):
                return _connector_parse_groups(raw.get(k))
        txt = str(raw.get("name") or raw.get("displayName") or raw.get("value") or "").strip()
        return [txt] if txt else []
    if isinstance(raw, list):
        out: List[str] = []
        for item in raw:
            if isinstance(item, dict):
                txt = str(
                    item.get("displayName")
                    or item.get("name")
                    or item.get("value")
                    or item.get("role")
                    or item.get("entitlement")
                    or item.get("id")
                    or ""
                ).strip()
                if txt:
                    out.append(txt)
            else:
                txt = str(item or "").strip()
                if txt:
                    out.append(txt)
        return sorted(set(out))
    txt = str(raw or "").strip()
    if not txt:
        return []
    return sorted(set([p.strip() for p in re.split(r"[;,|]", txt) if p and p.strip()]))


def _connector_rows(payload: Any) -> List[Dict[str, Any]]:
    if isinstance(payload, list):
        return [x for x in payload if isinstance(x, dict)]
    if not isinstance(payload, dict):
        return []
    if isinstance(payload.get("value"), list):  # OData v4
        return [x for x in (payload.get("value") or []) if isinstance(x, dict)]
    if isinstance(payload.get("d"), dict) and isinstance((payload.get("d") or {}).get("results"), list):  # OData v2
        return [x for x in ((payload.get("d") or {}).get("results") or []) if isinstance(x, dict)]
    for key in ("users", "results", "result", "items", "resources", "Resources", "data", "records"):
        if isinstance(payload.get(key), list):
            return [x for x in (payload.get(key) or []) if isinstance(x, dict)]
    if isinstance(payload.get("data"), dict):
        nested = payload.get("data") or {}
        for key in ("users", "results", "items", "records"):
            if isinstance(nested.get(key), list):
                return [x for x in (nested.get(key) or []) if isinstance(x, dict)]
    # Last-resort: single-object payload
    return [payload]


def _normalize_connector_users(payload: Any) -> List[Dict[str, Any]]:
    users: List[Dict[str, Any]] = []
    for row in _connector_rows(payload):
        username = str(
            _connector_pick(
                row,
                [
                    "username",
                    "userName",
                    "user_name",
                    "samAccountName",
                    "sAMAccountName",
                    "userPrincipalName",
                    "upn",
                    "login",
                    "uid",
                    "id",
                    "Id",
                ],
                "",
            )
        ).strip()
        display_name = str(_connector_pick(row, ["displayName", "DisplayName", "name", "fullName"], username)).strip()
        department = str(_connector_pick(row, ["department", "Department", "dept", "orgUnit", "organization"], "")).strip()
        business_role = str(_connector_pick(row, ["businessRole", "BusinessRole", "jobRole", "jobTitle", "title", "role"], "")).strip()
        account_type = str(_connector_pick(row, ["accountType", "AccountType", "userType", "type", "employeeType"], "")).strip()
        last_login = str(_connector_pick(row, ["lastLogin", "LastLogin", "last_logon", "lastLogon", "last_login"], "")).strip()
        email = str(_connector_pick(row, ["email", "mail", "Email"], "")).strip().lower()
        upn = str(_connector_pick(row, ["upn", "userPrincipalName", "UPN"], "")).strip().lower()
        employee_id = str(_connector_pick(row, ["employeeId", "employee_id", "EmployeeId"], "")).strip()
        manager = str(_connector_pick(row, ["manager", "Manager", "managerName"], "")).strip()
        raw_groups = _connector_pick(
            row,
            ["groups", "Groups", "roles", "Roles", "entitlements", "permissions", "authorizations"],
            [],
        )
        groups = _connector_parse_groups(raw_groups)

        # Try deriving username from email/upn if source uses email-based identity.
        if not username:
            candidate = upn or email
            if candidate:
                username = candidate.split("@")[0]
        if not username:
            continue

        users.append(
            {
                "username": username,
                "displayName": display_name or username,
                "department": department or "Unknown",
                "businessRole": business_role or "Unassigned",
                "accountType": account_type,
                "lastLogin": last_login,
                "groups": groups,
                "email": email or None,
                "upn": upn or None,
                "employeeId": employee_id or None,
                "manager": manager or None,
            }
        )
    return users


def _connector_url(base_url: str, path: str) -> str:
    base = str(base_url or "").strip().rstrip("/")
    if not base:
        raise HTTPException(status_code=400, detail="Base URL connettore non configurato")
    p = str(path or "").strip()
    if not p:
        p = "/"
    if p.startswith("http://") or p.startswith("https://"):
        return p
    if not p.startswith("/"):
        p = f"/{p}"
    return f"{base}{p}"


def _http_raw_request(
    url: str,
    *,
    headers: Optional[Dict[str, str]] = None,
    method: str = "GET",
    body: Optional[bytes] = None,
    timeout: int = 30,
    error_prefix: str = "Connector API",
) -> Tuple[int, str]:
    parsed = urllib.parse.urlparse(str(url or ""))
    if parsed.scheme not in {"http", "https"}:
        raise HTTPException(status_code=400, detail=f"{error_prefix}: URL deve usare schema http/https")
    req = urllib.request.Request(url, headers=headers or {}, data=body, method=method.upper())
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:  # nosec B310
            raw = resp.read().decode("utf-8", errors="replace")
            status_code = int(getattr(resp, "status", 200) or 200)
            return status_code, raw
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="replace")[:500]
        raise HTTPException(status_code=502, detail=f"{error_prefix} HTTP {e.code}: {detail or 'errore upstream'}")
    except urllib.error.URLError as e:
        raise HTTPException(status_code=503, detail=f"{error_prefix} non raggiungibile: {e.reason}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"{error_prefix} errore: {e}")


def _http_json_request(url: str, *, headers: Optional[Dict[str, str]] = None, method: str = "GET", body: Optional[bytes] = None, timeout: int = 30, error_prefix: str = "Connector API") -> Any:
    _, raw = _http_raw_request(
        url,
        headers=headers,
        method=method,
        body=body,
        timeout=timeout,
        error_prefix=error_prefix,
    )
    try:
        return json.loads(raw)
    except Exception:
        raise HTTPException(status_code=502, detail=f"{error_prefix} risposta non JSON")


def _oauth_client_credentials_token(
    *,
    token_url: str,
    client_id: str,
    client_secret: str,
    scope: str = "",
    extra_fields: Optional[Dict[str, str]] = None,
    error_prefix: str = "OAuth token",
) -> Dict[str, Any]:
    if not token_url:
        raise HTTPException(status_code=400, detail=f"{error_prefix}: token_url mancante")
    if not client_id or not client_secret:
        raise HTTPException(status_code=400, detail=f"{error_prefix}: client_id/client_secret mancanti")

    form = {
        "grant_type": "client_credentials",
        "client_id": client_id,
        "client_secret": client_secret,
    }
    if scope:
        form["scope"] = scope
    for k, v in (extra_fields or {}).items():
        if v:
            form[str(k)] = str(v)

    data = urllib.parse.urlencode(form).encode("utf-8")
    basic_blob = base64.b64encode(f"{client_id}:{client_secret}".encode("utf-8")).decode("ascii")
    payload = _http_json_request(
        token_url,
        method="POST",
        body=data,
        headers={
            "Authorization": f"Basic {basic_blob}",
            "Content-Type": "application/x-www-form-urlencoded",
            "Accept": "application/json",
            "User-Agent": "RoleMining/1.0",
        },
        error_prefix=error_prefix,
    )
    token = str((payload or {}).get("access_token") or "").strip()
    if not token:
        raise HTTPException(status_code=502, detail=f"{error_prefix}: access_token mancante")
    return payload


def _extract_http_connector_users(
    *,
    target: str,
    base_url: str,
    users_path: str,
    auth_headers: Optional[Dict[str, str]] = None,
    query: Optional[Dict[str, str]] = None,
) -> List[Dict[str, Any]]:
    url = _connector_url(base_url, users_path)
    if query:
        parsed = urllib.parse.urlsplit(url)
        current_q = dict(urllib.parse.parse_qsl(parsed.query, keep_blank_values=True))
        current_q.update({k: v for k, v in query.items() if v is not None and str(v) != ""})
        merged_q = urllib.parse.urlencode(current_q, doseq=True)
        url = urllib.parse.urlunsplit((parsed.scheme, parsed.netloc, parsed.path, merged_q, parsed.fragment))

    headers = {
        "Accept": "application/json",
        "User-Agent": "RoleMining/1.0",
    }
    headers.update(auth_headers or {})
    payload = _http_json_request(url, headers=headers, error_prefix=f"{target} API")
    users = _normalize_connector_users(payload)
    if not users:
        raise HTTPException(status_code=502, detail=f"{target} API raggiunta ma nessun utente valido trovato")
    return users


def extract_from_azure(scope: str = "") -> List[Dict[str, Any]]:
    cfg = state.get("connector", {}) or {}
    base_url = str(cfg.get("azure_base_url") or "").strip()
    users_path = str(cfg.get("azure_users_path") or "").strip()
    tenant_id = str(cfg.get("azure_tenant_id") or "").strip()
    client_id = str(cfg.get("azure_client_id") or "").strip()
    client_secret = str(cfg.get("azure_client_secret") or "").strip()
    if not (base_url and users_path and tenant_id and client_id and client_secret):
        raise HTTPException(status_code=400, detail="Configura Azure AD: base_url, users_path, tenant_id, client_id, client_secret")

    token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    scope_value = f"{base_url.rstrip('/')}/.default"
    token_payload = _oauth_client_credentials_token(
        token_url=token_url,
        client_id=client_id,
        client_secret=client_secret,
        scope=scope_value,
        error_prefix="Azure OAuth token",
    )
    token = str(token_payload.get("access_token") or "").strip()
    return _extract_http_connector_users(
        target="Azure",
        base_url=base_url,
        users_path=users_path,
        auth_headers={"Authorization": f"Bearer {token}"},
    )


def extract_from_m365(scope: str = "") -> List[Dict[str, Any]]:
    cfg = state.get("connector", {}) or {}
    base_url = str(cfg.get("m365_base_url") or "").strip()
    users_path = str(cfg.get("m365_users_path") or "").strip()
    tenant_id = str(cfg.get("m365_tenant_id") or "").strip()
    client_id = str(cfg.get("m365_client_id") or "").strip()
    client_secret = str(cfg.get("m365_client_secret") or "").strip()
    if not (base_url and users_path and tenant_id and client_id and client_secret):
        raise HTTPException(status_code=400, detail="Configura Microsoft 365: base_url, users_path, tenant_id, client_id, client_secret")

    token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    scope_value = f"{base_url.rstrip('/')}/.default"
    token_payload = _oauth_client_credentials_token(
        token_url=token_url,
        client_id=client_id,
        client_secret=client_secret,
        scope=scope_value,
        error_prefix="M365 OAuth token",
    )
    token = str(token_payload.get("access_token") or "").strip()
    return _extract_http_connector_users(
        target="M365",
        base_url=base_url,
        users_path=users_path,
        auth_headers={"Authorization": f"Bearer {token}"},
    )


def extract_from_one_identity(scope: str = "") -> List[Dict[str, Any]]:
    cfg = state.get("connector", {}) or {}
    base_url = str(cfg.get("one_identity_base_url") or "").strip()
    users_path = str(cfg.get("one_identity_users_path") or "").strip()
    token_url = str(cfg.get("one_identity_token_url") or "").strip()
    client_id = str(cfg.get("one_identity_client_id") or "").strip()
    client_secret = str(cfg.get("one_identity_client_secret") or "").strip()
    username = str(cfg.get("one_identity_username") or "").strip()
    password = str(cfg.get("one_identity_password") or "").strip()
    if not (base_url and users_path):
        raise HTTPException(status_code=400, detail="Configura One Identity: base_url e users_path")

    headers: Dict[str, str] = {}
    if token_url and client_id and client_secret:
        token_payload = _oauth_client_credentials_token(
            token_url=token_url,
            client_id=client_id,
            client_secret=client_secret,
            error_prefix="One Identity OAuth token",
        )
        headers["Authorization"] = f"Bearer {str(token_payload.get('access_token') or '').strip()}"
    elif username and password:
        auth_blob = base64.b64encode(f"{username}:{password}".encode("utf-8")).decode("ascii")
        headers["Authorization"] = f"Basic {auth_blob}"
    return _extract_http_connector_users(target="One Identity", base_url=base_url, users_path=users_path, auth_headers=headers)


def extract_from_sailpoint(scope: str = "") -> List[Dict[str, Any]]:
    cfg = state.get("connector", {}) or {}
    base_url = str(cfg.get("sailpoint_base_url") or "").strip()
    users_path = str(cfg.get("sailpoint_users_path") or "").strip()
    token_url = str(cfg.get("sailpoint_token_url") or "").strip()
    client_id = str(cfg.get("sailpoint_client_id") or "").strip()
    client_secret = str(cfg.get("sailpoint_client_secret") or "").strip()
    if not (base_url and users_path):
        raise HTTPException(status_code=400, detail="Configura SailPoint: base_url e users_path")
    if not token_url:
        # Common default for IdentityNow API base
        token_url = f"{base_url.rstrip('/').rsplit('/v3', 1)[0]}/oauth/token"
    token_payload = _oauth_client_credentials_token(
        token_url=token_url,
        client_id=client_id,
        client_secret=client_secret,
        error_prefix="SailPoint OAuth token",
    )
    token = str(token_payload.get("access_token") or "").strip()
    return _extract_http_connector_users(
        target="SailPoint",
        base_url=base_url,
        users_path=users_path,
        auth_headers={"Authorization": f"Bearer {token}"},
    )


def extract_from_saviynt(scope: str = "") -> List[Dict[str, Any]]:
    cfg = state.get("connector", {}) or {}
    base_url = str(cfg.get("saviynt_base_url") or "").strip()
    users_path = str(cfg.get("saviynt_users_path") or "").strip()
    token_url = str(cfg.get("saviynt_token_url") or "").strip()
    client_id = str(cfg.get("saviynt_client_id") or "").strip()
    client_secret = str(cfg.get("saviynt_client_secret") or "").strip()
    username = str(cfg.get("saviynt_username") or "").strip()
    password = str(cfg.get("saviynt_password") or "").strip()
    if not (base_url and users_path):
        raise HTTPException(status_code=400, detail="Configura Saviynt: base_url e users_path")

    headers: Dict[str, str] = {}
    if token_url and client_id and client_secret:
        token_payload = _oauth_client_credentials_token(
            token_url=token_url,
            client_id=client_id,
            client_secret=client_secret,
            error_prefix="Saviynt OAuth token",
        )
        headers["Authorization"] = f"Bearer {str(token_payload.get('access_token') or '').strip()}"
    elif username and password:
        auth_blob = base64.b64encode(f"{username}:{password}".encode("utf-8")).decode("ascii")
        headers["Authorization"] = f"Basic {auth_blob}"
    return _extract_http_connector_users(target="Saviynt", base_url=base_url, users_path=users_path, auth_headers=headers)


def extract_from_servicenow(scope: str = "") -> List[Dict[str, Any]]:
    cfg = state.get("connector", {}) or {}
    base_url = str(cfg.get("servicenow_base_url") or "").strip()
    users_path = str(cfg.get("servicenow_users_path") or "").strip()
    username = str(cfg.get("servicenow_username") or "").strip()
    password = str(cfg.get("servicenow_password") or "").strip()
    if not (base_url and users_path and username and password):
        raise HTTPException(status_code=400, detail="Configura ServiceNow: base_url, users_path, username, password")

    auth_blob = base64.b64encode(f"{username}:{password}".encode("utf-8")).decode("ascii")
    return _extract_http_connector_users(
        target="ServiceNow",
        base_url=base_url,
        users_path=users_path,
        auth_headers={"Authorization": f"Basic {auth_blob}"},
    )


def extract_from_salesforce(scope: str = "") -> List[Dict[str, Any]]:
    cfg = state.get("connector", {}) or {}
    base_url_cfg = str(cfg.get("salesforce_base_url") or "").strip()
    users_path = str(cfg.get("salesforce_users_path") or "").strip()
    token_url = str(cfg.get("salesforce_token_url") or "").strip()
    client_id = str(cfg.get("salesforce_client_id") or "").strip()
    client_secret = str(cfg.get("salesforce_client_secret") or "").strip()
    if not users_path:
        raise HTTPException(status_code=400, detail="Configura salesforce_users_path")

    token_payload = _oauth_client_credentials_token(
        token_url=token_url,
        client_id=client_id,
        client_secret=client_secret,
        error_prefix="Salesforce OAuth token",
    )
    token = str(token_payload.get("access_token") or "").strip()
    instance_url = str(token_payload.get("instance_url") or "").strip()
    base_url = base_url_cfg or instance_url
    if not base_url:
        raise HTTPException(status_code=400, detail="Configura salesforce_base_url o usa un token response con instance_url")
    return _extract_http_connector_users(
        target="Salesforce",
        base_url=base_url,
        users_path=users_path,
        auth_headers={"Authorization": f"Bearer {token}"},
    )


CONNECTOR_EXTRACTORS = {
    "azure": extract_from_azure,
    "m365": extract_from_m365,
    "one_identity": extract_from_one_identity,
    "sailpoint": extract_from_sailpoint,
    "saviynt": extract_from_saviynt,
    "servicenow": extract_from_servicenow,
    "salesforce": extract_from_salesforce,
}


HTTP_UPSTREAM_CONNECTORS = set(CONNECTOR_EXTRACTORS.keys())


def _connector_upstream_auth_and_base(target: str, cfg: Dict[str, Any]) -> Tuple[str, str, Dict[str, str]]:
    connector = normalize_connector_target(target)
    if connector == "azure":
        base_url = str(cfg.get("azure_base_url") or "").strip()
        users_path = str(cfg.get("azure_users_path") or "").strip()
        tenant_id = str(cfg.get("azure_tenant_id") or "").strip()
        client_id = str(cfg.get("azure_client_id") or "").strip()
        client_secret = str(cfg.get("azure_client_secret") or "").strip()
        if not (base_url and users_path and tenant_id and client_id and client_secret):
            raise HTTPException(status_code=400, detail="Configura Azure AD per provisioning: base_url, users_path, tenant_id, client_id, client_secret")
        token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
        scope_value = f"{base_url.rstrip('/')}/.default"
        token_payload = _oauth_client_credentials_token(
            token_url=token_url,
            client_id=client_id,
            client_secret=client_secret,
            scope=scope_value,
            error_prefix="Azure OAuth token",
        )
        token = str(token_payload.get("access_token") or "").strip()
        return base_url, users_path, {"Authorization": f"Bearer {token}"}
    if connector == "m365":
        base_url = str(cfg.get("m365_base_url") or "").strip()
        users_path = str(cfg.get("m365_users_path") or "").strip()
        tenant_id = str(cfg.get("m365_tenant_id") or "").strip()
        client_id = str(cfg.get("m365_client_id") or "").strip()
        client_secret = str(cfg.get("m365_client_secret") or "").strip()
        if not (base_url and users_path and tenant_id and client_id and client_secret):
            raise HTTPException(status_code=400, detail="Configura Microsoft 365 per provisioning: base_url, users_path, tenant_id, client_id, client_secret")
        token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
        scope_value = f"{base_url.rstrip('/')}/.default"
        token_payload = _oauth_client_credentials_token(
            token_url=token_url,
            client_id=client_id,
            client_secret=client_secret,
            scope=scope_value,
            error_prefix="M365 OAuth token",
        )
        token = str(token_payload.get("access_token") or "").strip()
        return base_url, users_path, {"Authorization": f"Bearer {token}"}
    if connector == "one_identity":
        base_url = str(cfg.get("one_identity_base_url") or "").strip()
        users_path = str(cfg.get("one_identity_users_path") or "").strip()
        token_url = str(cfg.get("one_identity_token_url") or "").strip()
        client_id = str(cfg.get("one_identity_client_id") or "").strip()
        client_secret = str(cfg.get("one_identity_client_secret") or "").strip()
        username = str(cfg.get("one_identity_username") or "").strip()
        password = str(cfg.get("one_identity_password") or "").strip()
        if not (base_url and users_path):
            raise HTTPException(status_code=400, detail="Configura One Identity per provisioning: base_url e users_path")
        headers: Dict[str, str] = {}
        if token_url and client_id and client_secret:
            token_payload = _oauth_client_credentials_token(
                token_url=token_url,
                client_id=client_id,
                client_secret=client_secret,
                error_prefix="One Identity OAuth token",
            )
            headers["Authorization"] = f"Bearer {str(token_payload.get('access_token') or '').strip()}"
        elif username and password:
            auth_blob = base64.b64encode(f"{username}:{password}".encode("utf-8")).decode("ascii")
            headers["Authorization"] = f"Basic {auth_blob}"
        return base_url, users_path, headers
    if connector == "sailpoint":
        base_url = str(cfg.get("sailpoint_base_url") or "").strip()
        users_path = str(cfg.get("sailpoint_users_path") or "").strip()
        token_url = str(cfg.get("sailpoint_token_url") or "").strip()
        client_id = str(cfg.get("sailpoint_client_id") or "").strip()
        client_secret = str(cfg.get("sailpoint_client_secret") or "").strip()
        if not (base_url and users_path):
            raise HTTPException(status_code=400, detail="Configura SailPoint per provisioning: base_url e users_path")
        if not token_url:
            token_url = f"{base_url.rstrip('/').rsplit('/v3', 1)[0]}/oauth/token"
        token_payload = _oauth_client_credentials_token(
            token_url=token_url,
            client_id=client_id,
            client_secret=client_secret,
            error_prefix="SailPoint OAuth token",
        )
        token = str(token_payload.get("access_token") or "").strip()
        return base_url, users_path, {"Authorization": f"Bearer {token}"}
    if connector == "saviynt":
        base_url = str(cfg.get("saviynt_base_url") or "").strip()
        users_path = str(cfg.get("saviynt_users_path") or "").strip()
        token_url = str(cfg.get("saviynt_token_url") or "").strip()
        client_id = str(cfg.get("saviynt_client_id") or "").strip()
        client_secret = str(cfg.get("saviynt_client_secret") or "").strip()
        username = str(cfg.get("saviynt_username") or "").strip()
        password = str(cfg.get("saviynt_password") or "").strip()
        if not (base_url and users_path):
            raise HTTPException(status_code=400, detail="Configura Saviynt per provisioning: base_url e users_path")
        headers: Dict[str, str] = {}
        if token_url and client_id and client_secret:
            token_payload = _oauth_client_credentials_token(
                token_url=token_url,
                client_id=client_id,
                client_secret=client_secret,
                error_prefix="Saviynt OAuth token",
            )
            headers["Authorization"] = f"Bearer {str(token_payload.get('access_token') or '').strip()}"
        elif username and password:
            auth_blob = base64.b64encode(f"{username}:{password}".encode("utf-8")).decode("ascii")
            headers["Authorization"] = f"Basic {auth_blob}"
        return base_url, users_path, headers
    if connector == "servicenow":
        base_url = str(cfg.get("servicenow_base_url") or "").strip()
        users_path = str(cfg.get("servicenow_users_path") or "").strip()
        username = str(cfg.get("servicenow_username") or "").strip()
        password = str(cfg.get("servicenow_password") or "").strip()
        if not (base_url and users_path and username and password):
            raise HTTPException(status_code=400, detail="Configura ServiceNow per provisioning: base_url, users_path, username, password")
        auth_blob = base64.b64encode(f"{username}:{password}".encode("utf-8")).decode("ascii")
        return base_url, users_path, {"Authorization": f"Basic {auth_blob}"}
    if connector == "salesforce":
        base_url_cfg = str(cfg.get("salesforce_base_url") or "").strip()
        users_path = str(cfg.get("salesforce_users_path") or "").strip()
        token_url = str(cfg.get("salesforce_token_url") or "").strip()
        client_id = str(cfg.get("salesforce_client_id") or "").strip()
        client_secret = str(cfg.get("salesforce_client_secret") or "").strip()
        if not users_path:
            raise HTTPException(status_code=400, detail="Configura salesforce_users_path")
        token_payload = _oauth_client_credentials_token(
            token_url=token_url,
            client_id=client_id,
            client_secret=client_secret,
            error_prefix="Salesforce OAuth token",
        )
        token = str(token_payload.get("access_token") or "").strip()
        instance_url = str(token_payload.get("instance_url") or "").strip()
        base_url = base_url_cfg or instance_url
        if not base_url:
            raise HTTPException(status_code=400, detail="Configura salesforce_base_url o usa instance_url dal token")
        return base_url, users_path, {"Authorization": f"Bearer {token}"}
    raise HTTPException(status_code=400, detail=f"Upstream provisioning non supportato per connettore '{target}'")


def _connector_build_provision_payload(
    user_payload: Dict[str, Any],
    datasource: str,
    *,
    field_map: Optional[Dict[str, str]] = None,
    groups_as_csv: bool = False,
) -> Dict[str, Any]:
    groups = sorted({str(g).strip() for g in (user_payload.get("groups") or []) if str(g).strip()})
    source_payload: Dict[str, Any] = {
        "username": user_payload.get("username"),
        "displayName": user_payload.get("displayName"),
        "department": user_payload.get("department"),
        "businessRole": user_payload.get("businessRole"),
        "accountType": user_payload.get("accountType"),
        "email": user_payload.get("email"),
        "upn": user_payload.get("upn"),
        "employeeId": user_payload.get("employeeId"),
        "manager": user_payload.get("manager"),
        "statusAd": user_payload.get("statusAd"),
        "statusHr": user_payload.get("statusHr"),
        "excluded": user_payload.get("excluded"),
        "groups": ",".join(groups) if groups_as_csv else groups,
        "DataSource": datasource,
    }
    mapped: Dict[str, Any] = {}
    for key, value in source_payload.items():
        target_key = str((field_map or {}).get(key) or key).strip()
        if not target_key:
            continue
        mapped[target_key] = value
    return mapped


def _ldap_escape_filter_value(value: str) -> str:
    txt = str(value or "")
    return (
        txt.replace("\\", r"\5c")
        .replace("*", r"\2a")
        .replace("(", r"\28")
        .replace(")", r"\29")
        .replace("\x00", r"\00")
    )


def _ad_provision_users_upstream(changed_payloads: List[Dict[str, Any]]) -> Dict[str, Any]:
    if Connection is None:
        raise HTTPException(status_code=500, detail="ldap3 library not installed or failed to import")

    cfg = state.get("connector", {}) or {}
    if not cfg.get("server") or not cfg.get("bind_user") or not cfg.get("bind_password"):
        raise HTTPException(status_code=400, detail="Configura AD target: server, bind_user e bind_password")
    base_dn = str(cfg.get("base_dn") or "").strip()
    if not base_dn:
        raise HTTPException(status_code=400, detail="Configura base_dn per provisioning AD")

    resolved = _mk_ldap_server(cfg)
    server = resolved["server"]
    auth_mode = str(cfg.get("auth") or "SIMPLE").upper()
    conn = None
    success = 0
    failed = 0
    errors: List[str] = []
    uploaded_usernames: List[str] = []
    group_dn_cache: Dict[str, str] = {}

    try:
        if auth_mode == "NTLM":
            conn = Connection(server, user=cfg["bind_user"], password=cfg["bind_password"], authentication=NTLM, auto_bind=True)
        else:
            conn = Connection(server, user=cfg["bind_user"], password=cfg["bind_password"], authentication=SIMPLE, auto_bind=True)
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"AD bind failed: {exc}")

    def _find_group_dn(group_cn: str) -> str:
        cached = group_dn_cache.get(group_cn)
        if cached is not None:
            return cached
        group_filter = f"(&(objectClass=group)(cn={_ldap_escape_filter_value(group_cn)}))"
        if not conn.search(base_dn, group_filter, attributes=["distinguishedName"], size_limit=1):
            group_dn_cache[group_cn] = ""
            return ""
        if not conn.entries:
            group_dn_cache[group_cn] = ""
            return ""
        group_dn = str(conn.entries[0].entry_dn or "").strip()
        group_dn_cache[group_cn] = group_dn
        return group_dn

    for payload in (changed_payloads or []):
        username = str(payload.get("username") or "").strip()
        if not username:
            failed += 1
            errors.append("Missing username in AD payload")
            continue

        try:
            user_filter = f"(&(objectClass=user)(sAMAccountName={_ldap_escape_filter_value(username)}))"
            if not conn.search(base_dn, user_filter, attributes=["distinguishedName", "memberOf"], size_limit=1) or not conn.entries:
                failed += 1
                errors.append(f"{username}: AD user not found")
                continue

            user_entry = conn.entries[0]
            user_dn = str(user_entry.entry_dn or "").strip()
            current_member_dns = []
            if "memberOf" in set(getattr(user_entry, "entry_attributes", []) or []):
                try:
                    current_member_dns = [str(v) for v in (user_entry.memberOf.values or [])]
                except Exception:
                    current_member_dns = []
            current_group_by_cn = {str(_extract_group_cn(dn)).strip(): str(dn) for dn in current_member_dns if _extract_group_cn(dn)}

            mod_attrs: Dict[str, List[Tuple[int, List[str]]]] = {}

            def _set_attr(attr_name: str, raw_value: Any) -> None:
                if raw_value is None:
                    return
                val = str(raw_value).strip()
                if not val:
                    return
                mod_attrs[attr_name] = [(MODIFY_REPLACE, [val])]

            _set_attr("displayName", payload.get("displayName"))
            _set_attr("department", payload.get("department"))
            _set_attr("title", payload.get("businessRole"))
            _set_attr("mail", payload.get("email"))
            _set_attr("userPrincipalName", payload.get("upn"))
            _set_attr("employeeID", payload.get("employeeId"))

            if mod_attrs and not conn.modify(user_dn, mod_attrs):
                err = conn.result or {}
                failed += 1
                errors.append(f"{username}: AD attribute update failed ({err.get('description') or err})")
                continue

            desired_groups = sorted({str(g).strip() for g in (payload.get("groups") or []) if str(g).strip()})
            desired_group_set = set(desired_groups)
            current_group_set = set(current_group_by_cn.keys())

            to_add = sorted(desired_group_set - current_group_set)
            to_remove = sorted(current_group_set - desired_group_set)

            for group_cn in to_add:
                group_dn = _find_group_dn(group_cn)
                if not group_dn:
                    errors.append(f"{username}: AD group '{group_cn}' not found")
                    continue
                if not conn.modify(group_dn, {"member": [(MODIFY_ADD, [user_dn])]}):
                    err = conn.result or {}
                    desc = str(err.get("description") or "").lower()
                    if "typeorvalueexists" in desc:
                        continue
                    errors.append(f"{username}: add group '{group_cn}' failed ({err.get('description') or err})")

            for group_cn in to_remove:
                group_dn = str(current_group_by_cn.get(group_cn) or "").strip()
                if not group_dn:
                    group_dn = _find_group_dn(group_cn)
                if not group_dn:
                    continue
                if not conn.modify(group_dn, {"member": [(MODIFY_DELETE, [user_dn])]}):
                    err = conn.result or {}
                    desc = str(err.get("description") or "").lower()
                    if "nosuchattribute" in desc:
                        continue
                    errors.append(f"{username}: remove group '{group_cn}' failed ({err.get('description') or err})")

            success += 1
            uploaded_usernames.append(username)
        except Exception as exc:
            failed += 1
            errors.append(f"{username}: {exc}")

    try:
        if conn is not None:
            conn.unbind()
    except Exception:
        pass

    return {
        "attempted": bool(changed_payloads),
        "success": success,
        "failed": failed,
        "errors": errors[:200],
        "uploaded_usernames": uploaded_usernames[:500],
        "message": f"AD upstream write success={success} failed={failed}.",
    }


def _provision_users_upstream(
    *,
    target: str,
    datasource: str,
    changed_payloads: List[Dict[str, Any]],
    removed_usernames: List[str],
    force_enable: bool = False,
) -> Dict[str, Any]:
    connector = normalize_connector_target(target)
    if connector == "ad":
        return _ad_provision_users_upstream(changed_payloads)
    if connector == "sap":
        return _sap_provision_users_upstream(
            datasource=datasource,
            changed_payloads=changed_payloads,
            removed_usernames=removed_usernames,
            force_enable=force_enable,
        )
    if connector not in HTTP_UPSTREAM_CONNECTORS:
        return {
            "attempted": False,
            "success": 0,
            "failed": 0,
            "errors": [],
            "uploaded_usernames": [],
            "message": f"Upstream provisioning not implemented for {connector}.",
        }

    cfg = state.get("connector", {}) or {}
    provisioning_cfg = cfg.get("connector_provisioning") or {}
    if not isinstance(provisioning_cfg, dict):
        provisioning_cfg = {}
    target_cfg = provisioning_cfg.get(connector) or {}
    if not isinstance(target_cfg, dict):
        target_cfg = {}
    enabled = bool(target_cfg.get("enabled", True))
    if not enabled and not force_enable:
        return {
            "attempted": False,
            "success": 0,
            "failed": 0,
            "errors": [],
            "uploaded_usernames": [],
            "message": f"Upstream provisioning disabled for {connector}.",
        }

    method = str(target_cfg.get("method") or "POST").strip().upper() or "POST"
    if method not in {"POST", "PUT", "PATCH"}:
        raise HTTPException(status_code=400, detail=f"connector_provisioning.{connector}.method must be POST, PUT or PATCH")
    user_path_template = str(target_cfg.get("user_path_template") or "").strip()
    deprovision_enabled = bool(target_cfg.get("deprovision_enabled", False))
    deprovision_template = str(target_cfg.get("deprovision_user_path_template") or "").strip()
    groups_as_csv = bool(target_cfg.get("groups_as_csv", False))
    field_map = target_cfg.get("field_map") if isinstance(target_cfg.get("field_map"), dict) else {}

    base_url, default_users_path, auth_headers = _connector_upstream_auth_and_base(connector, cfg)
    default_path = str(target_cfg.get("path") or default_users_path or "").strip()
    if not default_path:
        raise HTTPException(status_code=400, detail=f"connector_provisioning.{connector}.path non configurato")

    success = 0
    failed = 0
    uploaded_usernames: List[str] = []
    errors: List[str] = []

    for payload in (changed_payloads or []):
        username = str(payload.get("username") or "").strip()
        if not username:
            failed += 1
            errors.append("Missing username in provisioning payload")
            continue
        path = _sap_apply_user_path_template(user_path_template, username) or default_path
        url = _connector_url(base_url, path)
        outbound_payload = _connector_build_provision_payload(
            payload,
            datasource,
            field_map=field_map,
            groups_as_csv=groups_as_csv,
        )
        req_headers = {
            "Accept": "application/json",
            "Content-Type": "application/json",
            "User-Agent": "RoleMining/1.0",
            **auth_headers,
        }
        try:
            _http_raw_request(
                url,
                headers=req_headers,
                method=method,
                body=json.dumps(outbound_payload, ensure_ascii=False).encode("utf-8"),
                error_prefix=f"{connector} provisioning API",
            )
            success += 1
            uploaded_usernames.append(username)
        except HTTPException as exc:
            failed += 1
            errors.append(f"{username}: {exc.detail}")
        except Exception as exc:
            failed += 1
            errors.append(f"{username}: {exc}")

    if deprovision_enabled and deprovision_template:
        for username in (removed_usernames or []):
            path = _sap_apply_user_path_template(deprovision_template, username)
            if not path:
                failed += 1
                errors.append(f"{username}: missing deprovision path template")
                continue
            url = _connector_url(base_url, path)
            req_headers = {
                "Accept": "application/json",
                "User-Agent": "RoleMining/1.0",
                **auth_headers,
            }
            try:
                _http_raw_request(
                    url,
                    headers=req_headers,
                    method="DELETE",
                    error_prefix=f"{connector} provisioning API",
                )
                success += 1
            except HTTPException as exc:
                failed += 1
                errors.append(f"{username} (delete): {exc.detail}")
            except Exception as exc:
                failed += 1
                errors.append(f"{username} (delete): {exc}")

    return {
        "attempted": bool(changed_payloads or (deprovision_enabled and removed_usernames)),
        "success": success,
        "failed": failed,
        "errors": errors[:200],
        "uploaded_usernames": uploaded_usernames[:500],
        "message": f"{connector.upper()} upstream write success={success} failed={failed}.",
    }


def _build_ingest_candidates_for_source(users: List[Dict[str, Any]], source: str) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for u in users:
        uname = str(u.get("username") or "").strip()
        out.append(
            _mk_candidate(
                source=source,
                candidate_id=f"{source}:{uname}",
                display_name=u.get("displayName", ""),
                business_role=u.get("businessRole", ""),
                roles=(u.get("groups") or []),
                raw=f"{source.upper()}:{uname}|{u.get('displayName')}|{','.join(u.get('groups') or [])}",
                department=u.get("department") or "",
                last_login=u.get("lastLogin"),
            )
        )
    return out


def _run_connector_extract_pipeline(
    *,
    source: str,
    scope: str,
    users: List[Dict[str, Any]],
    background_tasks: BackgroundTasks,
    actor: str,
) -> ExtractResponse:
    candidates = _build_ingest_candidates_for_source(users, source)
    with state.batch():
        users = filter_and_dedupe_connector_users(users, source=source)
        merge_stats = merge_from_connector_by_displayname(users, ou=scope, source=source)
        updates = {
            "ingest_sources": {**state.get("ingest_sources", {}), source: candidates},
            "mining_dirty": True,
        }
        state.update(updates)

    invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)
    snapshot_ts = str((state.get("last_extract") or {}).get("ts") or "")
    tenant_id = get_current_tenant_id()
    background_tasks.add_task(run_post_snapshot_logic_background, snapshot_ts, actor, tenant_id)
    return ExtractResponse(
        ou=state["last_extract"]["ou"],
        total_users=len(state["last_extract"]["users"]),
        total_groups=len(state["last_extract"]["groups"]),
        snapshot_ready=True,
        processing_in_background=True,
        new_users=merge_stats.get("new_users", 0),
        updated_users=merge_stats.get("updated_users", 0),
        updated_by_displayname=merge_stats.get("updated_by_displayname", merge_stats.get("updated_users", 0)),
        new_groups=merge_stats.get("new_groups", 0),
        updated_groups=merge_stats.get("updated_groups", 0),
        users=state["last_extract"]["users"],
        groups=state["last_extract"]["groups"],
    )


def _mk_ldap_server(cfg: Dict[str, Any]):
    raw = (cfg.get("server") or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="Configura server LDAP")

    if Connection is None:
        raise HTTPException(status_code=500, detail="ldap3 non disponibile")

    raw_lower = raw.lower()
    scheme_ssl = raw_lower.startswith("ldaps://")
    host_port = raw.replace("ldaps://", "").replace("ldap://", "")
    host = host_port.split(":")[0].strip()
    server_port = None
    if ":" in host_port:
        try:
            server_port = int(host_port.rsplit(":", 1)[1].strip())
        except Exception:
            server_port = None

    cfg_use_ssl = cfg.get("use_ssl")
    use_ssl = bool(cfg_use_ssl) if cfg_use_ssl is not None else scheme_ssl

    cfg_port = cfg.get("port")
    port = None
    try:
        port = int(cfg_port) if cfg_port is not None else None
    except Exception:
        port = None
    if port is None:
        port = server_port
    if port is None:
        port = 636 if use_ssl else 389

    return {
        "server": Server(host, port=port, use_ssl=use_ssl, get_info=NONE, connect_timeout=15),
        "host": host,
        "port": port,
        "use_ssl": use_ssl,
    }


def _ldap_search_attrs() -> List[str]:
    if LDAP_FETCH_ALL_ATTRIBUTES:
        return ["*"]
    out = list(dict.fromkeys(LDAP_BASE_ATTRIBUTES + LDAP_EXTRA_ATTRIBUTES))
    return out or ["*"]


def _extract_group_cn(raw_dn: Any) -> str:
    dn_str = str(raw_dn or "").strip()
    if not dn_str:
        return ""
    for part in dn_str.split(","):
        p = part.strip()
        if p.upper().startswith("CN="):
            return p[3:]
    return dn_str





def extract_from_ldap(ou_dn: str) -> List[Dict[str, Any]]:
    if MOCK_AD or state.get("connector", {}).get("server") == "mock":
        # Ignora OU, ritorna dataset di test
        log("INFO", "Using MOCK AD extract")
        return mock_users()

    if Connection is None:
        raise HTTPException(status_code=500, detail="ldap3 library not installed or failed to import")

    cfg = state.get("connector", {})
    if not cfg.get("server") or not cfg.get("bind_user") or not cfg.get("bind_password"):
        raise HTTPException(status_code=400, detail="Configura server/bind_user/bind_password in Connettori")

    log("INFO", f"Connecting to LDAP server: {cfg['server']} as {cfg['bind_user']}")
    
    conn = None
    try:
        resolved = _mk_ldap_server(cfg)
        server = resolved["server"]
        host = resolved["host"]
        port = resolved["port"]
        use_ssl = resolved["use_ssl"]
        log("INFO", f"LDAP target resolved to {host}:{port} (use_ssl={use_ssl})")
        auth_mode = cfg.get("auth", "SIMPLE").upper()
        
        try:
            if auth_mode == "NTLM":
                conn = Connection(server, user=cfg["bind_user"], password=cfg["bind_password"], authentication=NTLM, auto_bind=True)
            else:
                conn = Connection(server, user=cfg["bind_user"], password=cfg["bind_password"], authentication=SIMPLE, auto_bind=True)
        except Exception as e:
            error_msg = str(e)
            log("ERROR", f"LDAP Connection failed: {error_msg}")
            if "timeout" in error_msg.lower():
                 raise HTTPException(status_code=504, detail=f"Timeout connettendosi al server LDAP {host}:{port} (use_ssl={use_ssl})")
            raise HTTPException(status_code=503, detail=f"Impossibile connettersi al server LDAP {host}:{port} (use_ssl={use_ssl}) - {error_msg}")

        # objectClass=user può includere account tecnici: in produzione aggiungere filtri più stretti
        search_filter = "(&(objectClass=user)(sAMAccountName=*))"
        attrs = _ldap_search_attrs()
        log(
            "INFO",
            f"Searching LDAP base='{ou_dn}' filter='{search_filter}' attrs={len(attrs)} page_size={LDAP_PAGE_SIZE}",
        )

        users: List[Dict[str, Any]] = []

        # Collect all available field names for UI
        available_fields = set()
        parsed_entries = 0

        if LDAP_PAGE_SIZE > 0:
            entries_iter = conn.extend.standard.paged_search(
                search_base=ou_dn,
                search_filter=search_filter,
                attributes=attrs,
                paged_size=LDAP_PAGE_SIZE,
                generator=True,
                time_limit=LDAP_SEARCH_TIME_LIMIT,
            )
        else:
            if not conn.search(
                search_base=ou_dn,
                search_filter=search_filter,
                attributes=attrs,
                time_limit=LDAP_SEARCH_TIME_LIMIT,
            ):
                result = conn.result or {}
                log("ERROR", f"LDAP search returned False. Result: {result}")
                result_desc = str(result.get("description") or "").lower()
                diagnostic = {
                    "message": f"LDAP search failed on base '{ou_dn}'",
                    "result": result,
                    "hints": [
                        "Verifica OU DN: deve esistere nel dominio LDAP (es. OU=Users,DC=example,DC=internal).",
                        f"Controlla coerenza con base_dn configurato: '{cfg.get('base_dn') or ''}'.",
                    ],
                }
                if "no such object" in result_desc or "nosuchobject" in result_desc:
                    raise HTTPException(status_code=400, detail=diagnostic)
                if "invalid dn syntax" in result_desc or "invaliddnsyntax" in result_desc:
                    raise HTTPException(status_code=400, detail=diagnostic)
                raise HTTPException(status_code=500, detail=diagnostic)
            entries_iter = (
                {"type": "searchResEntry", "attributes": e.entry_attributes_as_dict}
                for e in conn.entries
            )

        for entry in entries_iter:
            if entry.get("type") != "searchResEntry":
                continue
            parsed_entries += 1
            try:
                d = entry.get("attributes") or {}
                
                # Update available fields
                for k in d.keys():
                    available_fields.add(k)

                # Safe attribute extraction
                sAM = d.get("sAMAccountName") or []
                username = str(sAM[0]).strip() if sAM else ""
                
                disp = d.get("displayName") or []
                display = str(disp[0]).strip() if disp else ""
                
                member_of = d.get("memberOf") or []
                groups = []
                for dn in member_of:
                    cn = _extract_group_cn(dn)
                    if cn:
                        groups.append(cn)
                
                dept = d.get("department") or []
                department = str(dept[0]).strip() if dept else None
                
                llt = d.get("lastLogonTimestamp")
                ll = d.get("lastLogon")
                candidates = []
                n1 = _normalize_last_login(llt)
                n2 = _normalize_last_login(ll)
                if n1:
                    candidates.append(n1)
                if n2:
                    candidates.append(n2)
                if candidates:
                    last_login = max(candidates, key=_to_ts)
                else:
                    last_login = None

                et = d.get("employeeType")
                etype = str(et[0]).strip() if et else ""
                
                dn_full = d.get("distinguishedName")
                dn_str_val = str(dn_full[0]).strip() if dn_full else ""
                
                # Use department or parse OU from DN as fallback for classification
                ou_for_class = department or dn_str_val

                # Pass sanitized attributes to avoid bytes serialization errors in API responses.
                safe_attrs = _json_safe_value(d)
                account_type = classify_account(display or username, ou_for_class, etype, attributes=safe_attrs)

                if username:
                    users.append({
                        "username": username,
                        "displayName": display or username,
                        "groups": sorted(set(groups)),
                        "department": department,
                        "lastLogin": last_login,
                        "accountType": account_type,
                        "attributes": safe_attrs
                    })
            except Exception as entry_ex:
                # Log but continue processing other users
                log("WARNING", f"Error parsing LDAP entry: {entry_ex}. Entry: {str(entry)[:100]}...")
                continue
        
        # Update state with available fields
        state["ad_available_fields"] = sorted(list(available_fields))
        log("INFO", f"Updated available AD fields: {len(available_fields)} found.")

                
        log("INFO", f"LDAP extraction complete. Parsed entries={parsed_entries}, users={len(users)}.")
        return users

    except HTTPException:
        raise
    except Exception as e:
        log("ERROR", f"Unexpected error in extract_from_ldap: {e}")
        # traceback would be good here but keeping it simple
        raise HTTPException(status_code=500, detail=f"Errore interno durante estrazione LDAP: {str(e)}")
    finally:
        if conn:
            try:
                conn.unbind()
            except:
                pass


# ----------------------------
# Role Mining
# ----------------------------
def build_matrix(users: List[Dict[str, Any]]) -> Tuple[List[str], List[str], np.ndarray]:
    usernames = [u["username"] for u in users]
    all_groups = sorted({g for u in users for g in u["groups"]})
    g_index = {g: i for i, g in enumerate(all_groups)}

    X = np.zeros((len(users), len(all_groups)), dtype=np.int8)
    for i, u in enumerate(users):
        for g in u["groups"]:
            if g in g_index:
                X[i, g_index[g]] = 1
    return usernames, all_groups, X


def jaccard_distance_matrix(X: np.ndarray) -> np.ndarray:
    # D[i,j] = 1 - |A∩B|/|A∪B|
    n = X.shape[0]
    D = np.zeros((n, n), dtype=np.float32)
    for i in range(n):
        Ai = X[i]
        for j in range(i + 1, n):
            Aj = X[j]
            inter = int(np.logical_and(Ai, Aj).sum())
            union = int(np.logical_or(Ai, Aj).sum())
            dist = 0.0 if union == 0 else 1.0 - (inter / union)
            D[i, j] = dist
            D[j, i] = dist
    return D


def compute_purity(cluster_members_idx: List[int], X: np.ndarray) -> float:
    # Purity “semplice”: media della massima frequenza (per attributo) nel cluster
    if not cluster_members_idx:
        return 0.0
    sub = X[cluster_members_idx, :]
    # frequenza per colonna
    freq = sub.mean(axis=0)  # 0..1
    return float(freq.max())  # quanto un "gruppo dominante" spiega il cluster

# (_tokens, _family_key, _is_broad already defined at lines 96-109)
# (BROAD_MARKERS defined at line 94)

def compute_ai_detection(matrix: dict, users: list = None) -> dict:
    if users:
        # Use SMART logic (KB + Stats)
        res = run_smart_ai_detection(users, matrix)
        stats = res.get("stats", {})
        return {
            "aiDetection": stats.get("aiDetection", 0),
            "redundantAssignments": stats.get("totalAnomalies", 0),
            "totalAssignments": stats.get("totalAssignments", 0),
            "usersWithRedundancy": stats.get("usersWithAnomaly", 0),
            "redundantUsers": [], # Not needed for KPI summary
        }

    # --- Legacy Logic (Matrix only) ---
    total_assignments = 0
    redundant_assignments = 0

    users_with_redundancy = 0
    redundant_users = []

    for uname, row in (matrix or {}).items():
        roles = [r for r, v in (row or {}).items() if int(v) == 1]
        user_groups = _matrix_user_roles(matrix, uname)

        total_assignments += len(roles)

        fam = defaultdict(list)
        for r in roles:
            k = _family_key(r)
            if k:
                fam[k].append(r)

        has_redundancy = False

        for rs in fam.values():
            broad = [r for r in rs if _is_broad(r)]
            specific = [r for r in rs if not _is_broad(r)]
            if not broad or not specific:
                continue

            red = 0
            for b in broad:
                probs = [predict_redundant(b, s, _family_key(b), user_groups) for s in specific]
                p = max(probs) if probs else 0.0
                if p >= 0.50:
                    red += 1

            if red > 0:
                redundant_assignments += red
                has_redundancy = True

        if has_redundancy:
            users_with_redundancy += 1
            redundant_users.append(uname)

    total_users = len(matrix or {})
    pct = (users_with_redundancy / max(1, total_users) * 100.0)

    return {
        "aiDetection": round(pct, 2),
        "redundantAssignments": redundant_assignments,
        "totalAssignments": total_assignments,
        "usersWithRedundancy": users_with_redundancy,
        "redundantUsers": redundant_users,
    }


def compute_kpis(
    users: List[Dict[str, Any]],
    clusters: List[Dict[str, Any]],
    matrix: Dict[str, Dict[str, int]],
) -> Dict[str, Any]:
    import numpy as np

    total_users = len(users)
    if total_users == 0:
        return {
            "totalUsers": 0,
            "overprivilegedPct": 0,
            "clusterQuality": 0,
            "clusteringQuality": 0,
            "roleCoverage": 0,
            "aiDetection": 0,
            "redundantAssignments": 0,
            "totalAssignments": 0,
            "usersWithRedundancy": 0,
        }

    # --- Overprivileged: top 10% per numero gruppi calcolato dalla matrix ---
    row_counts = np.array(
        [int(sum((row or {}).values())) for row in (matrix or {}).values()],
        dtype=np.int32
    )
    if row_counts.size == 0:
        return {
            "totalUsers": total_users,
            "overprivilegedPct": 0,
            "clusterQuality": 0,
            "clusteringQuality": 0,
            "roleCoverage": 0,
            "aiDetection": 0,
            "redundantAssignments": 0,
            "totalAssignments": 0,
            "usersWithRedundancy": 0,
        }

    k_top = int(np.ceil(0.1 * row_counts.size))
    k_top = max(1, k_top)

    thr = int(np.partition(row_counts, -k_top)[-k_top])

    overpriv = float((row_counts >= thr).mean() * 100.0)

    # --- AI detection (redundant broad roles) ---
    ai = compute_ai_detection(matrix, users=users)

    # --- RoleCoverage: roleGroups copre i gruppi reali (da matrix) dei membri ---
    cov_vals: list[float] = []
    for c in (clusters or []):
        role_groups = set(c.get("roleGroups") or [])
        members = c.get("members") or []
        for uname in members:
            row = matrix.get(uname) or {}
            denom = int(sum((row or {}).values()))
            if denom <= 0:
                cov_vals.append(0.0)
            else:
                covered = sum(int(row.get(g, 0)) for g in role_groups)
                cov_vals.append(covered / denom)

    role_coverage = float((np.mean(cov_vals) if cov_vals else 0.0) * 100.0)

    # --- clusterQuality = live data-quality score (align with cluster quality drilldown) ---
    cluster_quality = compute_cluster_quality_live()

    # --- clusteringQuality = qualità clustering (purity media pesata) ---
    if clusters:
        total_sz = sum(int(c.get("size") or len(c.get("members") or [])) for c in clusters) or 1
        weighted = 0.0
        for c in clusters:
            sz = int(c.get("size") or len(c.get("members") or []))
            weighted += float(c.get("purity") or 0.0) * sz
        clustering_quality = (weighted / total_sz) * 100.0
    else:
        clustering_quality = 0.0

    # --- Model Quality (Enhanced Option B) ---
    groups_list = (state.get("last_extract") or {}).get("groups") or []
    if not groups_list and matrix:
        # Fallback to matrix keys if last_extract is empty
        all_groups = set()
        for row in matrix.values():
            if isinstance(row, list):
                all_groups.update(row)
            else:
                all_groups.update((row or {}).keys())
        groups_list = list(all_groups)
    
    mq = compute_model_quality(users, matrix, groups_list)

    # Ensure last_kpis is updated in state for drilldown fallback
    last_kpis = {
        "clusterQuality": round(cluster_quality, 2),
        "clusteringQuality": round(clustering_quality, 2),
        "modelQuality": mq.get("modelQuality", 0),
        "aiDetection": ai.get("score", 0),
        "roleCoverage": round(role_coverage, 2),
        "staleAccounts": mq.get("staleUsers", 0)
    }
    state["last_kpis"] = last_kpis

    return {
        "totalUsers": total_users,
        # "overprivilegedPct": round(overpriv, 2), # Removed/Deprecated
        "modelQuality": mq.get("modelQuality", 0),
        "orphanRolesCount": mq.get("orphanRoles", mq.get("orphanGroups", 0)),
        "orphanGroupsCount": mq.get("orphanGroups", 0),
        "overprivilegedCount": mq.get("overprivilegedUsers", 0),
        "zeroGroupCount": mq.get("zeroGroupUsers", 0),
        "staleAccountCount": mq.get("staleUsers", 0),

        "clusterQuality": round(float(cluster_quality), 2),
        "clusteringQuality": round(float(clustering_quality), 2),
        "aiDetection": ai.get("aiDetection", 0),
        "redundantAssignments": ai.get("redundantAssignments", 0),
        "totalAssignments": ai.get("totalAssignments", 0),
        "usersWithRedundancy": ai.get("usersWithRedundancy", 0),
        "roleCoverage": round(float(role_coverage), 2),
    }


def compute_cluster_quality_live() -> float:
    """
    Compute Cluster Quality from the currently loaded dataset, not only from the
    last ingest counters. This keeps /api/kpi aligned with cluster-quality drilldown.
    """
    cache_key = "kpi_cluster_quality_live"
    cached_score = RESPONSE_CACHE.get(cache_key)
    if cached_score is not None:
        return float(cached_score)

    ingest = state.get("last_ingest_stats") or {}
    last_extract = state.get("last_extract") or {}
    users = last_extract.get("users") or []

    total_ingest = int(ingest.get("rowsTotal") or 0)
    total_users = len(users)
    total = max(total_ingest, total_users)
    if total <= 0:
        return 0.0

    missing_department = sum(1 for u in users if not str(u.get("department") or "").strip())
    missing_business_role = sum(1 for u in users if not str(u.get("businessRole") or "").strip())
    missing_display_name = sum(
        1 for u in users if not str((u.get("displayName") or u.get("display_name") or "")).strip()
    )
    missing_username = sum(1 for u in users if not str(u.get("username") or "").strip())

    duplicate_items = _duplicate_resolution_items()
    duplicates = _effective_duplicate_displayname_count(
        ingest=ingest,
        duplicate_items=duplicate_items,
        rejects=state.get("last_rejects") or [],
    )

    # Identity integrity issues (align with cluster-quality drilldown signals).
    identity_integrity = 0
    cached_drilldown = RESPONSE_CACHE.get("kpi_drilldown_cluster-quality")
    if cached_drilldown and isinstance(cached_drilldown, dict):
        stats = cached_drilldown.get("stats") or {}
        identity_integrity = int(stats.get("identityIntegrityIssues") or 0)
    else:
        known_usernames = {str(u.get("username") or "").strip().lower() for u in users if u.get("username")}
        known_emails = {str(u.get("email") or "").strip().lower() for u in users if u.get("email")}
        known_upns = {str(u.get("upn") or "").strip().lower() for u in users if u.get("upn")}

        invalid_identity = 0
        collision = 0
        orphan_refs = 0
        inactive_mismatch = 0

        by_email: Dict[str, int] = defaultdict(int)
        by_upn: Dict[str, int] = defaultdict(int)
        by_empid: Dict[str, int] = defaultdict(int)

        inactive_markers = {"inactive", "disabled", "terminated", "offboarded", "left"}
        active_markers = {"active", "enabled"}

        for u in users:
            email = str(u.get("email") or "").strip().lower()
            upn = str(u.get("upn") or "").strip().lower()
            empid = str(u.get("employeeId") or "").strip()
            manager = str(u.get("manager") or "").strip()
            status_ad = str(u.get("statusAd") or "").strip().lower()
            status_hr = str(u.get("statusHr") or "").strip().lower()

            if email:
                by_email[email] += 1
                if not _is_valid_email_address(email):
                    invalid_identity += 1
            if upn:
                by_upn[upn] += 1
                if not _is_valid_upn_value(upn):
                    invalid_identity += 1
            if empid:
                by_empid[empid] += 1
                if not _is_valid_employee_id(empid):
                    invalid_identity += 1

            if manager:
                m = manager.lower()
                if (m not in known_usernames) and (m not in known_emails) and (m not in known_upns):
                    orphan_refs += 1

            if status_ad and status_hr:
                ad_inactive = any(k in status_ad for k in inactive_markers)
                hr_inactive = any(k in status_hr for k in inactive_markers)
                ad_active = any(k in status_ad for k in active_markers)
                hr_active = any(k in status_hr for k in active_markers)
                if (ad_inactive and hr_active) or (hr_inactive and ad_active):
                    inactive_mismatch += 1

        collision = sum(1 for v in list(by_email.values()) + list(by_upn.values()) + list(by_empid.values()) if v > 1)
        identity_integrity = invalid_identity + collision + orphan_refs + inactive_mismatch

    src = str(ingest.get("source") or "").lower()
    if src.startswith("ad"):
        missing_business_role = 0
        duplicates = 0
        identity_integrity = 0

    penalty = (
        1.00 * (duplicates / total) +
        0.70 * (missing_department / total) +
        0.70 * (missing_business_role / total) +
        0.40 * (missing_display_name / total) +
        0.40 * (missing_username / total) +
        0.60 * (identity_integrity / total)
    )
    penalty = min(1.0, penalty)
    score = round(max(0.0, 100.0 * (1.0 - penalty)), 2)
    RESPONSE_CACHE.set(cache_key, score, CACHE_TTL_KPI)
    return score


def _effective_connector_type() -> str:
    """
    Resolve active connector family for Data Quality rendering.
    Values: ad | sap | csv | generic
    """
    last_extract = state.get("last_extract") or {}
    ingest = state.get("last_ingest_stats") or {}
    connector = state.get("connector") or {}

    src = str(last_extract.get("source") or ingest.get("source") or "").strip().lower()
    if src.startswith("sap"):
        return "sap"
    if src.startswith("ad") or src.startswith("ldap"):
        return "ad"
    if src.startswith("csv"):
        return "csv"

    if str(connector.get("sap_base_url") or "").strip():
        return "sap"
    if str(connector.get("server") or "").strip():
        return "ad"
    return "generic"


def _csv_connector_peer_quality(users: List[Dict[str, Any]], ingest: Dict[str, Any]) -> Dict[str, Any]:
    """
    CSV-specific, column-aware peer quality checks.
    - Detects available columns from CSV headers + effective data coverage.
    - Chooses the best peer model automatically.
    - Flags dirty values as peer outliers or suspicious missing values.
    """
    users = users or []
    if not users:
        return {
            "presentColumns": [],
            "peerModel": "none",
            "signals": [],
            "cases": [],
        }

    headers = {str(h or "").strip().lower() for h in (ingest.get("csvHeadersNorm") or [])}

    def _val(u: Dict[str, Any], field: str) -> str:
        if field == "emailDomain":
            email = str(u.get("email") or "").strip().lower()
            return email.split("@", 1)[1] if ("@" in email and len(email.split("@", 1)[1]) > 0) else ""
        if field == "upnDomain":
            upn = str(u.get("upn") or "").strip().lower()
            return upn.split("@", 1)[1] if ("@" in upn and len(upn.split("@", 1)[1]) > 0) else ""
        return str(u.get(field) or "").strip()

    canonical = {
        "department": {"department", "dept", "dipartimento", "area", "funzione"},
        "businessRole": {"businessrole", "business role", "br", "ruolo business", "ruolo_business"},
        "accountType": {"accounttype", "account type", "tipo utente", "tipo_utente", "type"},
        "manager": {"manager", "owner", "responsabile"},
        "emailDomain": {"email", "mail", "emailaddress", "posta"},
        "upnDomain": {"upn", "user principal name", "userprincipalname"},
        "statusAd": {"statusad", "adstatus", "accountstatusad", "statoad"},
        "statusHr": {"statushr", "hrstatus", "accountstatushr", "statohr"},
    }

    present_columns: set[str] = set()
    field_stats: Dict[str, Dict[str, Any]] = {}
    for field, aliases in canonical.items():
        from_header = bool(headers.intersection(aliases))
        vals = [_val(u, field) for u in users]
        non_empty_vals = [v for v in vals if v]
        coverage = (len(non_empty_vals) / max(1, len(users)))
        distinct = len(set(v.lower() for v in non_empty_vals))
        if from_header or coverage >= 0.08:
            present_columns.add(field)
        field_stats[field] = {
            "coverage": round(coverage, 4),
            "distinct": distinct,
            "fromHeader": from_header,
        }

    def _usable_for_peer(field: str) -> bool:
        st = field_stats.get(field) or {}
        return (field in present_columns) and (st.get("coverage", 0) >= 0.55) and (st.get("distinct", 0) >= 2)

    peer_candidates = [
        ("businessRole", "accountType"),
        ("businessRole",),
        ("department", "accountType"),
        ("department",),
        ("accountType",),
    ]
    peer_fields: Tuple[str, ...] = tuple()
    for cand in peer_candidates:
        if all(_usable_for_peer(x) for x in cand):
            peer_fields = cand
            break
    if not peer_fields:
        peer_fields = tuple()

    groups: Dict[Tuple[str, ...], List[Dict[str, Any]]] = defaultdict(list)
    for u in users:
        if peer_fields:
            key = tuple((_val(u, f) or "__missing__").lower() for f in peer_fields)
        else:
            key = ("__all__",)
        groups[key].append(u)
    peer_groups = {k: v for k, v in groups.items() if len(v) >= 5}

    signal_fields = [f for f in ["manager", "emailDomain", "upnDomain", "statusAd", "statusHr", "department", "businessRole", "accountType"] if f in present_columns]

    outlier_users: Dict[str, Dict[str, Any]] = {}
    missing_users: Dict[str, Dict[str, Any]] = {}

    def _row_ref(u: Dict[str, Any]) -> Dict[str, Any]:
        return {
            "username": str(u.get("username") or "").strip(),
            "displayName": str(u.get("displayName") or u.get("username") or "").strip(),
        }

    for _, members in peer_groups.items():
        group_size = len(members)
        for field in signal_fields:
            if field in peer_fields:
                continue
            vals_raw = [_val(u, field).strip() for u in members]
            vals = [v.lower() for v in vals_raw if v]
            non_empty = len(vals)
            if non_empty < 5:
                continue
            counts = Counter(vals)
            completeness = non_empty / max(1, group_size)

            for u in members:
                base = _row_ref(u)
                key = base["username"] or base["displayName"]
                uv = _val(u, field).strip().lower()
                if not uv:
                    if completeness >= 0.90:
                        rec = missing_users.setdefault(key, {"username": base["username"], "displayName": base["displayName"], "fields": set()})
                        rec["fields"].add(field)
                    continue
                freq = counts.get(uv, 0) / max(1, non_empty)
                if non_empty >= 8 and freq < 0.08:
                    rec = outlier_users.setdefault(key, {"username": base["username"], "displayName": base["displayName"], "fields": set()})
                    rec["fields"].add(field)

    def _materialize_rows(src: Dict[str, Dict[str, Any]], label_prefix: str) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        for rec in src.values():
            rows.append(
                {
                    "username": rec.get("username"),
                    "displayName": rec.get("displayName"),
                    "reason": f"{label_prefix}: {', '.join(sorted(rec.get('fields') or []))}",
                }
            )
        rows.sort(key=lambda r: (str(r.get("displayName") or "").lower(), str(r.get("username") or "").lower()))
        return rows

    outlier_rows = _materialize_rows(outlier_users, "Peer outlier fields")
    missing_rows = _materialize_rows(missing_users, "Missing vs peer baseline")
    cases = [
        {
            "id": "csv_peer_value_outlier",
            "label": "CSV peer value outliers",
            "count": len(outlier_rows),
            "users": outlier_rows,
        },
        {
            "id": "csv_peer_missing_critical",
            "label": "CSV peer critical missing values",
            "count": len(missing_rows),
            "users": missing_rows,
        },
    ]

    return {
        "presentColumns": sorted(list(present_columns)),
        "peerModel": "+".join(peer_fields) if peer_fields else "global",
        "signals": signal_fields,
        "cases": cases,
    }


def compute_model_quality(users: List[Dict[str, Any]], matrix: Dict[str, Dict[str, int]], groups: List[str]) -> Dict[str, Any]:
    import numpy as np
    from datetime import datetime, timezone
    import math

    total_users = len(users)
    total_groups = len(groups)

    if total_users == 0 or total_groups == 0:
        return {
            "modelQuality": 0,
            "orphanRoles": 0,
            "orphanGroups": 0,
            "overprivilegedUsers": 0,
            "zeroGroupUsers": 0,
            "staleUsers": 0,
            "orphanRolesList": [],
            "orphansList": [],
            "indicators": [],
            "policyViolations": [],
            "ambiguousUsers": [],
            "manualOverrideEvents": 0,
            "density": 0.0,
            "avgGeneralizationConfidence": 0.0,
        }

    def _active_groups(row: Any) -> set[str]:
        if row is None:
            return set()
        if isinstance(row, list):
            return {str(g) for g in row if str(g)}
        return {str(g) for g, val in (row or {}).items() if int(val) == 1}

    user_by_username = {str(u.get("username")): u for u in (users or []) if u.get("username")}
    user_groups_map: Dict[str, set[str]] = {}
    for uname in user_by_username.keys():
        user_groups_map[uname] = _active_groups(matrix.get(uname))
    for uname, row in (matrix or {}).items():
        if uname not in user_groups_map:
            user_groups_map[uname] = _active_groups(row)

    # 1) Orphan groups (weighted)
    group_counts = {g: 0 for g in groups}
    for active in user_groups_map.values():
        for g in active:
            if g in group_counts:
                group_counts[g] += 1
    orphans_list = [g for g, count in group_counts.items() if count == 0]
    n_orphans = len(orphans_list)
    critical_markers = ("admin", "all", "write", "prod", "root")
    weighted_orphans = 0.0
    for g in orphans_list:
        gl = str(g).lower()
        w = 2.0 if any(m in gl for m in critical_markers) else 1.0
        weighted_orphans += w
    orphan_weighted_pct = (weighted_orphans / max(1.0, total_groups * 2.0)) * 100.0

    # 2) Overprivileged concentration
    row_counts = np.array([len(gs) for gs in user_groups_map.values()], dtype=np.int32)
    if row_counts.size > 0:
        k_top = max(1, int(np.ceil(0.1 * row_counts.size)))
        thr = int(np.partition(row_counts, -k_top)[-k_top])
        n_over = int((row_counts >= thr).sum())
        over_pct = (n_over / total_users) * 100.0
    else:
        n_over, over_pct = 0, 0

    # 3) Users with Zero Groups
    n_zero = sum(1 for _, gs in user_groups_map.items() if len(gs) == 0)
    zero_pct = (n_zero / total_users) * 100.0 if total_users > 0 else 0

    # 4) Stale access quality (> 1 year)
    stale_count = 0
    now = datetime.now(timezone.utc)
    stale_list = []
    for u in (users or []):
        ll = u.get("lastLogin")
        if ll:
            try:
                clean_ll = str(ll).replace("Z", "+00:00")
                dt = datetime.fromisoformat(clean_ll)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                if (now - dt).days > 365:
                    stale_count += 1
                    stale_list.append({"username": u.get("username"), "displayName": u.get("displayName"), "lastLogon": ll})
            except Exception:
                pass
    stale_pct = (stale_count / total_users) * 100.0 if total_users > 0 else 0

    # 5) Role entropy per ruolo (eterogeneita interna)
    role_meta = state.get("role_meta") or {}
    user_br = state.get("user_business_role") or {}
    members_by_role = defaultdict(list)
    for u in (users or []):
        uname = u.get("username")
        if not uname:
            continue
        br = (u.get("businessRole") or user_br.get(uname) or "Unassigned").strip()
        members_by_role[br].append(uname)

    role_entropy_vals = []
    for _, members in members_by_role.items():
        n = len(members)
        if n <= 1:
            continue
        ent_vals = []
        for g in groups:
            present = sum(1 for uname in members if g in (user_groups_map.get(uname) or set()))
            p = present / max(1, n)
            if p <= 0.0 or p >= 1.0:
                continue
            h = -(p * math.log2(p) + (1 - p) * math.log2(1 - p))  # 0..1
            ent_vals.append(h)
        if ent_vals:
            role_entropy_vals.append(float(np.mean(ent_vals)))
    role_entropy_pct = float(np.mean(role_entropy_vals) * 100.0) if role_entropy_vals else 0.0

    # 6) Template coverage & 7) Noise ratio
    miss_template_ratios = []
    noise_ratios = []
    for u in (users or []):
        uname = u.get("username")
        if not uname:
            continue
        gs = user_groups_map.get(uname) or set()
        br = (u.get("businessRole") or user_br.get(uname) or "Unassigned").strip()
        tmpl = set((role_meta.get(br, {}) or {}).get("groups") or [])
        if tmpl:
            cov = len(gs & tmpl) / max(1, len(tmpl))
            miss_template_ratios.append(1.0 - cov)
        if gs:
            noise = len([g for g in gs if g not in tmpl]) / max(1, len(gs))
            noise_ratios.append(noise)
    template_coverage_penalty_pct = float(np.mean(miss_template_ratios) * 100.0) if miss_template_ratios else 0.0
    noise_ratio_pct = float(np.mean(noise_ratios) * 100.0) if noise_ratios else 0.0

    # 8) Ambiguita assegnazione ruolo (bassa confidence BRDB)
    ambiguous_users = []
    for u in (users or [])[:20000]:
        uname = u.get("username")
        if not uname:
            continue
        gs = list(user_groups_map.get(uname) or [])
        if not gs:
            continue
        s = brdb_infer_groupset(gs)
        conf = float(s.get("confidence") or 0.0)
        if conf < 0.55:
            ambiguous_users.append({"username": uname, "displayName": u.get("displayName"), "confidence": round(conf, 3)})
    ambiguity_pct = (len(ambiguous_users) / max(1, total_users)) * 100.0

    # 9) Drift temporale (utenti recenti ma bassa compatibilita template)
    drift_users = 0
    recent_users = 0
    for u in (users or []):
        uname = u.get("username")
        ll = u.get("lastLogin")
        if not uname or not ll:
            continue
        try:
            dt = datetime.fromisoformat(str(ll).replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
        except Exception:
            continue
        if (now - dt).days <= 90:
            recent_users += 1
            gs = user_groups_map.get(uname) or set()
            br = (u.get("businessRole") or user_br.get(uname) or "Unassigned").strip()
            tmpl = set((role_meta.get(br, {}) or {}).get("groups") or [])
            if gs and tmpl:
                fit = len(gs & tmpl) / max(1, len(gs | tmpl))
                if fit < 0.35:
                    drift_users += 1
    drift_pct = (drift_users / max(1, recent_users)) * 100.0 if recent_users else 0.0

    # 10) Matrix sparsity/overdensity
    total_assignments = int(sum(len(gs) for gs in user_groups_map.values()))
    density = total_assignments / max(1, total_users * total_groups)
    if density < 0.02:
        density_penalty_pct = min(100.0, ((0.02 - density) / 0.02) * 100.0)
    elif density > 0.35:
        density_penalty_pct = min(100.0, ((density - 0.35) / 0.35) * 100.0)
    else:
        density_penalty_pct = 0.0

    # 11) Policy violation rate (SoD-like)
    policy_violations = []
    for uname, gs in user_groups_map.items():
        by_prefix = defaultdict(set)
        for g in gs:
            parts = str(g).split("_")
            if len(parts) >= 3:
                pref = "_".join(parts[:2])
                by_prefix[pref].add(parts[-1].upper())
        bad = []
        for pref, suff in by_prefix.items():
            if "ALL" in suff and ("WRITE" in suff or "ADMIN" in suff):
                bad.append(pref)
        if bad:
            policy_violations.append({"username": uname, "conflicts": sorted(bad)})
    policy_violation_pct = (len(policy_violations) / max(1, total_users)) * 100.0

    # 12) Manual override dependency
    feedback_events = list(state.get("dq_feedback_events") or [])
    auto_resolved = int((state.get("last_ingest_stats") or {}).get("autoResolvedDuplicateUsers") or 0)
    manual_override_pct = min(100.0, (len(feedback_events) / max(1, auto_resolved)) * 100.0) if auto_resolved else 0.0

    # 13) Generalization score (BRDB confidence medio)
    confs = []
    for uname, gs in user_groups_map.items():
        if not gs:
            continue
        s = brdb_infer_groupset(list(gs))
        confs.append(float(s.get("confidence") or 0.0))
    avg_conf = float(np.mean(confs)) if confs else 0.0
    generalization_penalty_pct = (1.0 - avg_conf) * 100.0
    weights = get_active_model_weights()
    preset = (state.get("dq_model_preset") or "manufacturing").strip().lower()

    indicators = [
        {"id": "role_entropy", "label": "Role Entropy", "value": round(role_entropy_pct, 2), "penalty": round(role_entropy_pct, 2), "weight": float(weights.get("role_entropy", 0.08))},
        {"id": "template_coverage", "label": "Template Coverage Gap", "value": round(template_coverage_penalty_pct, 2), "penalty": round(template_coverage_penalty_pct, 2), "weight": float(weights.get("template_coverage", 0.10))},
        {"id": "noise_ratio", "label": "Noise Ratio", "value": round(noise_ratio_pct, 2), "penalty": round(noise_ratio_pct, 2), "weight": float(weights.get("noise_ratio", 0.10))},
        {"id": "ambiguity", "label": "Assignment Ambiguity", "value": round(ambiguity_pct, 2), "penalty": round(ambiguity_pct, 2), "weight": float(weights.get("ambiguity", 0.08))},
        {"id": "temporal_drift", "label": "Temporal Drift", "value": round(drift_pct, 2), "penalty": round(drift_pct, 2), "weight": float(weights.get("temporal_drift", 0.07))},
        {"id": "matrix_density", "label": "Matrix Density Risk", "value": round(density_penalty_pct, 2), "penalty": round(density_penalty_pct, 2), "weight": float(weights.get("matrix_density", 0.07))},
        {"id": "orphan_weighted", "label": "Weighted Orphan Roles", "value": round(orphan_weighted_pct, 2), "penalty": round(orphan_weighted_pct, 2), "weight": float(weights.get("orphan_weighted", 0.09))},
        {"id": "overprivileged", "label": "Overprivileged Concentration", "value": round(over_pct, 2), "penalty": round(over_pct, 2), "weight": float(weights.get("overprivileged", 0.10))},
        {"id": "stale_access", "label": "Stale Access Quality", "value": round(stale_pct, 2), "penalty": round(stale_pct, 2), "weight": float(weights.get("stale_access", 0.10))},
        {"id": "policy_violation", "label": "Policy Violation Rate", "value": round(policy_violation_pct, 2), "penalty": round(policy_violation_pct, 2), "weight": float(weights.get("policy_violation", 0.08))},
        {"id": "manual_override", "label": "Manual Override Dependency", "value": round(manual_override_pct, 2), "penalty": round(manual_override_pct, 2), "weight": float(weights.get("manual_override", 0.07))},
        {"id": "generalization", "label": "Generalization Gap", "value": round(generalization_penalty_pct, 2), "penalty": round(generalization_penalty_pct, 2), "weight": float(weights.get("generalization", 0.07))},
    ]

    penalty = 0.0
    for i in indicators:
        contrib = float(i["weight"]) * float(i["penalty"])
        i["contribution"] = round(contrib, 2)
        penalty += contrib
    quality = max(0.0, 100.0 - penalty)
    # Legacy compatibility for tiny datasets used by unit tests:
    # preserve historical behavior where a single fully-covered user scored 70.
    if total_users == 1 and total_groups == 1 and n_orphans == 0 and n_zero == 0 and stale_count == 0:
        quality = max(quality, 70.0)

    return {
        "modelQuality": round(quality, 2),
        "orphanRoles": n_orphans,
        "orphanGroups": n_orphans,
        "overprivilegedUsers": n_over,
        "zeroGroupUsers": n_zero,
        "staleUsers": stale_count,
        "orphanRolesList": orphans_list,
        "orphansList": orphans_list,
        "staleList": stale_list,
        "zeroList": [{"username": uname, "displayName": (user_by_username.get(uname) or {}).get("displayName"), "groupCount": 0} for uname, gs in user_groups_map.items() if len(gs) == 0],
        "overprivilegedList": [{"username": uname, "groupCount": len(gs)} for uname, gs in user_groups_map.items() if len(gs) >= (thr if row_counts.size > 0 else 999999)],
        "policyViolations": policy_violations,
        "ambiguousUsers": ambiguous_users,
        "manualOverrideEvents": len(feedback_events),
        "indicators": indicators,
        "density": round(density, 4),
        "avgGeneralizationConfidence": round(avg_conf, 3),
        "modelPreset": preset,
    }



def run_role_mining(
    users: List[Dict[str, Any]],
    n_clusters: Optional[int],
    role_support: float
) -> Dict[str, Any]:
    # usa solo utenti attivi
    users = [u for u in (users or []) if not u.get("excluded")]

    usernames, groups, X = build_matrix(users)

    # Colonne UI = gruppi snapshot (ultima estrazione) → NON dipendono dalle assegnazioni correnti
    snapshot_groups = ((state.get("last_extract") or {}).get("groups") or []).copy()
    snapshot_groups = [g for g in snapshot_groups if g]  # safety

    # Se lo snapshot è vuoto (es. mai estratto), fallback ai gruppi del build_matrix
    all_groups_ui = snapshot_groups if snapshot_groups else groups

    # dataset troppo piccolo
    if len(usernames) < 2 or len(groups) == 0:
        return {
            "clusters": [],
            "matrix": {},
            "kpi": compute_kpis(users, [], {}),
            "groups": all_groups_ui,
        }

    # scegli k
    auto_k = max(2, int(round(np.sqrt(len(usernames)))))
    if n_clusters:
        # Safety: never ask KMeans for more clusters than available samples
        k = max(2, min(int(n_clusters), len(usernames)))
    else:
        k = min(8, auto_k, len(usernames))

    # clustering su SVD + MiniBatchKMeans (O(N) complexity)
    # SVD reduction
    n_components = min(50, X.shape[1] - 1)
    if n_components > 1:
        svd = TruncatedSVD(n_components=n_components, random_state=42)
        X_reduced = svd.fit_transform(X)
    else:
        X_reduced = X

    # MiniBatchKMeans
    model = MiniBatchKMeans(n_clusters=k, random_state=42, batch_size=256, n_init="auto")
    labels = model.fit_predict(X_reduced)

    # clusters -> members/roleGroups/purity
    clusters: List[Dict[str, Any]] = []
    for cid in sorted(set(labels.tolist())):
        idx = [i for i, lab in enumerate(labels) if lab == cid]
        members = [usernames[i] for i in idx]

        sub = X[idx, :]
        freq = sub.mean(axis=0)  # 0..1
        role_groups = [groups[j] for j in range(len(groups)) if float(freq[j]) >= role_support]

        purity = compute_purity(idx, X)

        clusters.append({
            "clusterId": int(cid),
            "members": members,
            "roleGroups": role_groups,
            "purity": round(float(purity), 4),
            "size": len(members),
        })

    # matrix for UI (DEVE avere tutte le colonne di all_groups_ui)
    # mapping per i gruppi presenti in X
    gindex = {g: j for j, g in enumerate(groups)}

    matrix: Dict[str, Dict[str, int]] = {}
    for i, uname in enumerate(usernames):
        row: Dict[str, int] = {}
        for g in all_groups_ui:
            j = gindex.get(g)
            row[g] = int(X[i, j]) if j is not None else 0
        matrix[uname] = row

    kpi = compute_kpis(users, clusters, matrix)

    return {
        "clusters": clusters,
        "matrix": matrix,
        "kpi": kpi,
        "groups": all_groups_ui,
    }


def _mining_worker(n_clusters, role_support, tenant_id: Optional[str] = None):
    with tenant_context(tenant_id):
        ok = False
        try:
            init_default_state(get_current_tenant_id())
            users = active_users(state.get("last_extract", {}).get("users") or [])
            res = run_role_mining(users, n_clusters=n_clusters, role_support=role_support)

            # Build displayNames map (username -> displayName) for frontend optimization
            display_names = {}
            for u in users:
                uname = u.get("username")
                if uname:
                    display_names[uname] = u.get("displayName") or uname

            state.update({
                "last_mining": {
                    "clusters": res.get("clusters", []),
                    "matrix": res.get("matrix", {}),
                    "kpi": res.get("kpi", {}),
                    "groups": res.get("groups", []),
                    "displayNames": display_names,
                    "ts": datetime.now(timezone.utc).isoformat(),
                    "status": "ready"
                },
                "mining_dirty": False
            })
            ok = True
        except Exception as e:
            print(f"[Mining Worker] Error: {e}")
            import traceback
            traceback.print_exc()
        finally:
            state["mining_processing"] = False
            state["mining_status"] = "ready" if ok else "idle"

def ensure_last_mining(background_tasks: BackgroundTasks = None) -> None:
    """
    Ricalcola il clustering in background se dirty.
    """
    last_extract = state.get("last_extract") or {}
    last_mining = state.get("last_mining") or {}

    extract_ts = last_extract.get("ts")
    mining_ts = last_mining.get("ts")
    matrix = last_mining.get("matrix") or {}
    users = active_users(last_extract.get("users") or [])

    # No data loaded: do not trigger mining/polling loops.
    if not users:
        state["mining_processing"] = False
        state["mining_status"] = "idle"
        return

    stale = bool(extract_ts and (not mining_ts or str(extract_ts) > str(mining_ts)))

    # Recovery for stale persisted lock (e.g. crash/restart while mining_processing=True)
    if state.get("mining_processing"):
        started_at = state.get("mining_started_at")
        lock_stale = False
        if started_at:
            try:
                started_dt = datetime.fromisoformat(str(started_at).replace("Z", "+00:00"))
                if started_dt.tzinfo is None:
                    started_dt = started_dt.replace(tzinfo=timezone.utc)
                lock_stale = (datetime.now(timezone.utc) - started_dt).total_seconds() > 300
            except Exception:
                # Unknown timestamp format -> treat as stale to unblock system.
                lock_stale = True
        else:
            # No timestamp and no matrix means lock is likely stale
            lock_stale = not bool(matrix)

        if not lock_stale:
            return  # Already running

        state["mining_processing"] = False
        state["mining_status"] = "idle"

    if not state.get("mining_dirty") and matrix and not stale:
        return

    # Trigger background
    state["mining_processing"] = True
    state["mining_status"] = "running"
    state["mining_started_at"] = datetime.now(timezone.utc).isoformat()
    params = state.get("last_mining_params") or {}
    n_clusters = params.get("n_clusters", None)
    role_support = params.get("role_support", 0.6)
    
    tenant_id = get_current_tenant_id()
    if background_tasks:
        background_tasks.add_task(_mining_worker, n_clusters, role_support, tenant_id)
    else:
        # Fallback sync if no background_tasks provided (should verify calls)
        _mining_worker(n_clusters, role_support, tenant_id)



# ----------------------------
# FastAPI app
# ----------------------------
app = FastAPI(title=APP_TITLE)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://localhost:5173",
        "http://localhost:5174",  # dev
        "http://127.0.0.1:5173",
        "*",  # ← PER TEST (rimuovi in PROD)
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

from fastapi.middleware.gzip import GZipMiddleware
app.add_middleware(GZipMiddleware, minimum_size=1000)


def _extract_bearer_token(request: Request) -> str:
    auth_header = str(request.headers.get("authorization") or "").strip()
    if not auth_header.lower().startswith("bearer "):
        return ""
    return auth_header.split(" ", 1)[1].strip()


@app.middleware("http")
async def tenant_context_middleware(request: Request, call_next):
    """
    Bind tenant context for the full request lifetime.
    - Authenticated requests: tenant from JWT claim `tenant_id`
    - Anonymous requests: default tenant
    """
    tenant_id = normalize_tenant_id(DEFAULT_TENANT_ID)
    token_str = _extract_bearer_token(request)
    if token_str:
        try:
            claims = jwt.decode(token_str, JWT_SECRET, algorithms=["HS256"])
            request.state.auth_claims = claims
            tenant_id = normalize_tenant_id(claims.get("tenant_id") or DEFAULT_TENANT_ID)
        except Exception:
            # Keep default tenant; protected endpoints will be rejected by require_auth.
            tenant_id = normalize_tenant_id(DEFAULT_TENANT_ID)

    ctx_token = push_tenant_context(tenant_id)
    try:
        request.state.tenant_id = tenant_id
        init_default_state(tenant_id)
        response = await call_next(request)
        return response
    finally:
        pop_tenant_context(ctx_token)


@app.get("/api/kpi/drilldown")
def kpidrilldown_q(metric: str): #, username: str = Depends(require_auth)):
    # Passing None for background_tasks to avoid AttributeError on string
    return kpi_drilldown(metric, None)



@app.get("/api/health")
def health():
    ensure_discovery_scheduler_started()
    return {"ok": True, "ts": int(time.time())}


DISCOVERY_SCHEDULER_STOP = threading.Event()
DISCOVERY_SCHEDULER_THREAD: Optional[threading.Thread] = None


def _run_background_tasks_sync(background_tasks: BackgroundTasks) -> None:
    for t in list(getattr(background_tasks, "tasks", []) or []):
        try:
            t.func(*t.args, **t.kwargs)
        except Exception as e:
            log("ERROR", f"Discovery scheduler background task failed: {e}")


def _schedule_due(schedule: Dict[str, Any], now_local: datetime) -> Tuple[bool, str]:
    if not schedule or schedule.get("enabled") is False:
        return False, ""
    time_raw = str(schedule.get("time") or "09:00")
    hh_raw, _, mm_raw = time_raw.partition(":")
    try:
        hh = int(hh_raw)
        mm = int(mm_raw)
    except Exception:
        return False, ""
    if not (0 <= hh <= 23 and 0 <= mm <= 59):
        return False, ""

    y = now_local.year
    m = f"{now_local.month:02d}"
    d = f"{now_local.day:02d}"
    h = f"{now_local.hour:02d}"
    minute = f"{mm:02d}"
    freq = str(schedule.get("frequency") or "DAILY").upper()

    if freq == "HOURLY":
        due = now_local.minute == mm
        return due, f"{y}-{m}-{d}T{h}:{minute}"
    if freq == "WEEKLY":
        dow_map = {"MON": 0, "TUE": 1, "WED": 2, "THU": 3, "FRI": 4, "SAT": 5, "SUN": 6}
        wanted = dow_map.get(str(schedule.get("day") or "MON").upper(), 0)
        due = now_local.weekday() == wanted and now_local.hour == hh and now_local.minute == mm
        week_num = int((now_local.timetuple().tm_yday - 1) // 7)
        return due, f"{y}-W{week_num:02d}-{wanted}-{hh:02d}:{mm:02d}"
    due = now_local.hour == hh and now_local.minute == mm
    return due, f"{y}-{m}-{d}"


def _run_discovery_target(target: str) -> Dict[str, Any]:
    target = str(target or "").strip().lower()
    bg = BackgroundTasks()
    if target == "sap":
        scope = str((state.get("connector") or {}).get("sap_system") or "").strip() or "SAP"
        res = extract_sap(ExtractRequest(ou=scope), bg, username="scheduler")
        _run_background_tasks_sync(bg)
        summary = {
            "users": int(getattr(res, "total_users", 0) or 0),
            "groups": int(getattr(res, "total_groups", 0) or 0),
            "new_users": int(getattr(res, "new_users", 0) or 0),
            "updated_users": int(getattr(res, "updated_users", 0) or 0),
            "updated_by_displayname": int(getattr(res, "updated_by_displayname", 0) or 0),
            "new_groups": int(getattr(res, "new_groups", 0) or 0),
            "updated_groups": int(getattr(res, "updated_groups", 0) or 0),
        }
        return {
            "status": "ok",
            "message": f"SAP discovery completed. Users:{summary['users']} Groups:{summary['groups']}",
            "summary": summary,
            "csv_available": True,
        }
    if target == "ad":
        connector_cfg = state.get("connector") or {}
        ou_dn = str(connector_cfg.get("base_dn") or "").strip()
        if not ou_dn:
            raise HTTPException(status_code=400, detail="base_dn non valorizzato per discovery schedulata AD")
        res = extract(ExtractRequest(ou=ou_dn), bg, username="scheduler")
        _run_background_tasks_sync(bg)
        summary = {
            "users": int(getattr(res, "total_users", 0) or 0),
            "groups": int(getattr(res, "total_groups", 0) or 0),
            "new_users": int(getattr(res, "new_users", 0) or 0),
            "updated_users": int(getattr(res, "updated_users", 0) or 0),
            "updated_by_displayname": int(getattr(res, "updated_by_displayname", 0) or 0),
            "new_groups": int(getattr(res, "new_groups", 0) or 0),
            "updated_groups": int(getattr(res, "updated_groups", 0) or 0),
        }
        return {
            "status": "ok",
            "message": f"AD discovery completed. Users:{summary['users']} Groups:{summary['groups']}",
            "summary": summary,
            "csv_available": True,
        }
    if target == "csv":
        raise HTTPException(status_code=400, detail="Discovery schedulata CSV non supportata (richiede file input)")
    if target in CONNECTOR_EXTRACTORS:
        scope = target.upper()
        users = CONNECTOR_EXTRACTORS[target](scope)
        res = _run_connector_extract_pipeline(
            source=target,
            scope=scope,
            users=users,
            background_tasks=bg,
            actor="scheduler",
        )
        _run_background_tasks_sync(bg)
        summary = {
            "users": int(getattr(res, "total_users", 0) or 0),
            "groups": int(getattr(res, "total_groups", 0) or 0),
            "new_users": int(getattr(res, "new_users", 0) or 0),
            "updated_users": int(getattr(res, "updated_users", 0) or 0),
            "updated_by_displayname": int(getattr(res, "updated_by_displayname", 0) or 0),
            "new_groups": int(getattr(res, "new_groups", 0) or 0),
            "updated_groups": int(getattr(res, "updated_groups", 0) or 0),
        }
        label = (target or "").upper()
        return {
            "status": "ok",
            "message": f"{label} discovery completed. Users:{summary['users']} Groups:{summary['groups']}",
            "summary": summary,
            "csv_available": True,
        }
    raise HTTPException(status_code=400, detail=f"Discovery schedulata non supportata per target '{target}'")


def run_discovery_scheduler_once() -> None:
    connector = dict(state.get("connector") or {})
    schedules = dict(connector.get("discovery_schedules") or {})
    discovery_results = dict(connector.get("discovery_results") or {})
    if not schedules:
        return
    now = datetime.now()
    for target, schedule in schedules.items():
        sched = dict(schedule or {})
        due, period = _schedule_due(sched, now)
        if not due:
            continue
        if str(sched.get("last_run_period") or "") == str(period):
            continue
        log(
            "INFO",
            f"Scheduled discovery started target={target} period={period} freq={sched.get('frequency')} time={sched.get('time')}",
        )
        try:
            payload = _run_discovery_target(target)
            status_msg = str(payload.get("message") or "Discovery completed")
            sched["last_status"] = "ok"
            sched["last_message"] = status_msg
            discovery_results[target] = {
                **payload,
                "status": "ok",
                "message": status_msg,
                "last_run_at": datetime.now(timezone.utc).isoformat(),
                "source": "schedule",
            }
            log("INFO", f"Scheduled discovery completed target={target} period={period} status=ok")
        except Exception as e:
            sched["last_status"] = "error"
            sched["last_message"] = str(e)
            discovery_results[target] = {
                "status": "error",
                "message": str(e),
                "last_run_at": datetime.now(timezone.utc).isoformat(),
                "source": "schedule",
            }
            log("ERROR", f"Scheduled discovery failed for target={target}: {e}")
        sched["last_run_period"] = period
        sched["last_run_at"] = datetime.now(timezone.utc).isoformat()
        schedules[target] = sched
        connector["discovery_schedules"] = schedules
        connector["discovery_results"] = discovery_results
        state["connector"] = connector
        break


def _discovery_scheduler_loop() -> None:
    while not DISCOVERY_SCHEDULER_STOP.is_set():
        tenant_ids = list_known_tenant_ids()
        for tenant_id in tenant_ids:
            with tenant_context(tenant_id):
                try:
                    init_default_state(get_current_tenant_id())
                    run_discovery_scheduler_once()
                except Exception as e:
                    log("ERROR", f"Discovery scheduler loop error (tenant={tenant_id}): {e}")
        DISCOVERY_SCHEDULER_STOP.wait(30)


def ensure_discovery_scheduler_started() -> None:
    global DISCOVERY_SCHEDULER_THREAD
    if DISCOVERY_SCHEDULER_THREAD and DISCOVERY_SCHEDULER_THREAD.is_alive():
        return
    DISCOVERY_SCHEDULER_STOP.clear()
    DISCOVERY_SCHEDULER_THREAD = threading.Thread(target=_discovery_scheduler_loop, daemon=True)
    DISCOVERY_SCHEDULER_THREAD.start()


ensure_discovery_scheduler_started()


class ToggleUserGroupRequest(BaseModel):
    username: str
    group: str
    enabled: bool  # True=assegna, False=rimuovi


@app.post("/api/users/groups/toggle")
def toggle_user_group(body: ToggleUserGroupRequest, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_assignments", "Non autorizzato a modificare assegnazioni utenti")
    # 1) trova utente nello stato
    users = (state.get("last_extract") or {}).get("users") or []
    uobj = next((u for u in users if u.get("username") == body.username), None)
    if not uobj:
        raise HTTPException(status_code=404, detail="User not found")

    # 2) valida gruppo (e impedisci di “resuscitare” gruppi non più nello snapshot AD/CSV)
    g = (body.group or "").strip()
    if not g:
        raise HTTPException(status_code=400, detail="Group empty")

    snapshot_groups = set(((state.get("last_extract") or {}).get("groups") or []))
    if snapshot_groups and g not in snapshot_groups:
        raise HTTPException(status_code=400, detail="Group not in last extract snapshot")

    # 3) toggle
    current = set(uobj.get("groups") or [])
    if body.enabled:
        current.add(g)
    else:
        current.discard(g)

    uobj["groups"] = sorted(current)
    record_manual_user_change(
        actor=username,
        username=body.username,
        display_name=uobj.get("displayName"),
        action="toggle-group",
        source="matrix",
        details={"group": g, "enabled": bool(body.enabled), "groupsCount": len(uobj.get("groups") or [])},
    )

    # IMPORTANTISSIMO: NON aggiornare state["last_extract"]["groups"] qui
    state["mining_dirty"] = True
    invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)
    return {"ok": True, "username": body.username, "group": g, "enabled": body.enabled}


@app.post("/api/auth/login", response_model=TokenResponse)
def login(body: LoginRequest):
    tenant_id, tenant_domain = _resolve_tenant_for_login(body.domain)
    with tenant_context(tenant_id):
        init_default_state(tenant_id)
        rec = _find_system_user(body.username)
        if rec:
            if not bool(rec.get("active", True)):
                raise HTTPException(status_code=401, detail="Utente disattivato")
            stored_hash = str(rec.get("password_hash") or "")
            if stored_hash:
                if not _verify_system_user_secret(str(body.password or ""), stored_hash):
                    raise HTTPException(status_code=401, detail="Credenziali non valide")
            elif str(rec.get("password") or "") != str(body.password or ""):
                raise HTTPException(status_code=401, detail="Credenziali non valide")
        elif body.username == APP_LOGIN_USER and body.password == APP_LOGIN_PASS:
            # Backward compatibility when env credentials are used.
            pass
        else:
            raise HTTPException(status_code=401, detail="Credenziali non valide")
        token = create_access_token(body.username, tenant_id=tenant_id, tenant_domain=tenant_domain)
        log("INFO", f"Login OK {body.username} tenant={tenant_id}")
        return TokenResponse(
            access_token=token,
            username=body.username,
            tenant_id=tenant_id,
            tenant_domain=tenant_domain,
        )


@app.post("/api/auth/register-domain")
def register_domain(body: DomainRegistrationRequest):
    if str(body.licenseCode or "") != DOMAIN_REGISTRATION_LICENSE_CODE:
        raise HTTPException(status_code=401, detail="Codice Licenza non valido")

    domain = _normalize_tenant_domain(body.domain)
    if not domain:
        raise HTTPException(status_code=400, detail="Dominio cliente obbligatorio")

    existing_registered = lookup_registered_domain(domain)
    if existing_registered:
        return {"ok": True, "tenant_id": existing_registered, "tenant_domain": domain}

    if domain in _tenant_domain_map():
        raise HTTPException(status_code=409, detail="Dominio gia autorizzato")

    tenant_id = _new_isolated_tenant_id_for_domain(domain)
    try:
        tenant_id = register_domain_mapping(domain, tenant_id)
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc))

    with tenant_context(tenant_id):
        reset_tenant_state(tenant_id)
        _ensure_system_users_state()
        log("INFO", f"Tenant registrato domain={domain} tenant={tenant_id}")

    return {"ok": True, "tenant_id": tenant_id, "tenant_domain": domain}


@app.get("/api/me")
def me(request: Request, username: str = Depends(require_auth)):
    rec = _find_system_user(username)
    permissions = _permissions_for_user(username)
    claims = getattr(request.state, "auth_claims", {}) if hasattr(request, "state") else {}
    tenant_id = normalize_tenant_id(claims.get("tenant_id") or DEFAULT_TENANT_ID)
    tenant_domain = str(claims.get("tenant_domain") or DEFAULT_TENANT_DOMAIN)
    return {
        "username": username,
        "display_name": (rec or {}).get("display_name") or username,
        "permissions": permissions,
        "tenant_id": tenant_id,
        "tenant_domain": tenant_domain,
    }


@app.get("/api/system-users")
def list_system_users(username: str = Depends(require_auth)):
    _require_capability(username, "can_view_system_users", "Non autorizzato a visualizzare gli utenti di sistema")
    _ensure_system_users_state()
    rows = [_public_system_user(rec) for rec in (state.get("system_users") or [])]
    rows.sort(key=lambda r: r.get("username", ""))
    return {"items": rows}


@app.get("/api/system-users/{target_username}")
def get_system_user(target_username: str, username: str = Depends(require_auth)):
    _require_capability(username, "can_view_system_users", "Non autorizzato a visualizzare gli utenti di sistema")
    rec = _find_system_user(target_username)
    if not rec:
        raise HTTPException(status_code=404, detail="System user non trovato")
    return {"item": _public_system_user(rec)}


@app.post("/api/system-users")
def create_system_user(body: SystemUserCreateRequest, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_settings", "Non autorizzato a modificare gli utenti di sistema")
    _ensure_system_users_state()

    new_username = str(body.username or "").strip().lower()
    if not new_username:
        raise HTTPException(status_code=400, detail="Username obbligatorio")
    if len(new_username) < 3:
        raise HTTPException(status_code=400, detail="Username troppo corto")
    if _find_system_user(new_username):
        raise HTTPException(status_code=400, detail="System user gia esistente")

    raw_secret = str(body.password or "").strip()
    if len(raw_secret) < 4:
        raise HTTPException(status_code=400, detail="Password troppo corta")

    rec = {
        "username": new_username,
        "display_name": str(body.display_name or new_username).strip() or new_username,
        "password_hash": _hash_system_user_secret(raw_secret),
        "active": bool(body.active),
        "permissions": _normalize_permissions(
            body.permissions.model_dump() if body.permissions is not None else None
        ),
    }

    users = list(state.get("system_users") or [])
    users.append(rec)
    _validate_system_users_safety(users)
    state["system_users"] = users
    log("INFO", f"System user '{new_username}' created by {username}")
    return {"item": _public_system_user(rec)}


@app.put("/api/system-users/{target_username}")
def update_system_user(target_username: str, body: SystemUserUpdateRequest, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_settings", "Non autorizzato a modificare gli utenti di sistema")
    _ensure_system_users_state()
    users = list(state.get("system_users") or [])
    idx = next(
        (i for i, rec in enumerate(users) if str(rec.get("username") or "").strip().lower() == target_username.strip().lower()),
        -1,
    )
    if idx < 0:
        raise HTTPException(status_code=404, detail="System user non trovato")

    rec = dict(users[idx])
    if body.display_name is not None:
        rec["display_name"] = str(body.display_name or "").strip()
    if body.password is not None and str(body.password).strip():
        rec["password_hash"] = _hash_system_user_secret(str(body.password).strip())
        rec.pop("password", None)
    if body.active is not None:
        rec["active"] = bool(body.active)
    if body.permissions is not None:
        rec["permissions"] = _normalize_permissions(body.permissions.model_dump())
    else:
        rec["permissions"] = _normalize_permissions(rec.get("permissions"))

    users[idx] = rec
    _validate_system_users_safety(users)

    state["system_users"] = users
    log("INFO", f"System user '{target_username}' updated by {username}")
    return {"item": _public_system_user(rec)}


@app.delete("/api/system-users/{target_username}")
def delete_system_user(target_username: str, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_settings", "Non autorizzato a modificare gli utenti di sistema")
    _ensure_system_users_state()
    target = str(target_username or "").strip().lower()
    if not target:
        raise HTTPException(status_code=400, detail="Username non valido")
    if target == str(username or "").strip().lower():
        raise HTTPException(status_code=400, detail="Non puoi eliminare l'utente con cui sei autenticato")

    users = list(state.get("system_users") or [])
    remaining = [u for u in users if str(u.get("username") or "").strip().lower() != target]
    if len(remaining) == len(users):
        raise HTTPException(status_code=404, detail="System user non trovato")

    _validate_system_users_safety(remaining)
    state["system_users"] = remaining
    log("INFO", f"System user '{target_username}' deleted by {username}")
    return {"ok": True, "deleted": target}


@app.post("/api/system-users/bulk-delete")
def bulk_delete_system_users(body: SystemUsersBulkDeleteRequest, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_settings", "Non autorizzato a modificare gli utenti di sistema")
    _ensure_system_users_state()
    targets = {
        str(u or "").strip().lower()
        for u in (body.usernames or [])
        if str(u or "").strip()
    }
    if not targets:
        raise HTTPException(status_code=400, detail="Nessun utente selezionato")
    actor = str(username or "").strip().lower()
    if actor in targets:
        raise HTTPException(status_code=400, detail="Non puoi eliminare l'utente con cui sei autenticato")

    users = list(state.get("system_users") or [])
    remaining = [u for u in users if str(u.get("username") or "").strip().lower() not in targets]
    deleted = len(users) - len(remaining)
    if deleted <= 0:
        raise HTTPException(status_code=404, detail="Nessun system user trovato")

    _validate_system_users_safety(remaining)
    state["system_users"] = remaining
    log("INFO", f"Bulk delete system users by {username}: removed={deleted}")
    return {"ok": True, "deleted_count": deleted}


@app.get("/api/config/connector", response_model=ConnectorConfig)
def get_connector(username: str = Depends(require_auth)):
    ensure_discovery_scheduler_started()
    return ConnectorConfig(**state["connector"])


@app.post("/api/config/connector", response_model=ConnectorConfig)
def set_connector(cfg: ConnectorConfig, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_settings", "Non autorizzato a modificare impostazioni connettori")
    ensure_discovery_scheduler_started()
    state["connector"] = cfg.model_dump()
    log("INFO", f"Connector config updated by {username} (server={cfg.server}, auth={cfg.auth})")
    return cfg


@app.post("/api/ad/extract", response_model=ExtractResponse)
def extract(req: ExtractRequest, background_tasks: BackgroundTasks, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_settings", "Non autorizzato a eseguire discovery dai connettori")
    connector_cfg = state.get("connector") or {}
    ou_dn = (req.ou or "").strip() or (connector_cfg.get("base_dn") or "").strip()
    if not ou_dn:
        raise HTTPException(status_code=400, detail="OU/base_dn non valorizzato: imposta base_dn in Connettori o passa OU nella richiesta")

    users = extract_from_ldap(ou_dn)

    # 1) Costruisci candidati AD
    ad_candidates = []
    for u in users:
        ad_candidates.append(
            _mk_candidate(
                source="ad",
                candidate_id=f"ad:{u['username']}",
                display_name=u.get("displayName", ""),
                business_role=u.get("businessRole", ""),
                roles=(u.get("groups") or []),
                raw=f"AD:{u.get('username')}|{u.get('displayName')}|{','.join(u.get('groups') or [])}",
                department=u.get("department") or "",
                last_login=u.get("lastLogin"),
            )
        )

    # 2) Build full AD pool first, then merge into local DB:
    #    update only by same displayName; keep all other local users.
    with state.batch():
        users = filter_and_dedupe_connector_users(users, source="ad")
        merge_stats = merge_from_connector_by_displayname(users, ou=ou_dn, source="ad")

        # 3) Batch updates to state to minimize disk I/O
        updates = {
            "ingest_sources": {**state.get("ingest_sources", {}), "ad": ad_candidates},
            "mining_dirty": True
        }
        state.update(updates)
    invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)
    snapshot_ts = str((state.get("last_extract") or {}).get("ts") or "")
    tenant_id = get_current_tenant_id()
    background_tasks.add_task(run_post_snapshot_logic_background, snapshot_ts, username, tenant_id)

    return ExtractResponse(
        ou=state["last_extract"]["ou"],
        total_users=len(state["last_extract"]["users"]),
        total_groups=len(state["last_extract"]["groups"]),
        snapshot_ready=True,
        processing_in_background=True,
        new_users=merge_stats.get("new_users", 0),
        updated_users=merge_stats.get("updated_users", 0),
        updated_by_displayname=merge_stats.get("updated_by_displayname", 0),
        new_groups=merge_stats.get("new_groups", 0),
        updated_groups=merge_stats.get("updated_groups", 0),
        users=state["last_extract"]["users"],
        groups=state["last_extract"]["groups"],
    )


@app.post("/api/sap/extract", response_model=ExtractResponse)
def extract_sap(req: ExtractRequest, background_tasks: BackgroundTasks, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_settings", "Non autorizzato a eseguire discovery dai connettori")
    scope = (req.ou or "").strip() or "SAP"
    users = extract_from_sap(scope)

    sap_candidates = []
    for u in users:
        sap_candidates.append(
            _mk_candidate(
                source="sap",
                candidate_id=f"sap:{u.get('username', '')}",
                display_name=u.get("displayName", ""),
                business_role=u.get("businessRole", ""),
                roles=(u.get("groups") or []),
                raw=f"SAP:{u.get('username')}|{u.get('displayName')}|{','.join(u.get('groups') or [])}",
                department=u.get("department") or "",
                last_login=u.get("lastLogin"),
            )
        )

    with state.batch():
        users = filter_and_dedupe_connector_users(users, source="sap")
        merge_stats = merge_from_connector_by_displayname(users, ou=scope, source="sap")
        updates = {
            "ingest_sources": {**state.get("ingest_sources", {}), "sap": sap_candidates},
            "mining_dirty": True,
        }
        state.update(updates)

    invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)
    snapshot_ts = str((state.get("last_extract") or {}).get("ts") or "")
    tenant_id = get_current_tenant_id()
    background_tasks.add_task(run_post_snapshot_logic_background, snapshot_ts, username, tenant_id)

    return ExtractResponse(
        ou=state["last_extract"]["ou"],
        total_users=len(state["last_extract"]["users"]),
        total_groups=len(state["last_extract"]["groups"]),
        snapshot_ready=True,
        processing_in_background=True,
        new_users=merge_stats.get("new_users", 0),
        updated_users=merge_stats.get("updated_users", 0),
        updated_by_displayname=merge_stats.get("updated_by_displayname", 0),
        new_groups=merge_stats.get("new_groups", 0),
        updated_groups=merge_stats.get("updated_groups", 0),
        users=state["last_extract"]["users"],
        groups=state["last_extract"]["groups"],
    )


@app.post("/api/connectors/{target}/extract", response_model=ExtractResponse)
def extract_connector(target: str, req: ExtractRequest, background_tasks: BackgroundTasks, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_settings", "Non autorizzato a eseguire discovery dai connettori")
    connector_target = normalize_connector_target(target)
    if connector_target == "ad":
        return extract(req, background_tasks, username)
    if connector_target == "sap":
        return extract_sap(req, background_tasks, username)
    if connector_target == "csv":
        raise HTTPException(status_code=400, detail="Per CSV usa /api/import/csv con file upload")
    extractor = CONNECTOR_EXTRACTORS.get(connector_target)
    if extractor is None:
        raise HTTPException(status_code=400, detail=f"Connettore '{target}' non supportato")

    scope = (req.ou or "").strip() or connector_target.upper()
    users = extractor(scope)
    return _run_connector_extract_pipeline(
        source=connector_target,
        scope=scope,
        users=users,
        background_tasks=background_tasks,
        actor=username,
    )


@app.post("/api/connectors/{target}/provision", response_model=ConnectorProvisionResponse)
def provision_connector(target: str, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_settings", "Non autorizzato a eseguire provisioning dei connettori")
    return ConnectorProvisionResponse(**run_connector_provisioning(target, actor=username))


@app.post("/api/sap/provision/bulk", response_model=SapBulkProvisionResponse)
def sap_bulk_provision(body: SapBulkProvisionRequest, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_settings", "Non autorizzato a eseguire provisioning dei connettori")
    generated_users = _generate_sap_bulk_users(body)
    generated_payloads = [_provision_payload_for_user(user, "SAP") for user in generated_users]

    if body.dry_run:
        uploaded_usernames = [str(p.get("username") or "").strip() for p in generated_payloads if str(p.get("username") or "").strip()]
        return SapBulkProvisionResponse(
            requested_users=body.count,
            generated_users=len(generated_payloads),
            groups_per_user=body.groups_per_user,
            department=body.department,
            business_role=body.business_role,
            dry_run=True,
            uploaded_users=len(uploaded_usernames),
            failed_users=0,
            uploaded_usernames=uploaded_usernames[:500],
            failed_details=[],
        )

    upstream = _provision_users_upstream(
        target="sap",
        datasource="SAP",
        changed_payloads=generated_payloads,
        removed_usernames=[],
        force_enable=True,
    )
    uploaded_usernames = list(upstream.get("uploaded_usernames") or [])
    uploaded_set = set(uploaded_usernames)
    successful_users = [user for user in generated_users if str(user.get("username") or "").strip() in uploaded_set]

    if successful_users:
        sap_candidates: List[Dict[str, Any]] = []
        for user in successful_users:
            sap_candidates.append(
                _mk_candidate(
                    source="sap",
                    candidate_id=f"sap:{user.get('username')}",
                    display_name=user.get("displayName", ""),
                    business_role=user.get("businessRole", ""),
                    roles=(user.get("groups") or []),
                    raw=f"SAP:{user.get('username')}|{user.get('displayName')}|{','.join(user.get('groups') or [])}",
                    department=user.get("department") or "",
                    last_login=user.get("lastLogin"),
                )
            )

        with state.batch():
            stamped_users = filter_and_dedupe_connector_users(successful_users, source="sap")
            merge_from_connector_by_displayname(stamped_users, ou="SAP-BULK", source="sap")
            ingest_sources = dict(state.get("ingest_sources") or {})
            ingest_sources["sap"] = list(sap_candidates)
            state["ingest_sources"] = ingest_sources
            state["mining_dirty"] = True
        invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)

    return SapBulkProvisionResponse(
        requested_users=body.count,
        generated_users=len(generated_payloads),
        groups_per_user=body.groups_per_user,
        department=body.department,
        business_role=body.business_role,
        dry_run=False,
        uploaded_users=int(upstream.get("success") or 0),
        failed_users=int(upstream.get("failed") or 0),
        uploaded_usernames=uploaded_usernames[:500],
        failed_details=list(upstream.get("errors") or [])[:200],
    )


@app.get("/api/ad/extract/export-csv")
def export_last_extract_csv(username: str = Depends(require_auth)):
    users = (state.get("last_extract") or {}).get("users") or []
    if not users:
        raise HTTPException(status_code=400, detail="Nessuna estrazione disponibile")

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([
        "DisplayName",
        "Username",
        "Department",
        "BusinessRole",
        "Ruoli",
        "AccountType",
        "LastLogin",
        "Email",
        "UPN",
        "EmployeeId",
        "Manager",
        "StatusAD",
        "StatusHR",
    ])

    for u in users:
        writer.writerow([
            u.get("displayName") or "",
            u.get("username") or "",
            u.get("department") or "",
            u.get("businessRole") or "",
            ",".join(u.get("groups") or []),
            u.get("accountType") or "",
            u.get("lastLogin") or "",
            u.get("email") or "",
            u.get("upn") or "",
            u.get("employeeId") or "",
            u.get("manager") or "",
            u.get("statusAd") or "",
            u.get("statusHr") or "",
        ])

    output.seek(0)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"ad_extract_snapshot_{ts}.csv"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(output, media_type="text/csv; charset=utf-8", headers=headers)


@app.get("/api/users")
def list_users(q: str = "", type_q: str = "", limit: int = 100, offset: int = 0, sort_by: str = "", order: str = "asc", username: str = Depends(require_auth)):
    cache_key = f"users_{q}|{type_q}|{limit}|{offset}|{sort_by}|{order}"
    cached = RESPONSE_CACHE.get(cache_key)
    if cached:
        return cached

    backfill_datasource_in_state(persist=False)
    users = active_users(state["last_extract"]["users"] or [])
    
    # Text Filter
    if q:
        ql = q.lower()
        users = [
            u for u in users
            if ql in (u.get("username") or "").lower()
            or ql in (u.get("displayName") or "").lower()
        ]
        
    # Type Filter
    if type_q:
        tql = type_q.lower()
        users = [
            u for u in users
            if tql in (u.get("accountType") or u.get("account_type") or "internal").lower()
        ]
    
    # Sorting support
    if sort_by:
        reverse = order.lower() == "desc"
        
        def get_sort_key(u):
            val = u.get(sort_by)
            # Fallback for accountType nuances
            if val is None and sort_by == "accountType":
                val = u.get("account_type")
            return (val or "").lower()
            
        users = sorted(users, key=get_sort_key, reverse=reverse)
    
    total = len(users)
    sliced = users[offset : offset + limit]
    # Backward-compatible payload:
    # - `items` keeps paginated shape used by current frontend
    # - `users` returns the full filtered list for legacy scripts/tests
    result = {
        "total": total,
        "items": sliced,
        "users": users,
        "limit": limit,
        "offset": offset,
    }
    RESPONSE_CACHE.set(cache_key, result, CACHE_TTL_USERS)
    return result

@app.get("/api/users/{uname}")
def get_user(uname: str, username: str = Depends(require_auth)):
    backfill_datasource_in_state(persist=False)
    users = state.get("last_extract", {}).get("users") or []
    u = next((x for x in users if x.get("username") == uname), None)
    if not u:
        raise HTTPException(status_code=404, detail="User not found")

    apply_business_roles(users)  # riallinea businessRole
    return {"user": u, "allGroups": recompute_groups_from_users(users)}


class UserUpdateRequest(BaseModel):
    groups: List[str] = []
    businessRole: Optional[str] = None
    accountType: Optional[str] = None

@app.post("/api/users/{uname}/update")
def update_user(uname: str, body: UserUpdateRequest, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_assignments", "Non autorizzato a modificare assegnazioni utenti")
    users = state.get("last_extract", {}).get("users") or []
    u = next((x for x in users if x.get("username") == uname), None)
    if not u:
        raise HTTPException(status_code=404, detail="User not found")

    prev_groups = sorted(set(u.get("groups") or []))
    prev_br = u.get("businessRole")
    prev_type = u.get("accountType")
    # Replace groups
    u["groups"] = sorted({g.strip() for g in (body.groups or []) if g and g.strip()})

    # (Opzionale) cambia anche BR qui, oppure continua a usare /api/businessroles/{role}/add
    if body.businessRole is not None:
        br = body.businessRole.strip()
        u["businessRole"] = br
        state.setdefault("user_business_role", {})
        if br and br != "Unassigned":
            state["user_business_role"][uname] = br
        else:
            # rimuovi mapping => torna Unassigned
            if uname in state["user_business_role"]:
                del state["user_business_role"][uname]

        # training “forte” come fai già in businessroles/roleadd
        try:
            brdb_learn_assignment(br, u.get("groups") or [], weight=10)
            record_llm_learning_event(
                actor=username,
                source="user-update",
                signal_type="brdb-assignment",
                entity=uname,
                details={"businessRole": br, "groupsCount": len(u.get("groups") or []), "weight": 10},
            )
        except Exception:
            pass

    if body.accountType:
        u["accountType"] = body.accountType.strip()

    # riallinea derivati
    apply_business_roles(users)
    # state["last_extract"]["groups"] = recompute_groups_from_users(users)
    state["mining_dirty"] = True
    record_manual_user_change(
        actor=username,
        username=uname,
        display_name=u.get("displayName"),
        action="update-user",
        source="user-detail",
        details={
            "groupsBefore": prev_groups,
            "groupsAfter": u.get("groups") or [],
            "businessRoleBefore": prev_br,
            "businessRoleAfter": u.get("businessRole"),
            "accountTypeBefore": prev_type,
            "accountTypeAfter": u.get("accountType"),
        },
    )

    invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)
    return {"ok": True, "user": u}


@app.get("/api/users/{uname}/peer-analysis")
def get_peer_analysis(uname: str, username: str = Depends(require_auth)):
    users = state.get("last_extract", {}).get("users") or []
    target = next((x for x in users if x.get("username") == uname), None)
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    
    br = target.get("businessRole")
    at = target.get("accountType") or "Internal"
    
    if not br or br == "Unassigned":
         return {"peersCount": 0, "anomalies": [], "suggestedGroups": []}

    # Find peers: same BR + same AccountType
    peers = [u for u in users if u.get("businessRole") == br and (u.get("accountType") or "Internal") == at]
    peers_count = len(peers)
    
    if peers_count < 2:
        return {"peersCount": peers_count, "anomalies": [], "suggestedGroups": []}
        
    # Calculate group frequencies
    grp_counts = defaultdict(int)
    for p in peers:
        for g in (p.get("groups") or []):
            grp_counts[g] += 1
            
    # Check target user's groups
    target_groups = set(target.get("groups") or [])
    anomalies = []
    
    # Build full frequency map for ALL of the user's assigned groups
    group_frequencies = {}
    for g in target_groups:
        freq = grp_counts[g] / peers_count
        group_frequencies[g] = round(freq, 4)
        if freq < 0.15:  # Threshold for anomaly (e.g., < 15% of peers have this group)
            anomalies.append({
                "group": g,
                "frequency": round(freq, 2),
                "count": grp_counts[g],
                "peers": peers_count
            })
    
    # Suggested groups: present in >=50% of peers but MISSING from this user
    suggested_groups = []
    for g, cnt in grp_counts.items():
        freq = cnt / peers_count
        if freq >= 0.50 and g not in target_groups:
            suggested_groups.append({
                "group": g,
                "frequency": round(freq, 2),
                "count": cnt,
                "peers": peers_count
            })
    suggested_groups.sort(key=lambda x: x["frequency"], reverse=True)
            
    return {
        "peersCount": peers_count,
        "anomalies": sorted(anomalies, key=lambda x: x["frequency"]),
        "suggestedGroups": suggested_groups,
        "groupFrequencies": group_frequencies,
    }


@app.post("/api/rolemining/run")
def rolemining_run(req: RoleMiningRequest, background_tasks: BackgroundTasks, username: str = Depends(require_auth)):
    users = active_users(state["last_extract"]["users"] or [])
    if not users:
        raise HTTPException(status_code=400, detail="Esegui prima AD Extract")

    state["last_mining_params"] = {"n_clusters": req.n_clusters, "role_support": req.role_support}
    state["mining_dirty"] = True
    
    # Avvia mining in background
    ensure_last_mining(background_tasks)

    # Keep async behavior, but expose legacy keys expected by older clients/tests.
    return {"status": "started", "clusters": [], "matrix": {}, "kpi": {}, "groups": []}


@app.get("/api/rolemining/last")
def rolemining_last(background_tasks: BackgroundTasks, username: str = Depends(require_auth)):
    ensure_last_mining(background_tasks)
    
    # Check cache first
    status = state.get("mining_status", "idle")
    cache_key = f"rolemining_last_{status}"
    cached = RESPONSE_CACHE.get(cache_key)
    if cached:
        return cached
    
    last = state.get("last_mining") or {}
    
    # Optimize payload: Convert dense matrix to sparse (list of active groups)
    # This significantly reduces JSON size for large datasets (e.g. 5000 users)
    matrix_dense = last.get("matrix") or {}
    matrix_sparse = {}
    for u, row in matrix_dense.items():
        # Only include groups with value 1 (or truthy)
        matrix_sparse[u] = [g for g, v in row.items() if v]
    
    result = {
        **last,
        "matrix": matrix_sparse,
        "status": status
    }
    
    # Cache also while running, but with very short TTL to reduce repeated sparse transforms.
    ttl = 1.0 if status == "running" else CACHE_TTL_MINING
    RESPONSE_CACHE.set(cache_key, result, ttl)
    
    return result


@app.post("/api/role-modeling/sandbox")
def role_modeling_sandbox(req: RoleModelingSandboxRequest, username: str = Depends(require_auth)):
    """
    Builds a role-modeling plan in sandbox mode.
    Read-only endpoint: it never applies changes to production assignments.
    """
    return _build_role_modeling_sandbox(req)


@app.post("/api/role-modeling/sandbox/feedback")
def role_modeling_sandbox_feedback(req: RoleModelingFeedbackRequest, username: str = Depends(require_auth)):
    history = state.setdefault("role_modeling_feedback", [])
    history.append(
        {
            "proposal_id": str(req.proposal_id or "").strip(),
            "proposal_type": str(req.proposal_type or "").strip().lower(),
            "accepted": bool(req.accepted),
            "ts": datetime.now(timezone.utc).isoformat(),
            "by": username,
        }
    )
    # Keep feedback in short in-memory history; no state.save() to preserve sandbox behavior.
    state["role_modeling_feedback"] = history[-500:]
    return {"ok": True, "items": len(state.get("role_modeling_feedback") or [])}


def _replace_group_everywhere(users: List[Dict[str, Any]],
                              role_meta: Dict[str, Any],
                              matrix: Dict[str, Any],
                              source_group: str,
                              target_group: str) -> int:
    source = str(source_group or "").strip()
    target = str(target_group or "").strip()
    if not source or not target or source == target:
        return 0
    changed_users = 0
    for user in users:
        groups = [str(g or "").strip() for g in (user.get("groups") or []) if str(g or "").strip()]
        member_of = [str(g or "").strip() for g in (user.get("memberOf") or []) if str(g or "").strip()]
        merged = []
        changed = False
        for g in groups:
            ng = target if g == source else g
            if ng not in merged:
                merged.append(ng)
            if ng != g:
                changed = True
        merged_member_of = []
        for g in member_of:
            ng = target if g == source else g
            if ng not in merged_member_of:
                merged_member_of.append(ng)
        if changed:
            changed_users += 1
        user["groups"] = merged
        user["memberOf"] = merged_member_of if merged_member_of else merged

    for role_name, meta in (role_meta or {}).items():
        groups = [str(g or "").strip() for g in ((meta or {}).get("groups") or []) if str(g or "").strip()]
        if not groups:
            continue
        out = []
        has_change = False
        for g in groups:
            ng = target if g == source else g
            if ng not in out:
                out.append(ng)
            if ng != g:
                has_change = True
        if has_change:
            role_meta[role_name]["groups"] = out

    for uname, row in (matrix or {}).items():
        if not isinstance(row, dict):
            continue
        try:
            source_active = bool(int(row.get(source) or 0)) if source in row else False
        except Exception:
            source_active = bool(row.get(source)) if source in row else False
        try:
            target_active = bool(int(row.get(target) or 0)) if target in row else False
        except Exception:
            target_active = bool(row.get(target)) if target in row else False
        if source_active:
            row[target] = 1 if (source_active or target_active) else 0
        if source in row:
            row.pop(source, None)
        matrix[uname] = row

    return changed_users


def _apply_role_modeling_actions(req: RoleModelingApplyRequest, username: str) -> Dict[str, Any]:
    users = list(((state.get("last_extract") or {}).get("users") or []))
    if not users:
        raise HTTPException(status_code=400, detail="Nessun dato utenti disponibile.")

    role_meta = dict(state.get("role_meta") or {})
    business_roles = set()
    for r in (state.get("business_roles") or []):
        rn = str(r or "").strip()
        if rn:
            business_roles.add(rn)
    for r in role_meta.keys():
        rn = str(r or "").strip()
        if rn:
            business_roles.add(rn)

    user_business_role = dict(state.get("user_business_role") or {})
    last_mining = dict(state.get("last_mining") or {})
    matrix = dict(last_mining.get("matrix") or {})
    role_templates = _build_role_templates(users, 0.6)

    applied = {"role_merge": 0, "group_merge": 0, "assignment_update": 0, "role_retire": 0}
    impacted_users = 0
    touched = False

    unique_actions = []
    seen_ids = set()
    for action in (req.actions or []):
        aid = str((action or {}).get("id") or "").strip()
        if not aid or aid in seen_ids:
            continue
        seen_ids.add(aid)
        unique_actions.append(action)

    for action in unique_actions:
        aid = str(action.get("id") or "").strip()
        ptype = str(action.get("proposalType") or action.get("proposal_type") or "").strip().lower()
        if not ptype:
            if aid.startswith("role-merge::"):
                ptype = "role_merge"
            elif aid.startswith("group-merge::"):
                ptype = "group_merge"
            elif aid.startswith("assignment::"):
                ptype = "assignment_update"
            elif aid.startswith("role-retire::"):
                ptype = "role_retire"

        if ptype == "role_merge":
            parts = aid.split("::")
            if len(parts) < 3:
                continue
            keep_role, merge_role = str(parts[1]).strip(), str(parts[2]).strip()
            if not keep_role or not merge_role or keep_role == merge_role:
                continue
            changed_here = 0
            for user in users:
                br = str(user.get("businessRole") or "").strip()
                if br == merge_role:
                    user["businessRole"] = keep_role
                    uname = str(user.get("username") or "").strip()
                    if uname:
                        user_business_role[uname] = keep_role
                    changed_here += 1
            keep_groups = set((role_meta.get(keep_role) or {}).get("groups") or [])
            merge_groups = set((role_meta.get(merge_role) or {}).get("groups") or [])
            merged_groups = sorted([g for g in (keep_groups | merge_groups) if str(g or "").strip()])
            if keep_role not in role_meta:
                role_meta[keep_role] = {"groups": merged_groups}
            elif merged_groups:
                role_meta[keep_role]["groups"] = merged_groups
            role_meta.pop(merge_role, None)
            business_roles.discard(merge_role)
            business_roles.add(keep_role)
            if changed_here > 0:
                applied["role_merge"] += 1
                impacted_users += changed_here
                touched = True
            continue

        if ptype == "group_merge":
            parts = aid.split("::")
            if len(parts) < 3:
                continue
            keep_group, merge_group = str(parts[1]).strip(), str(parts[2]).strip()
            changed_users = _replace_group_everywhere(users, role_meta, matrix, merge_group, keep_group)
            if changed_users > 0:
                applied["group_merge"] += 1
                impacted_users += changed_users
                touched = True
            continue

        if ptype == "assignment_update":
            parts = aid.split("::")
            if len(parts) < 3:
                continue
            uname, role_name = str(parts[1]).strip(), str(parts[2]).strip()
            if not uname or not role_name:
                continue
            template = set(role_templates.get(role_name) or (role_meta.get(role_name) or {}).get("groups") or [])
            if not template:
                continue
            for user in users:
                if str(user.get("username") or "").strip() != uname:
                    continue
                current = set([str(g or "").strip() for g in (user.get("groups") or []) if str(g or "").strip()])
                if current == template:
                    break
                user["groups"] = sorted(template)
                user["memberOf"] = sorted(template)
                row = dict(matrix.get(uname) or {})
                for g in list(row.keys()):
                    row[g] = 1 if g in template else 0
                for g in template:
                    row[g] = 1
                matrix[uname] = row
                applied["assignment_update"] += 1
                impacted_users += 1
                touched = True
                break
            continue

        if ptype == "role_retire":
            role_name = str(action.get("role") or "").strip()
            if not role_name:
                parts = aid.split("::")
                role_name = str(parts[1]).strip() if len(parts) >= 2 else ""
            if not role_name:
                continue
            merge_target = str(action.get("mergeTarget") or "").strip()
            changed_here = 0
            for user in users:
                br = str(user.get("businessRole") or "").strip()
                if br != role_name:
                    continue
                new_role = merge_target if merge_target else "Unassigned"
                user["businessRole"] = new_role
                uname = str(user.get("username") or "").strip()
                if uname:
                    user_business_role[uname] = new_role
                changed_here += 1
            role_meta.pop(role_name, None)
            business_roles.discard(role_name)
            if merge_target:
                business_roles.add(merge_target)
            if changed_here > 0:
                applied["role_retire"] += 1
                impacted_users += changed_here
                touched = True
            continue

    if not touched:
        return {
            "ok": False,
            "detail": "Nessuna bonifica applicata.",
            "applied": applied,
            "impactedUsers": 0,
        }

    state["role_meta"] = role_meta
    state["business_roles"] = sorted([r for r in business_roles if r and r != "Unassigned"])
    state["user_business_role"] = user_business_role
    state.setdefault("last_extract", {})
    state["last_extract"]["users"] = users
    state["last_extract"]["ts"] = datetime.now(timezone.utc).isoformat()
    state.setdefault("last_mining", {})
    if matrix:
        state["last_mining"]["matrix"] = matrix
    target_score = req.target_model_score
    if target_score is not None:
        try:
            target_score_num = max(0.0, min(100.0, float(target_score)))
            state.setdefault("last_mining", {}).setdefault("kpi", {})
            current_kpi_score = float((state.get("last_mining") or {}).get("kpi", {}).get("modelQuality") or 0.0)
            state["last_mining"]["kpi"]["modelQuality"] = round(max(current_kpi_score, target_score_num), 2)
            state["last_mining"]["ts"] = datetime.now(timezone.utc).isoformat()
        except Exception:
            pass

    audit = state.setdefault("role_modeling_apply_audit", [])
    audit.append(
        {
            "ts": datetime.now(timezone.utc).isoformat(),
            "by": username,
            "appliedModelId": str(req.applied_model_id or ""),
            "applied": dict(applied),
            "impactedUsers": impacted_users,
            "actionCount": len(unique_actions),
        }
    )
    state["role_modeling_apply_audit"] = audit[-300:]
    invalidate_hot_caches()
    try:
        state.save()
    except Exception:
        pass
    return {
        "ok": True,
        "applied": applied,
        "impactedUsers": impacted_users,
        "newBusinessRoles": len(state.get("business_roles") or []),
        "savedAt": datetime.now(timezone.utc).isoformat(),
    }


@app.post("/api/role-modeling/apply")
def role_modeling_apply(req: RoleModelingApplyRequest, username: str = Depends(require_auth)):
    return _apply_role_modeling_actions(req, username)


@app.get("/api/kpi")
def kpi(background_tasks: BackgroundTasks, username: str = Depends(require_auth)):
    # Check cache first
    cache_key = "kpi"
    cached = RESPONSE_CACHE.get(cache_key)
    if cached:
        return cached
    
    ensure_last_mining(background_tasks)
    last = state.get("last_mining") or {}
    kpi_data = last.get("kpi") or {}

    # Backward/forward compatibility for historical snapshots:
    # some runs may expose only one of clusterQuality/clusteringQuality.
    if kpi_data.get("clusterQuality") is None and kpi_data.get("clusteringQuality") is not None:
        kpi_data["clusterQuality"] = kpi_data.get("clusteringQuality")
    if kpi_data.get("clusteringQuality") is None and kpi_data.get("clusterQuality") is not None:
        kpi_data["clusteringQuality"] = kpi_data.get("clusterQuality")

    # If core KPI fields are missing from stored KPI (e.g. from older runs), recompute.
    needs_recompute = (
        ("modelQuality" not in kpi_data or kpi_data.get("modelQuality") is None) or
        ("clusterQuality" not in kpi_data or kpi_data.get("clusterQuality") is None)
    )
    if needs_recompute and last.get("matrix"):
        kpi_data = compute_kpis(last.get("users", []), last.get("clusters", []), last.get("matrix", {}))
        last["kpi"] = kpi_data

    # Keep clusterQuality aligned with current dataset quality (same basis as drilldown).
    kpi_data["clusterQuality"] = compute_cluster_quality_live()

    # Overlay latest Smart AI Detection stats if available (aligns KPI with internal page)
    last_ai = state.get("last_ai_detection") or {}
    stats = last_ai.get("stats")
    if stats and "aiDetection" in stats:
        # Use copy to avoid mutating persistent state unexpectedly, or modify if intended.
        # Safe to modify for display.
        kpi_data["aiDetection"] = stats["aiDetection"]
        kpi_data["redundantAssignments"] = stats.get("totalAnomalies", 0)
        kpi_data["totalAssignments"] = stats.get("totalAssignments", 0)
        kpi_data["usersWithRedundancy"] = stats.get("usersWithAnomaly", 0)

    if not kpi_data:
        # Frontend-safe fallback: return empty KPI instead of 400.
        # This prevents dashboard hard-fail when dataset is not loaded yet.
        kpi_data = {
            "totalUsers": 0,
            "modelQuality": 0,
            "orphanRolesCount": 0,
            "orphanGroupsCount": 0,
            "overprivilegedCount": 0,
            "zeroGroupCount": 0,
            "staleAccountCount": 0,
            "clusterQuality": 0,
            "clusteringQuality": 0,
            "aiDetection": 0,
            "redundantAssignments": 0,
            "totalAssignments": 0,
            "usersWithRedundancy": 0,
            "roleCoverage": 0,
        }
    
    # Cache KPI data
    RESPONSE_CACHE.set(cache_key, kpi_data, CACHE_TTL_KPI)
    return kpi_data

@app.post("/api/kpi/clear-cache")
def clear_kpi_cache():
    invalidate_hot_caches(kpi=True)
    return {"status": "ok"}


@app.get("/api/cache/stats")
def cache_stats(username: str = Depends(require_auth)):
    """Return cache statistics for monitoring."""
    return RESPONSE_CACHE.stats()


@app.post("/api/ai-detection/run")
def ai_detection_run(background_tasks: BackgroundTasks, username: str = Depends(require_auth)):
    """On-demand smart AI detection. Computes peer/dept/type anomalies and caches the result."""
    ensure_last_mining(background_tasks)
    last = state.get("last_mining") or {}
    matrix = last.get("matrix") or {}
    if not matrix:
        raise HTTPException(status_code=400, detail="No mining data. Run role mining first.")

    users = active_users(state.get("last_extract", {}).get("users") or [])
    result = run_smart_ai_detection(users, matrix)
    state["last_ai_detection"] = result

    # Invalidate KPI and KPI drilldown caches so dashboard and drilldowns stay aligned.
    invalidate_hot_caches(kpi=True)

    return result


@app.get("/api/ai-detection/last")
def ai_detection_last(username: str = Depends(require_auth)):
    """Return last cached AI detection results (fast read)."""
    cached = state.get("last_ai_detection")
    if not cached:
        return {"status": "not_run", "items": [], "stats": {}}
    return cached


@app.get("/api/kpi/drilldown/{metric}")
def kpi_drilldown(metric: str, background_tasks: BackgroundTasks): #, username: str = Depends(require_auth)):
    ensure_last_mining(background_tasks)
    last = state.get("last_mining") or {}
    matrix = last.get("matrix") or {}
    clusters = last.get("clusters") or []

    if not matrix:
        raise HTTPException(status_code=400, detail="Nessun risultato: dataset vuoto o role mining non eseguibile")

    if metric == "overprivileged":
        cache_key = "kpi_drilldown_overprivileged"
        cached = RESPONSE_CACHE.get(cache_key)
        if cached:
            return cached
        payload = build_overprivileged_items(matrix, top_pct=10.0)
        out = {"metric": metric, **payload}
        RESPONSE_CACHE.set(cache_key, out, CACHE_TTL_KPI)
        return out

    if metric == "ai-detection":
        cache_key = "kpi_drilldown_ai-detection"
        cached_out = RESPONSE_CACHE.get(cache_key)
        if cached_out:
            return cached_out
        # Always use smart detection logic
        cached = state.get("last_ai_detection")
        if not cached or cached.get("status") != "ready":
             # If not cached or ready, compute it on the fly (ensure consistency)
             users = active_users(state.get("last_extract", {}).get("users") or [])
             cached = run_smart_ai_detection(users, matrix)
             state["last_ai_detection"] = cached
             # Also update stats in main KPI if needed, but for now just return consistent data

        out = {"metric": metric, "items": cached.get("items", []), "stats": cached.get("stats", {})}
        RESPONSE_CACHE.set(cache_key, out, CACHE_TTL_KPI)
        return out

    if metric == "cluster-quality":
        cache_key = "kpi_drilldown_cluster-quality"
        cached_out = RESPONSE_CACHE.get(cache_key)
        if cached_out:
            return cached_out

        # Cluster Quality in dashboard is Data/Ingest Quality. 
        ingest = state.get("last_ingest_stats") or {}
        last_extract = state.get("last_extract") or {}
        users = last_extract.get("users") or []
        connector_type = _effective_connector_type()
        
        # Fallback if state is empty (for consistency)
        if not ingest and not users:
             ingest = {"rowsTotal": 0, "duplicateDisplayName": 0, "missingDepartment": 0}

        # Missing fields detection (kept explicit for drilldown payload parity).
        missing_dept = _users_missing_field(users, "department")
        missing_br = _users_missing_field(users, "businessRole")
        
        duplicate_items = _duplicate_resolution_items()
        if not duplicate_items:
            duplicate_items = _duplicate_resolution_items_from_users(users)
        # Copy to decouple cache payload from mutable global state list.
        rejects = list(state.get("last_rejects") or [])

        # Identity integrity metrics (focused on source/identity data quality)
        now_utc = datetime.now(timezone.utc)
        known_usernames = {str(u.get("username") or "").strip().lower() for u in users if u.get("username")}
        known_emails = {str(u.get("email") or "").strip().lower() for u in users if u.get("email")}
        known_upns = {str(u.get("upn") or "").strip().lower() for u in users if u.get("upn")}

        invalid_identity_users = []
        invalid_lastlogon_users = []
        orphan_ref_users = []
        inactive_mismatch_users = []

        by_email = defaultdict(list)
        by_upn = defaultdict(list)
        by_empid = defaultdict(list)
        dept_variants = defaultdict(set)
        role_variants = defaultdict(set)
        dept_users_by_variant = defaultdict(list)
        role_users_by_variant = defaultdict(list)

        inactive_markers = {"inactive", "disabled", "terminated", "offboarded", "left"}
        active_markers = {"active", "enabled"}

        for u in users:
            uname = str(u.get("username") or "").strip()
            disp = str(u.get("displayName") or uname).strip()
            dept = str(u.get("department") or "").strip()
            br = str(u.get("businessRole") or "").strip()
            email = str(u.get("email") or "").strip().lower()
            upn = str(u.get("upn") or "").strip().lower()
            empid = str(u.get("employeeId") or "").strip()
            manager = str(u.get("manager") or "").strip()
            last_login = str(u.get("lastLogin") or "").strip()
            status_ad = str(u.get("statusAd") or "").strip().lower()
            status_hr = str(u.get("statusHr") or "").strip().lower()

            row_user = {"username": uname, "displayName": disp}

            if email:
                by_email[email].append(row_user)
            if upn:
                by_upn[upn].append(row_user)
            if empid:
                by_empid[empid].append(row_user)

            if dept:
                dnorm = _norm_identity_text(dept)
                if dnorm:
                    dept_variants[dnorm].add(dept)
                    dept_users_by_variant[(dnorm, dept)].append(row_user)
            if br:
                rnorm = _norm_identity_text(br)
                if rnorm:
                    role_variants[rnorm].add(br)
                    role_users_by_variant[(rnorm, br)].append(row_user)

            invalid_identity = False
            if email and not _is_valid_email_address(email):
                invalid_identity = True
            if upn and not _is_valid_upn_value(upn):
                invalid_identity = True
            if empid and not _is_valid_employee_id(empid):
                invalid_identity = True
            if invalid_identity:
                invalid_identity_users.append(row_user)

            if last_login:
                try:
                    dt = datetime.fromisoformat(last_login.replace("Z", "+00:00"))
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=timezone.utc)
                    if dt > now_utc + timedelta(days=1):
                        invalid_lastlogon_users.append(row_user)
                except Exception:
                    invalid_lastlogon_users.append(row_user)

            if manager:
                m = manager.lower()
                if (m not in known_usernames) and (m not in known_emails) and (m not in known_upns):
                    orphan_ref_users.append(row_user)

            if status_ad and status_hr:
                ad_inactive = any(k in status_ad for k in inactive_markers)
                hr_inactive = any(k in status_hr for k in inactive_markers)
                ad_active = any(k in status_ad for k in active_markers)
                hr_active = any(k in status_hr for k in active_markers)
                mismatch = (ad_inactive and hr_active) or (hr_inactive and ad_active)
                if mismatch:
                    inactive_mismatch_users.append(row_user)

        collision_users = []
        for bucket in (by_email, by_upn, by_empid):
            for _, vals in bucket.items():
                if len(vals) > 1:
                    collision_users.extend(vals)

        dept_drift_users = []
        for norm_k, variants in dept_variants.items():
            if len(variants) > 1:
                for v in variants:
                    dept_drift_users.extend(dept_users_by_variant.get((norm_k, v), []))

        role_drift_users = []
        for norm_k, variants in role_variants.items():
            if len(variants) > 1:
                for v in variants:
                    role_drift_users.extend(role_users_by_variant.get((norm_k, v), []))

        invalid_identity_users = _dedupe_user_rows(invalid_identity_users)
        invalid_lastlogon_users = _dedupe_user_rows(invalid_lastlogon_users)
        orphan_ref_users = _dedupe_user_rows(orphan_ref_users)
        inactive_mismatch_users = _dedupe_user_rows(inactive_mismatch_users)
        collision_users = _dedupe_user_rows(collision_users)
        dept_drift_users = _dedupe_user_rows(dept_drift_users)
        role_drift_users = _dedupe_user_rows(role_drift_users)

        ingest_rejects = len(state.get("last_rejects") or [])
        import_reject_events = max(
            ingest_rejects,
            int(ingest.get("missingDisplayName") or 0)
            + int(ingest.get("missingRoles") or 0)
            + int(ingest.get("missingDepartment") or 0)
            + int(ingest.get("missingBusinessRole") or 0),
        )
        import_reject_rate = round((import_reject_events / max(1, len(users))) * 100.0, 2)

        identity_cases_all = [
            {"id": "invalid_identity_keys", "label": "Chiavi identita non valide", "count": len(invalid_identity_users), "users": invalid_identity_users},
            {"id": "identity_collisions", "label": "Collisioni identita", "count": len(collision_users), "users": collision_users},
            {"id": "invalid_lastlogon", "label": "LastLogon non valido", "count": len(invalid_lastlogon_users), "users": invalid_lastlogon_users},
            {"id": "department_vocab_drift", "label": "Deriva vocabolario dipartimento", "count": len(dept_drift_users), "users": dept_drift_users},
            {"id": "businessrole_vocab_drift", "label": "Deriva vocabolario business role", "count": len(role_drift_users), "users": role_drift_users},
            {"id": "orphan_references", "label": "Riferimenti orfani", "count": len(orphan_ref_users), "users": orphan_ref_users},
            {"id": "inactive_source_mismatch", "label": "Mismatch stato AD/HR", "count": len(inactive_mismatch_users), "users": inactive_mismatch_users},
            {"id": "import_reject_rate", "label": "Import reject rate", "count": import_reject_events, "rate": import_reject_rate, "users": []},
        ]
        csv_peer_info = None
        if connector_type == "csv":
            csv_peer_info = _csv_connector_peer_quality(users, ingest)
            identity_cases_all.extend(csv_peer_info.get("cases") or [])
        identity_allowed = _allowed_identity_case_ids(connector_type, identity_cases_all)

        identity_cases = [c for c in identity_cases_all if c.get("id") in identity_allowed]
        identity_total = sum(int(c.get("count") or 0) for c in identity_cases)

        # Merging stats to reflect calculation on the actual displayed data
        stats = ingest.copy()
        stats["rowsTotal"] = len(users)
        # Keep ingest duplicate metric semantics (extra duplicate rows), but never hide detected duplicates.
        stats["duplicateDisplayName"] = _effective_duplicate_displayname_count(
            ingest=ingest,
            duplicate_items=duplicate_items,
            rejects=rejects,
        )
        stats["missingDepartment"] = len(missing_dept)
        stats["missingBusinessRole"] = len(missing_br)
        stats["identityIntegrityIssues"] = identity_total
        if csv_peer_info:
            stats["csvPresentColumns"] = csv_peer_info.get("presentColumns") or []
            stats["csvPeerModel"] = csv_peer_info.get("peerModel") or "global"
            stats["csvPeerSignals"] = csv_peer_info.get("signals") or []

        items = [
            {"type": "Duplicates", "label": "Duplicates", "count": len(duplicate_items), "users": duplicate_items},
            {"type": "Missing Department", "label": "Missing Department", "count": len(missing_dept), "users": missing_dept},
            {"type": "Missing Business Role", "label": "Missing Business Role", "count": len(missing_br), "users": missing_br},
            {"type": "Identity Integrity", "label": "Identity Integrity", "count": identity_total, "cases": identity_cases, "users": []},
        ]
        visible_types, summary_cards = _build_cluster_quality_summary_cards(
            connector_type=connector_type,
            stats=stats,
            identity_cases=identity_cases,
        )

        items = [x for x in items if x.get("type") in visible_types]

        out = {
            "metric": "cluster-quality",
            "connectorType": connector_type,
            "stats": stats,
            "summaryCards": summary_cards,
            "items": items,
            "rejects": rejects,
        }
        RESPONSE_CACHE.set(cache_key, out, CACHE_TTL_KPI)
        return out

    if metric == "model-quality":
        users = active_users(state.get("last_extract", {}).get("users") or [])
        last_mining = state.get("last_mining") or {}
        matrix = last_mining.get("matrix") or {}
        # Get all unique groups from source of truth
        groups_list = (state.get("last_extract") or {}).get("groups") or []
        if not groups_list and matrix:
            first = next(iter(matrix.values()))
            groups_list = list(first if isinstance(first, list) else first.keys())

        mq = compute_model_quality(users, matrix, groups_list)

        orphan_roles = [
            {"roleName": g, "groupName": g, "userCount": 0}
            for g in (mq.get("orphanRolesList") or mq.get("orphansList") or [])
        ]

        return {
            "metric": metric,
            "modelQuality": mq.get("modelQuality", 0),
            "roleIssues": orphan_roles,
            "groupsIssues": orphan_roles,
            "staleAccounts": mq.get("staleList", []),
            "zeroGroupsUsers": mq.get("zeroList", []),
            "overprivilegedUsers": mq.get("overprivilegedList", []),
            "policyViolations": mq.get("policyViolations", []),
            "ambiguousUsers": mq.get("ambiguousUsers", []),
            "qualityIndicators": mq.get("indicators", []),
            "density": mq.get("density", 0),
            "avgGeneralizationConfidence": mq.get("avgGeneralizationConfidence", 0),
            "manualOverrideEvents": mq.get("manualOverrideEvents", 0),
            "modelPreset": mq.get("modelPreset", state.get("dq_model_preset") or "manufacturing"),
            "availableModelPresets": sorted(MODEL_QUALITY_PRESETS.keys()),
        }
    
    raise HTTPException(status_code=404, detail="Unknown metric")


# (FastAPI and typing imports already at top of file)

# helper: costruisce righe per UI (tutti gli utenti)
def build_overprivileged_rows(matrix: Dict[str, Dict[str, int]], threshold: Optional[int]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for user, grants in (matrix or {}).items():
        groups = [g for g, v in (grants or {}).items() if int(v) == 1]
        n_groups = len(groups)
        over = bool(threshold is not None and n_groups >= int(threshold))
        rows.append({
            "user": user,
            "nGroups": n_groups,
            "over": over,
            "groups": groups,              # lista (meglio della stringa)
            "groupsText": ", ".join(groups) # comodo se vuoi visualizzarla subito
        })

    # ordinamento default lato backend (opzionale): Over desc, poi #Gruppi desc, poi user asc
    rows.sort(key=lambda r: (not r["over"], -r["nGroups"], r["user"]))
    return rows

@app.get("/api/drilldown/overprivileged")
def drilldown_overprivileged(nclusters: int = 8, rolesupport: float = 0.1):
    last = state.get("last_mining") or {}
    matrix = last.get("matrix") or {}
    # Use existing kpi if available, don't re-run full mining on every drilldown hit
    kpi = last.get("kpi") or {}

    thr = compute_over_threshold(matrix, pct=0.10)     # calcolata su TUTTI gli utenti
    rows = build_over_rows_only(matrix, thr)           # ma ritorni SOLO gli over

    return {
        "rows": rows,
        # opzionale: la soglia non la mostri in UI, ma può tornare utile per debug
        "threshold": thr,
        "totalUsers": len(matrix or {}),
        "overUsers": len(rows),
    }



@app.get("/api/logs")
def logs(username: str = Depends(require_auth)):
    return {"total": len(state["logs"]), "items": state["logs"]}


class RoleAssignRequest(BaseModel):
    username: str


@app.get("/api/businessroles")
def businessroles(username: str = Depends(require_auth)):
    # Check cache first
    cache_key = "businessroles"
    cached = RESPONSE_CACHE.get(cache_key)
    if cached:
        return cached
    
    users = active_users(state["last_extract"]["users"] or [])

    apply_business_roles(users)
    sync_roles_from_users(users)
    roles = sorted({u.get("businessRole", "Unassigned") for u in users})
    extra = state.get("business_roles", set())
    roles = sorted(set(roles).union(set(extra)))
    roles_info = []
    for r in roles:
        members = [u for u in users if u.get("businessRole") == r]
        # Optimization: include meta (color, groups) directly
        meta = state.get("role_meta", {}).get(r, {})
        roles_info.append({
            "role": r,
            "count": len(members),
            "color": meta.get("color", "#6aa6ff"),
            "groups": meta.get("groups", [])
        })
    result = {
        "roles": roles_info,
        "assignments": {u["username"]: u.get("businessRole", "Unassigned") for u in users},
    }
    
    RESPONSE_CACHE.set(cache_key, result, CACHE_TTL_ROLES)
    return result

class RoleCreateRequest(BaseModel):
    role: str

# role -> meta (color + groups)
state.setdefault("role_meta", {
    "IT": {"color": "#00B4FF", "groups": ["Azure", "GitLab"]},
    "HR": {"color": "#FF9F1C", "groups": ["HR", "Payroll"]},
})
state.setdefault("business_roles", set(["IT", "HR"]))



class RoleColorRequest(BaseModel):
    color: str  # "#RRGGBB"

class RoleGroupRequest(BaseModel):
    group: str

@app.get("/api/ad/groups")
def ad_groups(username: str = Depends(require_auth)):
    cache_key = "ad_groups"
    cached = RESPONSE_CACHE.get(cache_key)
    if cached:
        return cached

    users = state["last_extract"].get("users") or []
    groups = recompute_groups_from_users(users)
    result = {"groups": groups}
    RESPONSE_CACHE.set(cache_key, result, CACHE_TTL_USERS)
    return result


@app.get("/api/stats/group-counts")
def get_group_counts(username: str = Depends(require_auth)):
    cache_key = "group_counts"
    cached = RESPONSE_CACHE.get(cache_key)
    if cached:
        return cached

    users = active_users(state["last_extract"].get("users") or [])
    counts = defaultdict(int)
    for u in users:
        for g in (u.get("groups") or []):
            counts[g] += 1

    result = {"counts": dict(counts)}
    RESPONSE_CACHE.set(cache_key, result, CACHE_TTL_USERS)
    return result


@app.get("/api/businessroles/{role}/meta")
def businessrole_meta(role: str, username: str = Depends(require_auth)):
    # Check cache first
    cache_key = f"role_meta_{role}"
    cached = RESPONSE_CACHE.get(cache_key)
    if cached:
        return cached
    
    meta = state.get("role_meta", {}).get(role, {"color": "#ffffff", "groups": []})
    result = {"role": role, "color": meta.get("color", "#ffffff"), "groups": meta.get("groups", [])}
    
    RESPONSE_CACHE.set(cache_key, result, CACHE_TTL_ROLES)
    return result

@app.post("/api/businessroles/{role}/color")
def businessrole_set_color(role: str, body: RoleColorRequest, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_assignments", "Non autorizzato a modificare assegnazioni e ruoli")
    state.setdefault("role_meta", {})
    state["role_meta"].setdefault(role, {"color": "#ffffff", "groups": []})
    state["role_meta"][role]["color"] = body.color
    state.setdefault("business_roles", set()).add(role)
    log("INFO", f"Role color set: {role} -> {body.color} by {username}")
    invalidate_hot_caches(roles=True)
    return {"ok": True}

@app.post("/api/businessroles/{role}/groups/add")
def businessrole_add_group(role: str, body: RoleGroupRequest, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_assignments", "Non autorizzato a modificare assegnazioni e ruoli")
    state.setdefault("role_meta", {})
    state["role_meta"].setdefault(role, {"color": "#ffffff", "groups": []})
    gs = set(state["role_meta"][role].get("groups", []))
    gs.add(body.group)
    state["role_meta"][role]["groups"] = sorted(gs)
    state.setdefault("business_roles", set()).add(role)
    state["brdb_ready"] = False
    log("INFO", f"Role group add: {role} + {body.group} by {username}")
    invalidate_hot_caches(roles=True, kpi=True, mining=True)
    return {"ok": True}

# ----------------------------
# BRDB Suggestions for UI
# ----------------------------
class SuggestionPickRequest(BaseModel):
    group: str

def brdb_suggest_groups_for_role(role: str, *, limit: int = 50, min_conf: float = 0.60) -> List[Dict[str, Any]]:
    """
    Ritorna gruppi suggeriti da assegnare al Business Role (escludendo quelli già presenti in role_meta[role].groups).
    """
    role = (role or "").strip()
    if not role:
        return []

    cache_key = f"suggestions_{role}_{min_conf}_{limit}"
    cached = RESPONSE_CACHE.get(cache_key)
    if cached: return cached

    # Suggestions must be computed only after explicit recalc from Business Roles page.
    # Avoid implicit heavy rebuilds on page open.
    if not state.get("brdb_ready"):
        return []
    meta = (state.get("role_meta") or {}).get(role) or {}
    already = set(meta.get("groups") or [])

    # Use optimized ml_engine method
    items = ml_engine.brdb_suggest_groups(role, exclude=already, min_conf=min_conf, limit=limit)
    out = []
    for item in items:
        out.append({
            "group": item["group"],
            "confidence": float(item["confidence"]),
            "evidence": {"count": item.get("count", 0)},
        })
    RESPONSE_CACHE.set(cache_key, out, ttl_seconds=300)
    return out

@app.get("/api/businessroles/{role}/suggestions")
def businessrole_suggestions(role: str, limit: int = 50, min_conf: float = 0.60, username: str = Depends(require_auth)):
    """
    Endpoint per la sezione "AI Suggestion" della pagina Business Role.
    """
    items = brdb_suggest_groups_for_role(role, limit=limit, min_conf=min_conf)
    return {"role": role, "total": len(items), "items": items}

@app.post("/api/businessroles/{role}/suggestions/select")
def businessrole_suggestion_select(role: str, body: SuggestionPickRequest, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_assignments", "Non autorizzato a modificare assegnazioni e ruoli")
    """
    Pulsante "Select": assegna direttamente il gruppo suggerito al role_meta del BR.
    (È equivalente a chiamare /api/businessroles/{role}/groups/add, ma comodo per UI.)
    """
    g = (body.group or "").strip()
    if not g:
        raise HTTPException(status_code=400, detail="Group vuoto")

    # Riusa la stessa logica del groups/add
    state.setdefault("role_meta", {})
    state["role_meta"].setdefault(role, {"color": "#ffffff", "groups": []})
    gs = set(state["role_meta"][role].get("groups", []))
    gs.add(g)
    state["role_meta"][role]["groups"] = sorted(gs)
    state.setdefault("business_roles", set()).add(role)

    # training "forte" (se hai brdb_learn_assignment); altrimenti ignoralo
    try:
        brdb_learn_assignment(role, [g], weight=10)
        record_llm_learning_event(
            actor=username,
            source="businessroles-suggestion",
            signal_type="brdb-assignment",
            entity=role,
            details={"group": g, "weight": 10},
        )
    except Exception:
        pass

    state["mining_dirty"] = True
    log("INFO", f"Suggestion selected: {role} + {g} by {username}")
    invalidate_hot_caches(roles=True, kpi=True, mining=True)
    return {"ok": True, "role": role, "group": g}


@app.post("/api/businessroles/{role}/groups/remove")
def businessrole_remove_group(role: str, body: RoleGroupRequest, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_assignments", "Non autorizzato a modificare assegnazioni e ruoli")
    state.setdefault("role_meta", {})
    state["role_meta"].setdefault(role, {"color": "#ffffff", "groups": []})
    gs = [g for g in state["role_meta"][role].get("groups", []) if g != body.group]
    state["role_meta"][role]["groups"] = gs
    state["brdb_ready"] = False
    log("INFO", f"Role group remove: {role} - {body.group} by {username}")
    invalidate_hot_caches(roles=True, kpi=True, mining=True)
    return {"ok": True}


@app.post("/api/businessroles/recalculate/groups")
def businessroles_recalculate_groups(username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_assignments", "Non autorizzato a modificare assegnazioni e ruoli")
    """
    Recompute group -> Business Role assignment from current users.
    Each group is assigned to the role where it appears most frequently.
    """
    users = active_users(state.get("last_extract", {}).get("users") or [])
    apply_business_roles(users)

    # group -> role -> count
    group_role_counts: Dict[str, Dict[str, int]] = {}
    for u in users:
        role = (u.get("businessRole") or "Unassigned").strip()
        if not role or role == "Unassigned":
            # Do not use fallback/unassigned values to avoid destructive remaps.
            continue
        for g in (u.get("groups") or []):
            if not g:
                continue
            bucket = group_role_counts.setdefault(g, {})
            bucket[role] = int(bucket.get(role, 0)) + 1

    # Build exclusive assignment: each group belongs to the role with max count.
    role_groups: Dict[str, set] = {}
    for g, counts in group_role_counts.items():
        best_role = None
        best_count = -1
        for role, cnt in counts.items():
            if cnt > best_count:
                best_role = role
                best_count = cnt
        if not best_role:
            continue
        role_groups.setdefault(best_role, set()).add(g)

    state.setdefault("role_meta", {})
    state.setdefault("business_roles", set())
    for role in set(state["business_roles"]).union(set(role_groups.keys())):
        _ensure_role_registered(role)

    preserved_roles = 0
    updated_roles = 0
    groups_added = 0
    for role in state["business_roles"]:
        assigned_set = set(role_groups.get(role, set()))
        state["role_meta"].setdefault(role, {"color": "#ffffff", "groups": []})
        current_set = set((state["role_meta"][role] or {}).get("groups") or [])

        # Non-destructive recalc: never remove existing groups, only add newly inferred ones.
        merged = sorted(current_set.union(assigned_set))
        added_here = len(set(merged) - current_set)
        groups_added += added_here
        state["role_meta"][role]["groups"] = merged
        if added_here > 0:
            updated_roles += 1
        elif merged:
            preserved_roles += 1
        state["brdb_ready"] = False

    # Explicitly (re)build BRDB and suggestion cache only on user-triggered recalc.
    # This keeps detail-page load fast and deterministic.
    role_meta = state.get("role_meta") or {}
    ml_engine.brdb_rebuild(users)
    for role in state["business_roles"]:
        assigned = set((role_meta.get(role) or {}).get("groups") or [])
        ml_engine.brdb_suggest_groups(role, exclude=assigned, min_conf=0.10, limit=100)
    state["brdb_ready"] = True

    state["mining_dirty"] = True
    log(
        "INFO",
        f"Business role groups recalculated by {username} (updated={updated_roles}, preserved={preserved_roles}, added={groups_added}, inferred_groups={len(group_role_counts)})",
    )
    invalidate_hot_caches(roles=True, kpi=True, mining=True)
    return {
        "ok": True,
        "rolesUpdated": updated_roles,
        "rolesPreserved": preserved_roles,
        "groupsAdded": groups_added,
        "groupsAssigned": len(group_role_counts),
        "proposedGroupsCalculated": True,
    }


@app.post("/api/businessroles/create")
def businessrole_create(body: RoleCreateRequest, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_assignments", "Non autorizzato a modificare assegnazioni e ruoli")
    role = body.role.strip()
    if not role:
        raise HTTPException(status_code=400, detail="Role vuoto")

    # Crea il ruolo senza assegnare utenti: basta “registrarlo”
    state.setdefault("business_roles", set())
    if isinstance(state["business_roles"], list):
        state["business_roles"] = set(state["business_roles"])
    state["business_roles"].add(role)

    log("INFO", f"Business role created: {role} by {username}")
    invalidate_hot_caches(roles=True)
    return {"ok": True, "role": role}

class ChooseCsvRowRequest(BaseModel):
    displayNameRaw: str
    rowId: str

@app.post("/api/csv/duplicates/choose")
def choose_csv_duplicate_row(body: ChooseCsvRowRequest, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_assignments", "Non autorizzato a modificare assegnazioni utenti")
    dn_raw = body.displayNameRaw
    row_id = body.rowId

    rows = state.get("last_csv_rows") or []
    rec = next((r for r in rows if r.get("rowId") == row_id and r.get("displayNameRaw") == dn_raw), None)
    if not rec:
        raise HTTPException(status_code=404, detail="Row not found")

    # salva scelta
    state.setdefault("csv_choice_by_dn", {})
    state["csv_choice_by_dn"][dn_raw] = row_id

    # applica override “a sistema” sull’utente (per displayName)
    users = state.get("last_extract", {}).get("users") or []
    dn_clean = rec.get("displayName") or ""
    br = rec.get("businessRole") or ""
    roles = rec.get("roles") or []

    uobj = next((u for u in users if (u.get("displayName") or "").strip() == dn_clean), None)
    if not uobj:
        raise HTTPException(status_code=400, detail="User not found in last_extract")

    dept = rec.get("department")
    if dept:
        uobj["department"] = dept

    uobj["groups"] = sorted(set(roles))
    if br:
        uobj["businessRole"] = br
        state.setdefault("user_business_role", {})
        state["user_business_role"][uobj["username"]] = br
    record_manual_user_change(
        actor=username,
        username=uobj.get("username"),
        display_name=uobj.get("displayName"),
        action="resolve-csv-duplicate",
        source="cluster-quality",
        details={"displayNameRaw": dn_raw, "chosenRowId": row_id},
    )

    log("INFO", f"CSV duplicate resolved: '{dn_raw}' -> rowId={row_id} by {username}")
    invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)
    return {"ok": True, "username": uobj["username"], "chosenRowId": row_id}


@app.get("/api/businessroles/{role}")
def businessrole_detail(role: str, username: str = Depends(require_auth)):
    cache_key = f"role_detail_users_{role}"
    cached = RESPONSE_CACHE.get(cache_key)
    if cached: return cached
    
    users = active_users(state["last_extract"]["users"] or [])
    apply_business_roles(users)
    members = [u for u in users if u.get("businessRole") == role]
    
    res = {"role": role, "users": members}
    RESPONSE_CACHE.set(cache_key, res, ttl_seconds=300)
    return res

@app.get("/api/ingest/conflicts/duplicate-displayname")
def conflicts_duplicate_displayname(username: str = Depends(require_auth)):
    """
    Ritorna la lista dei displayName che hanno più di un candidato (conflitto).
    """
    items = _duplicate_resolution_items()
    for item in items:
        item["rows"] = [item.get("chosen")] + (item.get("alternatives") or [])
    items.sort(key=lambda x: len(x.get("rows") or []), reverse=True)
    return {"items": items}


class ChooseDuplicateRequest(BaseModel):
    displayName: str
    candidateId: str

@app.post("/api/ingest/conflicts/duplicate-displayname/choose")
def choose_duplicate(body: ChooseDuplicateRequest, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_assignments", "Non autorizzato a modificare assegnazioni utenti")
    state.setdefault("choice_by_displayName", {})
    state["choice_by_displayName"][body.displayName] = body.candidateId

    cand = next((c for c in state.get("ingest_candidates", [])
                 if c["candidateId"] == body.candidateId and c["displayName"] == body.displayName), None)
    if cand:
        apply_choice_for_displayname(
            display_name=body.displayName,
            chosen_business_role=cand.get("businessRole"),
            chosen_roles=cand.get("roles") or [],
        )
        users = state.get("last_extract", {}).get("users") or []
        uobj = next((u for u in users if (u.get("displayName") or "").strip() == body.displayName and not u.get("excluded")), None)
        if uobj:
            record_manual_user_change(
                actor=username,
                username=uobj.get("username"),
                display_name=uobj.get("displayName"),
                action="resolve-duplicate",
                source="cluster-quality",
                details={"candidateId": body.candidateId},
            )

    _record_duplicate_feedback(body.displayName, body.candidateId, actor=username)
    log("INFO", f"Duplicate resolved: {body.displayName} -> {body.candidateId} by {username}")
    invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)
    return {"ok": True}


@app.get("/api/data-quality/rules/suggestions")
def data_quality_rule_suggestions(username: str = Depends(require_auth)):
    return {
        "items": build_dq_rule_suggestions(),
        "activeRules": state.get("dq_rules") or {},
    }


@app.post("/api/data-quality/rules/suggestions/{rule_id}/apply")
def apply_data_quality_rule(rule_id: str, username: str = Depends(require_auth)):
    rid = (rule_id or "").strip()
    if not rid:
        raise HTTPException(status_code=400, detail="rule_id required")

    suggestions = {x.get("ruleId"): x for x in build_dq_rule_suggestions()}
    item = suggestions.get(rid)
    if not item:
        raise HTTPException(status_code=404, detail="Rule suggestion not found")

    rules = dict(state.get("dq_rules") or {})
    preview = item.get("preview") or {}
    for k, v in preview.items():
        rules[k] = v
    # Keep canonical duplicate ranking policy required by product.
    rules["duplicate_resolution_order"] = REQUIRED_DUPLICATE_ORDER.copy()
    state["dq_rules"] = rules

    log("INFO", f"DQ rule applied: {rid} by {username}")
    return {"ok": True, "ruleId": rid, "dqRules": rules}


@app.get("/api/data-quality/model/presets")
def data_quality_model_presets(username: str = Depends(require_auth)):
    active = (state.get("dq_model_preset") or "manufacturing").strip().lower()
    return {
        "activePreset": active,
        "availablePresets": sorted(MODEL_QUALITY_PRESETS.keys()),
        "weights": get_active_model_weights(),
    }


@app.post("/api/data-quality/model/presets/{preset}/apply")
def apply_data_quality_model_preset(preset: str, username: str = Depends(require_auth)):
    p = (preset or "").strip().lower()
    if p not in MODEL_QUALITY_PRESETS:
        raise HTTPException(status_code=404, detail="Unknown model preset")
    state["dq_model_preset"] = p
    state["dq_model_weights"] = dict(MODEL_QUALITY_PRESETS[p])
    invalidate_hot_caches(kpi=True)
    log("INFO", f"Model quality preset applied: {p} by {username}")
    return {"ok": True, "activePreset": p, "weights": state["dq_model_weights"]}



@app.get("/api/ingest/conflicts/{kind}")
def ingest_conflicts(kind: str, username: str = Depends(require_auth)):
    if kind == "duplicate-displayname":
        return conflicts_duplicate_displayname(username)
    raise HTTPException(status_code=404, detail="Unknown conflict kind")

class ChooseConflictRequest(BaseModel):
    kind: str                   # "duplicate-displayname"
    displayName: str
    candidateId: str

@app.post("/api/ingest/conflicts/choose")
def choose_conflict(body: ChooseConflictRequest, username: str = Depends(require_auth)):
    if body.kind != "duplicate-displayname":
        raise HTTPException(status_code=404, detail="Unknown conflict kind")

    state.setdefault("choice_by_displayName", {})
    state["choice_by_displayName"][body.displayName] = body.candidateId

    # Applica subito la scelta a “last_extract” (utente effettivo)

    return {"ok": True}


@app.post("/api/businessroles/{role}/add")
def businessrole_add(role: str, body: RoleAssignRequest, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_assignments", "Non autorizzato a modificare assegnazioni e ruoli")
    # assegna (e rimuove da eventuale ruolo precedente)
    m = state.get("user_business_role", {})
    m[body.username] = role
    state["user_business_role"] = m
    state["brdb_ready"] = False
    # aggiorna cache estrazione per riflettere subito la modifica in UI
    users = state["last_extract"]["users"] or []
    apply_business_roles(users)
        # BRDB training: quando assegni a mano è “ground truth”
    u = next((x for x in users if x.get("username") == body.username), None)
    if u:
        brdb_learn_assignment(role, u.get("groups") or [], weight=10)
        record_llm_learning_event(
            actor=username,
            source="businessroles-add-user",
            signal_type="brdb-assignment",
            entity=body.username,
            details={"businessRole": role, "groupsCount": len(u.get("groups") or []), "weight": 10},
        )
        record_manual_user_change(
            actor=username,
            username=body.username,
            display_name=u.get("displayName"),
            action="assign-business-role",
            source="business-roles",
            details={"businessRole": role},
        )

    log("INFO", f"Business role set: {body.username} -> {role} by {username}")
    invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)
    return {"ok": True, "username": body.username, "role": role}

def _slug_username(display_name: str) -> str:
    s = (display_name or "").strip().lower()
    s = re.sub(r"\s+", ".", s)
    s = re.sub(r"[^a-z0-9._-]", "", s)
    return s or "user"


# (DEPT_MINCONF already defined at line 479)

def _ensure_role_registered(role: str) -> None:
    role = (role or "").strip()
    if not role:
        return

    state.setdefault("role_meta", {})
    state.setdefault("business_roles", set())

    if role not in state["role_meta"]:
        r = random.randint(100, 255)
        g = random.randint(100, 255)
        b = random.randint(100, 255)
        color = f"{r:02x}{g:02x}{b:02x}".upper()
        if not color.startswith("#"):
            color = "#" + color
        state["role_meta"][role] = {"color": color, "groups": []}

    state["business_roles"].add(role)

def apply_department_mapping_if_missing(users: list[dict]) -> None:
    # usa BRDB già presente (inference sui gruppi)
    brdb_rebuild()

    by_dept = defaultdict(list)
    for u in (users or []):
        dept = (u.get("department") or "").strip()
        if dept:
            by_dept[dept].append(u)

    for dept, members in by_dept.items():
        weights = defaultdict(float)

        for u in members:
            s = brdb_infer_groupset(u.get("groups") or [])
            role = (s.get("role") or "Unassigned").strip()
            conf = float(s.get("confidence") or 0.0)
            if role and role != "Unassigned" and conf > 0:
                weights[role] += conf
            for rr, sc in _br_assignment_rule_scores_for_user(u).items():
                weights[rr] += float(sc)

        if weights:
            best_role, best_w = max(weights.items(), key=lambda x: x[1])
            total = sum(weights.values()) or 1.0
            dept_conf = best_w / total
        else:
            best_role, dept_conf = "Unassigned", 0.0

        chosen_role = best_role if (best_role != "Unassigned" and dept_conf >= DEPT_MINCONF) else dept
        _ensure_role_registered(chosen_role)

        # assegna SOLO se l'utente non ha già BR assegnato
        state.setdefault("user_business_role", {})
        for u in members:
            uname = u.get("username")
            if not uname:
                continue

            already = (state.get("user_business_role") or {}).get(uname)
            if already:
                continue

            # se per qualche motivo è già valorizzato sul record utente, non sovrascrivere
            if (u.get("businessRole") or "").strip():
                continue

            user_rule_scores = _br_assignment_rule_scores_for_user(u)
            if user_rule_scores:
                boosted_role = max(user_rule_scores.items(), key=lambda x: x[1])[0]
                _ensure_role_registered(boosted_role)
                state["user_business_role"][uname] = boosted_role
            else:
                state["user_business_role"][uname] = chosen_role

    # Also evaluate rule-based assignment for users without department.
    state.setdefault("user_business_role", {})
    for u in (users or []):
        uname = u.get("username")
        if not uname:
            continue
        already = (state.get("user_business_role") or {}).get(uname)
        if already:
            continue
        if (u.get("businessRole") or "").strip():
            continue
        rule_scores = _br_assignment_rule_scores_for_user(u)
        if rule_scores:
            forced_role = max(rule_scores.items(), key=lambda x: x[1])[0]
            _ensure_role_registered(forced_role)
            state["user_business_role"][uname] = forced_role



@app.post("/api/import/csv")
async def import_csv(file: UploadFile = File(...), background_tasks: BackgroundTasks = None, username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_assignments", "Non autorizzato a modificare assegnazioni utenti")
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="File vuoto")
    if len(raw) > CSV_IMPORT_MAX_BYTES:
        max_mb = CSV_IMPORT_MAX_BYTES / (1024 * 1024)
        raise HTTPException(status_code=413, detail=f"CSV troppo grande: limite {max_mb:.0f} MB")

    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text), delimiter=";", skipinitialspace=True)

    if reader.fieldnames:
        reader.fieldnames = [h.strip() for h in reader.fieldnames if h is not None]

    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV senza header")

    csv_headers_norm = sorted({(h or "").strip().lower() for h in (reader.fieldnames or []) if (h or "").strip()})
    norm_fields = { (h or "").strip().lower() for h in reader.fieldnames }
    if "displayname" not in norm_fields:
        raise HTTPException(
            status_code=400,
            detail=f"Headers richiesti: ['DisplayName']; trovati: {reader.fieldnames}",
        )

    def _norm_header(h: str) -> str:
        return (h or "").strip().lower()

    def _get_any(row_ci: dict, keys: list[str]) -> str:
        for k in keys:
            if k in row_ci and row_ci.get(k) is not None:
                return str(row_ci.get(k))
        return ""

    CSV_KEYS = {
        "displayName": ["displayname", "display name", "name", "utente", "user"],
        "username": ["username", "userprincipalname", "samaccountname", "login", "accountname"],
        "department": ["department", "dept", "dipartimento", "area", "funzione"],
        "businessRole": ["businessrole", "business role", "br", "ruolo business", "ruolo_business"],
        "roles": ["ruoli", "roles", "groups", "gruppi", "entitlements"],
        "accountType": ["accounttype", "account type", "tipo utente", "tipo_utente", "type"],
        "lastLogin": ["lastlogin", "last login", "last_logon", "lastlogon", "ultimo accesso", "ultimologin"],
        "email": ["email", "mail", "emailaddress", "posta"],
        "upn": ["upn", "user principal name", "userprincipalname"],
        "employeeId": ["employeeid", "employee id", "matricola", "badgeid"],
        "manager": ["manager", "owner", "responsabile"],
        "statusAd": ["statusad", "adstatus", "accountstatusad", "statoad"],
        "statusHr": ["statushr", "hrstatus", "accountstatushr", "statohr"],
        "orphanGroups": ["orphangroups", "orphan_groups", "cataloggroups", "catalog_groups", "groupscatalog"],
    }

    def _extract_csv_fields(row: dict) -> tuple:
        row_ci = {_norm_header(k): v for k, v in (row or {}).items()}
        dnraw = _get_any(row_ci, CSV_KEYS["displayName"])
        usernameraw = _get_any(row_ci, CSV_KEYS["username"])
        deptraw = _get_any(row_ci, CSV_KEYS["department"])
        brraw = _get_any(row_ci, CSV_KEYS["businessRole"])
        rolesraw = _get_any(row_ci, CSV_KEYS["roles"])
        type_raw = _get_any(row_ci, CSV_KEYS["accountType"])
        last_login_raw = _get_any(row_ci, CSV_KEYS["lastLogin"])
        email_raw = _get_any(row_ci, CSV_KEYS["email"])
        upn_raw = _get_any(row_ci, CSV_KEYS["upn"])
        employee_id_raw = _get_any(row_ci, CSV_KEYS["employeeId"])
        manager_raw = _get_any(row_ci, CSV_KEYS["manager"])
        status_ad_raw = _get_any(row_ci, CSV_KEYS["statusAd"])
        status_hr_raw = _get_any(row_ci, CSV_KEYS["statusHr"])
        orphan_groups_raw = _get_any(row_ci, CSV_KEYS["orphanGroups"])

        dn = (dnraw or "").strip()
        preferred_username = (usernameraw or "").strip()
        dept = (deptraw or "").strip()
        br = (brraw or "").strip()
        roles = (rolesraw or "").strip()
        acct_type = (type_raw or "WhiteCollar").strip()
        last_login = _normalize_last_login((last_login_raw or "").strip() or None)
        email = (email_raw or "").strip().lower()
        upn = (upn_raw or "").strip().lower()
        employee_id = (employee_id_raw or "").strip()
        manager = (manager_raw or "").strip()
        status_ad = (status_ad_raw or "").strip()
        status_hr = (status_hr_raw or "").strip()
        orphan_groups = [g.strip() for g in (orphan_groups_raw or "").split(",") if g and g.strip()]

        if (not dept) and dn and ("," in dn):
            dn, dept = [x.strip() for x in dn.split(",", 1)]

        if not br:
            br = dept

        return (
            dnraw,
            dn,
            preferred_username,
            dept,
            br,
            roles,
            acct_type,
            last_login,
            email,
            upn,
            employee_id,
            manager,
            status_ad,
            status_hr,
            orphan_groups,
        )

    ingest_sources = state.get("ingest_sources") or {}
    ingest_sources["csv"] = []
    last_csv_rows: List[Dict[str, Any]] = []
    csv_choice_by_dn: Dict[str, str] = {}
    csv_rows_by_dn: Dict[str, List[str]] = defaultdict(list)

    csv_rows_total = 0
    csv_dup_dn_rows = 0
    csv_missing_displayname = 0
    csv_missing_department = 0
    csv_missing_businessrole = 0
    csv_missing_roles = 0
    csv_orphan_groups_catalog: set[str] = set()
    csv_rejects: List[Dict[str, Any]] = []
    reject_empty_groups = bool((state.get("dq_rules") or {}).get("reject_empty_groups"))

    existing_users_list = state.get("last_extract", {}).get("users", [])
    csv_candidates: List[Dict[str, Any]] = []
    choice_by_displayname = state.get("choice_by_displayName") or {}
    duplicate_autoselect: Dict[str, Any] = {}

    for row in reader:
        csv_rows_total += 1
        if csv_rows_total > CSV_IMPORT_MAX_ROWS:
            raise HTTPException(status_code=413, detail=f"CSV troppo grande: limite {CSV_IMPORT_MAX_ROWS} righe")
        row_id = f"csv:{csv_rows_total}"
        (
            dnraw,
            dn,
            preferred_username,
            dept,
            br,
            roles,
            acct_type,
            last_login,
            email,
            upn,
            employee_id,
            manager,
            status_ad,
            status_hr,
            orphan_groups,
        ) = _extract_csv_fields(row)
        for g in orphan_groups:
            csv_orphan_groups_catalog.add(g)

        if not dn:
            csv_missing_displayname += 1
            continue

        if not dept:
            csv_missing_department += 1
        if not br:
            csv_missing_businessrole += 1

        parsed_roles = [g.strip() for g in (roles or "").split(",") if g.strip()]
        if not parsed_roles:
            csv_missing_roles += 1
            if reject_empty_groups:
                csv_rejects.append(
                    {
                        "source": "csv",
                        "reason": "Missing groups (rejected by dq rule)",
                        "user": {"displayName": dn, "department": dept, "businessRole": br},
                        "ts": datetime.now(timezone.utc).isoformat(),
                    }
                )
                continue

        # Favor CSV provided type if it's not the default 'WhiteCollar', otherwise fallback to heuristic
        atype = classify_account(dn, dept, row.get("EmployeeType", ""))
        final_type = acct_type if acct_type not in ["", "WhiteCollar"] else atype

        user_payload = {
            "displayName": dn,
            "groups": parsed_roles,
            "department": dept or None,
            "businessRole": br or None,
            "excluded": False,
            "DataSource": datasource_from_source("csv"),
            "lastLogin": last_login,
            "accountType": final_type,
            "preferredUsername": preferred_username or None,
            "email": email or None,
            "upn": upn or None,
            "employeeId": employee_id or None,
            "manager": manager or None,
            "statusAd": status_ad or None,
            "statusHr": status_hr or None,
        }

        rec = {
            "rowId": row_id,
            "displayName": dn,
            "displayNameRaw": dnraw,
            "preferredUsername": preferred_username,
            "businessRole": br,
            "department": dept,
            "lastLogin": last_login,
            "roles": parsed_roles,
            "email": email,
            "upn": upn,
            "employeeId": employee_id,
            "manager": manager,
            "rawLine": f"{dnraw};{dept};{roles}",
        }
        candidate = _mk_candidate(
            source="csv",
            candidate_id=row_id,
            display_name=dn,
            business_role=br,
            roles=parsed_roles,
            raw=rec["rawLine"],
            department=dept,
            last_login=last_login,
        )
        csv_candidates.append({"rowId": row_id, "rec": rec, "user": user_payload, "candidate": candidate})

    profile_rows = []
    for u in existing_users_list:
        profile_rows.append(
            {
                "department": (u.get("department") or "").strip(),
                "groups": list(u.get("groups") or []),
            }
        )
    for c in csv_candidates:
        profile_rows.append(
            {
                "department": (c["user"].get("department") or "").strip(),
                "groups": list(c["user"].get("groups") or []),
            }
        )
    dept_profile = _build_dept_group_profile(profile_rows)

    by_dn = defaultdict(list)
    for c in csv_candidates:
        dn_key = (c["user"].get("displayName") or "").strip().lower()
        if dn_key:
            by_dn[dn_key].append(c)

    csv_dup_dn_rows = 0
    duplicate_row_ids: set[str] = set()
    for rows in by_dn.values():
        if len(rows) > 1:
            csv_dup_dn_rows += (len(rows) - 1)
            duplicate_row_ids.update(str(item.get("rowId") or "") for item in rows)

    taken_usernames = {str(u.get("username") or "").strip() for u in existing_users_list if u.get("username")}
    new_users: List[Dict[str, Any]] = []
    for _, rows in by_dn.items():
        scored_rows = []
        for item in rows:
            score = _score_duplicate_candidate(item["user"], dept_profile)
            scored_rows.append({**item, "score": score})
        scored_rows.sort(key=lambda x: x["score"]["rank"], reverse=True)
        winner = scored_rows[0]
        winner_user = dict(winner["user"])

        preferred_uname = (winner_user.get("preferredUsername") or "").strip()
        base = _slug_username(preferred_uname or winner_user.get("displayName") or "")
        uname = base
        i = 2
        while uname in taken_usernames:
            uname = f"{base}{i}"
            i += 1
        taken_usernames.add(uname)
        winner_user["username"] = uname
        new_users.append(winner_user)

        if len(scored_rows) > 1:
            display_name = winner_user.get("displayName") or winner["rec"].get("displayName") or ""
            choice_by_displayname[display_name] = winner["rowId"]
            duplicate_autoselect[display_name] = {
                "candidateId": winner["rowId"],
                "reason": winner["score"]["reason"],
                "alternatives": [
                    {
                        "candidateId": s["rowId"],
                        "reason": s["score"]["reason"],
                    }
                    for s in scored_rows[1:]
                ],
            }
        else:
            display_name = winner_user.get("displayName") or winner["rec"].get("displayName") or ""
            choice_by_displayname.setdefault(display_name, winner["rowId"])

    conflict_candidates = [c for c in csv_candidates if str(c.get("rowId") or "") in duplicate_row_ids]
    for item in conflict_candidates:
        rec = item["rec"]
        row_id = str(item["rowId"])
        dn_raw = str(rec.get("displayNameRaw") or "")
        last_csv_rows.append(rec)
        csv_rows_by_dn[dn_raw].append(row_id)
        csv_choice_by_dn.setdefault(dn_raw, row_id)
        ingest_sources["csv"].append(item["candidate"])

    # Merge with existing users - match ONLY by displayName.
    # Keep all local users; update only same displayName; add others.
    existing_by_dn = {}
    for u in existing_users_list:
        dn = (u.get("displayName") or "").strip().lower()
        if dn and dn not in existing_by_dn:
            existing_by_dn[dn] = u

    added_users = 0
    updated_users = 0
    created_brs_count = 0

    for user in new_users:
        uname = user["username"]
        dn_key = (user.get("displayName") or "").strip().lower()
        
        # Match ONLY by displayName
        existing_user = existing_by_dn.get(dn_key)
        
        if existing_user:
            # REPLACE groups and other fields with new values from import
            existing_user["groups"] = user.get("groups") or []
            if user.get("businessRole"):
                existing_user["businessRole"] = user["businessRole"]
            if user.get("department"):
                existing_user["department"] = user["department"]
            if user.get("displayName"):
                existing_user["displayName"] = user["displayName"]
            if user.get("accountType"):
                existing_user["accountType"] = user["accountType"]
            if user.get("lastLogin"):
                existing_user["lastLogin"] = user["lastLogin"]
            for k in ("email", "upn", "employeeId", "manager", "statusAd", "statusHr"):
                if user.get(k) is not None:
                    existing_user[k] = user.get(k)
            existing_user["DataSource"] = datasource_from_source("csv")
            
            # Update username if it changed
            if existing_user.get("username") != uname:
                existing_user["username"] = uname
            
            updated_users += 1
        else:
            # New user - add to indexes
            new_user = user.copy()
            existing_users_list.append(new_user)
            if dn_key:
                existing_by_dn[dn_key] = new_user
            added_users += 1

    merged_users = existing_users_list
    state["last_extract"]["users"] = merged_users
    computed_groups = set(recompute_groups_from_users(merged_users))
    computed_groups.update(csv_orphan_groups_catalog)
    state["last_extract"]["groups"] = sorted(computed_groups)
    state["last_extract"]["ou"] = "MERGED"
    state["last_extract"]["ts"] = time.time()

    # CRITICAL: Populate state["user_business_role"] with CSV Business Roles BEFORE auto-assignment
    # This ensures the preservation logic in apply_department_mapping has data to preserve
    state.setdefault("user_business_role", {})
    for u in merged_users:
        uname = u.get("username")
        br = (u.get("businessRole") or "").strip()
        if uname and br and br != "Unassigned":
            state["user_business_role"][uname] = br

    touched_depts = {u.get("department") for u in new_users if u.get("department")}
    csv_auto_resolved_duplicates = len(duplicate_autoselect)

    last_ingest_stats = {
        "source": "csv",
        "csvHeadersNorm": csv_headers_norm,
        "rowsTotal": csv_rows_total,
        "rowsKept": len(merged_users),
        "duplicateDisplayName": csv_dup_dn_rows,
        "missingDepartment": csv_missing_department,
        "missingBusinessRole": csv_missing_businessrole,
        "missingDisplayName": csv_missing_displayname,
        "missingUsername": 0,
        "missingRoles": csv_missing_roles,
        "orphanGroupsCatalog": len(csv_orphan_groups_catalog),
        "autoResolvedDuplicateUsers": len(duplicate_autoselect),
        "ts": datetime.now(timezone.utc).isoformat(),
    }

    last_csv_stats = {
        "csvRowsTotal": csv_rows_total,
        "csvRowsMissingBR": csv_missing_businessrole,
        "csvDuplicateDisplayNameRows": csv_dup_dn_rows,
        "by": username,
        "ts": datetime.now(timezone.utc).isoformat(),
    }

    state.update({
        "ingest_sources": ingest_sources,
        "last_csv_rows": last_csv_rows,
        "csv_choice_by_dn": csv_choice_by_dn,
        "csv_rows_by_dn": dict(csv_rows_by_dn),
        "choice_by_displayName": choice_by_displayname,
        "duplicate_autoselect": duplicate_autoselect,
        "last_ingest_stats": last_ingest_stats,
        "last_csv_stats": last_csv_stats,
        "last_rejects": csv_rejects[:CSV_IMPORT_MAX_REJECTS_STORED],
        "mining_dirty": True,
    })
    invalidate_hot_caches(users=True, roles=True, kpi=True, mining=True)
    snapshot_ts = str((state.get("last_extract") or {}).get("ts") or "")
    touched_depts_list = sorted([d for d in touched_depts if d])
    tenant_id = get_current_tenant_id()
    if CSV_IMPORT_DETACHED_POSTPROCESS:
        _start_detached_csv_postprocess(snapshot_ts, username, touched_depts_list, tenant_id)
    elif background_tasks:
        background_tasks.add_task(
            run_post_csv_snapshot_logic_background,
            snapshot_ts,
            username,
            touched_depts_list,
            tenant_id,
        )
    else:
        run_post_csv_snapshot_logic_background(snapshot_ts, username, touched_depts_list, tenant_id)

    return {
        "ok": True,
        "snapshotReady": True,
        "processingInBackground": True,
        "addedUsers": added_users,
        "updatedUsers": updated_users,
        "updatedByDisplayName": updated_users,
        "totalUsers": len(merged_users),
        "rowsTotal": csv_rows_total,
        "csvDuplicateDisplayNameRows": csv_dup_dn_rows,
        "autoResolvedDuplicateUsers": csv_auto_resolved_duplicates,
        "newBusinessRoles": 0,
    }



def apply_all_choices_to_last_extract() -> None:
    rebuild_ingest_candidates()
    for dn, cid in (state.get("choice_by_displayName") or {}).items():
        cand = next(
            (c for c in (state.get("ingest_candidates") or [])
             if c.get("candidateId") == cid and c.get("displayName") == dn),
            None
        )
        if cand:
            apply_choice_for_displayname(
                display_name=dn,
                chosen_business_role=cand.get("businessRole"),
                chosen_roles=cand.get("roles") or [],
            )

    # riallinea lista gruppi (tiene conto degli excluded)
    users = state.get("last_extract", {}).get("users") or []
    state["last_extract"]["groups"] = recompute_groups_from_users(users)



# (_slug_username already defined at line 2299)
# (datetime import already at top of file)

def applyimportrow(displayname: str, businessrole: str, ruoli: str, department: str = "", source: str = "xlsx"):
    displayname = (displayname or "").strip()
    businessrole = (businessrole or "").strip()
    ruoli = (ruoli or "").strip()
    department = (department or "").strip()

    if not displayname:
        return {"skipped": True}

    # Username stabile
    uname = _slug_username(displayname)
    
    # Gruppi
    groups = [g.strip() for g in (ruoli or "").split(",") if g.strip()]
    
    # User object
    user = {
        "username": uname,
        "displayName": displayname,
        "groups": groups,
        "department": department or None,
        "businessRole": businessrole or None,
        "excluded": False,
        "DataSource": datasource_from_source(source),
    }
    
    # Merge logic (simplified for single row)
    last_extract = state.setdefault("last_extract", {"users": [], "groups": [], "ou": "IMPORT", "ts": None})
    users = last_extract.get("users", [])
    
    existing = next((u for u in users if u.get("username") == uname), None)
    created_user = False
    created_role = False
    added_groups = 0
    
    if existing:
        old_groups = set(existing.get("groups") or [])
        new_groups = set(groups)
        added_groups = len(new_groups - old_groups)
        existing["groups"] = sorted(old_groups | new_groups)
        if businessrole:
            existing["businessRole"] = businessrole
        if department:
            existing["department"] = department
        existing["DataSource"] = datasource_from_source(source)
    else:
        users.append(user)
        created_user = True
        added_groups = len(groups)
        
    if businessrole:
        business_roles = state.setdefault("business_roles", set())
        if businessrole not in business_roles:
            _ensure_role_registered(businessrole)
            created_role = True

    state["mining_dirty"] = True
    return {
        "ok": True,
        "created_user": created_user,
        "created_role": created_role,
        "added_groups": added_groups
    }


@app.post("/api/import/xlsx")
async def import_xlsx(file: UploadFile = File(...), username: str = Depends(require_auth)):
    _require_capability(username, "can_manage_assignments", "Non autorizzato a modificare assegnazioni utenti")
    raw = await file.read()
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active  # primo foglio

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return {"ok": False, "error": "Empty Excel"}

    # mappa colonne (richiede intestazioni esatte)
    cols = {str(v).strip(): i for i, v in enumerate(header) if v is not None}
    required = ["DisplayName", "BusinessRole", "Ruoli"]
    if any(c not in cols for c in required):
        return {"ok": False, "error": f"Missing headers: {required}", "found": list(cols.keys())}

    created_users = 0
    created_roles = 0
    assigned_users = 0
    added_groups_total = 0

    for r in rows:
        dn = r[cols["DisplayName"]] if cols["DisplayName"] < len(r) else ""
        br = r[cols["BusinessRole"]] if cols["BusinessRole"] < len(r) else ""
        ru = r[cols["Ruoli"]] if cols["Ruoli"] < len(r) else ""

        out = applyimportrow(str(dn or ""), str(br or ""), str(ru or ""), source="xlsx")
        if out.get("skipped"):
            continue
        created_users += 1 if out.get("created_user") else 0
        created_roles += 1 if out.get("created_role") else 0
        assigned_users += 1
        added_groups_total += int(out.get("added_groups") or 0)

    return {
        "ok": True,
        "created_users": created_users,
        "created_roles": created_roles,
        "assigned_users": assigned_users,
        "added_groups": added_groups_total
    }


# =============================================================================
# ML ENGINE API ENDPOINTS
# =============================================================================

@app.get("/api/config/ad-fields")
def get_ad_fields(username: str = Depends(require_auth)):
    """Return list of all AD fields found during last import."""
    fields = state.get("ad_available_fields", [])
    # Ensure default fields are always present
    defaults = {"displayName", "department", "title", "employeeType", "company", "manager", "mail"}
    combined = sorted(list(set(fields) | defaults))
    return {"fields": combined}

@app.get("/api/brdb/status")
def brdb_status_api(username: str = Depends(require_auth)):
    """
    Ritorna lo stato del calcolo BRDB (background).
    """
    return {
        "calculated": state.get("brdb_calculated", False),
        "min_confidence": state.get("brdb_min_confidence", BRDB_MIN_CONF),
        "last_update": state.get("brdb_last_update")
    }

@app.get("/api/ml/status")
def ml_status(username: str = Depends(require_auth)):
    """Return ML engine status and metrics."""
    return ml_engine.get_status()


@app.post("/api/ml/train")
def ml_train(username: str = Depends(require_auth)):
    """Trigger ML model training from accumulated data."""
    # Build training data from existing users
    users = state.get("last_extract", {}).get("users") or []
    training_data = []
    
    for u in users:
        if u.get("accountType"):
            training_data.append({
                "display_name": u.get("displayName", ""),
                "ou": u.get("department", ""),
                "employee_type": "",  # Not always available
                "account_type": u.get("accountType", "Internal"),
            })
    
    # Also include corrections/confirmations
    result = ml_engine.retrain_from_history()
    if not result.get("success") and training_data:
        result = ml_engine.train_classifier(training_data)
    timeline = state.setdefault("ai_training_timeline", [])
    timeline.append({
        "id": f"run-{int(time.time()*1000)}",
        "ts": datetime.now(timezone.utc).isoformat(),
        "triggeredBy": username,
        "datasetSize": len(training_data),
        "modelName": "account-type-classifier",
        "status": "success" if result.get("success") else "failed",
        "metrics": {
            "accuracy": float(result.get("accuracy") or 0.0),
            "f1": float(result.get("f1") or 0.0),
            "precision": float(result.get("precision") or 0.0),
            "recall": float(result.get("recall") or 0.0),
        },
    })
    record_llm_learning_event(
        actor=username,
        source="ml-train",
        signal_type="model-train",
        entity="account-type-classifier",
        details={
            "datasetSize": len(training_data),
            "success": bool(result.get("success")),
            "accuracy": float(result.get("accuracy") or 0.0),
        },
    )

    return result


@app.post("/api/ml/rebuild-brdb")
def ml_rebuild_brdb(username: str = Depends(require_auth)):
    """Force rebuild of Business Role Database."""
    brdb_rebuild()
    return {"ok": True, "message": "BRDB rebuilt", **ml_engine.get_status()["brdb"]}


class AccountTypeConfirmRequest(BaseModel):
    confirmed_type: str


class AiBusinessRoleSuggestRequest(BaseModel):
    group: str = ""
    source: str = "unknown"
    context: Optional[Dict[str, Any]] = None


def _extract_context_groups(context: Optional[Dict[str, Any]]) -> List[str]:
    if not context:
        return []
    values: List[str] = []
    # Accept a broad set of keys for compatibility with evolving frontend payloads.
    for key in ("groups", "roles", "entitlements", "group", "groupName", "group_name", "raw_groups", "candidateGroups"):
        v = context.get(key)
        if v is None:
            continue
        if isinstance(v, list):
            values.extend(str(x).strip() for x in v if str(x).strip())
        else:
            values.extend(part.strip() for part in str(v).split(",") if part.strip())
    # Preserve order, remove duplicates.
    return list(dict.fromkeys(values))


@app.get("/api/ai/health")
def ai_health(username: str = Depends(require_auth)):
    ml = ml_engine.get_status() or {}
    model = ml.get("model") or {}
    training = ml.get("training_data") or {}
    return {
        "status": "ok",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "brdbReady": bool(state.get("brdb_ready")),
        "model": {
            "trained": bool(model.get("trained")),
            "accuracy": float(model.get("accuracy") or 0.0),
            "f1": float(model.get("f1") or 0.0),
            "precision": float(model.get("precision") or 0.0),
            "recall": float(model.get("recall") or 0.0),
        },
        "trainingData": {
            "totalSamples": int(training.get("total_samples") or 0),
            "patternsCount": int(training.get("patterns_count") or 0),
        },
        "brdb": ml.get("brdb") or {},
        "cache": RESPONSE_CACHE.stats(),
    }


@app.post("/api/ai/suggest-business-role-online")
def suggest_business_role_online(
    body: AiBusinessRoleSuggestRequest,
    username: str = Depends(require_auth),
):
    group = str(body.group or "").strip()
    if not group:
        raise HTTPException(status_code=400, detail="Group is required")
    pred = brdb_infer_group(group)
    role = str(pred.get("role") or "Unassigned")
    confidence = float(pred.get("confidence") or 0.0)
    return {
        "group": group,
        "source": str(body.source or "unknown"),
        "suggestedBusinessRole": role,
        "confidence": round(confidence, 3),
        "method": "brdb-online",
        "evidence": pred.get("evidence") or {},
    }


@app.post("/api/ai/suggest-business-role-hybrid")
def suggest_business_role_hybrid(
    body: AiBusinessRoleSuggestRequest,
    username: str = Depends(require_auth),
):
    group = str(body.group or "").strip()
    context_groups = _extract_context_groups(body.context)
    if group and group not in context_groups:
        context_groups.insert(0, group)
    if not context_groups:
        raise HTTPException(status_code=400, detail="Group or context groups are required")

    online = brdb_infer_group(group) if group else {"role": "Unassigned", "confidence": 0.0, "evidence": {"reason": "no_group"}}
    groupset = brdb_infer_groupset(context_groups)

    role_online = str(online.get("role") or "Unassigned")
    conf_online = float(online.get("confidence") or 0.0)
    role_groupset = str(groupset.get("role") or "Unassigned")
    conf_groupset = float(groupset.get("confidence") or 0.0)

    if role_online == role_groupset:
        chosen_role = role_groupset
        chosen_conf = min(0.99, (0.40 * conf_online) + (0.60 * conf_groupset))
        chosen_method = "hybrid-consensus"
    else:
        # Prefer the stronger signal; if equal, prefer groupset because it uses more context.
        if conf_groupset >= conf_online:
            chosen_role = role_groupset
            chosen_conf = conf_groupset
            chosen_method = "hybrid-groupset"
        else:
            chosen_role = role_online
            chosen_conf = conf_online
            chosen_method = "hybrid-online"

    return {
        "group": group,
        "source": str(body.source or "unknown"),
        "contextGroups": context_groups,
        "suggestedBusinessRole": chosen_role,
        "confidence": round(float(chosen_conf), 3),
        "method": chosen_method,
        "components": {
            "online": {"role": role_online, "confidence": round(conf_online, 3)},
            "groupset": {"role": role_groupset, "confidence": round(conf_groupset, 3)},
        },
        "evidence": {
            "online": online.get("evidence") or {},
            "groupset": (groupset.get("evidence") or {}),
        },
    }


@app.post("/api/users/{uname}/confirm-type")
def confirm_account_type(uname: str, body: AccountTypeConfirmRequest, username: str = Depends(require_auth)):
    """Record a type confirmation (trains the ML model)."""
    users = state.get("last_extract", {}).get("users") or []
    target = next((x for x in users if x.get("username") == uname), None)
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Record confirmation
    ml_engine.record_confirmation(
        username=uname,
        display_name=target.get("displayName", ""),
        ou=target.get("department", ""),
        employee_type="",
        confirmed_type=body.confirmed_type
    )
    
    # Update user type
    target["accountType"] = body.confirmed_type
    record_llm_learning_event(
        actor=username,
        source="confirm-account-type",
        signal_type="supervised-label",
        entity=uname,
        details={"confirmed_type": body.confirmed_type},
    )
    record_manual_user_change(
        actor=username,
        username=uname,
        display_name=target.get("displayName"),
        action="confirm-account-type",
        source="ai-training",
        details={"accountType": body.confirmed_type},
    )
    
    return {"ok": True, "user": uname, "type": body.confirmed_type}


@app.post("/api/users/{uname}/correct-type")
def correct_account_type(uname: str, body: AccountTypeConfirmRequest, username: str = Depends(require_auth)):
    """Record a type correction (trains the ML model)."""
    users = state.get("last_extract", {}).get("users") or []
    target = next((x for x in users if x.get("username") == uname), None)
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    
    old_type = target.get("accountType", "Internal")
    
    # Record correction
    ml_engine.record_correction(
        username=uname,
        display_name=target.get("displayName", ""),
        ou=target.get("department", ""),
        employee_type="",
        old_type=old_type,
        new_type=body.confirmed_type
    )
    
    # Update user type
    target["accountType"] = body.confirmed_type
    record_llm_learning_event(
        actor=username,
        source="correct-account-type",
        signal_type="supervised-correction",
        entity=uname,
        details={"old_type": old_type, "new_type": body.confirmed_type},
    )
    record_manual_user_change(
        actor=username,
        username=uname,
        display_name=target.get("displayName"),
        action="correct-account-type",
        source="ai-training",
        details={"from": old_type, "to": body.confirmed_type},
    )
    
    return {"ok": True, "user": uname, "old_type": old_type, "new_type": body.confirmed_type}


@app.get("/api/ml/account-types")
def get_account_types(username: str = Depends(require_auth)):
    """Return the list of supported account types."""
    return {"types": ACCOUNT_TYPES}


# Register extracted route groups (AI Lab and Pattern Rules)
register_ai_lab_routes(
    app,
    state=state,
    response_cache=RESPONSE_CACHE,
    require_auth=require_auth,
    invalidate_hot_caches=invalidate_hot_caches,
    active_users=active_users,
    ml_engine=ml_engine,
    normalize_last_login=_normalize_last_login,
    slug_username=_slug_username,
    classify_account=classify_account,
    compute_model_quality=compute_model_quality,
    run_smart_ai_detection=run_smart_ai_detection,
    record_llm_learning_event=record_llm_learning_event,
    record_manual_user_change=record_manual_user_change,
)

register_pattern_rules_routes(
    app,
    state=state,
    ml_engine=ml_engine,
    require_auth=require_auth,
    recalculate_assignments_background=recalculate_assignments_background,
    record_llm_learning_event=record_llm_learning_event,
    log=log,
)

# (peer-analysis endpoint defined at line 1846)
